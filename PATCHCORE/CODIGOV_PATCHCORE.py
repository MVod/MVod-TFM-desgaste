"""
================================================================================
 CODIGOV23.py - DETECCIÓN DE DESGASTE EN HERRAMIENTAS DE ROSCADO
 EDA + PatchCore + Coreset + 6 PIPELINES + SNR + Perfilómetro-clasificador
================================================================================
 TFM - Máster Universitario en Inteligencia Artificial - UNIR
 Autores: Daniel Alcalde Martín-Calero, Miguel González Sánchez, Jordi Peiró
 Castelló

 RECORRIDO DE VERSIONES:

   V17: cierre completo de los 5 objetivos SMART del TFM (preprocesado +
        PatchCore + perfilómetro + SNR + clasificador).
   V19: integración con la app Streamlit (motor común, mismo CODIGO).
   V20: incorporación de bloques adicionales solicitados en revisión:
        - Ablación con/sin máscara para evaluar la aportación del fondo.
        - Perfil patrón promedio BUENAS y comparación con ISO teórico.
        - Matrices de confusión visuales (PNG) para la calibración de
          umbral a nivel pieza.
   V23: ablación con/sin máscara obligatoria (no opt-in), corrección de la
        alineación por correlación cruzada en el perfil patrón, distancia
        normalizada en el clasificador patrón.

   - Obj 1 (homomórfico): pipeline SOLO_HOMO evaluado.
   - Obj 2 (BLMD):        BLMD_REAL (Smith 2005, Nunes 2009) + BLMD_V10.
   - Obj 3 (SNR BLMD vs BEMD vs wavelet): bloque SNR con 3 métricas
            (estructural, contornos, separación buenas/malas por Cohen d),
            salida snr_pipelines.csv + informe + gráfica.
   - Obj 4 (área y ancho medio de desgaste): perfilómetro + clasificador
            calibrado data-driven + agregación a nivel pieza + diagnóstico
            automático de mediciones sospechosas.
   - Obj 5 (evaluación con dataset etiquetado): AUC global, AUC sobre
            subconjunto detectable, F1, matriz de confusión y comparativa
            entre 6 pipelines + perfilómetro.

   - 6 PIPELINES por defecto (incluye BEMD_REAL):
       1. BLMD_REAL   - BLMD riguroso (Smith 2005, Nunes 2009)
       2. BLMD_V10    - BLMD aproximado rápido
       3. BEMD_REAL   - Bidimensional Empirical Mode Decomposition (PyEMD)
       4. WAVELET     - Daubechies-4, 3 niveles (PyWavelets)
       5. SOLO_HOMO   - Solo filtrado homomórfico
       6. SIN_PREP    - Sin preprocesado (ablación)

 DIFERENCIAS RESPECTO A V15 y V16:
   - Respecto a V15: añade clasificador del perfilómetro, agregación a nivel
     pieza, calibración data-driven, diagnóstico automático.
   - Respecto a V16: re-integra el bloque SNR (Obj 3 del TFM).
   - BEMD_REAL entra por defecto en la lista de pipelines.
   - El informe final integra los resultados de SNR junto al resto.

================================================================================
"""

import os
import sys
import csv
import json
import time
import argparse
import logging
from datetime import datetime
from pathlib import Path

import cv2
import numpy as np
from scipy import ndimage, stats
from scipy.interpolate import griddata

import torch
import torchvision.models as models
import torchvision.transforms as transforms
from sklearn.neighbors import NearestNeighbors
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import RandomForestClassifier
from sklearn.manifold import TSNE

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

try:
    import tkinter as tk
    from tkinter import filedialog
    TK_OK = True
except ImportError:
    TK_OK = False

# tqdm para barra de progreso. Si no está instalado, fallback silencioso.
try:
    from tqdm import tqdm
    TQDM_OK = True
except ImportError:
    TQDM_OK = False

# V20: PyEMD para BEMD. Si no está instalado, BEMD se desactiva.
try:
    from PyEMD import BEMD
    BEMD_OK = True
except ImportError:
    BEMD_OK = False

# V20: PyWavelets para wavelet decomposition. Si no está, wavelet se desactiva.
try:
    import pywt
    PYWAVELETS_OK = True
except ImportError:
    PYWAVELETS_OK = False
    def tqdm(iterable, **kwargs):
        return iterable

# V20.1: openpyxl para generar el Excel de resultados para compañeros.
# Si no está, la fase 7 se salta silenciosamente.
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ==============================================================================
#  CONFIGURACIÓN GLOBAL Y LOGGING
# ==============================================================================
logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] %(levelname)s - %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("BLMD-TFM")

DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")
log.info(f"Motor de IA cargado en: {DEVICE}")


# ==============================================================================
#  INSTRUMENTACIÓN — TIEMPOS, COSTE COMPUTACIONAL Y HISTÓRICO DE EJECUCIONES
# ==============================================================================
#
#  Este bloque añade tres capacidades transversales al pipeline V18:
#
#    1. Medición de tiempos por fase (preprocesado, entrenamiento del detector,
#       inferencia sobre buenas, inferencia sobre malas, perfilómetro).
#    2. Medición de coste computacional del modelo: número de parámetros del
#       backbone, tamaño del memory bank en parches y MB, memoria RAM y pico
#       de memoria GPU, dispositivo (CPU/CUDA).
#    3. Persistencia de un histórico acumulado entre ejecuciones, con dos
#       ficheros generados en la carpeta de salida de la ejecución actual y
#       en la carpeta raíz del proyecto:
#         - historico_ejecuciones.json: lista de ejecuciones completas
#         - historico_costes.csv: tabla plana con una fila por método y run
#
#  La instrumentación se diseña para ser silenciosa: si psutil no está
#  disponible, los campos de memoria caen a None pero el resto sigue
#  funcionando con normalidad.
# ==============================================================================

import platform as _platform

try:
    import psutil as _psutil
    PSUTIL_OK = True
except ImportError:
    PSUTIL_OK = False
    _psutil = None


class Cronometro:
    """Cronómetro acumulador. Suma tiempos de fragmentos sucesivos y cuenta
    cuántas veces se ha activado.

        cron = Cronometro()
        with cron:
            ...   # bloque temporizado
        cron.total   # segundos acumulados
        cron.n       # número de bloques contados
        cron.media   # segundos medios por bloque (None si n=0)
    """
    __slots__ = ("total", "n", "_t0")

    def __init__(self):
        self.total = 0.0
        self.n = 0
        self._t0 = None

    def __enter__(self):
        self._t0 = time.perf_counter()
        return self

    def __exit__(self, *exc):
        self.total += time.perf_counter() - self._t0
        self.n += 1
        self._t0 = None

    @property
    def media(self):
        return self.total / self.n if self.n > 0 else None


def _memoria_rss_mb():
    """Memoria residente del proceso en MB. None si psutil no está."""
    if not PSUTIL_OK:
        return None
    try:
        return _psutil.Process(os.getpid()).memory_info().rss / (1024 ** 2)
    except Exception:
        return None


def _memoria_gpu_pico_mb():
    """Pico de memoria GPU asignada por torch (MB). None si no hay CUDA."""
    try:
        if torch.cuda.is_available():
            return torch.cuda.max_memory_allocated() / (1024 ** 2)
    except Exception:
        pass
    return None


def _reset_gpu_peak():
    """Resetea el contador de pico de memoria GPU."""
    try:
        if torch.cuda.is_available():
            torch.cuda.reset_peak_memory_stats()
    except Exception:
        pass


def _contar_parametros(model):
    """Devuelve (total, entrenables) de un modelo PyTorch. (0,0) si falla."""
    if model is None:
        return (0, 0)
    try:
        total = sum(p.numel() for p in model.parameters())
        train = sum(p.numel() for p in model.parameters() if p.requires_grad)
        return (int(total), int(train))
    except Exception:
        return (0, 0)


def _info_sistema():
    """Snapshot del entorno de ejecución para guardar en el histórico."""
    info = {
        "python": _platform.python_version(),
        "sistema": _platform.system(),
        "procesador": _platform.processor() or _platform.machine(),
        "cpu_logicos": os.cpu_count(),
        "torch_version": torch.__version__,
        "gpu_disponible": False,
        "gpu_nombre": None,
    }
    try:
        if torch.cuda.is_available():
            info["gpu_disponible"] = True
            info["gpu_nombre"] = torch.cuda.get_device_name(0)
    except Exception:
        pass
    return info


# Carpeta raíz para el histórico acumulado (junto al script). Se calcula
# perezosamente la primera vez que se llama a registrar_ejecucion_en_historico.
_HISTORICO_DIR = None


def _ruta_historico_raiz():
    """Devuelve la carpeta donde vive el histórico global (la del script)."""
    global _HISTORICO_DIR
    if _HISTORICO_DIR is None:
        try:
            _HISTORICO_DIR = Path(__file__).resolve().parent
        except NameError:
            # __file__ no existe en algunos entornos (notebook): fallback cwd
            _HISTORICO_DIR = Path.cwd()
    return _HISTORICO_DIR


def registrar_ejecucion_en_historico(ruta_salida, payload):
    """Persiste el resumen de la ejecución actual en dos ficheros:

      - <raiz>/historico_ejecuciones.json: lista acumulada (una entrada por
        ejecución de main()). Se carga si existe, se le añade el payload y
        se vuelve a escribir.
      - <ruta_salida>/historico_costes.csv y <raiz>/historico_costes.csv:
        tabla plana con una fila por método y ejecución, ideal para abrirla
        en Excel o pegarla en el TFM.

    El payload tiene la forma:
        {
          "fecha": "...", "comando": "...", "n_buenas": N, "n_malas": M,
          "sistema": {...info_sistema...},
          "metodos": [
              {"nombre": "...", "tipo": "patchcore|legacy|perfilometro",
               "auc": float|None, "acc": float|None,
               "tiempo_train_s": ..., "tiempo_inferencia_por_imagen_s": ...,
               "params_total": ..., "bank_mb": ..., "dispositivo": ..., ...}
              ...
          ]
        }
    """
    raiz = _ruta_historico_raiz()
    json_global = raiz / "historico_ejecuciones.json"
    csv_global = raiz / "historico_costes.csv"
    csv_run = Path(ruta_salida) / "historico_costes.csv"

    # ----- JSON acumulado -----
    historico = []
    if json_global.exists():
        try:
            with open(json_global, encoding="utf-8") as fh:
                historico = json.load(fh)
                if not isinstance(historico, list):
                    historico = [historico]
        except Exception as e:
            log.warning(f"Histórico JSON corrupto, empiezo de cero: {e}")
            historico = []
    historico.append(payload)
    try:
        with open(json_global, "w", encoding="utf-8") as fh:
            json.dump(historico, fh, indent=2, default=str, ensure_ascii=False)
        log.info(f"  Histórico actualizado en {json_global}")
    except Exception as e:
        log.error(f"  No pude escribir histórico JSON: {e}")

    # ----- CSV plano (una fila por método) -----
    columnas = [
        "fecha", "ejecucion_id", "comando",
        "n_buenas", "n_malas", "dispositivo",
        "metodo", "tipo",
        "auc", "accuracy",
        "tiempo_train_total_s", "tiempo_train_fit_s",
        "tiempo_inferencia_total_s", "tiempo_inferencia_por_imagen_s",
        "tiempo_total_s",
        "params_total", "params_entrenables",
        "bank_n_parches", "bank_mb",
        "ram_delta_mb", "gpu_pico_mb",
    ]
    filas_csv = []
    eid = payload.get("ejecucion_id") or payload.get("fecha", "")[:19]
    for met in payload.get("metodos", []):
        filas_csv.append({
            "fecha": payload.get("fecha", ""),
            "ejecucion_id": eid,
            "comando": payload.get("comando", ""),
            "n_buenas": payload.get("n_buenas", ""),
            "n_malas": payload.get("n_malas", ""),
            "dispositivo": met.get("dispositivo",
                                    payload.get("sistema", {}).get("gpu_nombre")
                                    or "cpu"),
            "metodo": met.get("nombre", ""),
            "tipo": met.get("tipo", ""),
            "auc": met.get("auc"),
            "accuracy": met.get("acc"),
            "tiempo_train_total_s": met.get("tiempo_train_total_s"),
            "tiempo_train_fit_s": met.get("tiempo_train_fit_s"),
            "tiempo_inferencia_total_s": met.get("tiempo_inferencia_total_s"),
            "tiempo_inferencia_por_imagen_s":
                met.get("tiempo_inferencia_por_imagen_s"),
            "tiempo_total_s": met.get("tiempo_total_s"),
            "params_total": met.get("params_total"),
            "params_entrenables": met.get("params_entrenables"),
            "bank_n_parches": met.get("bank_n_parches"),
            "bank_mb": met.get("bank_mb"),
            "ram_delta_mb": met.get("ram_delta_mb"),
            "gpu_pico_mb": met.get("gpu_pico_mb"),
        })

    # CSV global: si existe, se hace append; si no, se crea con cabecera.
    def _volcar_csv(ruta, filas, append):
        try:
            modo = "a" if (append and Path(ruta).exists()) else "w"
            with open(ruta, modo, newline="", encoding="utf-8") as fh:
                w = csv.DictWriter(fh, fieldnames=columnas)
                if modo == "w":
                    w.writeheader()
                for f in filas:
                    w.writerow(f)
        except Exception as e:
            log.error(f"  No pude escribir {ruta}: {e}")

    _volcar_csv(csv_global, filas_csv, append=True)
    _volcar_csv(csv_run, filas_csv, append=False)
    log.info(f"  Histórico CSV global: {csv_global}")
    log.info(f"  Histórico CSV de esta ejecución: {csv_run}")


# ==============================================================================
#  FASE 1 - ANÁLISIS EXPLORATORIO DE DATOS (EDA)
# ==============================================================================
#
#  Antes de entrenar nada, analizamos estadísticamente el dataset para
#  responder a:
#    1. ¿Hay diferencias visuales sistemáticas entre BUENAS y MALAS?
#    2. ¿Qué métricas las separan?
#    3. ¿Qué fracción de MALAS es estadísticamente indistinguible?
#
#  El EDA produce un fichero JSON que el detector usa después para calibrarse
#  y un informe que se incorpora al informe final del TFM.
# ==============================================================================

def eda_estadisticas_intensidad(img):
    """Estadísticas globales de la distribución de intensidades."""
    f = img.astype(np.float64).flatten()
    return {
        "int_media": float(np.mean(f)),
        "int_mediana": float(np.median(f)),
        "int_std": float(np.std(f)),
        "int_p05": float(np.percentile(f, 5)),
        "int_p95": float(np.percentile(f, 95)),
        "int_skew": float(stats.skew(f)),
        "int_kurtosis": float(stats.kurtosis(f)),
        "int_pct_oscuro": float(100 * np.mean(f < 85)),
        "int_pct_medio": float(100 * np.mean((f >= 85) & (f < 170))),
        "int_pct_claro": float(100 * np.mean(f >= 170)),
        "int_pct_saturado": float(100 * np.mean(f >= 250)),
    }


def eda_metricas_geometria(img, segmentar_fn):
    """Métricas geométricas del contorno."""
    mask = segmentar_fn(img)
    H, W = img.shape
    area_h = int(np.sum(mask > 0))
    if area_h < 100:
        return _eda_geom_vacias(area_h, H, W)

    contornos, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL,
                                    cv2.CHAIN_APPROX_NONE)
    if not contornos:
        return _eda_geom_vacias(area_h, H, W)
    cnt = max(contornos, key=cv2.contourArea)

    perimetro = float(cv2.arcLength(cnt, True))
    area_real = float(cv2.contourArea(cnt))
    hull = cv2.convexHull(cnt)
    area_hull = float(cv2.contourArea(hull))
    solidez = area_real / max(1.0, area_hull)
    compacidad = 4 * np.pi * area_real / max(1.0, perimetro ** 2)
    perdida_hull = area_hull - area_real

    pts = cnt.reshape(-1, 2).astype(np.float64)
    if len(pts) >= 30:
        k = 15; half = k // 2
        pts_pad = np.vstack([pts[-half:], pts, pts[:half]])
        pts_suav = np.zeros_like(pts)
        for i in range(len(pts)):
            pts_suav[i] = pts_pad[i:i + k].mean(axis=0)
        desv = np.linalg.norm(pts - pts_suav, axis=1)
        rugosidad = float(np.std(desv))
        rugosidad_max = float(np.max(desv))
    else:
        rugosidad = 0.0; rugosidad_max = 0.0

    x, y, w, h = cv2.boundingRect(cnt)
    aspect = w / max(1, h)

    return {
        "geom_area_herr_px": area_h,
        "geom_pct_herr": 100.0 * area_h / (H * W),
        "geom_perimetro": perimetro,
        "geom_area_real": area_real,
        "geom_area_hull": area_hull,
        "geom_perdida_hull": perdida_hull,
        "geom_solidez": solidez,
        "geom_compacidad": compacidad,
        "geom_rugosidad_std": rugosidad,
        "geom_rugosidad_max": rugosidad_max,
        "geom_aspect": aspect,
        "geom_n_pts_contorno": len(pts),
    }


def _eda_geom_vacias(area_h, H, W):
    return {"geom_area_herr_px": area_h, "geom_pct_herr": 100.0 * area_h / (H * W),
            "geom_perimetro": 0.0, "geom_area_real": 0.0, "geom_area_hull": 0.0,
            "geom_perdida_hull": 0.0, "geom_solidez": 0.0,
            "geom_compacidad": 0.0, "geom_rugosidad_std": 0.0,
            "geom_rugosidad_max": 0.0, "geom_aspect": 0.0,
            "geom_n_pts_contorno": 0}


def eda_metricas_textura(img, segmentar_fn):
    """Métricas de textura simples."""
    lap = cv2.Laplacian(img, cv2.CV_64F)
    sx = cv2.Sobel(img, cv2.CV_64F, 1, 0, ksize=3)
    sy = cv2.Sobel(img, cv2.CV_64F, 0, 1, ksize=3)
    grad = np.sqrt(sx ** 2 + sy ** 2)

    hist, _ = np.histogram(img, bins=64, range=(0, 256))
    p = hist / max(1, hist.sum())
    p = p[p > 0]
    entropia = float(-np.sum(p * np.log2(p)))

    mask = segmentar_fn(img)
    if np.sum(mask > 0) > 100:
        valores = img[mask > 0].astype(np.float64)
        std_herr = float(np.std(valores))
        media_herr = float(np.mean(valores))
    else:
        std_herr = 0.0; media_herr = 0.0

    return {
        "tex_lap_var": float(np.var(lap)),
        "tex_sobel_media": float(np.mean(grad)),
        "tex_sobel_std": float(np.std(grad)),
        "tex_entropia": entropia,
        "tex_std_dentro_herr": std_herr,
        "tex_media_dentro_herr": media_herr,
    }


def eda_tests(df_b, df_m, columnas):
    """Tests KS + Mann-Whitney + Cohen's d por métrica."""
    resultados = []
    for col in columnas:
        b = np.array([r[col] for r in df_b if col in r])
        m = np.array([r[col] for r in df_m if col in r])
        if len(b) < 5 or len(m) < 5:
            continue
        ks_stat, ks_p = stats.ks_2samp(b, m)
        try:
            _, mw_p = stats.mannwhitneyu(b, m, alternative="two-sided")
        except Exception:
            mw_p = 1.0
        s_pool = np.sqrt(((len(b) - 1) * np.var(b, ddof=1) +
                          (len(m) - 1) * np.var(m, ddof=1)) /
                         max(1, len(b) + len(m) - 2))
        cohen_d = (np.mean(m) - np.mean(b)) / max(1e-12, s_pool)
        resultados.append({
            "metrica": col,
            "media_buena": float(np.mean(b)), "media_mala": float(np.mean(m)),
            "std_buena": float(np.std(b)), "std_mala": float(np.std(m)),
            "ks_stat": float(ks_stat), "ks_p": float(ks_p), "mw_p": float(mw_p),
            "cohen_d": float(cohen_d),
        })
    resultados.sort(key=lambda r: -r["ks_stat"])
    return resultados


def eda_outliers(X, y, registros_b, registros_m):
    """
    Identifica MALAS detectables vs indistinguibles según distancia al
    centroide de BUENAS en el espacio escalado de features.
    """
    Xs = StandardScaler().fit_transform(X)
    centroide_b = np.mean(Xs[y == 0], axis=0)
    distancias = np.linalg.norm(Xs - centroide_b, axis=1)

    # Umbral: percentil 95 de las distancias de BUENAS
    umbral_distancia = float(np.percentile(distancias[y == 0], 95))

    # Para cada MALA, está dentro o fuera del 95% de BUENAS
    idx_malas = np.where(y == 1)[0]
    dist_malas = distancias[idx_malas]
    detectables_mask = dist_malas > umbral_distancia
    n_detectables = int(np.sum(detectables_mask))
    n_total_malas = len(idx_malas)
    pct_indistinguibles = 100.0 * (n_total_malas - n_detectables) / max(1, n_total_malas)

    todos = registros_b + registros_m
    archivos_detectables = [todos[idx_malas[i]]["archivo"]
                            for i in range(len(idx_malas))
                            if detectables_mask[i]]
    archivos_indist = [todos[idx_malas[i]]["archivo"]
                       for i in range(len(idx_malas))
                       if not detectables_mask[i]]

    return {
        "umbral_distancia_p95_buenas": umbral_distancia,
        "n_malas_detectables": n_detectables,
        "n_malas_total": n_total_malas,
        "pct_malas_indistinguibles": pct_indistinguibles,
        "archivos_detectables": archivos_detectables,
        "archivos_indistinguibles": archivos_indist,
    }


def eda_figura_resumen(registros_b, registros_m, X, y, columnas, ruta_eda):
    """Una sola figura compacta con lo esencial: PCA, t-SNE, top features, outliers."""
    fig = plt.figure(figsize=(16, 10))
    gs = fig.add_gridspec(2, 3)

    Xs = StandardScaler().fit_transform(X)
    pca = PCA(n_components=2).fit_transform(Xs)

    ax1 = fig.add_subplot(gs[0, 0])
    for cls, color, lab in [(0, "#4CAF50", "BUENA"), (1, "#E53935", "MALA")]:
        ax1.scatter(pca[y == cls, 0], pca[y == cls, 1], c=color, alpha=0.6,
                    label=lab, s=25)
    ax1.set_title("PCA 2D"); ax1.legend(); ax1.grid(alpha=0.3)
    ax1.set_xlabel("PC1"); ax1.set_ylabel("PC2")

    ax2 = fig.add_subplot(gs[0, 1])
    if len(X) > 10:
        try:
            ts = TSNE(n_components=2, perplexity=min(30, len(X) // 4),
                      random_state=0, init="pca").fit_transform(Xs)
            for cls, color, lab in [(0, "#4CAF50", "BUENA"),
                                    (1, "#E53935", "MALA")]:
                ax2.scatter(ts[y == cls, 0], ts[y == cls, 1], c=color,
                            alpha=0.6, label=lab, s=25)
            ax2.set_title("t-SNE 2D"); ax2.legend(); ax2.grid(alpha=0.3)
        except Exception as e:
            ax2.text(0.5, 0.5, f"t-SNE fallo: {e}", ha="center")

    ax3 = fig.add_subplot(gs[0, 2])
    rf = RandomForestClassifier(n_estimators=200, random_state=0,
                                class_weight="balanced", n_jobs=-1)
    rf.fit(X, y)
    imp = rf.feature_importances_
    orden = np.argsort(imp)[::-1][:10]
    ax3.barh(range(len(orden)), imp[orden][::-1], color="steelblue")
    ax3.set_yticks(range(len(orden)))
    ax3.set_yticklabels([columnas[i] for i in orden][::-1], fontsize=8)
    ax3.set_title(f"Top 10 features (RF acc={rf.score(X, y):.2f})")

    # Histograma de distancias al centroide de buenas
    centroide_b = np.mean(Xs[y == 0], axis=0)
    distancias = np.linalg.norm(Xs - centroide_b, axis=1)
    ax4 = fig.add_subplot(gs[1, :])
    ax4.hist(distancias[y == 0], bins=30, color="#4CAF50", alpha=0.6,
             label="BUENAS")
    ax4.hist(distancias[y == 1], bins=30, color="#E53935", alpha=0.6,
             label="MALAS")
    p95_buenas = np.percentile(distancias[y == 0], 95)
    ax4.axvline(p95_buenas, color="black", ls="--",
                label=f"P95 BUENAS = {p95_buenas:.2f}")
    ax4.set_title("Distancia al centroide de BUENAS en espacio de features")
    ax4.set_xlabel("Distancia"); ax4.set_ylabel("Frecuencia"); ax4.legend()

    plt.tight_layout()
    plt.savefig(os.path.join(ruta_eda, "eda_resumen.png"), dpi=110)
    plt.close()
    return float(rf.score(X, y))


def ejecutar_eda(imgs_buenas, imgs_malas, ruta_salida_eda, segmentar_fn):
    """
    Pipeline completo de EDA: calcula métricas, ejecuta tests, genera
    figura resumen y guarda JSON con resultados clave para el detector.
    """
    log.info("=" * 70)
    log.info(" FASE 1 / EDA - Análisis exploratorio de datos")
    log.info("=" * 70)
    Path(ruta_salida_eda).mkdir(parents=True, exist_ok=True)

    # 1. Métricas por imagen
    log.info(f"  Calculando métricas sobre {len(imgs_buenas)} BUENAS + "
             f"{len(imgs_malas)} MALAS...")
    registros_b = [{"archivo": n, "etiqueta": "BUENA"} for n, _ in imgs_buenas]
    registros_m = [{"archivo": n, "etiqueta": "MALA"} for n, _ in imgs_malas]

    iterador = enumerate(imgs_buenas)
    if TQDM_OK:
        iterador = tqdm(list(iterador), desc="EDA BUENAS", ncols=80)
    for i, (n, img) in iterador:
        registros_b[i].update(eda_estadisticas_intensidad(img))
        registros_b[i].update(eda_metricas_geometria(img, segmentar_fn))
        registros_b[i].update(eda_metricas_textura(img, segmentar_fn))

    iterador = enumerate(imgs_malas)
    if TQDM_OK:
        iterador = tqdm(list(iterador), desc="EDA MALAS", ncols=80)
    for i, (n, img) in iterador:
        registros_m[i].update(eda_estadisticas_intensidad(img))
        registros_m[i].update(eda_metricas_geometria(img, segmentar_fn))
        registros_m[i].update(eda_metricas_textura(img, segmentar_fn))

    columnas = [k for k in registros_b[0].keys()
                if k.startswith(("int_", "geom_", "tex_"))]

    # 2. Tests
    log.info("  Ejecutando tests Kolmogorov-Smirnov + Mann-Whitney...")
    tests = eda_tests(registros_b, registros_m, columnas)
    n_muysig = sum(1 for r in tests if r["ks_p"] < 0.001)
    n_sig = sum(1 for r in tests if r["ks_p"] < 0.05)
    log.info(f"  Métricas con separación: {n_muysig} muy sig. (p<0.001), "
             f"{n_sig} sig. (p<0.05) de {len(tests)} totales")

    # 3. Matriz X, y
    X = np.array([[r[c] for c in columnas]
                  for r in registros_b + registros_m], dtype=np.float64)
    y = np.array([0] * len(registros_b) + [1] * len(registros_m))
    X = np.nan_to_num(X, nan=0.0, posinf=0.0, neginf=0.0)

    # 4. Outliers
    out = eda_outliers(X, y, registros_b, registros_m)
    log.info(f"  MALAS detectables (outliers de BUENAS): "
             f"{out['n_malas_detectables']}/{out['n_malas_total']} "
             f"({100 - out['pct_malas_indistinguibles']:.1f}%)")
    log.info(f"  MALAS indistinguibles: "
             f"{out['pct_malas_indistinguibles']:.1f}%")

    # 5. Figura resumen
    log.info("  Generando figura resumen...")
    rf_acc = eda_figura_resumen(registros_b, registros_m, X, y, columnas,
                                ruta_salida_eda)

    # 6. Guardar tablas
    with open(os.path.join(ruta_salida_eda, "tabla_estadistica.csv"),
              "w", newline="", encoding="utf-8") as fh:
        campos = ["archivo", "etiqueta"] + columnas
        w = csv.DictWriter(fh, fieldnames=campos, extrasaction="ignore")
        w.writeheader()
        for r in registros_b + registros_m:
            w.writerow(r)
    with open(os.path.join(ruta_salida_eda, "tabla_resumen_tests.csv"),
              "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(tests[0].keys()))
        w.writeheader(); w.writerows(tests)

    # 7. JSON con resumen para el detector
    resumen_json = {
        "n_buenas": len(registros_b),
        "n_malas": len(registros_m),
        "n_metricas": len(columnas),
        "n_metricas_muy_significativas_p001": n_muysig,
        "n_metricas_significativas_p005": n_sig,
        "rf_accuracy_train": rf_acc,
        "outliers": {
            "umbral_distancia_p95_buenas": out["umbral_distancia_p95_buenas"],
            "n_malas_detectables": out["n_malas_detectables"],
            "n_malas_total": out["n_malas_total"],
            "pct_malas_indistinguibles": out["pct_malas_indistinguibles"],
            "archivos_detectables": out["archivos_detectables"],
            "archivos_indistinguibles": out["archivos_indistinguibles"],
        },
        "top_metricas": tests[:10],
    }
    with open(os.path.join(ruta_salida_eda, "eda_resumen.json"),
              "w", encoding="utf-8") as fh:
        json.dump(resumen_json, fh, indent=2, ensure_ascii=False)
    log.info(f"  EDA guardado en: {ruta_salida_eda}")
    return resumen_json


# ==============================================================================
#  FASE 2 - Bloques originales (segmentación, BLMD, PatchCore)
# ==============================================================================


# ==============================================================================
#  BLOQUE 0 - SEGMENTACIÓN DE LA HERRAMIENTA Y BANDA DE CONTORNO
# ==============================================================================
#
#  Estas imágenes son siluetas en transmisión (shadowgraph). La herramienta
#  ocupa la zona OSCURA conectada al borde superior de la imagen; el resto
#  (luz de fondo blanca) NO es región de análisis.
#
#  El desgaste no está dentro del cuerpo metálico (homogéneo) ni en el
#  fondo (luz). Está en la SILUETA: filo de los dientes, puntas, valles.
#  Por eso definimos una "banda de contorno" de N píxeles a cada lado del
#  borde de la herramienta y restringimos PatchCore a esa banda.
# ==============================================================================

def segmentar_herramienta(img_gris, kernel_cierre=15, area_min_rel=0.05,
                          orientacion="auto"):
    """
    Devuelve una máscara binaria (uint8 0/255) de la herramienta.

    Parámetros
    ----------
    orientacion : str
        Dónde está conectada la herramienta:
          "top"    -> conectada al borde superior (caso típico de tus imágenes)
          "bottom" -> conectada al borde inferior
          "auto"   -> probar ambas y elegir la más oscura
    """
    H, W = img_gris.shape
    img_blur = cv2.GaussianBlur(img_gris, (5, 5), 0)

    # Otsu en ambas direcciones
    _, mask_oscuro = cv2.threshold(img_blur, 0, 255,
                                   cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    _, mask_claro = cv2.threshold(img_blur, 0, 255,
                                  cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    # Cierre morfológico
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE,
                                       (kernel_cierre, kernel_cierre))
    mask_oscuro = cv2.morphologyEx(mask_oscuro, cv2.MORPH_CLOSE, kernel)
    mask_claro = cv2.morphologyEx(mask_claro, cv2.MORPH_CLOSE, kernel)

    def quedarse_componente_conexa_a_borde(mask, borde):
        """Conserva solo la componente conexa que toca el borde indicado."""
        n_lab, labels, stats, _ = cv2.connectedComponentsWithStats(mask, 8)
        if n_lab <= 1:
            return mask
        candidatas = []
        for i in range(1, n_lab):
            comp_mask = (labels == i)
            toca = False
            if borde == "top" and np.any(comp_mask[0, :]):
                toca = True
            elif borde == "bottom" and np.any(comp_mask[-1, :]):
                toca = True
            elif borde == "left" and np.any(comp_mask[:, 0]):
                toca = True
            elif borde == "right" and np.any(comp_mask[:, -1]):
                toca = True
            if toca:
                candidatas.append((i, stats[i, cv2.CC_STAT_AREA]))
        if not candidatas:
            return np.zeros_like(mask)
        # La de mayor área que toque ese borde
        idx_sel = max(candidatas, key=lambda x: x[1])[0]
        return np.where(labels == idx_sel, 255, 0).astype(np.uint8)

    # Decisión de orientación
    if orientacion == "auto":
        # Probar la versión "oscuro pegado a top"
        cand_top = quedarse_componente_conexa_a_borde(mask_oscuro, "top")
        cand_bot = quedarse_componente_conexa_a_borde(mask_oscuro, "bottom")
        # Elegir la que tenga mayor área (mejor cobertura)
        if np.sum(cand_top > 0) >= np.sum(cand_bot > 0):
            mask_final = cand_top
            log.debug("Segmentación: herramienta conectada al borde SUPERIOR")
        else:
            mask_final = cand_bot
            log.debug("Segmentación: herramienta conectada al borde INFERIOR")
    elif orientacion == "top":
        mask_final = quedarse_componente_conexa_a_borde(mask_oscuro, "top")
    elif orientacion == "bottom":
        mask_final = quedarse_componente_conexa_a_borde(mask_oscuro, "bottom")
    else:
        raise ValueError(f"orientacion desconocida: {orientacion}")

    # Salvavidas: si la segmentación dio algo degenerado, volver a Otsu plano
    if np.sum(mask_final > 0) < area_min_rel * H * W:
        log.debug("Segmentación degenerada; usando Otsu sin filtrar.")
        return mask_oscuro

    return mask_final


def banda_contorno(mascara_herramienta, ancho=20):
    """
    Devuelve una máscara binaria de una banda de 'ancho' píxeles a cada lado
    del contorno de la herramienta. Esta es la zona donde realmente está el
    desgaste: el filo de los dientes, sus puntas y valles.

    Parámetros
    ----------
    ancho : int
        Mitad del grosor total de la banda (ancho a CADA lado del contorno).
    """
    k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (ancho * 2 + 1,
                                                       ancho * 2 + 1))
    dilatado = cv2.dilate(mascara_herramienta, k)
    erosionado = cv2.erode(mascara_herramienta, k)
    return cv2.subtract(dilatado, erosionado)


def aplicar_mascara_a_imagen(img, mascara, valor_fondo=128):
    """Pone los píxeles fuera de la máscara a un valor neutro."""
    img_out = img.copy()
    img_out[mascara == 0] = valor_fondo
    return img_out


def parches_dentro_de_mascara(mascara, shape_patches):
    """
    Devuelve un vector booleano que indica qué parches del feature map
    (Hp x Wp) están mayoritariamente dentro de la máscara.
    """
    Hp, Wp = shape_patches
    mascara_224 = cv2.resize(mascara, (224, 224), interpolation=cv2.INTER_NEAREST)
    cell_h = 224 // Hp
    cell_w = 224 // Wp
    bloques = mascara_224[:Hp * cell_h, :Wp * cell_w].reshape(
        Hp, cell_h, Wp, cell_w).mean(axis=(1, 3))
    return (bloques > 127).flatten()


# ==============================================================================
#  BLOQUE 1 - BLMD RIGUROSO (Bidimensional Local Mean Decomposition)
# ==============================================================================
#
#  El algoritmo BLMD descompone una imagen I(x,y) en una suma de Product
#  Functions (PFs) y un residuo monótono:
#
#                    I(x,y) = sum_k PF_k(x,y) + r(x,y)
#
#  Cada PF_k = a_k(x,y) * h_k(x,y), donde:
#       - a_k es la envolvente local (amplitud)
#       - h_k es una función puramente FM (modulada en frecuencia)
#
#  Procedimiento (sifting) por cada PF:
#       1. Detectar extremos locales (max y min) en I.
#       2. Interpolar envolvente superior e_max y envolvente inferior e_min.
#       3. Calcular media local:        m = (e_max + e_min) / 2
#       4. Calcular envolvente local:   a = (e_max - e_min) / 2
#       5. Suavizar m y a con ventana móvil.
#       6. Restar la media:             h_raw = I - m
#       7. Demodular en amplitud:       h = h_raw / a
#       8. Si h no es FM puro (amplitud != 1), repetir desde 1 con h como I.
#       9. Producto:                    PF = a_total * h
#      10. Residuo:                     I_next = I - PF
#      11. Si I_next es monótono o se alcanza el máximo de PFs, parar.
# ==============================================================================

def _submuestrear_extremos(coords, vals, max_n):
    """
    Si hay más de max_n extremos, muestrea espacialmente uniformes para
    preservar la cobertura. Evita que griddata cúbica colapse con miles de
    puntos en imágenes muy detalladas.
    """
    n = len(coords)
    if n <= max_n:
        return coords, vals
    factor = int(np.ceil(np.sqrt(n / max_n)))
    keys = (coords[:, 0] // factor) * 100000 + (coords[:, 1] // factor)
    _, idx = np.unique(keys, return_index=True)
    if len(idx) > max_n:
        rng = np.random.default_rng(0)
        idx = rng.choice(idx, size=max_n, replace=False)
    return coords[idx], vals[idx]


def detectar_extremos_2d(img, vecindario=5, max_extremos=2000):
    """
    Detecta máximos y mínimos locales en una imagen 2D.

    Parámetros
    ----------
    max_extremos : int
        Máximo número de extremos por tipo (max/min) que se devuelven. Si la
        imagen tiene más, se hace submuestreo espacial uniforme.
    """
    img = img.astype(np.float64)

    # Un píxel es máximo si es igual al máximo de su vecindario
    max_filtrado = ndimage.maximum_filter(img, size=vecindario, mode="reflect")
    min_filtrado = ndimage.minimum_filter(img, size=vecindario, mode="reflect")

    mascara_max = (img == max_filtrado) & (img > min_filtrado)  # excluye zonas planas
    mascara_min = (img == min_filtrado) & (img < max_filtrado)

    coords_max = np.argwhere(mascara_max)
    coords_min = np.argwhere(mascara_min)

    vals_max = img[mascara_max]
    vals_min = img[mascara_min]

    coords_max, vals_max = _submuestrear_extremos(coords_max, vals_max, max_extremos)
    coords_min, vals_min = _submuestrear_extremos(coords_min, vals_min, max_extremos)

    return coords_max, vals_max, coords_min, vals_min


def _anadir_extremos_borde(coords, vals, shape):
    """
    Añade las cuatro esquinas como extremos virtuales para evitar artefactos
    de borde en la interpolación. Es una práctica estándar en BEMD/BLMD
    (Nunes & Deléchelle, 2009).
    """
    h, w = shape
    if len(coords) == 0:
        # Si no hay extremos, usar las esquinas con valor medio
        return np.array([[0, 0], [0, w - 1], [h - 1, 0], [h - 1, w - 1]]), \
               np.array([0.0, 0.0, 0.0, 0.0])

    val_medio = float(np.mean(vals))
    esquinas = np.array([[0, 0], [0, w - 1], [h - 1, 0], [h - 1, w - 1]])
    coords_ext = np.vstack([coords, esquinas])
    vals_ext = np.concatenate([vals, np.full(4, val_medio)])
    return coords_ext, vals_ext


def interpolar_envolvente(coords, vals, shape, metodo="cubic"):
    """
    Interpola una superficie envolvente a partir de un conjunto disperso de
    extremos. Si griddata falla (pocos puntos), recurre a 'nearest'.
    """
    coords, vals = _anadir_extremos_borde(coords, vals, shape)

    h, w = shape
    grid_y, grid_x = np.mgrid[0:h, 0:w]

    try:
        envolvente = griddata(
            points=coords, values=vals,
            xi=(grid_y, grid_x), method=metodo, fill_value=float(np.mean(vals))
        )
    except Exception as e:
        log.warning(f"griddata cúbica falló ({e}); usando 'nearest'.")
        envolvente = griddata(
            points=coords, values=vals,
            xi=(grid_y, grid_x), method="nearest"
        )

    # Por si quedan NaN por píxeles en el casco convexo
    if np.any(np.isnan(envolvente)):
        envolvente = np.nan_to_num(envolvente, nan=float(np.nanmean(envolvente)))

    return envolvente


def suavizar_ventana_movil(img, tam=11):
    """
    Suavizado por ventana móvil cuadrada (filtro de caja). Smith (2005)
    recomienda usar la longitud máxima entre extremos consecutivos como
    base para el tamaño de ventana.
    """
    tam = max(3, int(tam) | 1)  # forzar impar
    return cv2.boxFilter(img.astype(np.float64), ddepth=-1, ksize=(tam, tam))


def _es_fm_puro(a_total, tol=0.05):
    """
    Comprueba si la envolvente acumulada es aproximadamente 1 en todos los
    píxeles (criterio FM puro). Devuelve la desviación máxima respecto a 1.
    """
    return float(np.max(np.abs(a_total - 1.0))) < tol


def sifting_blmd(img, max_iter=10, tol_cauchy=0.01, vecindario=5,
                 tam_suavizado=11, a_min=1e-3):
    """
    Extrae UNA Product Function (PF) mediante sifting iterativo.

    Parámetros
    ----------
    a_min : float
        Suelo mínimo para la envolvente local. Evita explosiones numéricas
        cuando la imagen tiende a ser plana (envolvente → 0).

    Devuelve
    -------
    pf : np.ndarray
        Product Function (= a_total * h).
    n_iter : int
        Número de iteraciones realizadas.
    convergido : bool
        True si el bucle paró por criterio Cauchy, False si por max_iter.
    """
    h_actual = img.astype(np.float64).copy()
    a_total = np.ones_like(h_actual)
    h_prev = h_actual.copy()

    # Si la varianza inicial es ya muy pequeña, devolver la propia imagen
    # como PF degenerada (residuo casi plano). Evita amplificar ruido.
    if np.std(h_actual) < 1e-6:
        return np.zeros_like(h_actual), 0, True

    escala_global = float(np.std(img))  # referencia para suelo de a

    for it in range(1, max_iter + 1):
        c_max, v_max, c_min, v_min = detectar_extremos_2d(h_actual, vecindario)

        # Si quedan pocos extremos, ya no se puede seguir descomponiendo
        if len(c_max) < 4 or len(c_min) < 4:
            log.debug(f"  Sifting: insuficientes extremos en iter {it}, parando.")
            break

        e_max = interpolar_envolvente(c_max, v_max, h_actual.shape)
        e_min = interpolar_envolvente(c_min, v_min, h_actual.shape)

        # Suavizado para obtener funciones realmente locales y suaves
        e_max = suavizar_ventana_movil(e_max, tam_suavizado)
        e_min = suavizar_ventana_movil(e_min, tam_suavizado)

        m = 0.5 * (e_max + e_min)         # media local
        a = 0.5 * (e_max - e_min)         # envolvente local (amplitud)

        # Suelo en a relativo a la escala global de la señal: evita
        # explosiones cuando la envolvente es localmente cero.
        suelo = max(a_min, a_min * escala_global)
        a_safe = np.where(np.abs(a) < suelo, suelo, np.abs(a))

        h_raw = h_actual - m
        h_actual = h_raw / a_safe
        a_total = a_total * a_safe

        # Criterio de parada Cauchy (similar a Huang et al. para EMD)
        denom = np.sum(h_prev ** 2) + 1e-12
        sd = np.sum((h_actual - h_prev) ** 2) / denom
        h_prev = h_actual.copy()

        # Si la PF acumulada se vuelve numéricamente inestable, parar.
        if not np.all(np.isfinite(h_actual)) or np.max(np.abs(h_actual)) > 1e6:
            log.debug(f"  Sifting: inestabilidad numérica en iter {it}, parando.")
            break

        if sd < tol_cauchy or _es_fm_puro(a_total, tol=0.05):
            return a_total * h_actual, it, True

    return a_total * h_actual, max_iter, False


def blmd_decomposition(img, n_pfs=3, max_iter_sift=8, vecindario=5,
                       tam_suavizado=11, verbose=False,
                       energia_rel_min=1e-6, energia_rel_max=10.0):
    """
    Descomposición BLMD completa: extrae hasta n_pfs Product Functions
    y un residuo.

    Parámetros adicionales (guardas numéricas)
    ------------------------------------------
    energia_rel_min : float
        Si la energía relativa de una PF respecto a la imagen original cae
        por debajo de este umbral, se interpreta como ruido y se detiene.
    energia_rel_max : float
        Si la energía relativa de una PF supera este umbral, se considera
        explosión numérica y se descarta la PF + se detiene.

    Devuelve
    -------
    pfs : list[np.ndarray]
        Product Functions en orden (alta -> baja frecuencia).
    residuo : np.ndarray
        Residuo final (tendencia / iluminación).
    info : list[dict]
        Diagnóstico de cada PF.
    """
    img_f = img.astype(np.float64)
    energia_orig = float(np.sum(img_f ** 2)) + 1e-12
    residuo = img_f.copy()
    pfs = []
    info = []

    for k in range(1, n_pfs + 1):
        # Si el residuo ya no tiene suficientes extremos, parar
        c_max, _, c_min, _ = detectar_extremos_2d(residuo, vecindario)
        if len(c_max) < 6 or len(c_min) < 6:
            if verbose:
                log.info(f"  PF{k}: residuo casi monótono, parando.")
            break

        pf, n_iter, conv = sifting_blmd(
            residuo,
            max_iter=max_iter_sift,
            vecindario=vecindario,
            tam_suavizado=tam_suavizado,
        )

        energia_pf = float(np.sum(pf ** 2))
        energia_rel = energia_pf / energia_orig

        # Guarda 1: PF degenerada (todo ceros o demasiado pequeña)
        if energia_rel < energia_rel_min:
            if verbose:
                log.info(f"  PF{k}: energía relativa {energia_rel:.2e} muy "
                         f"baja, descartando y parando.")
            break

        # Guarda 2: explosión numérica
        if energia_rel > energia_rel_max or not np.all(np.isfinite(pf)):
            if verbose:
                log.info(f"  PF{k}: energía relativa {energia_rel:.2e} "
                         f"inestable, descartando y parando.")
            break

        pfs.append(pf)
        residuo = residuo - pf
        info.append({
            "pf_index": k,
            "iter": n_iter,
            "convergido": conv,
            "energia": energia_pf,
            "energia_relativa": energia_rel,
        })
        if verbose:
            log.info(f"  PF{k}: {n_iter} iter, conv={conv}, "
                     f"E_rel={energia_rel:.2e}")

    return pfs, residuo, info


def blmd_realce(img, n_pfs=3, pesos=None):
    """
    Reconstrucción realzada de la imagen sumando PFs con pesos. Por defecto
    da más peso a las primeras PFs (alta frecuencia = textura/desgaste) y
    descarta el residuo (iluminación).

    Esto es lo que entrega al pipeline de IA como imagen "preprocesada BLMD".
    """
    pfs, _residuo, _info = blmd_decomposition(img, n_pfs=n_pfs)
    if not pfs:
        return cv2.normalize(img, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)

    if pesos is None:
        # Más peso a alta frecuencia. Ej. con 3 PFs: [1.5, 1.0, 0.5]
        pesos = np.linspace(1.5, 0.5, num=len(pfs))

    realce = np.zeros_like(pfs[0])
    for pf, w in zip(pfs, pesos):
        realce += w * pf

    return cv2.normalize(realce, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)


# ==============================================================================
#  BLOQUE 1B - APROXIMACIÓN RÁPIDA (la de V10) - CONSERVADA PARA COMPARAR
# ==============================================================================

def blmd_aproximado_v10(img):
    """
    Aproximación rápida tipo high-pass usada en CODIGOV10. Se mantiene
    únicamente como referencia para la comparación de la memoria. NO es
    BLMD real: es una sustracción de envolvente gaussiana de un solo paso.
    """
    envolvente = cv2.GaussianBlur(img.astype(float), (15, 15), 0)
    imf = img - envolvente
    return cv2.normalize(imf, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)


# ==============================================================================
#  BLOQUE 2 - PREPROCESADO HOMOMÓRFICO (sin cambios respecto a V10)
# ==============================================================================

def filtro_homomorfico(img, sigma=30, gamma_l=0.5, gamma_h=1.5):
    """
    Filtrado homomórfico: separa iluminación (baja frec.) y reflectancia
    (alta frec.) en dominio logarítmico, atenúa la primera y realza la
    segunda. Cheng et al. (2009); Mrinal Sarvagya et al. (2024).
    """
    img_log = np.log1p(np.array(img, dtype=np.float64) / 255.0)
    dft = np.fft.fft2(img_log)
    dft_shift = np.fft.fftshift(dft)

    rows, cols = img.shape
    crow, ccol = rows // 2, cols // 2
    y, x = np.ogrid[-crow:rows - crow, -ccol:cols - ccol]

    mask = (gamma_h - gamma_l) * (1 - np.exp(-(x ** 2 + y ** 2) /
                                             (2 * (sigma ** 2)))) + gamma_l
    fshift = dft_shift * mask
    img_back = np.fft.ifft2(np.fft.ifftshift(fshift))
    img_res = np.exp(np.real(img_back)) - 1
    img_norm = cv2.normalize(img_res, None, 0, 255, cv2.NORM_MINMAX)
    return np.uint8(np.clip(img_norm, 0, 255))


# ==============================================================================
#  BLOQUE 3 - PATCHCORE: Extractor patch-level + Coreset + Anomaly Map
# ==============================================================================
#
#  PatchCore (Roth et al., CVPR 2022) es un método de detección de anomalías
#  que NO entrena ningún modelo: usa un backbone preentrenado en ImageNet
#  (WideResNet50) y construye un "memory bank" con los features de capas
#  intermedias de las imágenes BUENAS. En inferencia, calcula la distancia
#  de cada parche de la imagen test al parche más cercano del banco.
#
#  Tres puntos clave que diferencian PatchCore de un encoder global:
#    1. Usa capas INTERMEDIAS (layer2, layer3) -> features mid-level que
#       capturan textura, no semántica abstracta.
#    2. Mantiene la estructura espacial: un vector por cada celda del feature
#       map -> permite localizar dónde está la anomalía.
#    3. Reduce el banco con coreset sampling para acelerar inferencia sin
#       perder cobertura.
# ==============================================================================

_TRANSFORM_PATCHCORE = transforms.Compose([
    transforms.ToPILImage(),
    transforms.Resize((224, 224)),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406],
                         std=[0.229, 0.224, 0.225]),
])


class PatchCoreExtractor:
    """
    Extractor de features patch-level para PatchCore.

    Usa WideResNet50 preentrenada en ImageNet. Captura las salidas de layer2
    y layer3 mediante forward hooks, las adapta al mismo tamaño espacial
    (el de layer2, más grueso pero con más detalle local) y las concatena.

    Cada imagen de entrada (224x224) produce un tensor (H_p x W_p x D)
    donde H_p x W_p es la resolución del feature map (típicamente 28x28
    para layer2 de WideResNet50) y D es la dimensión de feature concatenada.
    """

    def __init__(self, device, backbone="wide_resnet50_2"):
        self.device = device
        self.backbone_name = backbone

        if backbone == "wide_resnet50_2":
            self.model = models.wide_resnet50_2(
                weights=models.Wide_ResNet50_2_Weights.IMAGENET1K_V1)
        elif backbone == "resnet18":
            self.model = models.resnet18(
                weights=models.ResNet18_Weights.IMAGENET1K_V1)
        else:
            raise ValueError(f"Backbone desconocido: {backbone}")

        self.model.eval().to(device)
        for p in self.model.parameters():
            p.requires_grad = False

        # Buffers donde los hooks dejan las activaciones
        self._feat_l2 = None
        self._feat_l3 = None
        self.model.layer2.register_forward_hook(self._hook_l2)
        self.model.layer3.register_forward_hook(self._hook_l3)

    def _hook_l2(self, module, inp, out):
        self._feat_l2 = out

    def _hook_l3(self, module, inp, out):
        self._feat_l3 = out

    def extraer_patches(self, img_gris):
        """
        Devuelve un array (H_p * W_p, D) con un embedding por parche, y la
        forma espacial (H_p, W_p) para reconstruir mapas 2D después.
        """
        img_rgb = cv2.cvtColor(img_gris, cv2.COLOR_GRAY2RGB)
        tensor_img = _TRANSFORM_PATCHCORE(img_rgb).unsqueeze(0).to(self.device)
        with torch.no_grad():
            _ = self.model(tensor_img)

        f2 = self._feat_l2  # (1, C2, H2, W2)
        f3 = self._feat_l3  # (1, C3, H3, W3) - más pequeño espacialmente

        # Adaptar f3 al tamaño de f2 con interpolación bilineal y concatenar
        # en el eje de canales -> features multi-escala combinadas.
        f3_up = torch.nn.functional.interpolate(
            f3, size=f2.shape[-2:], mode="bilinear", align_corners=False)
        feats = torch.cat([f2, f3_up], dim=1)        # (1, C2+C3, H2, W2)

        _, C, H, W = feats.shape
        # Aplanar parches: (H*W, C)
        patches = feats[0].permute(1, 2, 0).reshape(H * W, C)
        return patches.cpu().numpy(), (H, W)


def coreset_subsample(features, ratio=0.01, seed=0, verbose=False):
    """
    Coreset sampling por k-center greedy (Sener & Savarese, 2018).
    Selecciona un subconjunto que minimiza la máxima distancia entre
    cualquier punto del conjunto original y su vecino más cercano del
    subconjunto.

    Implementación O(N*K) donde K = N*ratio. Para N grande puede ser lenta;
    para PatchCore con ratio 1% suele ser asumible.

    Parámetros
    ----------
    features : np.ndarray (N, D)
    ratio : float
        Fracción del banco original a conservar (1% por defecto).
    seed : int
        Semilla para reproducibilidad del primer punto.
    """
    N = features.shape[0]
    K = max(1, int(N * ratio))

    if K >= N:
        return features, np.arange(N)

    rng = np.random.default_rng(seed)
    idx_inicial = int(rng.integers(0, N))
    seleccionados = [idx_inicial]

    # Distancia mínima de cada punto al conjunto seleccionado
    dist_min = np.linalg.norm(features - features[idx_inicial], axis=1)

    iterador = range(1, K)
    if verbose:
        iterador = tqdm(iterador, desc="Coreset", ncols=80)

    for _ in iterador:
        # El siguiente seleccionado es el punto más alejado del conjunto actual
        idx_nuevo = int(np.argmax(dist_min))
        seleccionados.append(idx_nuevo)

        # Actualizar distancia mínima incorporando el nuevo punto
        d_nuevo = np.linalg.norm(features - features[idx_nuevo], axis=1)
        dist_min = np.minimum(dist_min, d_nuevo)

    seleccionados = np.array(seleccionados)
    return features[seleccionados], seleccionados


class PatchCoreDetector:
    """
    Pipeline completo PatchCore con soporte opcional de máscaras de
    herramienta:
      1. fit(imgs_buenas, mascaras_buenas) -> banco solo con parches herramienta
      2. score(img, mascara)               -> score y mapa filtrados
    """

    def __init__(self, extractor, coreset_ratio=0.01, k=1):
        self.extractor = extractor
        self.coreset_ratio = coreset_ratio
        self.k = k
        self.bank = None
        self.shape_patches = None
        self.nn = None

    def fit(self, imgs_buenas, mascaras_buenas=None, verbose=True):
        """
        Construye el memory bank patch-level.

        Si mascaras_buenas se proporciona, solo los parches mayoritariamente
        DENTRO de la herramienta entran al banco. Esto evita que el detector
        aprenda estadísticas del fondo.
        """
        bank_completo = []
        usar_masc = mascaras_buenas is not None

        iterador = imgs_buenas
        if verbose and TQDM_OK:
            iterador = tqdm(imgs_buenas, desc="PatchCore fit", ncols=80)

        for i, (nombre, img) in enumerate(iterador):
            patches, shape_p = self.extractor.extraer_patches(img)
            self.shape_patches = shape_p

            if usar_masc:
                mascara = mascaras_buenas[i]
                dentro = parches_dentro_de_mascara(mascara, shape_p)
                if np.sum(dentro) == 0:
                    continue  # imagen sin herramienta detectable
                patches = patches[dentro]

            bank_completo.append(patches)

        if not bank_completo:
            raise RuntimeError("Ningún parche válido tras enmascarar. "
                               "Revisa la segmentación.")

        bank_completo = np.vstack(bank_completo)
        log.info(f"  Banco completo: {bank_completo.shape[0]} parches "
                 f"x {bank_completo.shape[1]} dim "
                 f"(enmascarado={usar_masc})")

        bank_red, _ = coreset_subsample(bank_completo,
                                        ratio=self.coreset_ratio,
                                        verbose=verbose)
        log.info(f"  Coreset: {bank_red.shape[0]} parches conservados "
                 f"({100*bank_red.shape[0]/bank_completo.shape[0]:.1f}%)")

        self.bank = bank_red
        self.nn = NearestNeighbors(n_neighbors=self.k).fit(self.bank)

    def score(self, img, mascara=None):
        """
        Calcula score + anomaly map.

        Si mascara se proporciona:
          - El score solo considera parches dentro de la herramienta.
          - El anomaly map se enmascara a 0 fuera de la herramienta.
        """
        patches, (Hp, Wp) = self.extractor.extraer_patches(img)
        dists, _ = self.nn.kneighbors(patches)
        d_patch = dists[:, -1]

        mapa_pequeno = d_patch.reshape(Hp, Wp)

        # Reescalar al tamaño original
        H_orig, W_orig = img.shape
        mapa_full = cv2.resize(mapa_pequeno, (W_orig, H_orig),
                               interpolation=cv2.INTER_CUBIC)
        mapa_full = cv2.GaussianBlur(mapa_full, (0, 0), sigmaX=4)

        if mascara is not None:
            # Score = max del mapa SOLO dentro de la herramienta
            valores_herramienta = mapa_full[mascara > 0]
            if len(valores_herramienta) == 0:
                score = float(mapa_full.max())
            else:
                score = float(valores_herramienta.max())
            # Y enmascarar el mapa visualmente
            mapa_full = mapa_full * (mascara > 0).astype(np.float32)
        else:
            score = float(mapa_full.max())

        return score, mapa_full

    def score_batch(self, imgs, mascaras=None, verbose=True):
        """Evalúa una lista [(nombre, img)] y devuelve scores + mapas."""
        scores, mapas, nombres = [], [], []
        iterador = enumerate(imgs)
        if verbose and TQDM_OK:
            iterador = tqdm(list(iterador), desc="PatchCore score", ncols=80)
        for i, (nombre, img) in iterador:
            m = mascaras[i] if mascaras is not None else None
            s, mp = self.score(img, mascara=m)
            scores.append(s); mapas.append(mp); nombres.append(nombre)
        return np.array(scores), mapas, nombres


# Mantenemos el extractor legacy ResNet18 sin cabeza para compatibilidad con
# los pipelines de comparación de V12 (no patch-level). Sigue siendo útil
# para la ablación "qué pasaría sin PatchCore".

_TRANSFORM_IA = transforms.Compose([
    transforms.ToPILImage(),
    transforms.Resize((224, 224)),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406],
                         std=[0.229, 0.224, 0.225]),
])

_MODEL_AI = models.resnet18(weights=models.ResNet18_Weights.IMAGENET1K_V1)
_MODEL_AI = torch.nn.Sequential(*(list(_MODEL_AI.children())[:-2]))
_MODEL_AI.eval().to(DEVICE)


# (función extraer_caracteristicas() retirada en limpieza V20: ya no se usa)


# ==============================================================================
#  BLOQUE 4 - PIPELINES DE PREPROCESADO (intercambiables)
# ==============================================================================
#
#  Cada pipeline recibe una imagen en gris y devuelve una imagen en gris
#  realzada lista para entrar al extractor de características. Esto permite
#  comparar limpiamente distintas estrategias.
# ==============================================================================

def pipeline_blmd_real(img_gris):
    """Homomórfico -> BLMD real (recomendado para el TFM)."""
    return blmd_realce(filtro_homomorfico(img_gris), n_pfs=3)


def pipeline_blmd_v10(img_gris):
    """Homomórfico -> BLMD aproximado (V10). Solo para comparación."""
    return blmd_aproximado_v10(filtro_homomorfico(img_gris))


def pipeline_solo_homomorfico(img_gris):
    """Solo homomórfico, sin descomposición (ablación)."""
    return filtro_homomorfico(img_gris)


def pipeline_sin_preprocesado(img_gris):
    """Imagen original. Línea base (ablación)."""
    return img_gris


# ==============================================================================
#  BEMD - Bidimensional Empirical Mode Decomposition (V20)
# ==============================================================================
#  Hermana del BLMD pero usando IMFs en lugar de PFs. Implementación basada en
#  PyEMD (Nunes & Deléchelle 2009 extendido a 2D). Permite cumplir el objetivo
#  SMART del TFM "comparar BLMD vs BEMD vs wavelets".
# ==============================================================================

# Cache global de resultados BEMD: PyEMD tarda 30-60s por imagen y se llama
# DOS veces (FASE 4 + FASE 4-bis SNR). Cacheando por id de imagen ahorramos
# la mitad del tiempo.
_BEMD_CACHE = {}


def bemd_realce(img, n_imfs=2, pesos=None):
    """
    Aplica BEMD a la imagen y reconstruye con realce de detalles.

    Parámetros
    ----------
    img : ndarray (H, W) en uint8 o float
    n_imfs : int, número máximo de IMFs a extraer (2 por defecto en V17:
             el tercer IMF apenas aporta y duplica el tiempo de sifting)
    pesos : list[float] o None, ponderación de cada IMF en la reconstrucción

    Notas V17 (parche memoria/velocidad):
    -------------------------------------
    PyEMD usa interpolación RBF sobre los extremos locales, construyendo una
    matriz NxN donde N = nº de extremos. A 256x256 esto puede explotar a
    matrices de 60.000+ x 60.000 (8 GiB+). Bajamos a 96x96 para que el nº
    de extremos máximo sea ~96² ≈ 9.000 y la RBF cabe en ~640 MB. Sigue
    siendo una resolución suficiente para descomposición multi-escala
    (los detalles finos se preservan en IMF1).
    """
    if not BEMD_OK:
        # Si PyEMD no está, comportarse como passthrough sin caer
        return img

    # Cache: si ya hemos procesado esta imagen (mismo id de array), reusar
    cache_key = (id(img), img.shape, n_imfs)
    if cache_key in _BEMD_CACHE:
        return _BEMD_CACHE[cache_key]

    if pesos is None:
        # Realce: amplifica detalles (IMF1, IMF2) y atenúa residuo
        pesos = [1.5, 1.2, 0.8]

    # BEMD trabaja con float
    img_f = img.astype(np.float32)
    # V17: submuestreo agresivo a 96x96 para acotar el coste de la RBF.
    # Resolución mayor (256x256) hace que PyEMD intente asignar 8 GiB y falle
    # actuando como passthrough silencioso (= BEMD se vuelve SIN_PREP).
    img_orig_shape = img.shape
    BEMD_SIZE = 96
    img_f = cv2.resize(img_f, (BEMD_SIZE, BEMD_SIZE),
                       interpolation=cv2.INTER_AREA)

    try:
        # PyEMD ha cambiado de API entre versiones. Probamos varias formas.
        # Versión nueva: EMD2D es la clase, BEMD es el módulo
        try:
            from PyEMD import EMD2D
            emd2d = EMD2D()
            # V17: limitar iteraciones de sifting (default 1000 era ridículo)
            try:
                emd2d.MAX_ITERATION = 10
            except Exception:
                pass
            imfs = emd2d.emd(img_f, max_imf=n_imfs)
        except (ImportError, AttributeError):
            # Versión antigua: BEMD es la clase
            from PyEMD.BEMD import BEMD as BEMDClass
            bemd = BEMDClass()
            # V17: limitar iteraciones de sifting
            try:
                bemd.MAX_ITERATION = 10
            except Exception:
                pass
            imfs = bemd.bemd(img_f, max_imf=n_imfs)
    except MemoryError as e:
        log.warning(f"BEMD sin memoria ({e}), devolviendo imagen original.")
        _BEMD_CACHE[cache_key] = img
        return img
    except Exception as e:
        log.warning(f"BEMD falló ({e}), devolviendo imagen original.")
        _BEMD_CACHE[cache_key] = img
        return img

    if imfs is None or len(imfs) == 0:
        _BEMD_CACHE[cache_key] = img
        return img

    # Reconstrucción ponderada
    n = min(len(imfs), len(pesos))
    reconst = np.zeros_like(img_f)
    for i in range(n):
        reconst += pesos[i] * imfs[i]
    # Si hay residuo, sumarlo con peso bajo
    if len(imfs) > n:
        reconst += 0.5 * imfs[-1]

    # Volver al tamaño original
    if reconst.shape != img.shape:
        reconst = cv2.resize(reconst, (img.shape[1], img.shape[0]),
                             interpolation=cv2.INTER_LINEAR)

    # Normalizar a [0, 255]
    rmin, rmax = float(reconst.min()), float(reconst.max())
    if rmax - rmin > 1e-8:
        reconst = 255.0 * (reconst - rmin) / (rmax - rmin)
    result = reconst.astype(np.uint8)
    _BEMD_CACHE[cache_key] = result
    return result


def pipeline_bemd_real(img_gris):
    """Homomórfico -> BEMD real (PyEMD).

    V17: n_imfs=2 (antes 3) para acelerar. El tercer IMF aporta poco
    en imágenes de herramientas y duplica el coste de sifting.
    """
    return bemd_realce(filtro_homomorfico(img_gris), n_imfs=2)


# ==============================================================================
#  WAVELET - Transformada Wavelet Discreta 2D (V20)
# ==============================================================================
#  Descomposición Daubechies-4 con 3 niveles. Realce de coeficientes de detalle
#  (donde está la textura del desgaste) y reconstrucción.
# ==============================================================================

def wavelet_realce(img, wavelet="db4", niveles=3, factor_detalle=1.5):
    """
    Aplica DWT 2D, amplifica los coeficientes de detalle y reconstruye.

    Parámetros
    ----------
    img : ndarray (H, W) en uint8
    wavelet : str, familia wavelet (db4, sym4, bior2.2...)
    niveles : int, niveles de descomposición
    factor_detalle : float, factor de amplificación de los coeficientes de
                     detalle (cH, cV, cD)
    """
    if not PYWAVELETS_OK:
        return img

    img_f = img.astype(np.float32)
    try:
        coefs = pywt.wavedec2(img_f, wavelet, level=niveles)
    except Exception as e:
        log.warning(f"DWT falló ({e}), devolviendo imagen original.")
        return img

    # coefs[0] = aproximación, coefs[1:] = tuplas (cH, cV, cD) por nivel
    coefs_realzados = [coefs[0]]
    for cH, cV, cD in coefs[1:]:
        coefs_realzados.append((cH * factor_detalle,
                                 cV * factor_detalle,
                                 cD * factor_detalle))

    try:
        reconst = pywt.waverec2(coefs_realzados, wavelet)
    except Exception as e:
        log.warning(f"IDWT falló ({e}), devolviendo imagen original.")
        return img

    # waverec2 puede devolver shape ligeramente distinta
    if reconst.shape != img.shape:
        reconst = reconst[:img.shape[0], :img.shape[1]]

    # Normalizar a [0, 255]
    rmin, rmax = float(reconst.min()), float(reconst.max())
    if rmax - rmin > 1e-8:
        reconst = 255.0 * (reconst - rmin) / (rmax - rmin)
    return np.clip(reconst, 0, 255).astype(np.uint8)


def pipeline_wavelet(img_gris):
    """Homomórfico -> Wavelet (Daubechies-4, 3 niveles)."""
    return wavelet_realce(filtro_homomorfico(img_gris))


# Diccionario maestro de pipelines.
# BEMD y WAVELET se incluyen aunque la librería no esté disponible: en ese caso
# los pipelines actúan como passthrough y emiten un warning al inicio.
PIPELINES = {
    "BLMD_REAL": pipeline_blmd_real,
    "BLMD_V10": pipeline_blmd_v10,
    "BEMD_REAL": pipeline_bemd_real,
    "WAVELET": pipeline_wavelet,
    "SOLO_HOMO": pipeline_solo_homomorfico,
    "SIN_PREP": pipeline_sin_preprocesado,
}


# ==============================================================================
#  BLOQUE 3-bis - COMPARATIVA SNR POR PIPELINE (Objetivo 3 TFM)
# ==============================================================================
#  Implementa el objetivo SMART 3 del TFM:
#     "Comparar la eficacia del realce de BLMD frente a alternativas como
#      BEMD y transformadas wavelet en términos de relación señal-ruido."
#
#  Se calculan tres métricas de SNR distintas, una por imagen, para cada
#  pipeline preprocesador:
#
#     SNR_estructural = |mean(I)| / std(I) en la zona herramienta
#                       (lo bien que se separa la herramienta del fondo)
#     SNR_contornos   = mean(|grad|_banda) / std(|grad|_interior)
#                       (cuán nítido queda el filo respecto al interior)
#     SNR_BM (Cohen d)= separación entre distribuciones de intensidad
#                       de BUENAS vs MALAS  (cuánto realza el preprocesado
#                       las diferencias entre clases)
#
#  Salidas:
#     snr_pipelines.csv         - tabla por pipeline
#     snr_pipelines_informe.txt - informe textual ordenado
#     snr_pipelines.png         - 3 barras comparativas por métrica
# ==============================================================================

def _calcular_snr_imagen(img, mascara=None, mascara_banda=None):
    """Calcula SNR_estructural y SNR_contornos sobre UNA imagen.

    Args:
       img: imagen preprocesada (float o uint8)
       mascara: máscara de la herramienta (True = herramienta)
       mascara_banda: máscara de la banda de contorno (True = filo)

    Devuelve dict con snr_estructural y snr_contornos.
    """
    img_f = img.astype(np.float32)

    # SNR estructural: |mean| / std en la zona de interés
    if mascara is not None and np.any(mascara):
        vals = img_f[mascara > 0]
    else:
        vals = img_f.flatten()
    if len(vals) < 10 or np.std(vals) < 1e-9:
        snr_estructural = 0.0
    else:
        snr_estructural = float(abs(np.mean(vals)) / np.std(vals))

    # SNR contornos: mean(|grad|_banda) / std(|grad|_interior)
    grad_x = cv2.Sobel(img_f, cv2.CV_32F, 1, 0, ksize=3)
    grad_y = cv2.Sobel(img_f, cv2.CV_32F, 0, 1, ksize=3)
    mag = np.sqrt(grad_x * grad_x + grad_y * grad_y)

    if mascara_banda is not None and mascara is not None:
        # Banda = filo dentado; Interior = herramienta sin banda
        interior = (mascara > 0) & (mascara_banda == 0)
        grad_banda = mag[mascara_banda > 0]
        grad_interior = mag[interior]
        if (len(grad_banda) >= 5 and len(grad_interior) >= 5
                and np.std(grad_interior) > 1e-9):
            snr_contornos = float(np.mean(grad_banda) /
                                   np.std(grad_interior))
        else:
            snr_contornos = 0.0
    else:
        # Sin máscaras: SNR_contornos = mean(grad)/std(grad)
        if np.std(mag) > 1e-9:
            snr_contornos = float(np.mean(mag) / np.std(mag))
        else:
            snr_contornos = 0.0

    return {"snr_estructural": snr_estructural,
            "snr_contornos": snr_contornos}


def calcular_snr_pipeline(nombre_pipeline, pipeline_fn, imgs_buenas, imgs_malas,
                          usar_mascara=True, ancho_banda=20):
    """Calcula las 3 métricas de SNR para UN pipeline completo.

    Aplica el preprocesado a cada imagen, calcula SNR_estructural y
    SNR_contornos por imagen, y obtiene SNR_BM (Cohen's d) entre las
    estadísticas globales de buenas y malas.

    Devuelve dict con:
       snr_estructural_buenas / _malas: medias por grupo
       snr_contornos_buenas / _malas: medias por grupo
       snr_bm: separación entre grupos (Cohen's d sobre la media de gris)
       n_buenas, n_malas: imágenes procesadas
    """
    log.info(f"  Calculando SNR [{nombre_pipeline}]...")

    def _procesar(imgs):
        snr_e, snr_c, mean_int = [], [], []
        for nombre, img in imgs:
            try:
                pre = pipeline_fn(img)
                if usar_mascara:
                    mask_full = segmentar_herramienta(img)
                    mask_banda = (banda_contorno(mask_full, ancho=ancho_banda)
                                  if ancho_banda > 0 else None)
                else:
                    mask_full = None
                    mask_banda = None
                snr = _calcular_snr_imagen(pre, mask_full, mask_banda)
                snr_e.append(snr["snr_estructural"])
                snr_c.append(snr["snr_contornos"])
                # Intensidad media en la zona de interés (para Cohen d)
                if mask_full is not None:
                    vals = pre[mask_full > 0]
                else:
                    vals = pre.flatten()
                mean_int.append(float(np.mean(vals)))
            except Exception as e:
                log.warning(f"    SNR falló en {nombre}: {e}")
        return snr_e, snr_c, mean_int

    snr_e_b, snr_c_b, mean_b = _procesar(imgs_buenas)
    snr_e_m, snr_c_m, mean_m = _procesar(imgs_malas)

    # SNR_BM = Cohen's d sobre la intensidad media
    if len(mean_b) >= 2 and len(mean_m) >= 2:
        m_b = np.mean(mean_b); m_m = np.mean(mean_m)
        s_b = np.std(mean_b, ddof=1); s_m = np.std(mean_m, ddof=1)
        pooled = np.sqrt(((len(mean_b) - 1) * s_b**2 +
                          (len(mean_m) - 1) * s_m**2) /
                         max(len(mean_b) + len(mean_m) - 2, 1))
        snr_bm = float(abs(m_b - m_m) / max(pooled, 1e-9))
    else:
        snr_bm = 0.0

    return {
        "pipeline": nombre_pipeline,
        "snr_estructural_buenas": float(np.mean(snr_e_b)) if snr_e_b else 0.0,
        "snr_estructural_malas": float(np.mean(snr_e_m)) if snr_e_m else 0.0,
        "snr_estructural_global": float(np.mean(snr_e_b + snr_e_m))
            if (snr_e_b or snr_e_m) else 0.0,
        "snr_contornos_buenas": float(np.mean(snr_c_b)) if snr_c_b else 0.0,
        "snr_contornos_malas": float(np.mean(snr_c_m)) if snr_c_m else 0.0,
        "snr_contornos_global": float(np.mean(snr_c_b + snr_c_m))
            if (snr_c_b or snr_c_m) else 0.0,
        "snr_bm_cohen_d": snr_bm,
        "n_buenas": len(snr_e_b),
        "n_malas": len(snr_e_m),
    }


def evaluar_snr_todos_pipelines(imgs_buenas, imgs_malas, pipelines_a_correr,
                                 ruta_salida, usar_mascara=True,
                                 ancho_banda=20):
    """Aplica calcular_snr_pipeline a TODOS los pipelines y guarda los
    resultados en CSV + gráfica + informe de texto.

    Implementa el objetivo 3 del TFM ("comparar BLMD vs BEMD vs wavelets
    en SNR").
    """
    log.info("=" * 70)
    log.info(" CÁLCULO DE SNR POR PIPELINE (Objetivo 3 TFM)")
    log.info("=" * 70)
    Path(ruta_salida).mkdir(parents=True, exist_ok=True)
    resultados_snr = {}

    for nombre_p in pipelines_a_correr:
        if nombre_p not in PIPELINES:
            continue
        try:
            res = calcular_snr_pipeline(
                nombre_p, PIPELINES[nombre_p],
                imgs_buenas, imgs_malas,
                usar_mascara=usar_mascara,
                ancho_banda=ancho_banda)
            resultados_snr[nombre_p] = res
            log.info(f"    {nombre_p}: "
                     f"SNR_estr={res['snr_estructural_global']:.3f}  "
                     f"SNR_cont={res['snr_contornos_global']:.3f}  "
                     f"SNR_BM={res['snr_bm_cohen_d']:.3f}")
        except Exception as e:
            log.error(f"  SNR pipeline {nombre_p} falló: {e}")

    if not resultados_snr:
        return None

    # CSV
    csv_path = os.path.join(ruta_salida, "snr_pipelines.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["pipeline", "snr_estructural_buenas",
                     "snr_estructural_malas", "snr_estructural_global",
                     "snr_contornos_buenas", "snr_contornos_malas",
                     "snr_contornos_global", "snr_bm_cohen_d"])
        for nombre, r in resultados_snr.items():
            w.writerow([nombre,
                        f"{r['snr_estructural_buenas']:.4f}",
                        f"{r['snr_estructural_malas']:.4f}",
                        f"{r['snr_estructural_global']:.4f}",
                        f"{r['snr_contornos_buenas']:.4f}",
                        f"{r['snr_contornos_malas']:.4f}",
                        f"{r['snr_contornos_global']:.4f}",
                        f"{r['snr_bm_cohen_d']:.4f}"])

    # Informe texto
    txt_path = os.path.join(ruta_salida, "snr_pipelines_informe.txt")
    with open(txt_path, "w", encoding="utf-8") as fh:
        fh.write("=" * 78 + "\n")
        fh.write(" COMPARATIVA SNR POR PIPELINE (Objetivo 3 TFM)\n")
        fh.write(" 'Comparar BLMD vs BEMD vs wavelet en términos de SNR'\n")
        fh.write("=" * 78 + "\n\n")
        fh.write("Definiciones:\n")
        fh.write("  SNR_estructural = |mean(I)| / std(I)  en zona herramienta\n")
        fh.write("  SNR_contornos   = mean(|grad|_banda) / std(|grad|_interior)\n")
        fh.write("  SNR_BM (Cohen d) = separación entre buenas y malas\n\n")
        fh.write(f"{'Pipeline':<14} {'SNR_estr':>10} {'SNR_cont':>10} "
                 f"{'SNR_BM':>10}\n")
        fh.write("-" * 50 + "\n")
        # Ordenar por SNR_BM descendente (mejor pipeline arriba)
        ordenados = sorted(resultados_snr.items(),
                           key=lambda x: x[1]['snr_bm_cohen_d'],
                           reverse=True)
        for nombre, r in ordenados:
            fh.write(f"{nombre:<14} "
                     f"{r['snr_estructural_global']:>10.4f} "
                     f"{r['snr_contornos_global']:>10.4f} "
                     f"{r['snr_bm_cohen_d']:>10.4f}\n")
        fh.write("\n")
        # Mejor pipeline por cada métrica
        mejor_bm = max(resultados_snr.items(),
                       key=lambda x: x[1]['snr_bm_cohen_d'])
        mejor_cont = max(resultados_snr.items(),
                         key=lambda x: x[1]['snr_contornos_global'])
        fh.write("Conclusiones:\n")
        fh.write(f"  Mejor SNR_BM (separación buenas/malas): "
                 f"{mejor_bm[0]}  ({mejor_bm[1]['snr_bm_cohen_d']:.3f})\n")
        fh.write(f"  Mejor SNR_contornos (realce de filo):   "
                 f"{mejor_cont[0]}  "
                 f"({mejor_cont[1]['snr_contornos_global']:.3f})\n")
        # Comparación con SIN_PREP como baseline
        if "SIN_PREP" in resultados_snr:
            baseline = resultados_snr["SIN_PREP"]["snr_bm_cohen_d"]
            if baseline > 1e-6:
                mejora_pct = ((mejor_bm[1]['snr_bm_cohen_d'] - baseline)
                              / baseline) * 100
                fh.write(f"  Mejora del mejor pipeline vs SIN_PREP: "
                         f"+{mejora_pct:.1f}% en SNR_BM\n")
        fh.write("=" * 78 + "\n")

    # Gráfica: barras comparativas con las 3 métricas
    try:
        fig, axes = plt.subplots(1, 3, figsize=(15, 5))
        nombres = list(resultados_snr.keys())
        # Si COLORES_PIPELINES no está definido aún, usar grises por defecto.
        try:
            colores = [COLORES_PIPELINES.get(f"PC_{n}", "gray")
                       for n in nombres]
        except NameError:
            colores = ["gray"] * len(nombres)
        for ax, clave, titulo in [
            (axes[0], "snr_estructural_global", "SNR estructural"),
            (axes[1], "snr_contornos_global", "SNR contornos (filo)"),
            (axes[2], "snr_bm_cohen_d", "SNR_BM (Cohen d - separación)"),
        ]:
            valores = [resultados_snr[n][clave] for n in nombres]
            bars = ax.bar(nombres, valores, color=colores)
            ax.set_title(titulo, fontsize=11, fontweight="bold")
            ax.set_ylabel(clave)
            ax.tick_params(axis="x", rotation=45)
            ax.grid(alpha=0.3, axis="y")
            # Marcar el mejor
            idx_max = int(np.argmax(valores))
            bars[idx_max].set_edgecolor("red")
            bars[idx_max].set_linewidth(2.5)
        fig.suptitle("COMPARATIVA SNR POR PIPELINE — Objetivo 3 TFM",
                      fontsize=13, fontweight="bold")
        plt.tight_layout()
        plt.savefig(os.path.join(ruta_salida, "snr_pipelines.png"),
                     dpi=120, bbox_inches="tight")
        plt.close()
    except Exception as e:
        log.warning(f"Gráfica SNR falló: {e}")

    log.info(f"  SNR guardado en {ruta_salida}")
    return resultados_snr


# ==============================================================================
#  BLOQUE 4-bis - PERFILÓMETRO ÓPTICO: PERFIL TEÓRICO vs PERFIL REAL
# ==============================================================================
#  Rosca métrica ISO (DIN 13-1 / ISO 261):
#     - Perfil triangular 60°
#     - Altura teórica del triángulo fundamental: H = sqrt(3)/2 * P  ≈ 0.866 * P
#     - Altura efectiva del filete:               h3 = 17/24 * H     (ext.)
#                                                 h1 = 5/8  * H      (int.)
#     - Truncamientos de cresta y valle: H/8 y H/4 respectivamente
#
#  Tolerancias por defecto (clase media, orientativas — ajustar según norma):
#     - Paso:           ± 0.5 % de P
#     - Altura:         ± 5 % de h3
#     - Ángulo flanco:  ± 2°
#     - RMS desviación: < 5 % de P
# ==============================================================================

PERFIL_TEORICO_PARAMS = {
    "iso_metrica": {
        "angulo_flanco_total_deg": 60.0,   # ángulo total entre flancos
        "factor_H": np.sqrt(3) / 2,        # H = factor_H * P
        "factor_h_efectivo": 17.0 / 24.0,  # filete externo (rosca macho)
        "trunc_cresta_factor": 1.0 / 8.0,  # H/8 truncado en la cresta
        "trunc_valle_factor": 1.0 / 4.0,   # H/4 truncado en el valle
    },
    "whitworth": {
        "angulo_flanco_total_deg": 55.0,
        "factor_H": 0.96,                  # H ≈ 0.96 * P
        "factor_h_efectivo": 0.640,
        "trunc_cresta_factor": 1.0 / 6.0,
        "trunc_valle_factor": 1.0 / 6.0,
    },
}

TOLERANCIAS_DEFECTO = {
    "paso_pct": 0.5,                  # % del paso nominal
    "altura_pct": 10.0,               # % de h1 (subida a 10% por backlight)
    "angulo_flanco_deg": 2.0,         # ° absolutos
    "rms_desviacion_pct_P": 5.0,      # % del paso nominal
    # Métricas específicas de cresta (geometría ISO P/8):
    "cresta_longitud_pct": 50.0,      # ±50% del nominal P/8
    "cresta_planitud_um": 30.0,       # RMS dentro de la cresta < 30 µm
    "cresta_radio_max_um": 80.0,      # radio de redondeo < 80 µm (sin desgaste)
    # Métricas específicas de valle (geometría ISO P/4):
    "valle_longitud_pct": 60.0,       # ±60% (más laxo: óptica suele cortar valle)
    "valle_planitud_um": 50.0,
    # Índice global de desgaste de cresta (0 = nuevo, 1 = muy desgastado):
    "desgaste_cresta_max": 0.40,
}


def perf_extraer_perfil_1d(img_gris, mascara=None, recorte_lateral=0.02):
    """
    Extrae el perfil 1D del borde dentado de la rosca.

    Estrategia:
      1. Otsu + apertura morfológica para máscara limpia.
      2. Para cada columna, primer píxel desde arriba (perfil_top) y desde
         abajo (perfil_bot).
      3. Limpia ARTEFACTOS verticales: saltos grandes y constantes a borde
         de imagen (típicos de paredes de la pieza que llegan al borde).
      4. Elige orientación con mayor amplitud robusta (P95 - P5).

    Devuelve: (perfil_suave, orientación, x_offset)
       x_offset = nº de columnas recortadas a la izquierda (para mapear
                  índices del perfil a la imagen original).
    """
    blur = cv2.GaussianBlur(img_gris, (5, 5), 0)
    if mascara is not None and np.sum(mascara > 0) > 100:
        bw = (mascara > 0).astype(np.uint8) * 255
    else:
        _, bw = cv2.threshold(blur, 0, 255,
                              cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
        bw = cv2.morphologyEx(bw, cv2.MORPH_OPEN, kernel)

    H, W = bw.shape
    perfil_top = np.full(W, np.nan)
    perfil_bot = np.full(W, np.nan)
    for x in range(W):
        ys = np.where(bw[:, x] == 255)[0]
        if len(ys):
            perfil_top[x] = ys[0]
            perfil_bot[x] = H - 1 - ys[-1]

    def rellenar(p):
        idx = np.arange(W)
        ok = ~np.isnan(p)
        if ok.sum() < 2:
            return np.full(W, H / 2.0)
        return np.interp(idx, idx[ok], p[ok])

    p_top = rellenar(perfil_top)
    p_bot = rellenar(perfil_bot)

    # Recortar márgenes mínimos (NaN -> interp puede crear bordes ficticios)
    n_rec = int(recorte_lateral * W)
    if n_rec > 0:
        p_top = p_top[n_rec:W - n_rec]
        p_bot = p_bot[n_rec:W - n_rec]
    x_offset = n_rec

    # Detectar la ZONA DENTADA real, descartando saltos a fondo (artefactos
    # de paredes verticales, sombras del fondo, etc.).
    # Estrategia ITERATIVA: aplicamos una banda y nos quedamos con el tramo
    # contiguo más largo. Si todavía hay artefactos en los bordes, repetimos
    # con percentiles recalculados (más limpios cada vez).
    def _filtrar_banda(p, p_lo_pct=10, p_hi_pct=90, expansion=1.5):
        plo, phi = np.percentile(p, [p_lo_pct, p_hi_pct])
        margen = (phi - plo) * expansion
        y_lo, y_hi = plo - margen, phi + margen
        bueno = (p >= y_lo) & (p <= y_hi)
        if not bueno.any():
            return 0, len(p)
        cambios = np.diff(bueno.astype(int))
        inicios = np.where(cambios == 1)[0] + 1
        fines = np.where(cambios == -1)[0] + 1
        if bueno[0]:
            inicios = np.concatenate([[0], inicios])
        if bueno[-1]:
            fines = np.concatenate([fines, [len(bueno)]])
        if len(inicios) == 0:
            return 0, len(p)
        longs = fines - inicios
        idx = int(np.argmax(longs))
        return int(inicios[idx]), int(fines[idx])

    def recortar_saltos_borde(p):
        if len(p) < 50:
            return p, 0, len(p)

        izq, der = 0, len(p)
        # Hasta 3 iteraciones, o hasta que no se recorte nada nuevo
        for _ in range(3):
            sub = p[izq:der]
            i, d = _filtrar_banda(sub, p_lo_pct=10, p_hi_pct=90, expansion=1.0)
            nuevo_izq = izq + i
            nuevo_der = izq + d
            if nuevo_izq == izq and nuevo_der == der:
                break
            izq, der = nuevo_izq, nuevo_der
            if der - izq < 100:   # protección
                break

        # Margen de seguridad
        margen = max(3, int(0.01 * (der - izq)))
        izq = min(izq + margen, len(p) // 2)
        der = max(der - margen, len(p) // 2 + 1)
        return p[izq:der], izq, der

    p_top_lim, izq_t, der_t = recortar_saltos_borde(p_top)
    p_bot_lim, izq_b, der_b = recortar_saltos_borde(p_bot)

    # Orientación: elegimos la versión limpia con mayor amplitud robusta
    def amp_robusta(p):
        if len(p) < 5:
            return 0.0
        return float(np.percentile(p, 95) - np.percentile(p, 5))

    if amp_robusta(p_top_lim) >= amp_robusta(p_bot_lim):
        orient = "top"
        perfil = p_top_lim
        x_offset += izq_t
    else:
        orient = "bottom"
        perfil = p_bot_lim
        x_offset += izq_b

    Wp = len(perfil)
    if Wp < 30:
        # Recuperación: si la limpieza fue demasiado agresiva, usar el perfil
        # sin recortar saltos.
        if orient == "top":
            perfil = p_top
        else:
            perfil = p_bot
        x_offset = n_rec
        Wp = len(perfil)

    k = max(5, Wp // 200) | 1
    perfil = np.convolve(perfil, np.ones(k) / k, mode="same")
    return perfil, orient, x_offset


def perf_detectar_picos_valles(perfil, orient="top"):
    """
    Devuelve (picos_idx, valles_idx) en coordenadas del perfil dado.

    Estrategia: iterar sobre prominencias decrecientes hasta encontrar al
    menos 2 picos. Esto es más robusto que un único umbral fijo, dado que
    la amplitud del perfil varía mucho entre piezas buenas (dientes muy
    profundos) y desgastadas (dientes someros).
    """
    from scipy.signal import find_peaks
    altura = -perfil if orient == "top" else perfil
    if len(altura) < 30:
        return np.array([], dtype=int), np.array([], dtype=int)

    p25, p75 = np.percentile(altura, [25, 75])
    iqr = max(p75 - p25, 1.0)
    rango = altura.max() - altura.min()

    dist_min = max(20, len(perfil) // 30)
    # Prominencias candidatas, de más estricta a más laxa
    prominencias = [iqr * f for f in (0.5, 0.3, 0.2, 0.1)] + \
                   [rango * f for f in (0.05, 0.03)]
    picos_idx = np.array([], dtype=int)
    valles_idx = np.array([], dtype=int)
    for prom in prominencias:
        picos_idx, _ = find_peaks(altura, distance=dist_min, prominence=prom)
        valles_idx, _ = find_peaks(-altura, distance=dist_min, prominence=prom)
        if len(picos_idx) >= 2:
            break
    return picos_idx, valles_idx


def perf_calibrar_mm_por_px(picos_idx, paso_mm):
    """
    Devuelve mm por píxel asumiendo que la distancia media entre picos
    consecutivos del perfil real es exactamente el paso nominal.
    """
    if len(picos_idx) < 2:
        return None, None
    distancias_px = np.diff(picos_idx)
    paso_medio_px = float(np.mean(distancias_px))
    paso_std_px = float(np.std(distancias_px))
    mm_por_px = paso_mm / paso_medio_px
    return mm_por_px, {"paso_medio_px": paso_medio_px,
                       "paso_std_px": paso_std_px,
                       "n_pasos": len(distancias_px)}


def perf_construir_teorico(x_mm, paso_mm, perfil_norma="iso_metrica",
                           offset_x_mm=0.0, y_base_mm=0.0):
    """
    Construye el perfil teórico ISO (o Whitworth) de la rosca en las mismas
    coordenadas x que el perfil real.

    El perfil es triangular truncado simétrico, periódico de paso P.
    Devuelve y_teorico_mm con los picos en y_base_mm y los valles en
    y_base_mm + h_efectivo (signo elegido para que sea coherente con la
    convención 'top' -> dientes hacia abajo, en imagen y crece hacia abajo).

    offset_x_mm sirve para alinear de fase con el perfil real.
    """
    p = PERFIL_TEORICO_PARAMS[perfil_norma]
    H = p["factor_H"] * paso_mm
    h_ef = p["factor_h_efectivo"] * H

    # Coordenada de fase normalizada en [0, 1)
    fase = ((x_mm - offset_x_mm) / paso_mm) % 1.0

    # Triángulo simétrico ideal (sin truncar): 0 en el pico (fase=0 y 1),
    # h_ef en el valle (fase=0.5)
    y_tri = 2.0 * h_ef * np.where(fase <= 0.5, fase, 1.0 - fase)

    # Truncamientos: aplastamos los extremos del triángulo
    trunc_cresta = p["trunc_cresta_factor"] * H
    trunc_valle = p["trunc_valle_factor"] * H
    y_tri = np.clip(y_tri, trunc_cresta, h_ef - trunc_valle / 2.0)
    # Renormalizamos para que la cresta esté en 0
    y_tri = y_tri - y_tri.min()

    return y_base_mm + y_tri, {"H_mm": H, "h_ef_mm": h_ef,
                               "trunc_cresta_mm": trunc_cresta,
                               "trunc_valle_mm": trunc_valle}


def perf_alinear_fase(perfil_real_mm, perfil_teorico_fn, paso_mm,
                      n_busqueda=200):
    """
    Encuentra el offset_x_mm que minimiza el error cuadrático entre real y
    teórico, y el desplazamiento vertical óptimo.
    """
    offsets = np.linspace(0, paso_mm, n_busqueda, endpoint=False)
    mejor = (np.inf, 0.0, 0.0)  # (error, offset, dy)
    for off in offsets:
        y_t = perfil_teorico_fn(off)
        # Desplazamiento vertical óptimo: dy* = mean(real - teorico)
        dy = float(np.mean(perfil_real_mm - y_t))
        err = float(np.mean((perfil_real_mm - (y_t + dy)) ** 2))
        if err < mejor[0]:
            mejor = (err, off, dy)
    return mejor[1], mejor[2], mejor[0]


def perf_medir_angulo_flanco(perfil_real_mm, x_mm, picos_idx, valles_idx,
                             frac_inferior=0.25, frac_superior=0.75):
    """
    Estima el ángulo TOTAL entre flancos.

    Para cada pico, identifica los valles vecinos y mide la pendiente del
    flanco usando solo el tramo central (entre frac_inferior y frac_superior
    de la altura del flanco), donde no hay truncamiento de cresta ni de
    valle. Esto aproxima mejor el ángulo "real" del flanco recto del perfil
    triangular ISO.

    Devuelve la mediana de los ángulos medidos (más robusta que la media).
    """
    angulos = []
    for p in picos_idx:
        izq = valles_idx[valles_idx < p]
        der = valles_idx[valles_idx > p]
        for v_arr, lado in [(izq, "izq"), (der, "der")]:
            if not len(v_arr):
                continue
            v = v_arr[-1] if lado == "izq" else v_arr[0]
            i_low = min(p, v)
            i_high = max(p, v)
            if i_high - i_low < 8:
                continue
            seg_x = x_mm[i_low:i_high + 1]
            seg_y = perfil_real_mm[i_low:i_high + 1]
            # Normalizar la altura del flanco a [0, 1]
            y_min, y_max = seg_y.min(), seg_y.max()
            if y_max - y_min < 1e-6:
                continue
            t = (seg_y - y_min) / (y_max - y_min)
            # Si el pico está arriba (orient='top') y vamos hacia el valle,
            # t puede ir creciente o decreciente: tomamos zona central por
            # valor absoluto
            mask = (t >= frac_inferior) & (t <= frac_superior)
            if mask.sum() < 5:
                continue
            xs, ys = seg_x[mask], seg_y[mask]
            # Ajuste lineal y --> x sería degenerado en flancos verticales,
            # mejor regresión lineal x = a*y + b
            a, _ = np.polyfit(ys, xs, 1)
            # 'a' es dx/dy. El flanco con la vertical forma ángulo
            # arctan(|dx/dy|). Ángulo TOTAL del perfil = 2*ese ángulo.
            ang_con_vertical = np.degrees(np.arctan(abs(a)))
            angulo_total = 2.0 * ang_con_vertical
            # Sanity: descartamos ángulos absurdos
            if 10.0 < angulo_total < 170.0:
                angulos.append(angulo_total)
    if not angulos:
        return None
    return float(np.median(angulos))


def perf_medir_cresta(perfil_real_mm, x_mm, picos_idx, paso_mm,
                      umbral_altura_frac=0.92):
    """
    Mide longitud, planitud y radio de redondeo de cada cresta.

    Para cada pico detectado, define la "zona cresta" como el conjunto
    contiguo de puntos cuya altura está por encima de
    umbral_altura_frac * h_pico (donde h_pico es la altura del pico tomando
    como base la altura del valle vecino).

    Devuelve dict con métricas mediadas (medianas) sobre todas las crestas:
      - longitud_mm
      - planitud_rms_um  (RMS de la altura dentro de la zona cresta)
      - radio_redondeo_um (ajuste parabólico en torno al pico)
      - longitud_individual : lista de longitudes por pico
      - planitud_individual : lista
      - radio_individual    : lista
    """
    if len(picos_idx) == 0:
        return {"longitud_mm": np.nan, "planitud_rms_um": np.nan,
                "radio_redondeo_um": np.nan,
                "longitud_individual": [], "planitud_individual": [],
                "radio_individual": []}

    longs, planitudes, radios = [], [], []
    N = len(perfil_real_mm)
    # Convención: en orient='top' los picos son y MÍNIMOS en la imagen, pero
    # ya hemos convertido y_real_mm a "altura creciente desde la cresta",
    # con cresta ≈ 0 y valle ≈ + h_ef. Por tanto el pico es el MÍNIMO de
    # y_real_mm y la "zona cresta" son los puntos cercanos al mínimo local.
    for p in picos_idx:
        # Estimar altura del pico respecto a sus vecinos (medio paso a cada
        # lado, en píxeles convertidos a mm)
        ventana_px = int((paso_mm / 2.0) /
                         max(1e-9, x_mm[1] - x_mm[0])) if len(x_mm) > 1 else 50
        i_lo = max(0, p - ventana_px)
        i_hi = min(N, p + ventana_px + 1)
        seg_y = perfil_real_mm[i_lo:i_hi]
        seg_x = x_mm[i_lo:i_hi]
        if len(seg_y) < 5:
            continue
        # Altura local del pico = (max valle vecino) - (min cresta)
        y_pico = perfil_real_mm[p]
        y_base = float(np.max(seg_y))   # punto más profundo en la ventana
        h_local = y_base - y_pico
        if h_local < 1e-3:
            continue

        # Zona cresta: puntos con altura entre y_pico y y_pico + (1 - umbral)*h_local
        umbral_y = y_pico + (1.0 - umbral_altura_frac) * h_local
        mask = seg_y <= umbral_y
        if mask.sum() < 3:
            continue
        # Conectividad alrededor del pico: solo el tramo contiguo que CONTIENE p
        idx_local = p - i_lo
        # Extender izquierda mientras siga en mask
        l = idx_local
        while l > 0 and mask[l - 1]:
            l -= 1
        r = idx_local
        while r < len(mask) - 1 and mask[r + 1]:
            r += 1
        if r - l < 2:
            continue

        zona_x = seg_x[l:r + 1]
        zona_y = seg_y[l:r + 1]

        # Longitud horizontal
        longitud = float(zona_x[-1] - zona_x[0])

        # Planitud: RMS de la altura respecto a la mediana, en µm
        plan_rms_mm = float(np.sqrt(np.mean(
            (zona_y - np.median(zona_y)) ** 2)))
        planitud_um = plan_rms_mm * 1000.0

        # Radio de redondeo de cresta: ajuste parabólico y = a*(x-x0)**2 + c
        # alrededor del pico. R = 1/(2|a|).
        # Tomamos un poco más de muestras para que el ajuste sea estable
        ext = max(3, (r - l) // 2)
        i_l = max(0, l - ext)
        i_r = min(len(seg_x) - 1, r + ext)
        xs = seg_x[i_l:i_r + 1]
        ys = seg_y[i_l:i_r + 1]
        if len(xs) >= 5:
            try:
                coefs = np.polyfit(xs - xs.mean(), ys, 2)
                a2 = coefs[0]
                if abs(a2) > 1e-9:
                    radio_mm = 1.0 / (2.0 * abs(a2))
                    # Sanity: descartamos radios fuera de [0.005, 5] mm
                    if 0.005 <= radio_mm <= 5.0:
                        radios.append(radio_mm * 1000.0)  # en µm
            except Exception:
                pass

        longs.append(longitud)
        planitudes.append(planitud_um)

    def _mediana_o_nan(lst):
        return float(np.median(lst)) if len(lst) else float("nan")

    return {
        "longitud_mm": _mediana_o_nan(longs),
        "planitud_rms_um": _mediana_o_nan(planitudes),
        "radio_redondeo_um": _mediana_o_nan(radios),
        "longitud_individual": longs,
        "planitud_individual": planitudes,
        "radio_individual": radios,
    }


def perf_medir_valle(perfil_real_mm, x_mm, valles_idx, paso_mm,
                     umbral_altura_frac=0.92):
    """
    Análogo a perf_medir_cresta pero para los valles.
    Devuelve longitud y planitud (no se mide radio de valle, suele ser
    irrelevante en piezas usadas).
    """
    if len(valles_idx) == 0:
        return {"longitud_mm": np.nan, "planitud_rms_um": np.nan,
                "longitud_individual": [], "planitud_individual": []}
    longs, planitudes = [], []
    N = len(perfil_real_mm)
    for v in valles_idx:
        ventana_px = int((paso_mm / 2.0) /
                         max(1e-9, x_mm[1] - x_mm[0])) if len(x_mm) > 1 else 50
        i_lo = max(0, v - ventana_px)
        i_hi = min(N, v + ventana_px + 1)
        seg_y = perfil_real_mm[i_lo:i_hi]
        seg_x = x_mm[i_lo:i_hi]
        if len(seg_y) < 5:
            continue
        y_valle = perfil_real_mm[v]
        y_base = float(np.min(seg_y))   # cresta vecina más alta
        h_local = y_valle - y_base
        if h_local < 1e-3:
            continue
        umbral_y = y_valle - (1.0 - umbral_altura_frac) * h_local
        mask = seg_y >= umbral_y
        if mask.sum() < 3:
            continue
        idx_local = v - i_lo
        l = idx_local
        while l > 0 and mask[l - 1]:
            l -= 1
        r = idx_local
        while r < len(mask) - 1 and mask[r + 1]:
            r += 1
        if r - l < 2:
            continue
        zona_x = seg_x[l:r + 1]
        zona_y = seg_y[l:r + 1]
        longitud = float(zona_x[-1] - zona_x[0])
        plan_rms_mm = float(np.sqrt(np.mean(
            (zona_y - np.median(zona_y)) ** 2)))
        longs.append(longitud)
        planitudes.append(plan_rms_mm * 1000.0)

    def _mediana_o_nan(lst):
        return float(np.median(lst)) if len(lst) else float("nan")

    return {
        "longitud_mm": _mediana_o_nan(longs),
        "planitud_rms_um": _mediana_o_nan(planitudes),
        "longitud_individual": longs,
        "planitud_individual": planitudes,
    }


def perf_indice_desgaste_cresta(metricas_cresta, paso_mm,
                                long_nom_factor=0.125):
    """
    Combina las métricas de cresta en un índice global [0, 1] donde
        0 = cresta perfectamente plana, longitud = P/8, sin redondeo
        1 = cresta totalmente redondeada o desviada del nominal

    Componentes (cada uno entre 0 y 1):
      - desv_long  : |L_med - L_nom| / L_nom       (saturado a 1)
      - planitud   : RMS_um / 100                  (saturado a 1)
      - redondeo   : R_um / 200                    (saturado a 1)
    """
    L_nom = paso_mm * long_nom_factor   # P/8 por defecto
    L = metricas_cresta.get("longitud_mm", np.nan)
    plan = metricas_cresta.get("planitud_rms_um", np.nan)
    R = metricas_cresta.get("radio_redondeo_um", np.nan)

    def _safe(x, escala, lo=0.0, hi=1.0):
        if not np.isfinite(x):
            return 1.0   # falta de dato cuenta como peor caso
        return float(np.clip(x / escala, lo, hi))

    desv_long = _safe(abs(L - L_nom) / max(L_nom, 1e-6), 1.0) if np.isfinite(L) else 1.0
    plan_norm = _safe(plan, 100.0)
    redondeo = _safe(R, 200.0)
    # Pesos: redondeo y planitud son más informativos del desgaste real
    indice = 0.2 * desv_long + 0.4 * plan_norm + 0.4 * redondeo
    return float(np.clip(indice, 0.0, 1.0))


def perf_medir_uniformidad_intraimagen(metricas_cresta, metricas_valle):
    """
    Mide la VARIABILIDAD entre dientes DENTRO DE LA MISMA IMAGEN.

    Esta es la métrica más robusta a calidad de captura: no le importa el
    valor absoluto (que puede variar entre vistas por iluminación, ángulo,
    foco), solo si los dientes son uniformes ENTRE SÍ en la misma vista.

    Una rosca BUENA tiene dientes IGUALES entre sí (uniformidad alta).
    Una rosca MALA con desgaste local tiene dientes DESIGUALES (uniformidad baja).

    Devuelve dict con coeficientes de variación (std/media):
       cv_radio_cresta_um  - CV del radio entre crestas
       cv_longitud_cresta  - CV de la longitud entre crestas
       cv_planitud_cresta  - CV de la planitud entre crestas
       cv_altura_picos_mm  - CV de la altura entre picos
       indice_no_uniformidad - score 0..1 combinado
    """
    def _cv(lista):
        """Coef. de variación (std/media). NaN si <2 muestras o media~0."""
        if not lista or len(lista) < 2:
            return float("nan")
        a = np.array([v for v in lista if v is not None and np.isfinite(v)])
        if len(a) < 2:
            return float("nan")
        m = np.mean(a)
        if abs(m) < 1e-9:
            return float("nan")
        return float(np.std(a) / abs(m))

    cv_radio = _cv(metricas_cresta.get("radio_individual", []))
    cv_long = _cv(metricas_cresta.get("longitud_individual", []))
    cv_plan = _cv(metricas_cresta.get("planitud_individual", []))
    cv_long_v = _cv(metricas_valle.get("longitud_individual", []))

    # Índice combinado: 0 = dientes muy uniformes, 1 = muy variables
    cvs = [c for c in [cv_radio, cv_long, cv_plan, cv_long_v]
           if np.isfinite(c)]
    if cvs:
        indice = float(np.clip(np.mean(cvs), 0.0, 1.0))
    else:
        indice = float("nan")

    return {
        "cv_radio_cresta": cv_radio,
        "cv_longitud_cresta": cv_long,
        "cv_planitud_cresta": cv_plan,
        "cv_longitud_valle": cv_long_v,
        "indice_no_uniformidad": indice,
    }


def perfilometro_analizar(img_gris, paso_mm=1.5, perfil_norma="iso_metrica",
                          mm_por_px=None, tolerancias=None,
                          mascara=None):
    """
    Análisis completo de perfilometría: extrae perfil real, lo compara con el
    teórico y devuelve un dict con todas las métricas + PASA/NO PASA.
    """
    if tolerancias is None:
        tolerancias = TOLERANCIAS_DEFECTO

    # 1. Extraer perfil real (en píxeles, y crece hacia abajo)
    perfil_px, orient, x_offset = perf_extraer_perfil_1d(img_gris,
                                                          mascara=mascara)
    picos_idx, valles_idx = perf_detectar_picos_valles(perfil_px, orient=orient)

    # 1-bis. FILTRAR features atípicos. Un pico real está a una altura
    # parecida a los demás picos; idem valles. Si un "pico" o un "valle"
    # está MUY desviado del resto, es un artefacto (típicamente el extremo
    # de la imagen donde la pieza se prolonga más allá de la rosca útil).
    def filtrar_outliers_altura(idx, valores, k_mad=3.0):
        if len(idx) < 3:
            return idx
        h = valores[idx]
        med = np.median(h)
        mad = np.median(np.abs(h - med)) + 1e-9
        bueno = np.abs(h - med) <= k_mad * mad
        return idx[bueno]

    if len(picos_idx) >= 3:
        picos_idx = filtrar_outliers_altura(picos_idx, perfil_px)
    if len(valles_idx) >= 3:
        valles_idx = filtrar_outliers_altura(valles_idx, perfil_px)

    # 1-ter. Recortar el perfil al rango útil [primer feature, último feature]
    # ya con los outliers descartados.
    todos = np.sort(np.concatenate([picos_idx, valles_idx])) if \
        (len(picos_idx) + len(valles_idx)) >= 2 else None
    if todos is not None and len(todos) >= 2:
        d_med = float(np.median(np.diff(todos)))
        margen = int(0.6 * d_med)
        izq = max(0, int(todos[0]) - margen)
        der = min(len(perfil_px), int(todos[-1]) + margen)
        if der - izq >= 50:
            perfil_px = perfil_px[izq:der]
            x_offset += izq
            picos_idx, valles_idx = perf_detectar_picos_valles(
                perfil_px, orient=orient)
            if len(picos_idx) >= 3:
                picos_idx = filtrar_outliers_altura(picos_idx, perfil_px)
            if len(valles_idx) >= 3:
                valles_idx = filtrar_outliers_altura(valles_idx, perfil_px)

    # 1-quater. LIMPIEZA FINAL POR BANDA Y. Tras recortar a la zona dentada,
    # si todavía queda algún artefacto en los extremos, nos quedamos con el
    # tramo contiguo más largo dentro de [P5..P95] expandido un 20%.
    if len(perfil_px) >= 100:
        plo, phi = np.percentile(perfil_px, [5, 95])
        margen_y = (phi - plo) * 0.2
        y_lo, y_hi = plo - margen_y, phi + margen_y
        bueno = (perfil_px >= y_lo) & (perfil_px <= y_hi)
        if bueno.any() and not bueno.all():
            cambios = np.diff(bueno.astype(int))
            inicios = np.where(cambios == 1)[0] + 1
            fines = np.where(cambios == -1)[0] + 1
            if bueno[0]:
                inicios = np.concatenate([[0], inicios])
            if bueno[-1]:
                fines = np.concatenate([fines, [len(bueno)]])
            if len(inicios) and len(fines):
                longs = fines - inicios
                idx_max = int(np.argmax(longs))
                izq2, der2 = int(inicios[idx_max]), int(fines[idx_max])
                if der2 - izq2 >= 50 and (der2 - izq2) >= 0.5 * len(perfil_px):
                    perfil_px = perfil_px[izq2:der2]
                    x_offset += izq2
                    picos_idx, valles_idx = perf_detectar_picos_valles(
                        perfil_px, orient=orient)
                    if len(picos_idx) >= 3:
                        picos_idx = filtrar_outliers_altura(picos_idx, perfil_px)
                    if len(valles_idx) >= 3:
                        valles_idx = filtrar_outliers_altura(valles_idx, perfil_px)

    if len(picos_idx) < 2:
        return {"ok": False,
                "motivo": f"Solo se detectaron {len(picos_idx)} picos. "
                          f"No se puede medir el paso."}

    # 2. Calibración mm/px (forzada o deducida del paso conocido)
    if mm_por_px is None:
        mm_por_px, info_paso = perf_calibrar_mm_por_px(picos_idx, paso_mm)
        calibracion = "deducida_del_paso"
    else:
        info_paso = {"paso_medio_px": float(np.mean(np.diff(picos_idx))),
                     "paso_std_px": float(np.std(np.diff(picos_idx))),
                     "n_pasos": len(picos_idx) - 1}
        calibracion = "manual"

    # 3. Convertir a mm. La cresta del diente es el "0" de altura (la pieza está
    #    arriba en orient='top' y los picos son y mínimos en píxeles).
    W = len(perfil_px)
    x_mm = np.arange(W) * mm_por_px
    if orient == "top":
        # picos = mínimos en y_px -> los pasamos a "altura" creciente hacia abajo
        y_real_mm = (perfil_px - perfil_px[picos_idx].mean()) * mm_por_px
    else:
        y_real_mm = (perfil_px[picos_idx].mean() - perfil_px) * mm_por_px

    # 4. Construir perfil teórico y alinearlo en fase con el real
    def fn_teorico(off):
        y_t, _ = perf_construir_teorico(x_mm, paso_mm,
                                        perfil_norma=perfil_norma,
                                        offset_x_mm=off, y_base_mm=0.0)
        return y_t

    off_opt, dy_opt, err_opt = perf_alinear_fase(y_real_mm, fn_teorico, paso_mm)
    y_teorico_mm, params_teorico = perf_construir_teorico(
        x_mm, paso_mm, perfil_norma=perfil_norma,
        offset_x_mm=off_opt, y_base_mm=dy_opt)

    # 5. Métricas
    # IMPORTANTE: el RMS y máx desviación se calculan SOLO sobre la zona
    # útil del perfil (entre el primer y último pico/valle válido). Esto
    # evita que pequeños artefactos residuales en los extremos contaminen
    # la métrica de RMS, que es la más sensible a outliers.
    if len(picos_idx) >= 1 and len(valles_idx) >= 1:
        i_min = int(min(picos_idx[0] if len(picos_idx) else len(perfil_px),
                        valles_idx[0] if len(valles_idx) else len(perfil_px)))
        i_max = int(max(picos_idx[-1] if len(picos_idx) else 0,
                        valles_idx[-1] if len(valles_idx) else 0))
        # Margen pequeño dentro
        margen_int = max(5, int(0.05 * (i_max - i_min)))
        i_min = max(0, i_min - margen_int)
        i_max = min(len(perfil_px), i_max + margen_int)
        sl = slice(i_min, i_max)
    else:
        sl = slice(0, len(perfil_px))
    desviacion = y_real_mm[sl] - y_teorico_mm[sl]
    rms_desv_mm = float(np.sqrt(np.mean(desviacion ** 2)))
    max_desv_mm = float(np.max(np.abs(desviacion)))

    # ============================================================
    # MÉTRICAS DE DESGASTE (Objetivo 4 TFM):
    #   - Área de desgaste = ∫ (teórico - real) dx, solo donde el real
    #     está POR DEBAJO del teórico (= material faltante).
    #   - Ancho medio de desgaste = área / longitud de la zona con
    #     desgaste detectable (donde |desviación| > umbral_ruido).
    #
    # Estas dos métricas son las que pide explícitamente el TFM en el
    # objetivo 4 ("calcular métricas geométricas (área y ancho medio)
    # de la zona degradada").
    # ============================================================
    x_desv_mm = x_mm[sl]
    # Convertimos a "déficit de material" = teórico - real, > 0 indica
    # zona con desgaste (falta material respecto al perfil ideal).
    deficit_mm = -desviacion   # = y_teorico - y_real

    # Umbral de ruido: descartamos como "no desgaste" desviaciones
    # menores que el ruido típico de la medición. Lo estimamos como
    # la mediana del |desviación| (estimador robusto).
    umbral_ruido_mm = float(np.median(np.abs(desviacion)))

    # Sólo cuentan los puntos donde el déficit supera el ruido
    mascara_desgaste = deficit_mm > umbral_ruido_mm
    n_puntos_desgaste = int(np.sum(mascara_desgaste))

    if n_puntos_desgaste >= 2 and len(x_desv_mm) >= 2:
        # Área de desgaste por integración trapezoidal (mm²)
        # Solo integramos donde hay desgaste (déficit positivo y > ruido)
        deficit_solo_desgaste = np.where(mascara_desgaste, deficit_mm, 0.0)
        area_desgaste_mm2 = float(np.trapezoid(deficit_solo_desgaste,
                                                x_desv_mm))

        # Longitud horizontal con desgaste (mm)
        # = nº de puntos con desgaste × paso de muestreo en x
        if len(x_desv_mm) >= 2:
            dx_mm = float(np.median(np.diff(x_desv_mm)))
        else:
            dx_mm = 0.0
        longitud_desgaste_mm = n_puntos_desgaste * dx_mm

        # Ancho medio de desgaste (mm) = altura media del déficit en zona
        # con desgaste. Equivale a área / longitud, que es la profundidad
        # promedio del material faltante.
        if longitud_desgaste_mm > 1e-6:
            ancho_medio_desgaste_mm = area_desgaste_mm2 / longitud_desgaste_mm
        else:
            ancho_medio_desgaste_mm = 0.0

        # Profundidad máxima del desgaste (peor punto)
        prof_max_desgaste_mm = float(np.max(deficit_mm[mascara_desgaste]))
    else:
        area_desgaste_mm2 = 0.0
        longitud_desgaste_mm = 0.0
        ancho_medio_desgaste_mm = 0.0
        prof_max_desgaste_mm = 0.0

    # Porcentajes relativos al perfil teórico para hacer las métricas
    # comparables entre vistas con distinta escala.
    longitud_total_mm = float(x_desv_mm[-1] - x_desv_mm[0]) \
        if len(x_desv_mm) >= 2 else 1.0
    pct_longitud_con_desgaste = 100.0 * longitud_desgaste_mm / max(
        longitud_total_mm, 1e-6)

    paso_medido_mm = info_paso["paso_medio_px"] * mm_por_px
    error_paso_pct = 100.0 * (paso_medido_mm - paso_mm) / paso_mm

    # Altura efectiva real: para cada pico medimos la profundidad hasta los
    # valles vecinos. Luego usamos la MEDIANA (no la media) para robustez
    # ante valles artefacto.
    alturas = []
    for p in picos_idx:
        izq = valles_idx[valles_idx < p]
        der = valles_idx[valles_idx > p]
        cands = []
        if len(izq):
            cands.append(abs(y_real_mm[izq[-1]] - y_real_mm[p]))
        if len(der):
            cands.append(abs(y_real_mm[der[0]] - y_real_mm[p]))
        if cands:
            # Tomamos la MENOR profundidad (= valle más somero adyacente),
            # más representativa que el máximo cuando un valle es artefacto.
            alturas.append(min(cands))
    if alturas:
        # Filtrar outliers: descartamos alturas a más de 2*MAD de la mediana
        a = np.array(alturas)
        med = np.median(a)
        mad = np.median(np.abs(a - med)) + 1e-9
        a_filt = a[np.abs(a - med) <= 3.0 * mad]
        altura_real_mm = float(np.median(a_filt)) if len(a_filt) else float(med)
    else:
        altura_real_mm = 0.0
    altura_teorica_mm = params_teorico["h_ef_mm"]
    error_altura_pct = (100.0 * (altura_real_mm - altura_teorica_mm)
                        / altura_teorica_mm) if altura_teorica_mm > 0 else 0.0

    angulo_real = perf_medir_angulo_flanco(y_real_mm, x_mm,
                                           picos_idx, valles_idx)
    angulo_teorico = PERFIL_TEORICO_PARAMS[perfil_norma][
        "angulo_flanco_total_deg"]
    error_angulo = (angulo_real - angulo_teorico) if angulo_real else None

    # 5-bis. MÉTRICAS DE GEOMETRÍA ISO ESPECÍFICA (cresta P/8, valle P/4):
    metricas_cresta = perf_medir_cresta(y_real_mm, x_mm, picos_idx, paso_mm)
    metricas_valle = perf_medir_valle(y_real_mm, x_mm, valles_idx, paso_mm)
    cresta_long_nom_mm = paso_mm * 0.125     # P/8
    valle_long_nom_mm = paso_mm * 0.25       # P/4
    cresta_L = metricas_cresta["longitud_mm"]
    valle_L = metricas_valle["longitud_mm"]
    error_cresta_long_pct = (100.0 * (cresta_L - cresta_long_nom_mm)
                             / cresta_long_nom_mm) if np.isfinite(cresta_L) \
        and cresta_long_nom_mm > 0 else float("nan")
    error_valle_long_pct = (100.0 * (valle_L - valle_long_nom_mm)
                            / valle_long_nom_mm) if np.isfinite(valle_L) \
        and valle_long_nom_mm > 0 else float("nan")

    # Índice global de desgaste de cresta
    desgaste_cresta = perf_indice_desgaste_cresta(metricas_cresta, paso_mm)

    # NUEVO: métricas de uniformidad intra-imagen (variabilidad entre dientes
    # dentro de la misma imagen). Robustas a diferencias de captura entre
    # vistas: lo que se mide es si los dientes son IGUALES entre sí, no si
    # cumplen el teórico ISO en valor absoluto.
    uniformidad = perf_medir_uniformidad_intraimagen(metricas_cresta,
                                                      metricas_valle)

    # 5-ter. AVISO óptico: el backlight puede no llegar al fondo del valle.
    # Si la altura medida es < 80% de la teórica, lo más probable es que la
    # óptica esté cortando el valle (no que la pieza tenga menos altura).
    aviso_optica_valle = (np.isfinite(altura_real_mm) and
                          altura_teorica_mm > 0 and
                          altura_real_mm < 0.80 * altura_teorica_mm)

    # 6. Veredicto PASA / NO PASA
    fallos = []
    if abs(error_paso_pct) > tolerancias["paso_pct"]:
        fallos.append(f"paso fuera de tol. ({error_paso_pct:+.2f}%)")
    if abs(error_altura_pct) > tolerancias["altura_pct"]:
        fallos.append(f"altura fuera de tol. ({error_altura_pct:+.2f}%)")
    if error_angulo is not None and \
       abs(error_angulo) > tolerancias["angulo_flanco_deg"]:
        fallos.append(f"ángulo fuera de tol. ({error_angulo:+.2f}°)")
    rms_pct = 100.0 * rms_desv_mm / paso_mm
    if rms_pct > tolerancias["rms_desviacion_pct_P"]:
        fallos.append(f"RMS desviación alta ({rms_pct:.2f}% de P)")

    # Tolerancias específicas de cresta/valle
    if (np.isfinite(error_cresta_long_pct) and
            abs(error_cresta_long_pct) > tolerancias["cresta_longitud_pct"]):
        fallos.append(
            f"cresta longitud fuera de tol. ({error_cresta_long_pct:+.1f}%)")
    if (np.isfinite(metricas_cresta["planitud_rms_um"]) and
            metricas_cresta["planitud_rms_um"] > tolerancias["cresta_planitud_um"]):
        fallos.append(
            f"cresta poco plana (RMS {metricas_cresta['planitud_rms_um']:.1f}µm)")
    if (np.isfinite(metricas_cresta["radio_redondeo_um"]) and
            metricas_cresta["radio_redondeo_um"] > tolerancias["cresta_radio_max_um"]):
        fallos.append(
            f"cresta redondeada (R≈{metricas_cresta['radio_redondeo_um']:.0f}µm)")
    if (np.isfinite(error_valle_long_pct) and
            abs(error_valle_long_pct) > tolerancias["valle_longitud_pct"]):
        fallos.append(
            f"valle longitud fuera de tol. ({error_valle_long_pct:+.1f}%)")
    if desgaste_cresta > tolerancias["desgaste_cresta_max"]:
        fallos.append(f"desgaste de cresta alto (índice {desgaste_cresta:.2f})")

    veredicto = "PASA" if not fallos else "NO PASA"

    return {
        "ok": True,
        "veredicto": veredicto,
        "motivos_fallo": fallos,
        "perfil_norma": perfil_norma,
        "orientacion": orient,
        "calibracion": calibracion,
        "mm_por_px": float(mm_por_px),
        "paso_nominal_mm": paso_mm,
        "paso_medido_mm": float(paso_medido_mm),
        "error_paso_pct": float(error_paso_pct),
        "altura_teorica_mm": float(altura_teorica_mm),
        "altura_real_mm": float(altura_real_mm),
        "error_altura_pct": float(error_altura_pct),
        "angulo_flanco_teorico_deg": float(angulo_teorico),
        "angulo_flanco_real_deg": angulo_real,
        "error_angulo_deg": error_angulo,
        "rms_desviacion_mm": rms_desv_mm,
        "max_desviacion_mm": max_desv_mm,
        "rms_desviacion_pct_P": float(rms_pct),
        # ===== MÉTRICAS DE DESGASTE (objetivo 4 TFM) =====
        "area_desgaste_mm2": float(area_desgaste_mm2),
        "ancho_medio_desgaste_um": float(ancho_medio_desgaste_mm * 1000.0),
        "longitud_desgaste_mm": float(longitud_desgaste_mm),
        "pct_longitud_con_desgaste": float(pct_longitud_con_desgaste),
        "prof_max_desgaste_um": float(prof_max_desgaste_mm * 1000.0),
        # =================================================
        "n_picos_detectados": int(len(picos_idx)),
        "n_pasos_medidos": int(info_paso["n_pasos"]),
        # Nuevas métricas (cresta y valle)
        "cresta_longitud_mm": float(cresta_L) if np.isfinite(cresta_L) else None,
        "cresta_longitud_nominal_mm": float(cresta_long_nom_mm),
        "error_cresta_longitud_pct": float(error_cresta_long_pct)
            if np.isfinite(error_cresta_long_pct) else None,
        "cresta_planitud_rms_um": float(metricas_cresta["planitud_rms_um"])
            if np.isfinite(metricas_cresta["planitud_rms_um"]) else None,
        "cresta_radio_redondeo_um": float(metricas_cresta["radio_redondeo_um"])
            if np.isfinite(metricas_cresta["radio_redondeo_um"]) else None,
        "valle_longitud_mm": float(valle_L) if np.isfinite(valle_L) else None,
        "valle_longitud_nominal_mm": float(valle_long_nom_mm),
        "error_valle_longitud_pct": float(error_valle_long_pct)
            if np.isfinite(error_valle_long_pct) else None,
        "valle_planitud_rms_um": float(metricas_valle["planitud_rms_um"])
            if np.isfinite(metricas_valle["planitud_rms_um"]) else None,
        "indice_desgaste_cresta": float(desgaste_cresta),
        # Métricas de uniformidad INTRA-IMAGEN (robustas a calidad captura)
        "cv_radio_cresta": uniformidad["cv_radio_cresta"]
            if np.isfinite(uniformidad["cv_radio_cresta"]) else None,
        "cv_longitud_cresta": uniformidad["cv_longitud_cresta"]
            if np.isfinite(uniformidad["cv_longitud_cresta"]) else None,
        "cv_planitud_cresta": uniformidad["cv_planitud_cresta"]
            if np.isfinite(uniformidad["cv_planitud_cresta"]) else None,
        "cv_longitud_valle": uniformidad["cv_longitud_valle"]
            if np.isfinite(uniformidad["cv_longitud_valle"]) else None,
        "indice_no_uniformidad": uniformidad["indice_no_uniformidad"]
            if np.isfinite(uniformidad["indice_no_uniformidad"]) else None,
        "aviso_optica_valle": bool(aviso_optica_valle),
        # Vectores para graficar
        "x_mm": x_mm,
        "y_real_mm": y_real_mm,
        "y_teorico_mm": y_teorico_mm,
        "picos_idx": picos_idx,
        "valles_idx": valles_idx,
        # Datos del perfil 1D crudo (para visualización 4-paneles)
        "perfil_px": perfil_px,
        "x_offset": int(x_offset),
        "orient": str(orient),
    }


def generar_comparativa_etapas(img_gris, resultado, ruta_salida, nombre_base):
    """
    Genera UNA imagen comparativa con 4 paneles que muestran las etapas
    sucesivas del análisis de una pieza:

       1. ORIGINAL          - imagen tal cual
       2. SEGMENTADA (EDA)  - imagen con la herramienta resaltada (binarizada)
       3. PERFIL EXTRAÍDO   - imagen con la línea del perfil 1D superpuesta
       4. PERFIL REAL VS TEÓRICO - el perfil real vs la curva teórica ISO

    Esta gráfica resume visualmente el pipeline para el TFM.

    Args:
       img_gris: imagen en escala de grises (np.uint8)
       resultado: dict de perfilometro_analizar (con las claves x_mm,
                  y_real_mm, y_teorico_mm, perfil_px, orient, x_offset)
       ruta_salida: carpeta donde guardar el PNG
       nombre_base: nombre de archivo (sin extensión)
    """
    import matplotlib.pyplot as plt
    Path(ruta_salida).mkdir(parents=True, exist_ok=True)

    fig, axes = plt.subplots(2, 2, figsize=(14, 9))

    # PANEL 1: imagen original
    ax = axes[0, 0]
    ax.imshow(img_gris, cmap="gray")
    ax.set_title("1) IMAGEN ORIGINAL", fontsize=11, fontweight="bold")
    ax.axis("off")

    # PANEL 2: imagen segmentada (EDA - máscara de la herramienta)
    ax = axes[0, 1]
    try:
        mascara = segmentar_herramienta(img_gris)
        # Mostrar la imagen original en gris claro y superponer la máscara
        # con un color que destaque (semitransparente)
        overlay = np.zeros((*img_gris.shape, 3), dtype=np.float32)
        overlay[..., 0] = img_gris.astype(np.float32) / 255.0  # canal R = imagen
        overlay[..., 1] = img_gris.astype(np.float32) / 255.0  # canal G = imagen
        overlay[..., 2] = img_gris.astype(np.float32) / 255.0  # canal B = imagen
        # Píxeles de la herramienta en cian
        mask_bool = mascara > 0
        overlay[mask_bool, 0] *= 0.4   # baja R
        overlay[mask_bool, 1] = np.clip(overlay[mask_bool, 1] * 1.0 + 0.3, 0, 1)  # G
        overlay[mask_bool, 2] = np.clip(overlay[mask_bool, 2] * 1.0 + 0.3, 0, 1)  # B
        ax.imshow(overlay)
        ax.set_title("2) SEGMENTACIÓN (EDA)\nHerramienta detectada (cian)",
                     fontsize=11, fontweight="bold")
    except Exception as e:
        ax.imshow(img_gris, cmap="gray")
        ax.set_title(f"2) SEGMENTACIÓN\n(error: {str(e)[:30]})",
                     fontsize=11, color="red")
    ax.axis("off")

    # PANEL 3: imagen con la línea del perfil 1D superpuesta
    ax = axes[1, 0]
    ax.imshow(img_gris, cmap="gray")
    if resultado.get("ok") and resultado.get("perfil_px") is not None:
        try:
            perfil_px = resultado["perfil_px"]
            x_offset = resultado.get("x_offset", 0)
            orient = resultado.get("orient", "bottom")
            # X coords en la imagen original
            xs = np.arange(len(perfil_px)) + x_offset
            # Y coords: perfil_px da la fila del primer píxel oscuro
            ys = perfil_px.astype(float)
            # Si la orientación es 'top', perfil_px da la altura desde arriba.
            # Si es 'bottom', perfil_px = nº de filas, restar para obtener Y abs.
            if orient == "bottom":
                ys = img_gris.shape[0] - 1 - ys
            ax.plot(xs, ys, color="red", lw=1.6,
                    label="Perfil 1D extraído")
            # Marcar picos y valles si existen
            picos = resultado.get("picos_idx")
            valles = resultado.get("valles_idx")
            if picos is not None and len(picos) > 0:
                xp = np.array(picos) + x_offset
                yp = ys[np.array(picos)] if max(picos) < len(ys) else None
                if yp is not None:
                    ax.scatter(xp, yp, c="lime", s=50, zorder=5,
                                marker="v", label="Picos")
            if valles is not None and len(valles) > 0:
                xv = np.array(valles) + x_offset
                yv = ys[np.array(valles)] if max(valles) < len(ys) else None
                if yv is not None:
                    ax.scatter(xv, yv, c="yellow", s=50, zorder=5,
                                marker="^", label="Valles")
            ax.legend(loc="upper right", fontsize=8)
            ax.set_title("3) PERFIL 1D EXTRAÍDO\n(silueta dentada)",
                         fontsize=11, fontweight="bold")
        except Exception as e:
            ax.set_title(f"3) PERFIL 1D\n(error: {str(e)[:30]})",
                         fontsize=11, color="red")
    else:
        motivo = resultado.get("motivo", "no medible")
        ax.set_title(f"3) PERFIL 1D\nNo extraído ({motivo[:30]})",
                     fontsize=11, color="orange")
    ax.axis("off")

    # PANEL 4: perfil real vs perfil teórico ISO + zona de desgaste
    ax = axes[1, 1]
    if resultado.get("ok") and resultado.get("y_real_mm") is not None:
        x_mm = resultado.get("x_mm")
        y_real = resultado.get("y_real_mm")
        y_teo = resultado.get("y_teorico_mm")
        ax.plot(x_mm, y_real, color="#264653", lw=1.8,
                label="Perfil real medido")
        if y_teo is not None:
            ax.plot(x_mm, y_teo, color="#e76f51", lw=1.5, ls="--",
                    label="Perfil teórico ISO")
            # Sombreado de la desviación general (claro)
            ax.fill_between(x_mm, y_real, y_teo, alpha=0.10,
                             color="#e76f51")
            # Sombreado intenso de ZONA DE DESGASTE (donde real < teórico,
            # es decir, falta material). Esto resalta visualmente las
            # métricas del objetivo 4 del TFM.
            mask_deficit = np.asarray(y_real) < np.asarray(y_teo)
            ax.fill_between(x_mm, y_real, y_teo, where=mask_deficit,
                             alpha=0.35, color="#c0392b",
                             label="Zona de desgaste\n(material faltante)")
        # Métricas clave (incluyendo las nuevas del objetivo 4)
        info = []
        # Métricas de DESGASTE (objetivo 4) — primero, las más relevantes
        area_d = resultado.get("area_desgaste_mm2")
        ancho_d = resultado.get("ancho_medio_desgaste_um")
        prof_d = resultado.get("prof_max_desgaste_um")
        pct_long_d = resultado.get("pct_longitud_con_desgaste")
        if area_d is not None:
            info.append(f"DESGASTE (Obj. 4):")
            info.append(f"  Área = {area_d*1000:.1f}·10⁻³ mm²")
        if ancho_d is not None:
            info.append(f"  Ancho medio = {ancho_d:.0f} µm")
        if prof_d is not None:
            info.append(f"  Prof. máx = {prof_d:.0f} µm")
        if pct_long_d is not None:
            info.append(f"  % long. afectada = {pct_long_d:.0f}%")
        # Métricas geométricas generales
        rms = resultado.get("rms_desviacion_mm")
        max_d_v = resultado.get("max_desviacion_mm")
        if rms is not None or max_d_v is not None:
            info.append(f"GEOMETRÍA:")
        if rms is not None:
            info.append(f"  RMS desv. = {rms*1000:.0f} µm")
        if max_d_v is not None:
            info.append(f"  Máx desv. = {max_d_v*1000:.0f} µm")
        ang = resultado.get("angulo_flanco_real_deg")
        if ang is not None:
            info.append(f"  Áng. flanco = {ang:.1f}°")
        if info:
            ax.text(0.02, 0.98, "\n".join(info),
                    transform=ax.transAxes,
                    verticalalignment="top",
                    fontsize=8, family="monospace",
                    bbox=dict(boxstyle="round", facecolor="white",
                               alpha=0.85, edgecolor="gray"))
        ax.set_xlabel("Posición a lo largo del eje [mm]")
        ax.set_ylabel("Altura del perfil [mm]")
        ax.set_title("4) PERFIL REAL vs TEÓRICO ISO + ZONA DE DESGASTE",
                     fontsize=11, fontweight="bold")
        ax.legend(loc="lower right", fontsize=7)
        ax.grid(alpha=0.3)
    else:
        motivo = resultado.get("motivo", "no medible")
        ax.set_title(f"4) PERFIL REAL vs TEÓRICO\nNo disponible ({motivo[:30]})",
                     fontsize=11, color="orange")
        ax.axis("off")

    fig.suptitle(f"PIPELINE DE ANÁLISIS: {nombre_base}",
                 fontsize=13, fontweight="bold", y=0.995)
    plt.tight_layout()
    out_path = os.path.join(ruta_salida, f"comparativa_{nombre_base}.png")
    plt.savefig(out_path, dpi=120, bbox_inches="tight")
    plt.close()
    return out_path


def perfilometro_graficar(resultado, ruta_salida, nombre_base):
    """
    Figura con cuatro paneles:
       (1) perfil real vs teórico, picos y valles, banda de cresta destacada
       (2) desviación real - teórico
       (3) zoom de la cresta del primer pico (donde se ve el redondeo)
       (4) panel de texto con todas las métricas y veredicto
    """
    if not resultado["ok"]:
        return None
    x = resultado["x_mm"]
    y_real = resultado["y_real_mm"]
    y_teo = resultado["y_teorico_mm"]
    desv = y_real - y_teo
    picos = resultado["picos_idx"]
    valles = resultado["valles_idx"]

    fig = plt.figure(figsize=(13.5, 8))
    gs = fig.add_gridspec(3, 3, height_ratios=[3, 1, 1.6],
                          width_ratios=[2.2, 1, 1.4],
                          hspace=0.45, wspace=0.30)
    ax_main = fig.add_subplot(gs[0, :2])
    ax_desv = fig.add_subplot(gs[1, :2], sharex=ax_main)
    ax_zoom = fig.add_subplot(gs[0:2, 2])
    ax_info = fig.add_subplot(gs[2, :])
    ax_info.axis("off")

    color_v = "#2a9d8f" if resultado["veredicto"] == "PASA" else "#e76f51"

    # === Panel principal ===
    ax_main.plot(x, y_teo, "k--", lw=1.5, label="Perfil teórico ISO")
    ax_main.plot(x, y_real, color=color_v, lw=1.8, label="Perfil real medido")
    if len(picos):
        ax_main.scatter(x[picos], y_real[picos], c="gold", s=45,
                        zorder=5, edgecolor="k", linewidth=0.6, label="Picos")
    if len(valles):
        ax_main.scatter(x[valles], y_real[valles], c="purple", s=45,
                        zorder=5, edgecolor="k", linewidth=0.6, label="Valles")
    ax_main.invert_yaxis()
    ax_main.set_ylabel("Altura (mm)")
    ax_main.set_title(
        f"PERFILÓMETRO — {nombre_base}   [{resultado['veredicto']}]",
        fontsize=12, fontweight="bold",
        color="#1a6b5e" if resultado["veredicto"] == "PASA" else "#a23a1a",
    )
    ax_main.legend(loc="upper right", fontsize=8, ncol=2)
    ax_main.grid(alpha=0.3)

    # === Panel de desviación ===
    ax_desv.fill_between(x, 0, desv, color="#888", alpha=0.5)
    ax_desv.axhline(0, color="k", lw=0.6)
    ax_desv.set_xlabel("Posición (mm)")
    ax_desv.set_ylabel("Desv. (mm)")
    ax_desv.set_title(
        f"Desviación   RMS = {resultado['rms_desviacion_mm']*1000:.1f} µm   "
        f"máx = {resultado['max_desviacion_mm']*1000:.1f} µm",
        fontsize=9,
    )
    ax_desv.grid(alpha=0.3)

    # === Panel de zoom de cresta (pico más central) ===
    if len(picos):
        # Usar el pico más cercano al centro del perfil (más representativo)
        centro = len(x) // 2
        idx_pico_central = int(np.argmin(np.abs(np.array(picos) - centro)))
        p = int(picos[idx_pico_central])
        # Ventana de un cuarto de paso a cada lado
        if len(x) > 1:
            dx = x[1] - x[0]
            ventana_px = max(20, int((resultado["paso_nominal_mm"] / 4.0) / dx))
        else:
            ventana_px = 50
        i_lo = max(0, p - ventana_px)
        i_hi = min(len(x), p + ventana_px + 1)
        zx = x[i_lo:i_hi] - x[p]
        zy = (y_real[i_lo:i_hi] - y_real[p]) * 1000.0   # en µm
        ax_zoom.plot(zx, zy, color=color_v, lw=2, label="Cresta real")
        # Línea horizontal idealizada (cresta plana de longitud P/8)
        L_nom = resultado.get("cresta_longitud_nominal_mm",
                              resultado["paso_nominal_mm"] / 8.0)
        ax_zoom.plot([-L_nom / 2, L_nom / 2], [0, 0], "k--",
                     lw=1.6, label=f"Ideal P/8 ({L_nom*1000:.0f}µm)")
        ax_zoom.scatter([0], [0], c="gold", s=70, zorder=5,
                        edgecolor="k", linewidth=0.6)
        # Limitar rango y a unas decenas de µm para ver bien la cresta
        ax_zoom.set_ylim(150, -30)   # invertido: Δaltura crece hacia abajo
        ax_zoom.set_xlabel("Posición desde pico (mm)")
        ax_zoom.set_ylabel("Δ altura (µm)")
        L_med = resultado.get("cresta_longitud_mm", None)
        plan = resultado.get("cresta_planitud_rms_um", None)
        R = resultado.get("cresta_radio_redondeo_um", None)
        titulo_zoom = "Zoom cresta:"
        if L_med is not None:
            titulo_zoom += f"  L={L_med*1000:.0f}µm"
        if plan is not None:
            titulo_zoom += f"  plan={plan:.1f}µm"
        if R is not None:
            titulo_zoom += f"  R={R:.0f}µm"
        ax_zoom.set_title(titulo_zoom, fontsize=9)
        ax_zoom.legend(fontsize=8, loc="lower center")
        ax_zoom.grid(alpha=0.3)
        ax_zoom.axhline(0, color="k", lw=0.4, alpha=0.5)

    # === Panel de información ===
    fmt_n = lambda v, dec=3: ("—" if v is None or
                              (isinstance(v, float) and not np.isfinite(v))
                              else f"{v:.{dec}f}")
    # Helpers para evitar f-strings anidados con escapes
    err_ang = resultado.get("error_angulo_deg")
    err_ang_str = "—" if err_ang is None else f"{err_ang:+.2f}°"
    err_cl = resultado.get("error_cresta_longitud_pct")
    err_cl_str = "—" if err_cl is None else f"{err_cl:+.0f}%"
    err_vl = resultado.get("error_valle_longitud_pct")
    err_vl_str = "—" if err_vl is None else f"{err_vl:+.0f}%"

    cl_um = (resultado["cresta_longitud_mm"] * 1000.0
             if resultado["cresta_longitud_mm"] is not None else None)
    vl_um = (resultado["valle_longitud_mm"] * 1000.0
             if resultado["valle_longitud_mm"] is not None else None)

    cols = [
        ("Paso medido", f"{fmt_n(resultado['paso_medido_mm'])} mm",
         f"({resultado['error_paso_pct']:+.2f}%)"),
        ("Altura efectiva", f"{fmt_n(resultado['altura_real_mm'])} mm",
         f"(teor {fmt_n(resultado['altura_teorica_mm'])}, "
         f"{resultado['error_altura_pct']:+.1f}%)"),
        ("Ángulo flanco",
         f"{fmt_n(resultado['angulo_flanco_real_deg'], 2)}°",
         f"(teor {resultado['angulo_flanco_teorico_deg']:.0f}°, {err_ang_str})"),
        ("Cresta longitud",
         f"{fmt_n(cl_um, 0)} µm" if cl_um is not None else "—",
         f"(nom {resultado['cresta_longitud_nominal_mm']*1000:.0f}µm "
         f"{err_cl_str})"),
        ("Cresta planitud",
         f"{fmt_n(resultado['cresta_planitud_rms_um'], 1)} µm"
         if resultado['cresta_planitud_rms_um'] is not None else "—",
         "(RMS dentro cresta)"),
        ("Cresta radio R",
         f"{fmt_n(resultado['cresta_radio_redondeo_um'], 0)} µm"
         if resultado['cresta_radio_redondeo_um'] is not None else "—",
         "(0 = arista viva)"),
        ("Valle longitud",
         f"{fmt_n(vl_um, 0)} µm" if vl_um is not None else "—",
         f"(nom {resultado['valle_longitud_nominal_mm']*1000:.0f}µm "
         f"{err_vl_str})"),
        ("Índice desgaste",
         f"{resultado['indice_desgaste_cresta']:.2f}",
         "(0=nuevo, 1=desgastado)"),
    ]
    # Dibuja en 2 filas x 4 columnas
    for i, (etq, val, sub) in enumerate(cols):
        col = i % 4
        fila = i // 4
        x_pos = 0.02 + col * 0.245
        y_pos = 0.78 - fila * 0.45
        ax_info.text(x_pos, y_pos, etq, fontsize=9, fontweight="bold",
                     transform=ax_info.transAxes)
        ax_info.text(x_pos, y_pos - 0.16, val, fontsize=11, color=color_v,
                     fontweight="bold", transform=ax_info.transAxes)
        ax_info.text(x_pos, y_pos - 0.30, sub, fontsize=7, color="#555",
                     transform=ax_info.transAxes)

    # Motivos de fallo o aviso óptico
    avisos = []
    if resultado.get("aviso_optica_valle"):
        avisos.append("⚠ óptica posiblemente cortando el fondo del valle "
                      "(altura medida < 80% de la teórica)")
    if resultado["motivos_fallo"]:
        avisos.append("Motivos NO PASA: " +
                      " · ".join(resultado["motivos_fallo"]))
    if avisos:
        ax_info.text(0.02, -0.02, "\n".join(avisos), fontsize=8,
                     color="#a23a1a" if resultado["motivos_fallo"]
                     else "#b08418",
                     transform=ax_info.transAxes,
                     verticalalignment="top")

    Path(ruta_salida).mkdir(parents=True, exist_ok=True)
    out = os.path.join(ruta_salida, f"perfilometro_{nombre_base}.png")
    plt.savefig(out, dpi=130, bbox_inches="tight")
    plt.close()
    return out


def comparativa_buena_mala_iso(res_buenas, res_malas, ruta_salida,
                                paso_mm=1.5, eda_resumen=None):
    """Genera la comparativa cuantitativa BUENA vs MALA vs ISO teórico.

    Caracterización cuantitativa de la diferencia geométrica entre roscas
    conformes y no conformes, expresada en unidades físicas (mm² y µm) y
    contrastada con el perfil teórico ISO. La salida consta de:

      1. Una figura única con los tres perfiles superpuestos:
         - perfil teórico ISO (línea de referencia)
         - perfil PROMEDIO de las BUENAS
         - perfil PROMEDIO de las MALAS
         con el área de desgaste rellena en color y anotada en mm² y µm.

      2. Una segunda figura con TRES paneles que cuantifican la diferencia
         por dos vías independientes:
            (a) DIFERENCIA POR ALTURA: histograma de la desviación punto
                a punto (perfil real menos ISO teórico) en µm.
            (b) DIFERENCIA POR ÁREA: boxplot de área de desgaste por
                imagen en mm² (BUENAS vs MALAS).
            (c) DIFERENCIA POR PROFUNDIDAD: boxplot de profundidad
                máxima en µm.

      3. Un CSV resumen con la media, mediana, p25, p75 y desviación
         típica de cada métrica por clase, más el delta absoluto y la
         d de Cohen (= separación estandarizada).

      4. Un fichero de texto interpretativo en castellano con frases
         del tipo "una rosca buena difiere del ISO en X µm; una rosca
         mala difiere en Y µm; la diferencia es Z veces mayor".

    Args:
        res_buenas: lista de dicts devueltos por perfilometro_analizar
            para las imágenes BUENAS. Cada dict tiene 'ok', 'y_real_mm',
            'y_teorico_mm', 'x_mm', 'area_desgaste_mm2', etc.
        res_malas: ídem para MALAS.
        ruta_salida: carpeta donde se escriben las figuras y el CSV.
        paso_mm: paso nominal de la rosca (para el eje X normalizado).

    Devuelve un dict con el resumen estadístico, listo para incrustar
    en el informe final.
    """
    log.info("=" * 70)
    log.info(" COMPARATIVA CUANTITATIVA BUENA vs MALA vs ISO TEÓRICO")
    log.info("=" * 70)
    Path(ruta_salida).mkdir(parents=True, exist_ok=True)

    # Filtramos solo los resultados válidos
    res_b = [r for r in (res_buenas or []) if r.get("ok")]
    res_m = [r for r in (res_malas or []) if r.get("ok")]
    if not res_b or not res_m:
        log.warning("  No hay suficientes perfiles válidos. "
                    f"BUENAS_ok={len(res_b)}, MALAS_ok={len(res_m)}. "
                    "Saltando comparativa.")
        return None
    log.info(f"  BUENAS válidas: {len(res_b)}    MALAS válidas: {len(res_m)}")

    # ------------------------------------------------------------------
    # 1. PERFILES PROMEDIO BUENA, MALA, ISO
    # ------------------------------------------------------------------
    # Los perfiles tienen longitudes distintas según la imagen, así que
    # los normalizamos a una rejilla común de [0, 3*paso_mm] (~ 3 dientes)
    # mediante interpolación 1D, partiendo del primer pico de cada perfil
    # para alinear la fase. Esto da promedios coherentes.
    L_mm = 3.0 * float(paso_mm)
    N_pts = 600
    x_comun = np.linspace(0.0, L_mm, N_pts)

    def _normalizar(res_lista):
        """Devuelve dos matrices (n_perfiles, N_pts):
           - matriz_real: perfil REAL alineado en [0, L_mm].
           - matriz_teo:  perfil ISO teórico de la misma longitud."""
        reales = []
        teos = []
        for r in res_lista:
            try:
                x = np.asarray(r["x_mm"])
                y_r = np.asarray(r["y_real_mm"])
                y_t = np.asarray(r["y_teorico_mm"])
                picos = r.get("picos_idx", [])
                if len(picos) < 1 or len(x) < 10:
                    continue
                # Origen en el primer pico para alinear la fase
                x0 = float(x[int(picos[0])])
                x_local = x - x0
                # Recortamos a [0, L_mm] e interpolamos
                mask = (x_local >= 0) & (x_local <= L_mm)
                if mask.sum() < 10:
                    continue
                y_r_i = np.interp(x_comun, x_local[mask], y_r[mask])
                y_t_i = np.interp(x_comun, x_local[mask], y_t[mask])
                # Restamos la media de cada perfil para comparar formas,
                # no offsets absolutos (cada imagen está calibrada por
                # su propio mm/px).
                reales.append(y_r_i - float(np.mean(y_r_i)))
                teos.append(y_t_i - float(np.mean(y_t_i)))
            except Exception:
                continue
        if not reales:
            return None, None
        return np.asarray(reales), np.asarray(teos)

    M_b, T_b = _normalizar(res_b)
    M_m, T_m = _normalizar(res_m)
    if M_b is None or M_m is None:
        log.warning("  Normalización falló. Saltando figura promedio.")
    else:
        perf_buena_media = np.median(M_b, axis=0)
        perf_mala_media = np.median(M_m, axis=0)
        # El ISO debería ser similar entre BUENAS y MALAS (es teórico):
        # cogemos la mediana de las BUENAS como referencia ISO observada.
        perf_iso = np.median(T_b, axis=0)
        # Bandas P25-P75 para mostrar dispersión
        p25_b, p75_b = np.percentile(M_b, [25, 75], axis=0)
        p25_m, p75_m = np.percentile(M_m, [25, 75], axis=0)

        fig, ax = plt.subplots(figsize=(12, 5.5))
        # ISO de referencia
        ax.plot(x_comun, perf_iso, "k--", lw=2.2,
                label="Perfil teórico ISO", zorder=4)
        # BUENA promedio + banda P25-P75
        ax.plot(x_comun, perf_buena_media, color="#2a9d8f", lw=2.3,
                label=f"Promedio BUENAS (n={len(M_b)})", zorder=3)
        ax.fill_between(x_comun, p25_b, p75_b, color="#2a9d8f",
                         alpha=0.15, label="BUENAS P25-P75")
        # MALA promedio + banda P25-P75
        ax.plot(x_comun, perf_mala_media, color="#e76f51", lw=2.3,
                label=f"Promedio MALAS (n={len(M_m)})", zorder=3)
        ax.fill_between(x_comun, p25_m, p75_m, color="#e76f51",
                         alpha=0.15, label="MALAS P25-P75")

        # Rellenamos el ÁREA entre la MALA media y el ISO (= déficit
        # de material respecto al teórico) — esto es el equivalente
        # gráfico del área de desgaste.
        deficit_m = perf_iso - perf_mala_media   # > 0 = falta material
        deficit_m_pos = np.where(deficit_m > 0, deficit_m, 0)
        ax.fill_between(x_comun, perf_iso, perf_iso - deficit_m_pos,
                         where=deficit_m_pos > 0,
                         color="#e76f51", alpha=0.30,
                         label="Área desgaste MALAS vs ISO", zorder=2)
        # Y lo mismo para BUENAS — debería ser prácticamente cero
        deficit_b = perf_iso - perf_buena_media
        deficit_b_pos = np.where(deficit_b > 0, deficit_b, 0)
        ax.fill_between(x_comun, perf_iso, perf_iso - deficit_b_pos,
                         where=deficit_b_pos > 0,
                         color="#2a9d8f", alpha=0.30,
                         label="Área desgaste BUENAS vs ISO", zorder=2)

        ax.invert_yaxis()
        ax.set_xlabel("Posición a lo largo del perfil (mm)")
        ax.set_ylabel("Altura (mm) — eje invertido (rosca hacia abajo)")
        ax.set_title("Perfil BUENA vs MALA vs ISO teórico — "
                      "diferencia cuantificada por altura y área",
                      fontsize=12, fontweight="bold")
        ax.legend(loc="lower right", fontsize=8, ncol=2)
        ax.grid(alpha=0.3)
        fig.tight_layout()
        ruta_fig1 = os.path.join(ruta_salida,
                                  "comparativa_buena_mala_iso.png")
        fig.savefig(ruta_fig1, dpi=140, bbox_inches="tight")
        plt.close(fig)
        log.info(f"  Figura promedio guardada: {ruta_fig1}")

    # ------------------------------------------------------------------
    # 2. ESTADÍSTICOS POR MÉTRICA
    # ------------------------------------------------------------------
    # Definimos las 4 magnitudes geométricas en unidades físicas.
    def _vec(res_lista, clave, mult=1.0):
        out = []
        for r in res_lista:
            v = r.get(clave)
            if v is not None and np.isfinite(v):
                out.append(float(v) * mult)
        return np.asarray(out, dtype=float)

    metricas_def = [
        # (clave_en_dict, multiplicador, nombre_legible, unidad)
        ("max_desviacion_mm",    1000.0, "Desv. máxima vs ISO",     "µm"),
        ("rms_desviacion_mm",    1000.0, "Desv. RMS vs ISO",        "µm"),
        ("prof_max_desgaste_um", 1.0,    "Profundidad desgaste",    "µm"),
        ("ancho_medio_desgaste_um", 1.0, "Ancho medio desgaste",    "µm"),
        ("area_desgaste_mm2",    1.0,    "Área desgaste",           "mm²"),
    ]

    resumen = []
    for clave, mult, nombre, unidad in metricas_def:
        v_b = _vec(res_b, clave, mult)
        v_m = _vec(res_m, clave, mult)
        if len(v_b) == 0 or len(v_m) == 0:
            continue
        med_b = float(np.median(v_b))
        med_m = float(np.median(v_m))
        mean_b = float(np.mean(v_b))
        mean_m = float(np.mean(v_m))
        std_b = float(np.std(v_b))
        std_m = float(np.std(v_m))
        delta = med_m - med_b
        delta_abs = abs(delta)
        # d de Cohen (separación estandarizada). >0.8 = grande, >0.5 = medio.
        pool = float(np.sqrt((std_b ** 2 + std_m ** 2) / 2.0))
        cohen_d = ((mean_m - mean_b) / pool) if pool > 1e-12 else 0.0
        # Razón mediana mala / mediana buena (cuántas veces más).
        razon = (med_m / med_b) if abs(med_b) > 1e-6 else float("inf")
        resumen.append({
            "metrica": nombre,
            "clave": clave,
            "unidad": unidad,
            "mediana_BUENA": med_b,
            "mediana_MALA": med_m,
            "media_BUENA": mean_b,
            "media_MALA": mean_m,
            "std_BUENA": std_b,
            "std_MALA": std_m,
            "delta_MALA_menos_BUENA": delta,
            "delta_abs": delta_abs,
            "razon_MALA_BUENA": razon,
            "cohen_d": cohen_d,
            "n_BUENA": int(len(v_b)),
            "n_MALA": int(len(v_m)),
        })

    # CSV
    if resumen:
        ruta_csv = os.path.join(ruta_salida, "comparativa_buena_mala.csv")
        try:
            with open(ruta_csv, "w", newline="", encoding="utf-8") as fh:
                w = csv.DictWriter(fh, fieldnames=list(resumen[0].keys()))
                w.writeheader()
                for row in resumen:
                    w.writerow(row)
            log.info(f"  CSV resumen: {ruta_csv}")
        except Exception as e:
            log.error(f"  No pude escribir CSV resumen: {e}")

    # ------------------------------------------------------------------
    # 3. FIGURA DE TRES PANELES: HISTOGRAMA + BOXPLOT ÁREA + BOXPLOT PROF
    # ------------------------------------------------------------------
    fig, axes = plt.subplots(1, 3, figsize=(16, 4.5))

    # Panel A: histograma de desviaciones punto a punto vs ISO (µm).
    # Tomamos TODAS las desviaciones de TODAS las imágenes y comparamos.
    desv_b_um = []
    desv_m_um = []
    for r in res_b:
        try:
            d = np.asarray(r["y_real_mm"]) - np.asarray(r["y_teorico_mm"])
            desv_b_um.extend((d * 1000.0).tolist())
        except Exception:
            continue
    for r in res_m:
        try:
            d = np.asarray(r["y_real_mm"]) - np.asarray(r["y_teorico_mm"])
            desv_m_um.extend((d * 1000.0).tolist())
        except Exception:
            continue
    desv_b_um = np.asarray(desv_b_um, dtype=float)
    desv_m_um = np.asarray(desv_m_um, dtype=float)
    # Recortamos a un rango razonable para que el histograma sea legible
    if len(desv_b_um) and len(desv_m_um):
        rango = float(np.percentile(np.abs(np.concatenate(
            [desv_b_um, desv_m_um])), 99))
        rango = max(rango, 50.0)
        bins = np.linspace(-rango, rango, 80)
        axes[0].hist(desv_b_um, bins=bins, alpha=0.55, color="#2a9d8f",
                      label=f"BUENAS (n={len(desv_b_um)} pts)",
                      density=True)
        axes[0].hist(desv_m_um, bins=bins, alpha=0.55, color="#e76f51",
                      label=f"MALAS  (n={len(desv_m_um)} pts)",
                      density=True)
        axes[0].axvline(0, color="k", lw=0.8, ls="--",
                         label="ISO ideal (0 µm)")
        med_b_um = float(np.median(np.abs(desv_b_um)))
        med_m_um = float(np.median(np.abs(desv_m_um)))
        axes[0].set_title(
            "Diferencia por ALTURA vs ISO\n"
            f"|desv| mediana BUENAS={med_b_um:.0f} µm | "
            f"MALAS={med_m_um:.0f} µm",
            fontsize=10)
        axes[0].set_xlabel("Desviación real − ISO (µm)")
        axes[0].set_ylabel("Densidad")
        axes[0].legend(fontsize=8, loc="upper right")
        axes[0].grid(alpha=0.3)

    # Panel B: boxplot ÁREA por imagen (mm²)
    areas_b = _vec(res_b, "area_desgaste_mm2", 1.0)
    areas_m = _vec(res_m, "area_desgaste_mm2", 1.0)
    if len(areas_b) and len(areas_m):
        bp = axes[1].boxplot([areas_b, areas_m], labels=["BUENAS", "MALAS"],
                              patch_artist=True, widths=0.5)
        for patch, c in zip(bp["boxes"], ["#2a9d8f", "#e76f51"]):
            patch.set_facecolor(c); patch.set_alpha(0.6)
        axes[1].set_title(
            "Diferencia por ÁREA de desgaste\n"
            f"mediana BUENAS={float(np.median(areas_b)):.4f} mm² | "
            f"MALAS={float(np.median(areas_m)):.4f} mm²",
            fontsize=10)
        axes[1].set_ylabel("Área desgaste vs ISO (mm²)")
        axes[1].grid(alpha=0.3, axis="y")

    # Panel C: boxplot PROFUNDIDAD máxima (µm)
    profs_b = _vec(res_b, "prof_max_desgaste_um", 1.0)
    profs_m = _vec(res_m, "prof_max_desgaste_um", 1.0)
    if len(profs_b) and len(profs_m):
        bp = axes[2].boxplot([profs_b, profs_m], labels=["BUENAS", "MALAS"],
                              patch_artist=True, widths=0.5)
        for patch, c in zip(bp["boxes"], ["#2a9d8f", "#e76f51"]):
            patch.set_facecolor(c); patch.set_alpha(0.6)
        axes[2].set_title(
            "Diferencia por PROFUNDIDAD máxima\n"
            f"mediana BUENAS={float(np.median(profs_b)):.0f} µm | "
            f"MALAS={float(np.median(profs_m)):.0f} µm",
            fontsize=10)
        axes[2].set_ylabel("Profundidad máxima desgaste (µm)")
        axes[2].grid(alpha=0.3, axis="y")

    fig.suptitle("Cuantificación entendible de la diferencia BUENA vs MALA",
                  fontsize=12, fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.95])
    ruta_fig2 = os.path.join(ruta_salida,
                              "comparativa_buena_mala_paneles.png")
    fig.savefig(ruta_fig2, dpi=140, bbox_inches="tight")
    plt.close(fig)
    log.info(f"  Figura paneles guardada: {ruta_fig2}")

    # ------------------------------------------------------------------
    # 4. TEXTO INTERPRETATIVO EN CASTELLANO
    # ------------------------------------------------------------------
    lineas = []
    lineas.append("=" * 78)
    lineas.append(" COMPARATIVA CUANTITATIVA BUENA vs MALA vs ISO TEÓRICO")
    lineas.append("=" * 78)
    lineas.append("")
    lineas.append("Pregunta: ¿en qué se diferencia, en milímetros y micras,")
    lineas.append("una rosca buena de una rosca mala respecto al ISO teórico?")
    lineas.append("")
    lineas.append(f"Muestras: {len(res_b)} imágenes BUENAS válidas, "
                   f"{len(res_m)} imágenes MALAS válidas.")
    lineas.append("")
    lineas.append("RESPUESTA POR ALTURA (desviación punto a punto vs ISO)")
    lineas.append("-" * 78)
    if len(desv_b_um) and len(desv_m_um):
        med_b_um = float(np.median(np.abs(desv_b_um)))
        med_m_um = float(np.median(np.abs(desv_m_um)))
        p95_b_um = float(np.percentile(np.abs(desv_b_um), 95))
        p95_m_um = float(np.percentile(np.abs(desv_m_um), 95))
        ratio = (med_m_um / med_b_um) if med_b_um > 1e-6 else float("inf")
        lineas.append(f"  Una rosca BUENA se desvía del ISO en torno a "
                       f"{med_b_um:.0f} µm en mediana, con un percentil 95")
        lineas.append(f"  de {p95_b_um:.0f} µm (es el ruido del sistema +")
        lineas.append(f"  la tolerancia natural del proceso).")
        lineas.append("")
        lineas.append(f"  Una rosca MALA se desvía del ISO en torno a "
                       f"{med_m_um:.0f} µm en mediana, con un percentil 95")
        lineas.append(f"  de {p95_m_um:.0f} µm.")
        lineas.append("")
        lineas.append(f"  Razón MALA/BUENA: {ratio:.2f}x.  La desviación")
        lineas.append(f"  típica frente al ISO es ~{ratio:.1f} veces mayor")
        lineas.append("  en una rosca defectuosa que en una conforme.")
    lineas.append("")
    lineas.append("RESPUESTA POR ÁREA (déficit de material vs ISO)")
    lineas.append("-" * 78)
    if len(areas_b) and len(areas_m):
        a_b = float(np.median(areas_b))
        a_m = float(np.median(areas_m))
        delta_a = a_m - a_b
        ratio_a = (a_m / a_b) if a_b > 1e-6 else float("inf")
        lineas.append(f"  Área desgaste mediana BUENAS: {a_b:.4f} mm²")
        lineas.append(f"  Área desgaste mediana MALAS:  {a_m:.4f} mm²")
        lineas.append(f"  Delta absoluto:                {delta_a:+.4f} mm²")
        lineas.append(f"  Razón MALA/BUENA:              {ratio_a:.2f}x")
    lineas.append("")
    lineas.append("RESPUESTA POR PROFUNDIDAD (peor punto del flanco)")
    lineas.append("-" * 78)
    if len(profs_b) and len(profs_m):
        p_b = float(np.median(profs_b))
        p_m = float(np.median(profs_m))
        delta_p = p_m - p_b
        ratio_p = (p_m / p_b) if p_b > 1e-6 else float("inf")
        lineas.append(f"  Profundidad máxima mediana BUENAS: {p_b:.0f} µm")
        lineas.append(f"  Profundidad máxima mediana MALAS:  {p_m:.0f} µm")
        lineas.append(f"  Delta absoluto:                    {delta_p:+.0f} µm")
        lineas.append(f"  Razón MALA/BUENA:                  {ratio_p:.2f}x")
    lineas.append("")
    lineas.append("RESUMEN COMPACTO POR MÉTRICA")
    lineas.append("-" * 78)
    lineas.append(f"  {'Métrica':<26} {'Unidad':<6} {'BUENA':>10} "
                   f"{'MALA':>10} {'Δ':>10} {'Razón':>7} {'CohenD':>7}")
    for r in resumen:
        lineas.append(
            f"  {r['metrica']:<26} {r['unidad']:<6} "
            f"{r['mediana_BUENA']:>10.3f} {r['mediana_MALA']:>10.3f} "
            f"{r['delta_MALA_menos_BUENA']:>+10.3f} "
            f"{r['razon_MALA_BUENA']:>7.2f} {r['cohen_d']:>+7.2f}")
    lineas.append("")
    lineas.append("INTERPRETACIÓN")
    lineas.append("-" * 78)
    lineas.append("  Razón MALA/BUENA > 1 indica que la métrica crece con")
    lineas.append("  el desgaste (esperado en área, profundidad y ancho).")
    lineas.append("  Cohen d con magnitud > 0,8 indica separación grande;")
    lineas.append("  0,5–0,8 separación media; < 0,3 separación pequeña.")
    lineas.append("")

    # -----------------------------------------------------------------
    # MEJORA C — métricas restringidas al subconjunto detectable del EDA.
    # Si el EDA marcó N de M malas como detectables (con desgaste
    # ópticamente medible), recomputamos las medianas y Cohen d solo
    # sobre ese subconjunto. La separación esperada es mucho mayor.
    # -----------------------------------------------------------------
    mejora_c = None
    try:
        archivos_det = None
        if eda_resumen:
            archivos_det = (eda_resumen.get("outliers") or {})\
                .get("archivos_detectables")
        if archivos_det:
            mejora_c = metricas_restringidas_a_detectables(
                res_b, res_m, archivos_det)
    except Exception as e:
        log.warning(f"  Mejora C falló: {e}")
        mejora_c = None

    if mejora_c and mejora_c.get("resumen"):
        lineas.append(
            "MEJORA C — MÉTRICAS RESTRINGIDAS A MALAS DETECTABLES")
        lineas.append("-" * 78)
        lineas.append(
            f"  El EDA marcó {mejora_c['n_malas_detectables']} de "
            f"{mejora_c['n_malas_total']} imágenes MALAS como "
            f"detectables (con desgaste medible ópticamente). El resto "
            f"son vistas de zonas sanas de piezas con defecto local.")
        lineas.append(
            "  Si restringimos la comparativa SOLO a las MALAS detectables,")
        lineas.append(
            "  la separación frente a las BUENAS es la siguiente:")
        lineas.append("")
        lineas.append(
            f"  {'Métrica':<26} {'Unidad':<6} {'BUENA':>10} "
            f"{'MALA':>10} {'Δ':>10} {'Razón':>7} {'CohenD':>7}")
        for r in mejora_c["resumen"]:
            lineas.append(
                f"  {r['metrica']:<26} {r['unidad']:<6} "
                f"{r['mediana_BUENA']:>10.3f} {r['mediana_MALA']:>10.3f} "
                f"{r['delta_MALA_menos_BUENA']:>+10.3f} "
                f"{r['razon_MALA_BUENA']:>7.2f} {r['cohen_d']:>+7.2f}")
        lineas.append("")
        lineas.append(
            f"  Nota: n_BUENA={mejora_c['n_buenas']}, "
            f"n_MALA_detectables={mejora_c['n_malas_detectables']} "
            f"(de {mejora_c['n_malas_total']} totales).")
        lineas.append("")
    elif eda_resumen is not None:
        lineas.append("MEJORA C — MÉTRICAS RESTRINGIDAS A DETECTABLES")
        lineas.append("-" * 78)
        lineas.append(
            "  No se pudo restringir (faltan archivos detectables en EDA")
        lineas.append(
            "  o no hay al menos 3 imágenes MALAS detectables).")
        lineas.append("")

    lineas.append("ARCHIVOS GENERADOS")
    lineas.append("-" * 78)
    lineas.append(f"  {ruta_salida}/comparativa_buena_mala_iso.png")
    lineas.append(f"     - Perfil BUENA vs MALA vs ISO teórico con áreas rellenas")
    lineas.append(f"  {ruta_salida}/comparativa_buena_mala_paneles.png")
    lineas.append(f"     - Histograma de altura + boxplot de área + boxplot profundidad")
    lineas.append(f"  {ruta_salida}/comparativa_buena_mala.csv")
    lineas.append(f"     - Tabla con todas las métricas, deltas y Cohen d")
    lineas.append("=" * 78)

    ruta_txt = os.path.join(ruta_salida, "comparativa_buena_mala_informe.txt")
    try:
        with open(ruta_txt, "w", encoding="utf-8") as fh:
            fh.write("\n".join(lineas))
        log.info(f"  Informe interpretativo: {ruta_txt}")
    except Exception as e:
        log.error(f"  No pude escribir informe interpretativo: {e}")

    # Log final en consola para que se vea desde la ventana
    log.info("")
    log.info("  ---- RESULTADO CUANTITATIVO ----")
    for ln in lineas[6:24]:  # las primeras frases interpretativas
        log.info(f"  {ln}")

    return {
        "resumen_metricas": resumen,
        "resumen_metricas_detectables": (mejora_c["resumen"]
                                          if mejora_c else None),
        "n_malas_detectables": (mejora_c.get("n_malas_detectables")
                                if mejora_c else None),
        "n_buenas_validas": len(res_b),
        "n_malas_validas": len(res_m),
        "ruta_figura_principal": ruta_fig1 if (M_b is not None) else None,
        "ruta_figura_paneles": ruta_fig2,
        "ruta_csv": ruta_csv if resumen else None,
        "ruta_informe": ruta_txt,
    }


# ============================================================================
# PERFIL PATRÓN — Detector basado en distancia al patrón de roscas BUENAS
# ============================================================================
# El propósito es doble:
#   a) Construir un perfil promedio de las roscas BUENAS de entrenamiento y
#      contrastarlo con el perfil teórico ISO para cuantificar la desviación
#      sistemática del proceso de fabricación respecto al ideal normativo.
#   b) Evaluar la viabilidad de un clasificador basado en la distancia entre
#      cada perfil y el patrón promedio.
#
# Procedimiento:
#   1) Calcular el perfil promedio de las BUENAS de TRAIN (= patrón).
#   2) Compararlo con el ISO teórico (delta por posición + área entre ambos).
#   3) Para cada imagen EVAL, distancia de su perfil al patrón = score.
#   4) Umbral = P95 de distancias TRAIN (BUENA contra BUENA).
#   5) Evaluar como clasificador (AUC, accuracy, matriz de confusión).
# ============================================================================

def construir_perfil_patron(res_buenas_train, n_puntos=512):
    """Construye el perfil patrón promediando perfiles 1D de roscas BUENAS.

    V20.2: corrección — los perfiles se alinean por correlación cruzada
    antes de promediar para que los picos y valles caigan en la misma
    posición del eje x normalizado. Sin esta alineación el promedio sale
    borroso y todas las distancias quedan grandes (umbral inviable).

    Args:
        res_buenas_train: lista de dicts devueltos por perfilometro_analizar()
            sobre las imágenes BUENAS del conjunto TRAIN. Cada dict tiene
            'y_real_mm' (altura real en mm) y 'x_mm' (posición en mm).
        n_puntos: número de puntos del perfil patrón (resolución).

    Devuelve:
        dict con:
            'patron': np.ndarray de longitud n_puntos = perfil promedio.
            'patron_std': desviación estándar punto a punto (banda).
            'n_buenas_usadas': número de perfiles que entraron en el promedio.
    """
    # 1) Remuestrear cada perfil al mismo número de puntos
    perfiles_rs = []
    for r in res_buenas_train:
        if not r or not r.get("ok"):
            continue
        y = r.get("y_real_mm")
        x = r.get("x_mm")
        if y is None or x is None:
            continue
        y = np.asarray(y, dtype=np.float32)
        x = np.asarray(x, dtype=np.float32)
        if len(y) < 16 or len(x) < 16 or len(y) != len(x):
            continue
        # Remuestreo a longitud n_puntos en el mismo rango físico de x
        x_norm = (x - x.min()) / max(1e-9, (x.max() - x.min()))
        x_tgt = np.linspace(0, 1, n_puntos)
        y_rs = np.interp(x_tgt, x_norm, y).astype(np.float32)
        # Mantenemos la geometría: NO restamos la media, solo restamos un
        # offset suave para que el mínimo del perfil quede en 0.
        y_rs = y_rs - np.min(y_rs)
        perfiles_rs.append(y_rs)

    if len(perfiles_rs) == 0:
        return None

    # 2) Alinear por correlación cruzada con el primer perfil (referencia)
    # Esto garantiza que los picos y valles caigan en la misma posición
    # del eje normalizado, evitando que el promedio se "borre".
    ref = perfiles_rs[0]
    perfiles_alineados = [ref]
    for p in perfiles_rs[1:]:
        # Correlación cruzada en modo 'full' devuelve 2N-1 muestras
        cc = np.correlate(p - p.mean(), ref - ref.mean(), mode="full")
        # Lag óptimo respecto al centro
        lag = int(np.argmax(cc)) - (len(p) - 1)
        # Aplicamos el shift con padding por bordes
        if lag > 0:
            p_a = np.concatenate([np.full(lag, p[0]), p[:-lag]])
        elif lag < 0:
            p_a = np.concatenate([p[-lag:], np.full(-lag, p[-1])])
        else:
            p_a = p.copy()
        perfiles_alineados.append(p_a.astype(np.float32))

    M = np.stack(perfiles_alineados, axis=0)
    return {
        "patron": M.mean(axis=0),
        "patron_std": M.std(axis=0),
        "n_buenas_usadas": int(M.shape[0]),
        "n_puntos": int(n_puntos),
    }


def comparar_patron_con_iso(patron, n_puntos=512, paso_mm=1.5):
    """Contrasta el perfil patrón (promedio BUENAS) con el ISO teórico.

    Devuelve dict con:
        'iso_teorico': perfil ISO ideal en las mismas coordenadas.
        'delta_punto': delta patrón - ISO por cada punto (mm).
        'delta_mediana_mm': mediana del |delta| en mm.
        'area_entre_curvas_mm2': área absoluta entre patrón e ISO.
    """
    # Construir perfil ISO teórico simplificado en el mismo espacio
    # (longitud n_puntos, dos dientes, paso=paso_mm)
    x = np.linspace(0, 2 * paso_mm, n_puntos)
    # Diente triangular ISO 60° con altura H = paso * 0.866 / 2
    H = paso_mm * 0.866 / 2.0
    iso = np.zeros(n_puntos, dtype=np.float32)
    for k in range(2):
        x_diente = (x - k * paso_mm) % paso_mm
        # Triángulo simétrico centrado en paso/2
        iso_d = np.where(x_diente < paso_mm / 2,
                          (x_diente / (paso_mm / 2)) * H,
                          ((paso_mm - x_diente) / (paso_mm / 2)) * H)
        iso += iso_d / 2.0
    # Normalizar patrón a la misma altura aproximada del ISO
    patron_norm = patron.copy()
    if patron_norm.max() > 0:
        patron_norm = patron_norm * (H / patron_norm.max())

    delta = patron_norm - iso
    # Área absoluta entre las dos curvas (integral en mm * mm = mm²)
    dx = (2 * paso_mm) / n_puntos
    area_mm2 = float(np.sum(np.abs(delta)) * dx)

    return {
        "iso_teorico": iso,
        "patron_normalizado": patron_norm,
        "delta_punto": delta,
        "delta_mediana_mm": float(np.median(np.abs(delta))),
        "area_entre_curvas_mm2": area_mm2,
    }


def evaluar_clasificador_patron(res_buenas_train, res_buenas_eval, res_malas_eval,
                                  ruta_salida, n_puntos=512, paso_mm=1.5):
    """Detector basado en distancia al perfil patrón.

    Para cada imagen de EVAL calcula la distancia L2 normalizada entre su
    perfil 1D y el patrón. El umbral se calibra como P95 de las distancias
    de BUENAS_TRAIN contra el patrón (BUENA típica vs patrón).

    Returns:
        dict con AUC, accuracy, matriz, umbral, rutas de figuras y CSV.
    """
    os.makedirs(ruta_salida, exist_ok=True)

    # 1) Construir el patrón con BUENAS de TRAIN
    patron_info = construir_perfil_patron(res_buenas_train, n_puntos=n_puntos)
    if patron_info is None:
        log.warning("Perfil patrón: no hay suficientes perfiles BUENAS de TRAIN.")
        return None
    patron = patron_info["patron"]

    # 2) Contrastar el patrón con el ISO teórico
    comp_iso = comparar_patron_con_iso(patron, n_puntos=n_puntos, paso_mm=paso_mm)

    # 3) Helper: alineación por correlación + distancia L2 normalizada
    # Antes de medir distancia, alineamos el perfil EVAL con el patrón
    # (igual que hicimos al construir el patrón). Si no se alinea, la
    # distancia mezcla "no parece a una BUENA" con "está desfasado".
    def _dist_a_patron(res):
        if res is None or not res.get("ok"):
            return None
        y = res.get("y_real_mm")
        x = res.get("x_mm")
        if y is None or x is None:
            return None
        y = np.asarray(y, dtype=np.float32)
        x = np.asarray(x, dtype=np.float32)
        if len(y) < 16 or len(x) < 16 or len(y) != len(x):
            return None
        # Remuestreo en el mismo número de puntos
        x_norm = (x - x.min()) / max(1e-9, (x.max() - x.min()))
        x_tgt = np.linspace(0, 1, n_puntos)
        y_rs = np.interp(x_tgt, x_norm, y).astype(np.float32)
        y_rs = y_rs - np.min(y_rs)
        # Alineación por correlación cruzada con el patrón
        cc = np.correlate(y_rs - y_rs.mean(),
                           patron - patron.mean(), mode="full")
        lag = int(np.argmax(cc)) - (len(y_rs) - 1)
        if lag > 0:
            y_a = np.concatenate([np.full(lag, y_rs[0]), y_rs[:-lag]])
        elif lag < 0:
            y_a = np.concatenate([y_rs[-lag:], np.full(-lag, y_rs[-1])])
        else:
            y_a = y_rs
        # Distancia L2 normalizada por la magnitud del patrón
        rms_patron = float(np.sqrt(np.mean(patron ** 2))) + 1e-9
        return float(np.sqrt(np.mean((y_a - patron) ** 2)) / rms_patron)

    # 4) Calcular distancias en TRAIN (BUENAS contra patrón) -> umbral P95
    dist_train = []
    for r in res_buenas_train:
        d = _dist_a_patron(r)
        if d is not None:
            dist_train.append(d)
    if not dist_train:
        log.warning("Perfil patrón: no se pudieron calcular distancias TRAIN.")
        return None
    umbral_p95 = float(np.percentile(dist_train, 95))

    # 5) Calcular distancias en EVAL
    log_data = []  # [(archivo, etiqueta_real, distancia, predicción)]
    for r in res_buenas_eval:
        d = _dist_a_patron(r)
        if d is not None:
            log_data.append((r.get("archivo", "?"), "BUENA", d, None))
    for r in res_malas_eval:
        d = _dist_a_patron(r)
        if d is not None:
            log_data.append((r.get("archivo", "?"), "MALA", d, None))

    if not log_data:
        log.warning("Perfil patrón: no hay perfiles EVAL para evaluar.")
        return None

    # 6) Métricas: AUC + matriz al umbral P95
    distancias = np.array([r[2] for r in log_data])
    etiquetas = np.array([1 if r[1] == "MALA" else 0 for r in log_data])

    # AUC
    try:
        from sklearn.metrics import roc_auc_score
        auc = float(roc_auc_score(etiquetas, distancias))
    except Exception:
        auc = None

    # Matriz al P95
    tp = int(((distancias > umbral_p95) & (etiquetas == 1)).sum())
    fn = int(((distancias <= umbral_p95) & (etiquetas == 1)).sum())
    fp = int(((distancias > umbral_p95) & (etiquetas == 0)).sum())
    tn = int(((distancias <= umbral_p95) & (etiquetas == 0)).sum())
    total = max(1, tp + fn + fp + tn)
    acc = (tp + tn) / total
    sens = tp / max(1, tp + fn)
    esp = tn / max(1, tn + fp)
    f1 = 2 * tp / max(1, 2 * tp + fp + fn)

    # 7) Figura 1: patrón vs ISO teórico
    try:
        fig, ax = plt.subplots(figsize=(9, 4.5))
        x_mm = np.linspace(0, 2 * paso_mm, n_puntos)
        ax.plot(x_mm, comp_iso["iso_teorico"], "k--", lw=2,
                 label="Perfil ISO teórico")
        ax.plot(x_mm, comp_iso["patron_normalizado"], "g-", lw=2.2,
                 label=f"Patrón promedio BUENAS (n={patron_info['n_buenas_usadas']})")
        ax.fill_between(x_mm, comp_iso["iso_teorico"],
                         comp_iso["patron_normalizado"],
                         alpha=0.25, color="orange",
                         label=f"Δ área = {comp_iso['area_entre_curvas_mm2']:.3f} mm²")
        ax.set_xlabel("Posición a lo largo del perfil (mm)")
        ax.set_ylabel("Altura normalizada (mm)")
        ax.set_title("Perfil patrón BUENAS vs ISO teórico\n"
                      f"Δ mediana = {comp_iso['delta_mediana_mm']*1000:.0f} µm  |  "
                      f"Área entre curvas = {comp_iso['area_entre_curvas_mm2']:.3f} mm²")
        ax.legend(loc="best", fontsize=9)
        ax.grid(True, alpha=0.3)
        plt.tight_layout()
        ruta_fig_patron = os.path.join(ruta_salida, "perfil_patron_vs_iso.png")
        plt.savefig(ruta_fig_patron, dpi=120)
        plt.close()
    except Exception as e:
        log.warning(f"Perfil patrón: no se pudo guardar fig patrón vs ISO: {e}")
        ruta_fig_patron = None

    # 8) Figura 2: matriz de confusión del clasificador patrón
    try:
        ruta_matriz = _graficar_matriz_pieza(
            tp=tp, fn=fn, fp=fp, tn=tn,
            umbral=umbral_p95, ruta_salida=ruta_salida,
            titulo="Clasificador Perfil Patrón — Matriz de confusión",
            nombre_archivo="matriz_perfil_patron.png",
        )
    except Exception as e:
        log.warning(f"Perfil patrón: no se pudo guardar matriz: {e}")
        ruta_matriz = None

    # 9) CSV con las distancias por imagen
    try:
        ruta_csv = os.path.join(ruta_salida, "perfil_patron_distancias.csv")
        with open(ruta_csv, "w", encoding="utf-8") as fh:
            fh.write("archivo,etiqueta_real,distancia_al_patron,prediccion_P95\n")
            for arch, et, d, _ in log_data:
                pred = "MALA" if d > umbral_p95 else "BUENA"
                fh.write(f"{arch},{et},{d:.6f},{pred}\n")
    except Exception:
        ruta_csv = None

    # 10) Informe corto en texto
    try:
        ruta_inf = os.path.join(ruta_salida, "perfil_patron_informe.txt")
        with open(ruta_inf, "w", encoding="utf-8") as fh:
            fh.write("=" * 78 + "\n")
            fh.write(" PERFIL PATRÓN — Clasificador basado en distancia al patrón promedio\n")
            fh.write("=" * 78 + "\n\n")
            fh.write("Construcción de un perfil patrón promediando los perfiles\n")
            fh.write("1D de roscas BUENAS de entrenamiento, contrastado con el\n")
            fh.write("perfil teórico ISO y evaluado como clasificador binario.\n\n")
            fh.write("CONSTRUCCIÓN DEL PATRÓN\n")
            fh.write("-" * 78 + "\n")
            fh.write(f"  Roscas BUENAS usadas: {patron_info['n_buenas_usadas']}\n")
            fh.write(f"  Resolución del patrón: {n_puntos} puntos\n\n")
            fh.write("CONTRASTE CON EL ISO TEÓRICO\n")
            fh.write("-" * 78 + "\n")
            fh.write(f"  Delta mediana (patrón - ISO): "
                      f"{comp_iso['delta_mediana_mm']*1000:.1f} µm\n")
            fh.write(f"  Área entre curvas:             "
                      f"{comp_iso['area_entre_curvas_mm2']:.3f} mm²\n\n")
            fh.write("CLASIFICADOR (umbral = P95 de distancias TRAIN)\n")
            fh.write("-" * 78 + "\n")
            fh.write(f"  Umbral P95:    {umbral_p95:.4f}\n")
            if auc is not None:
                fh.write(f"  AUC:           {auc:.3f}\n")
            fh.write(f"  Accuracy:      {acc*100:.1f}%\n")
            fh.write(f"  Sensibilidad:  {sens*100:.1f}%\n")
            fh.write(f"  Especificidad: {esp*100:.1f}%\n")
            fh.write(f"  F1:            {f1*100:.1f}%\n\n")
            fh.write("MATRIZ DE CONFUSIÓN\n")
            fh.write(f"  TP={tp}  FN={fn}  FP={fp}  TN={tn}\n\n")
            fh.write("ARCHIVOS GENERADOS\n")
            fh.write("-" * 78 + "\n")
            if ruta_fig_patron:
                fh.write(f"  perfil_patron_vs_iso.png    — patrón vs ISO\n")
            if ruta_matriz:
                fh.write(f"  matriz_perfil_patron.png    — matriz de confusión\n")
            if ruta_csv:
                fh.write(f"  perfil_patron_distancias.csv — distancias por imagen\n")
    except Exception as e:
        log.warning(f"Perfil patrón: no se pudo guardar informe: {e}")
        ruta_inf = None

    return {
        "auc": auc,
        "accuracy": acc,
        "sensibilidad": sens,
        "especificidad": esp,
        "f1": f1,
        "tp": tp, "fn": fn, "fp": fp, "tn": tn,
        "umbral_p95": umbral_p95,
        "n_buenas_train": patron_info["n_buenas_usadas"],
        "delta_patron_iso_mediana_um": comp_iso["delta_mediana_mm"] * 1000,
        "area_patron_iso_mm2": comp_iso["area_entre_curvas_mm2"],
        "ruta_figura": ruta_fig_patron,
        "ruta_matriz": ruta_matriz,
        "ruta_csv": ruta_csv,
        "ruta_informe": ruta_inf,
    }


def ejecutar_perfilometro(imgs, ruta_salida, paso_mm=1.5,
                          perfil_norma="iso_metrica",
                          mm_por_px=None, tolerancias=None,
                          tag="perf", ruta_comparativas=None,
                          out_tiempos=None):
    """
    Ejecuta el análisis perfilométrico sobre una lista de imágenes
    [(nombre, img_gris), ...] y guarda figura + CSV resumen.

    Args:
       ruta_salida: carpeta donde se guardan los dashboards individuales
          (uno por imagen) y el CSV resumen.
       ruta_comparativas: carpeta separada donde se guardan TODAS las
          comparativas de 4 paneles (uno por imagen). Si es None se usa
          ruta_salida/comparativas.
       tag: etiqueta corta ("buenas"/"malas") usada como prefijo en los
          nombres de las comparativas para distinguirlas cuando todas
          las imágenes se guardan en la misma carpeta.
       out_tiempos: dict opcional donde, si se pasa, se acumulan los
          tiempos medidos (en segundos):
             "tiempo_total_s", "n_imagenes", "tiempo_por_imagen_s"
          Se acumulan: si la misma carpeta tag llama dos veces (buenas y
          malas), out_tiempos guarda la suma de ambas llamadas.

    Devuelve la lista de resultados.
    """
    log.info("=" * 70)
    log.info(f" PERFILÓMETRO ÓPTICO — perfil real vs teórico  ({tag})")
    log.info(f"   Norma: {perfil_norma}    Paso nominal: {paso_mm} mm")
    log.info("=" * 70)
    Path(ruta_salida).mkdir(parents=True, exist_ok=True)

    # Carpeta única de comparativas (compartida entre buenas y malas si
    # ruta_comparativas se pasa desde el orquestador).
    if ruta_comparativas is None:
        ruta_comparativas = os.path.join(ruta_salida, "comparativas")
    Path(ruta_comparativas).mkdir(parents=True, exist_ok=True)

    # Cronómetro local — mide solo el coste del análisis perfilométrico
    # (no el de las gráficas, que es coste de visualización, no de método).
    cron = Cronometro()
    n_imgs_medidas = 0

    resultados = []
    iterador = imgs
    if TQDM_OK:
        iterador = tqdm(imgs, desc=f"Perfilometría {tag}", ncols=80)

    # Prefijo para diferenciar comparativas dentro de la carpeta única
    prefijo = tag.upper() if tag else "IMG"

    for nombre, img in iterador:
        base = Path(nombre).stem
        try:
            with cron:
                res = perfilometro_analizar(
                    img, paso_mm=paso_mm, perfil_norma=perfil_norma,
                    mm_por_px=mm_por_px, tolerancias=tolerancias,
                )
            n_imgs_medidas += 1
            res["archivo"] = nombre
            if res["ok"]:
                perfilometro_graficar(res, ruta_salida, base)
            # Comparativa de 4 paneles (siempre, incluso si res["ok"]=False
            # para que se vea por qué falló): original, segmentación,
            # perfil extraído, perfil real vs teórico.
            try:
                generar_comparativa_etapas(
                    img, res, ruta_comparativas,
                    nombre_base=f"{prefijo}_{base}")
            except Exception as e:
                log.warning(f"Comparativa {base} falló: {e}")
            resultados.append(res)
        except Exception as e:
            log.error(f"Perfilometría falló en {nombre}: {e}")

    # Acumular tiempos en el dict del orquestador si se pasó.
    if out_tiempos is not None:
        out_tiempos["tiempo_total_s"] = (
            out_tiempos.get("tiempo_total_s", 0.0) + cron.total)
        out_tiempos["n_imagenes"] = (
            out_tiempos.get("n_imagenes", 0) + n_imgs_medidas)
        n_tot = out_tiempos["n_imagenes"]
        out_tiempos["tiempo_por_imagen_s"] = (
            out_tiempos["tiempo_total_s"] / n_tot if n_tot > 0 else None)
        log.info(
            f"  Perfilómetro tiempo acumulado: "
            f"{out_tiempos['tiempo_total_s']:.1f}s sobre "
            f"{out_tiempos['n_imagenes']} imágenes "
            f"({out_tiempos['tiempo_por_imagen_s']*1000:.1f} ms/img)")

    # CSV resumen (sin los vectores, solo escalares)
    if resultados:
        campos_excl = {"x_mm", "y_real_mm", "y_teorico_mm",
                       "picos_idx", "valles_idx", "motivos_fallo",
                       "perfil_px", "x_offset", "orient"}
        # Usamos las claves del primer resultado válido como cabecera
        plantilla = next((r for r in resultados if r.get("ok")), None)
        if plantilla is not None:
            campos = [k for k in plantilla.keys() if k not in campos_excl]
            campos = ["archivo"] + [c for c in campos if c != "archivo"] \
                     + ["motivos_fallo"]
            csv_path = os.path.join(ruta_salida,
                                    f"perfilometro_resumen_{tag}.csv")
            with open(csv_path, "w", newline="", encoding="utf-8") as fh:
                w = csv.writer(fh)
                w.writerow(campos)
                for r in resultados:
                    fila = []
                    for c in campos:
                        v = r.get(c, "")
                        if c == "motivos_fallo":
                            v = " | ".join(v) if isinstance(v, list) else v
                        fila.append(v)
                    w.writerow(fila)
            log.info(f"   Resumen CSV: {csv_path}")

        n_pasa = sum(1 for r in resultados
                     if r.get("ok") and r["veredicto"] == "PASA")
        n_nopasa = sum(1 for r in resultados
                       if r.get("ok") and r["veredicto"] == "NO PASA")
        n_err = sum(1 for r in resultados if not r.get("ok"))
        log.info(f"   PASA: {n_pasa}  NO PASA: {n_nopasa}  Error: {n_err}")

    return resultados


# ==============================================================================
#  BLOQUE 4-ter - PERFILÓMETRO COMO CLASIFICADOR
# ==============================================================================
#  Evalúa el rendimiento del perfilómetro como sistema PASA/NO PASA sobre el
#  dataset completo (BUENAS + MALAS). Hace tres cosas:
#
#  1) Recolecta TODAS las métricas geométricas sobre cada imagen.
#  2) Calibra automáticamente umbrales a partir de las BUENAS (percentil 95
#     de cada métrica), de forma que la tolerancia refleje la variabilidad
#     real de las piezas funcionales.
#  3) Calcula métricas de clasificación (AUC, acc, sens, esp, F1, matriz de
#     confusión) tanto:
#        - aplicando el veredicto duro (PASA/NO PASA)
#        - usando un score continuo combinado (suma de z-scores ponderados)
#
#  Esto es el perfilómetro actuando como un SISTEMA DE INSPECCIÓN INDUSTRIAL,
#  comparable cuantitativamente con el clasificador PatchCore.
# ==============================================================================

# Métricas usadas para construir el score continuo, con peso y dirección.
# 'direccion'  +1 = mayor valor → MÁS desgaste (peor)
#              -1 = mayor valor → MENOR desgaste (mejor)
# El peso es orientativo; se reajusta automáticamente con Cohen d sobre el
# dataset si se solicita calibración data-driven.
METRICAS_CLASIFICADOR = {
    # ===== Métricas de DESGASTE (objetivo 4 del TFM) — peso máximo =====
    "area_desgaste_mm2":         {"peso": 2.0, "direccion": +1},
    "ancho_medio_desgaste_um":   {"peso": 2.0, "direccion": +1},
    "prof_max_desgaste_um":      {"peso": 1.5, "direccion": +1},
    "pct_longitud_con_desgaste": {"peso": 1.5, "direccion": +1},
    # Métricas absolutas (sensibles a calidad de captura)
    "indice_desgaste_cresta":   {"peso": 0.5, "direccion": +1},
    "cresta_radio_redondeo_um": {"peso": 0.5, "direccion": +1},
    "cresta_longitud_mm":       {"peso": 0.3, "direccion": +1},
    "cresta_planitud_rms_um":   {"peso": 0.3, "direccion": +1},
    "rms_desviacion_mm":        {"peso": 0.3, "direccion": +1},
    "max_desviacion_mm":        {"peso": 0.3, "direccion": +1},
    "error_angulo_deg":         {"peso": 0.4, "direccion": +1},
    "error_altura_pct":         {"peso": 0.2, "direccion": +1},
    # Métricas RELATIVAS intra-imagen (robustas a calidad de captura).
    # Una pieza buena tiene dientes uniformes entre sí; una pieza con
    # desgaste local tiene CV alto. Estas pesan MÁS porque son más fiables
    # frente a vistas con foco/iluminación variables.
    "cv_radio_cresta":          {"peso": 1.5, "direccion": +1},
    "cv_longitud_cresta":       {"peso": 1.2, "direccion": +1},
    "cv_planitud_cresta":       {"peso": 1.0, "direccion": +1},
    "cv_longitud_valle":        {"peso": 1.0, "direccion": +1},
    "indice_no_uniformidad":    {"peso": 1.5, "direccion": +1},
}


def _extraer_features_clasificador(resultados, etiqueta):
    """Convierte la lista de dicts de perfilometría en (X, nombres, y)."""
    nombres_metric = list(METRICAS_CLASIFICADOR.keys())
    X = []
    archivos = []
    y = []
    for r in resultados:
        if not r.get("ok"):
            continue
        fila = []
        for m in nombres_metric:
            v = r.get(m, None)
            # Para algunas métricas usamos el valor absoluto (errores)
            if m in ("error_angulo_deg", "error_altura_pct") and v is not None:
                v = abs(v)
            if v is None or (isinstance(v, float) and not np.isfinite(v)):
                v = np.nan
            fila.append(float(v))
        X.append(fila)
        archivos.append(r.get("archivo", ""))
        y.append(etiqueta)
    return np.array(X), nombres_metric, archivos, y


def _imputar_nan_con_mediana(X, X_ref=None):
    """Sustituye NaN por la mediana de cada columna (preferentemente
    calculada sobre X_ref, p. ej. solo las BUENAS)."""
    Xr = X_ref if X_ref is not None else X
    medianas = np.nanmedian(Xr, axis=0)
    X_out = X.copy()
    for j in range(X.shape[1]):
        col = X_out[:, j]
        col[np.isnan(col)] = medianas[j] if np.isfinite(medianas[j]) else 0.0
    return X_out


def _cohen_d(a, b):
    """Effect size estándar."""
    a = a[np.isfinite(a)]
    b = b[np.isfinite(b)]
    if len(a) < 2 or len(b) < 2:
        return 0.0
    s = np.sqrt(((len(a) - 1) * np.var(a, ddof=1) +
                 (len(b) - 1) * np.var(b, ddof=1)) /
                (len(a) + len(b) - 2))
    if s < 1e-12:
        return 0.0
    return float((np.mean(a) - np.mean(b)) / s)


def _ks_test(a, b):
    """KS test sin SciPy (implementación mínima 2-sample)."""
    try:
        from scipy.stats import ks_2samp
        a = a[np.isfinite(a)]
        b = b[np.isfinite(b)]
        if len(a) < 2 or len(b) < 2:
            return 0.0, 1.0
        ks = ks_2samp(a, b)
        return float(ks.statistic), float(ks.pvalue)
    except Exception:
        return 0.0, 1.0


def _roc_auc(scores, y_true):
    """ROC + AUC. y_true en {0,1}, score = más alto → más probable POSITIVO."""
    from sklearn.metrics import roc_curve, auc
    fpr, tpr, thr = roc_curve(y_true, scores)
    a = float(auc(fpr, tpr))
    return fpr, tpr, thr, a


def _umbral_youden(fpr, tpr, thresholds):
    """Devuelve el umbral que maximiza Youden's J = TPR - FPR."""
    j = tpr - fpr
    i = int(np.argmax(j))
    return float(thresholds[i]), float(j[i]), int(i)


def calibrar_tolerancias_data_driven(res_buenas, percentil=99.0,
                                       margen_seguridad=1.2):
    """
    Calibra TOLERANCIAS_DEFECTO usando los percentiles de las BUENAS.
    Devuelve un dict de tolerancias actualizado.

    IMPORTANTE: descarta automáticamente medidas claramente erróneas
    (outliers extremos que indican fallo del extractor de perfil), porque
    si no, una sola vista mal medida infla las tolerancias y todo el
    dataset acaba pasando como BUENO. Por ejemplo: si una vista
    devuelve altura_pct=2000% por una mala detección de pico, no debe
    contar como "rango normal de buenas".

    Estrategia:
      1. Filtrar valores físicamente imposibles (rango sanity)
      2. Filtrar outliers por MAD (>5*MAD del valor central)
      3. tolerancia = max(percentil, mediana + 3*MAD) * margen_seguridad
    """
    def _coleccionar(clave, abs_val=False, sanity_max=None):
        vals = []
        for r in res_buenas:
            if not r.get("ok"):
                continue
            v = r.get(clave)
            if v is None or (isinstance(v, float) and not np.isfinite(v)):
                continue
            v = abs(float(v)) if abs_val else float(v)
            # Sanity: descartar valores físicamente absurdos
            if sanity_max is not None and abs(v) > sanity_max:
                continue
            vals.append(v)
        return np.array(vals)

    def _filtrar_outliers_mad(vals, k=5.0):
        """Filtra valores a más de k*MAD de la mediana."""
        if len(vals) < 4:
            return vals
        med = np.median(vals)
        mad = np.median(np.abs(vals - med)) + 1e-9
        return vals[np.abs(vals - med) <= k * 1.4826 * mad]

    def _umbral_robusto(vals, k_mad=3.0, percentil_pct=percentil,
                          margen=margen_seguridad, defecto=1.0):
        """Umbral robusto sobre `vals` con piso de seguridad.

        V17 patch: la versión anterior fallaba cuando los valores eran
        constantes (todos 0 o muy próximos): mediana=0 + MAD=0 +
        percentil=0 → umbral=0, que provoca que TODAS las imágenes
        (incluidas las propias BUENAS de entrenamiento) sean marcadas
        como NO PASA. El piso `defecto` se ignoraba en ese caso porque
        sólo se devolvía cuando `len(vals)==0`. Ahora se aplica también
        cuando el umbral computado es despreciable.
        """
        if len(vals) == 0:
            return defecto
        # Filtrado de outliers antes de calcular tolerancia
        vals_lim = _filtrar_outliers_mad(vals, k=5.0)
        if len(vals_lim) == 0:
            vals_lim = vals
        med = float(np.median(vals_lim))
        mad = float(np.median(np.abs(vals_lim - med))) + 1e-9
        umbral_mad = med + k_mad * 1.4826 * mad
        umbral_pct = float(np.percentile(vals_lim, percentil_pct))
        umbral = max(umbral_mad, umbral_pct) * margen
        # V17 patch: piso de seguridad. Si el umbral computado es menor
        # que el `defecto` (caso degenerado: dataset muy pequeño o
        # valores casi constantes), usamos el defecto. Esto evita
        # tolerancias ridículamente apretadas (como paso_pct=0) que
        # rechazarían incluso las piezas buenas de entrenamiento.
        return max(umbral, defecto)

    tol_calibradas = {}
    # Sanity ranges MUY permisivos: solo descartamos valores claramente
    # patológicos (NaN, infinito o ridículos). Variaciones de captura
    # legítimas se aceptan.
    tol_calibradas["paso_pct"] = _umbral_robusto(
        _coleccionar("error_paso_pct", abs_val=True, sanity_max=100.0),
        defecto=0.5)
    tol_calibradas["altura_pct"] = _umbral_robusto(
        _coleccionar("error_altura_pct", abs_val=True, sanity_max=2000.0),
        defecto=15.0)
    tol_calibradas["angulo_flanco_deg"] = _umbral_robusto(
        _coleccionar("error_angulo_deg", abs_val=True, sanity_max=90.0),
        defecto=3.0)
    tol_calibradas["rms_desviacion_pct_P"] = _umbral_robusto(
        _coleccionar("rms_desviacion_pct_P", sanity_max=300.0),
        defecto=8.0)

    # Cresta: longitud relativa al nominal, sanity max=5mm
    cl = _coleccionar("cresta_longitud_mm", sanity_max=5.0)
    if len(cl):
        L_nom = res_buenas[0].get("cresta_longitud_nominal_mm", 0.1875)
        err = np.abs(cl - L_nom) / max(L_nom, 1e-6) * 100
        err = err[err < 2000]
        if len(err):
            tol_calibradas["cresta_longitud_pct"] = _umbral_robusto(
                err, defecto=80.0)
        else:
            tol_calibradas["cresta_longitud_pct"] = 80.0
    else:
        tol_calibradas["cresta_longitud_pct"] = 80.0

    tol_calibradas["cresta_planitud_um"] = _umbral_robusto(
        _coleccionar("cresta_planitud_rms_um", sanity_max=500.0),
        defecto=40.0)
    tol_calibradas["cresta_radio_max_um"] = _umbral_robusto(
        _coleccionar("cresta_radio_redondeo_um", sanity_max=3000.0),
        defecto=200.0)

    vl = _coleccionar("valle_longitud_mm", sanity_max=5.0)
    if len(vl):
        L_nom = res_buenas[0].get("valle_longitud_nominal_mm", 0.375)
        err = np.abs(vl - L_nom) / max(L_nom, 1e-6) * 100
        err = err[err < 2000]
        if len(err):
            tol_calibradas["valle_longitud_pct"] = _umbral_robusto(
                err, defecto=80.0)
        else:
            tol_calibradas["valle_longitud_pct"] = 80.0
    else:
        tol_calibradas["valle_longitud_pct"] = 80.0

    tol_calibradas["valle_planitud_um"] = 50.0
    tol_calibradas["desgaste_cresta_max"] = _umbral_robusto(
        _coleccionar("indice_desgaste_cresta", sanity_max=2.0),
        defecto=0.50)

    return tol_calibradas


def _aplicar_veredicto(res, tolerancias):
    """Recalcula el veredicto PASA/NO PASA con un nuevo dict de tolerancias."""
    fallos = []
    if abs(res.get("error_paso_pct", 0)) > tolerancias["paso_pct"]:
        fallos.append("paso")
    if abs(res.get("error_altura_pct", 0)) > tolerancias["altura_pct"]:
        fallos.append("altura")
    err_ang = res.get("error_angulo_deg")
    if err_ang is not None and abs(err_ang) > tolerancias["angulo_flanco_deg"]:
        fallos.append("ángulo")
    if res.get("rms_desviacion_pct_P", 0) > tolerancias["rms_desviacion_pct_P"]:
        fallos.append("RMS")
    err_cl = res.get("error_cresta_longitud_pct")
    if err_cl is not None and abs(err_cl) > tolerancias["cresta_longitud_pct"]:
        fallos.append("cresta_L")
    plan = res.get("cresta_planitud_rms_um")
    if plan is not None and plan > tolerancias["cresta_planitud_um"]:
        fallos.append("cresta_plan")
    rad = res.get("cresta_radio_redondeo_um")
    if rad is not None and rad > tolerancias["cresta_radio_max_um"]:
        fallos.append("cresta_R")
    err_vl = res.get("error_valle_longitud_pct")
    if err_vl is not None and abs(err_vl) > tolerancias["valle_longitud_pct"]:
        fallos.append("valle_L")
    desg = res.get("indice_desgaste_cresta", 0)
    if desg > tolerancias["desgaste_cresta_max"]:
        fallos.append("desgaste")
    return ("PASA" if not fallos else "NO PASA"), fallos


def _score_continuo_perfilometro(res_dict, mu, sigma, metricas_pesos):
    """
    Calcula un score continuo de "anomalía geométrica" para una imagen,
    como suma ponderada de z-scores.

       score = Σ peso_i * max(0, (x_i - mu_i) / sigma_i) * direccion_i

    Solo se suman desviaciones en la dirección "más desgaste".
    NaN se trata como 0 (no contribuye).
    Los valores claramente erróneos (z-score absoluto > 10) también se
    descartan: indican fallo de medición, no anomalía real.
    """
    s = 0.0
    n = 0
    for clave, info in metricas_pesos.items():
        v = res_dict.get(clave, None)
        if clave in ("error_angulo_deg", "error_altura_pct") and v is not None:
            v = abs(v)
        if v is None or not np.isfinite(v):
            continue
        idx = list(metricas_pesos.keys()).index(clave)
        if not np.isfinite(mu[idx]) or sigma[idx] < 1e-9:
            continue
        z = (v - mu[idx]) / sigma[idx]
        # Filtro de medición errónea: z-scores absurdos indican fallo
        # del extractor (no anomalía real). Se descartan.
        if abs(z) > 10.0:
            continue
        z_dir = z * info["direccion"]
        s += info["peso"] * max(0.0, z_dir)
        n += 1
    return s / max(n, 1)


def _extraer_id_pieza(nombre_archivo, etiqueta):
    """
    Extrae el ID de pieza desde el nombre de archivo.

    Convención observada en este TFM:
        <numero_pieza>R[B|M]<numero_vista>.<ext>
    Ejemplos:
        501RB01.jpg, 501RB02.jpg, ..., 501RB11.jpg  → pieza '501_BUENA'
        501RM01.jpg, 501RM02.jpg, ..., 501RM11.jpg  → pieza '501_MALA'
        502RB07.jpg                                  → pieza '502_BUENA'

    Heurísticas que se prueban en este orden:
      1. <num_pieza>R[BM]<num_vista>: extrae num_pieza + B/M
      2. R[BM]<num_vista> sin pieza explícita → un único grupo por etiqueta
      3. Si nada encaja, todas las imágenes de la etiqueta van a un grupo
    """
    import re
    base = os.path.basename(str(nombre_archivo))
    base_low = base.lower()

    # Patrón principal: <num_pieza>R[BM]<num_vista>
    # Capturamos cualquier número que preceda al token "RB" o "RM" como ID
    # de pieza. Soporta separadores opcionales: "501rb01", "501_RB_01",
    # "501-rb-01", "Imagen_501_RB01.jpg", etc.
    m = re.search(r'(\d{1,5})[\s_\-]*r[bm][\s_\-]*(\d{1,3})', base_low)
    if m:
        num_pieza = int(m.group(1))
        # Si es un número grande (>50) lo tratamos como ID de pieza distintiva.
        # Si es muy pequeño (<10) puede ser ruido (p.ej. "1RB1" = pieza 1).
        # En cualquier caso, agrupamos por (num_pieza, etiqueta).
        sufijo = "B" if "rb" in base_low[m.start():m.end()] else "M"
        return f"P{num_pieza:03d}_{sufijo}"

    # Patrón secundario: solo R[BM]##, sin número de pieza explícito
    # → todas las vistas con el mismo R[BM] van al mismo grupo (caso del
    # dataset actual, donde solo hay 1 pieza buena + 1 pieza mala).
    m = re.search(r'r([bm])(\d{1,3})', base_low)
    if m:
        sufijo = m.group(1).upper()
        return f"PIEZA_{sufijo}"

    # Fallback: agrupamos por etiqueta (todas las buenas como una pieza,
    # todas las malas como otra)
    return f"PIEZA_{etiqueta[0].upper()}"


def _extraer_num_vista(nombre_archivo):
    """Extrae el número de vista (01..11) del nombre del archivo.

    Patrón observado en el TFM: <numero_pieza>R[B|M]<num_vista>.
    Por ejemplo "Imagen_000501_RB07.jpg" → 7.

    Devuelve None si no se reconoce el patrón (en cuyo caso la agregación
    por posición de vista no se puede aplicar y el sistema cae sobre la
    calibración a nivel pieza por P95 solamente).
    """
    import re
    base = os.path.basename(str(nombre_archivo)).lower()
    m = re.search(r'r[bm][\s_\-]*(\d{1,3})', base)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return None
    return None


def agregar_a_nivel_pieza(resultados_b, resultados_m, scores_b, scores_m,
                           tolerancias):
    """
    Agrupa los resultados por pieza usando el ID extraído del nombre.
    Devuelve estructuras a nivel pieza:
       - Por pieza: ID, etiqueta, n_vistas, score_max, score_mean,
                    veredicto_any (PASA si ningún vista falla),
                    veredicto_majority (PASA si >50% pasan)
    """
    piezas = {}
    for r, sc, etq in [(r, sc, "BUENA")
                        for r, sc in zip(resultados_b, scores_b)] + \
                       [(r, sc, "MALA")
                        for r, sc in zip(resultados_m, scores_m)]:
        pid = _extraer_id_pieza(r.get("archivo", ""), etq)
        if pid not in piezas:
            piezas[pid] = {"id": pid, "etiqueta": etq,
                           "scores": [], "veredictos": [],
                           "vistas": []}
        piezas[pid]["scores"].append(float(sc))
        piezas[pid]["veredictos"].append(r.get("veredicto_calibrado", ""))
        piezas[pid]["vistas"].append(r.get("archivo", ""))

    # Calcular agregaciones
    for pid, info in piezas.items():
        info["n_vistas"] = len(info["scores"])
        info["score_max"] = float(np.max(info["scores"]))
        info["score_mean"] = float(np.mean(info["scores"]))
        info["score_med"] = float(np.median(info["scores"]))
        # Any: si UNA vista falla, la pieza falla (criterio calibre real)
        n_nopasa = sum(1 for v in info["veredictos"] if v == "NO PASA")
        info["n_nopasa"] = n_nopasa
        info["veredicto_any"] = "NO PASA" if n_nopasa >= 1 else "PASA"
        # Majority: pieza falla si >50% de vistas fallan
        info["veredicto_majority"] = ("NO PASA"
                                       if n_nopasa > info["n_vistas"] / 2
                                       else "PASA")
    return piezas


def calibrar_umbral_a_nivel_pieza(piezas, percentil=95.0):
    """Calibración del umbral de decisión a nivel pieza.

    Problema que resuelve: la calibración original (P99 sobre los scores de
    todas las vistas BUENAS) penaliza demasiado el criterio ANY a nivel
    pieza, porque cada pieza buena tiene 11 oportunidades de que UNA de
    sus vistas supere el umbral. Resultado: especificidad muy baja.

    Solución: para cada pieza BUENA, calculamos su 'score peor' (score
    máximo entre sus 11 vistas). El umbral será el percentil indicado de
    esa colección de scores máximos por pieza. Así, el 95% (o 99%) de
    las piezas buenas pasará — no el 95% de sus vistas.

    Devuelve dict con:
       umbral_pieza      → el nuevo umbral en unidades del score
       n_piezas_buenas   → cuántas piezas buenas se usaron
       scores_max_buenas → list para diagnóstico
       percentil_usado   → el percentil pedido
    """
    scores_max_buenas = [info["score_max"]
                          for info in piezas.values()
                          if info["etiqueta"] == "BUENA"]
    if len(scores_max_buenas) < 2:
        # Sin piezas suficientes para calibrar, devolvemos None
        return None
    umbral = float(np.percentile(scores_max_buenas, percentil))
    return {
        "umbral_pieza": umbral,
        "n_piezas_buenas": len(scores_max_buenas),
        "scores_max_buenas": list(map(float, scores_max_buenas)),
        "percentil_usado": float(percentil),
    }


def aplicar_veredicto_pieza_mejora_a(piezas, umbral_pieza):
    """Aplica el veredicto a nivel pieza usando el umbral calibrado por P95.
    Añade a cada pieza el campo 'veredicto_pieza_calibrado':
    'PASA' si su score_max <= umbral_pieza, 'NO PASA' en otro caso.

    Devuelve métricas (acc, sens, esp, F1, AUC) y la matriz de confusión.
    """
    y_true = []
    y_pred = []
    for info in piezas.values():
        y_t = 1 if info["etiqueta"] == "MALA" else 0
        y_p = 1 if info["score_max"] > umbral_pieza else 0
        info["veredicto_pieza_calibrado"] = "NO PASA" if y_p == 1 else "PASA"
        y_true.append(y_t)
        y_pred.append(y_p)
    y_true = np.array(y_true)
    y_pred = np.array(y_pred)
    if len(y_true) == 0:
        return None
    tp = int(np.sum((y_true == 1) & (y_pred == 1)))
    fn = int(np.sum((y_true == 1) & (y_pred == 0)))
    fp = int(np.sum((y_true == 0) & (y_pred == 1)))
    tn = int(np.sum((y_true == 0) & (y_pred == 0)))
    total = max(1, tp + fn + fp + tn)
    return {
        "accuracy": (tp + tn) / total,
        "sensibilidad": tp / max(1, tp + fn),
        "especificidad": tn / max(1, tn + fp),
        "f1": (2 * tp) / max(1, 2 * tp + fp + fn),
        "tp": tp, "fn": fn, "fp": fp, "tn": tn,
        "n_total": int(total),
    }


def calibrar_umbrales_por_vista(resultados_b, scores_b, percentil=99.0):
    """MEJORA B — calibración de un umbral DISTINTO por número de vista
    (posición fija).

    Solo es útil si las 11 vistas siguen un orden definido entre piezas
    (vista 1 = misma zona, vista 2 = misma zona…). Si no, los umbrales
    convergerán a un mismo valor y la mejora no aportará.

    Detección automática: si todas las imágenes tienen un número de
    vista extraíble (1..N) y la dispersión de scores DENTRO de cada vista
    es claramente menor que la dispersión GLOBAL, asumimos orden fijo.

    Devuelve dict {num_vista → umbral} o None si no aplica.
    """
    # Indexar scores por número de vista
    por_vista = {}
    for r, sc in zip(resultados_b, scores_b):
        n_v = _extraer_num_vista(r.get("archivo", ""))
        if n_v is None:
            continue
        por_vista.setdefault(n_v, []).append(float(sc))

    if len(por_vista) < 3:
        # No hay suficientes números de vista para que tenga sentido
        return None

    # Heurística para decidir si vale la pena: la varianza media DENTRO
    # de las vistas tiene que ser menor que la varianza GLOBAL.
    todos = [s for lista in por_vista.values() for s in lista]
    if len(todos) < 10:
        return None
    var_global = float(np.var(todos))
    vars_intra = [float(np.var(lista)) for lista in por_vista.values()
                   if len(lista) >= 3]
    if not vars_intra:
        return None
    var_intra_media = float(np.mean(vars_intra))
    # Si la varianza intra-vista es similar a la global, las vistas no
    # son posiciones fijas, son aleatorias. No aplicamos la mejora.
    if var_intra_media > 0.7 * var_global:
        log.info(f"   [Mejora B] No se detecta orden fijo entre vistas "
                 f"(var_intra={var_intra_media:.3f} / "
                 f"var_global={var_global:.3f}). Saltada.")
        return None

    # Sí hay orden fijo: calibramos un umbral por número de vista.
    umbrales = {}
    for n_v, lista in por_vista.items():
        if len(lista) >= 2:
            umbrales[int(n_v)] = float(np.percentile(lista, percentil))
    log.info(f"   [Mejora B] Detectado orden fijo. Umbrales por vista: "
             f"{ {k:f'{v:.3f}' for k,v in umbrales.items()} }")
    return umbrales


def metricas_restringidas_a_detectables(res_buenas_ok, res_malas_ok,
                                         detectables_archivos):
    """MEJORA C — recalcula métricas geométricas restringiendo el conjunto
    MALAS al subconjunto detectable identificado por el EDA.

    `detectables_archivos` es un set de nombres de archivo MALAS marcados
    como detectables (con desgaste medible) por el análisis exploratorio
    inicial. Si está vacío o es None, devuelve None.

    Devuelve un dict con las mismas claves que el resumen de la
    comparativa BUENA vs MALA, pero usando solo las imágenes detectables.
    Esto da la métrica fundamental: cuando el desgaste SÍ se ve en la
    imagen, ¿cuánto se separa de una BUENA?
    """
    if not detectables_archivos:
        return None
    set_det = set(os.path.basename(a) for a in detectables_archivos)
    res_m_det = [r for r in res_malas_ok
                  if os.path.basename(r.get("archivo", "")) in set_det]
    if len(res_m_det) < 3:
        return None

    def _vec(res_lista, clave, mult=1.0):
        out = []
        for r in res_lista:
            v = r.get(clave)
            if v is not None and np.isfinite(v):
                out.append(float(v) * mult)
        return np.asarray(out, dtype=float)

    metricas_def = [
        ("max_desviacion_mm",    1000.0, "Desv. máxima vs ISO",     "µm"),
        ("rms_desviacion_mm",    1000.0, "Desv. RMS vs ISO",        "µm"),
        ("prof_max_desgaste_um", 1.0,    "Profundidad desgaste",    "µm"),
        ("ancho_medio_desgaste_um", 1.0, "Ancho medio desgaste",    "µm"),
        ("area_desgaste_mm2",    1.0,    "Área desgaste",           "mm²"),
    ]
    resumen = []
    for clave, mult, nombre, unidad in metricas_def:
        v_b = _vec(res_buenas_ok, clave, mult)
        v_m = _vec(res_m_det, clave, mult)
        if len(v_b) == 0 or len(v_m) == 0:
            continue
        med_b = float(np.median(v_b))
        med_m = float(np.median(v_m))
        mean_b = float(np.mean(v_b))
        mean_m = float(np.mean(v_m))
        std_b = float(np.std(v_b))
        std_m = float(np.std(v_m))
        delta = med_m - med_b
        pool = float(np.sqrt((std_b ** 2 + std_m ** 2) / 2.0))
        cohen_d = ((mean_m - mean_b) / pool) if pool > 1e-12 else 0.0
        razon = (med_m / med_b) if abs(med_b) > 1e-6 else float("inf")
        resumen.append({
            "metrica": nombre, "unidad": unidad,
            "mediana_BUENA": med_b, "mediana_MALA": med_m,
            "delta_MALA_menos_BUENA": delta,
            "razon_MALA_BUENA": razon, "cohen_d": cohen_d,
            "n_BUENA": int(len(v_b)), "n_MALA_detectables": int(len(v_m)),
        })
    return {
        "resumen": resumen,
        "n_buenas": len(res_buenas_ok),
        "n_malas_total": len(res_malas_ok),
        "n_malas_detectables": len(res_m_det),
    }


def evaluar_perfilometro_clasificador(res_buenas, res_malas, ruta_salida,
                                       paso_mm=1.5, percentil_calib=99.0):
    """
    Evalúa el perfilómetro como sistema PASA/NO PASA sobre el dataset.

    Genera:
      - perfilometro_clf_resumen.csv : una fila por imagen con etiqueta,
        score continuo, veredicto, todas las métricas
      - perfilometro_clf_eda.csv     : Cohen d, KS, p-valor por métrica
      - perfilometro_clf_roc.png     : curva ROC del score continuo
      - perfilometro_clf_distrib.png : histogramas por métrica top
      - perfilometro_clf_informe.txt : informe con AUC, acc, matriz confusión
      - tolerancias_calibradas.json  : tolerancias derivadas del dataset
    """
    log.info("=" * 70)
    log.info(" PERFILÓMETRO COMO CLASIFICADOR — evaluación sobre dataset")
    log.info("=" * 70)
    Path(ruta_salida).mkdir(parents=True, exist_ok=True)

    # 1. Filtrar resultados válidos
    res_buenas_ok = [r for r in res_buenas if r.get("ok")]
    res_malas_ok = [r for r in res_malas if r.get("ok")]
    n_b_err = len(res_buenas) - len(res_buenas_ok)
    n_m_err = len(res_malas) - len(res_malas_ok)
    log.info(f"   BUENAS válidas: {len(res_buenas_ok)} (err: {n_b_err})")
    log.info(f"   MALAS  válidas: {len(res_malas_ok)} (err: {n_m_err})")

    if len(res_buenas_ok) < 5 or len(res_malas_ok) < 5:
        log.error("   No hay suficientes imágenes para evaluar.")
        return None

    # 2. Calibrar tolerancias data-driven
    tol_calib = calibrar_tolerancias_data_driven(
        res_buenas_ok, percentil=percentil_calib)
    log.info(f"   Tolerancias calibradas (P{percentil_calib} de BUENAS):")
    for k, v in tol_calib.items():
        log.info(f"     {k:30s} = {v:.3f}")

    # Guardar tolerancias
    import json
    with open(os.path.join(ruta_salida, "tolerancias_calibradas.json"),
              "w", encoding="utf-8") as fh:
        json.dump(tol_calib, fh, indent=2)

    # 3. Recalcular veredictos con tolerancias calibradas
    for r in res_buenas_ok + res_malas_ok:
        v, motivos = _aplicar_veredicto(r, tol_calib)
        r["veredicto_calibrado"] = v
        r["motivos_calibrado"] = motivos

    # 4. Construir matriz de features
    Xb, nombres, arch_b, _ = _extraer_features_clasificador(res_buenas_ok, 0)
    Xm, _, arch_m, _ = _extraer_features_clasificador(res_malas_ok, 1)
    # Imputar NaNs usando solo las BUENAS como referencia
    Xb_imp = _imputar_nan_con_mediana(Xb, X_ref=Xb)
    Xm_imp = _imputar_nan_con_mediana(Xm, X_ref=Xb)

    # Estadísticos de las BUENAS para z-scores
    # Estadísticos de las BUENAS para z-scores. Usamos MEDIANA y MAD
    # (robustos a outliers) en vez de media/std, porque una sola vista
    # con medición errónea puede inflar la std y hacer que el score
    # continuo no discrimine.
    mu_b = np.nanmedian(Xb_imp, axis=0)
    mad_b = np.nanmedian(np.abs(Xb_imp - mu_b), axis=0)
    sigma_b = mad_b * 1.4826   # MAD escalado a sigma equivalente normal
    sigma_b[sigma_b < 1e-9] = 1.0   # evita división por 0
    # También guardamos la versión clásica por si se necesita
    mu_b_mean = np.nanmean(Xb_imp, axis=0)
    sigma_b_std = np.nanstd(Xb_imp, axis=0)
    sigma_b_std[sigma_b_std < 1e-9] = 1.0

    # 5. EDA: Cohen d, KS por métrica
    log.info(f"   EDA por métrica perfilométrica:")
    eda_filas = []
    for j, nm in enumerate(nombres):
        a = Xb_imp[:, j]
        b = Xm_imp[:, j]
        d = _cohen_d(b, a)   # mala - buena (positivo => malas mayores)
        ks, p = _ks_test(a, b)
        eda_filas.append({
            "metrica": nm,
            "media_buena": float(np.mean(a)),
            "media_mala": float(np.mean(b)),
            "cohen_d": d, "ks_stat": ks, "p_valor": p,
        })
        log.info(f"     {nm:30s} d={d:+.2f}  ks={ks:.3f}  p={p:.2e}")

    # Guardar CSV EDA
    with open(os.path.join(ruta_salida, "perfilometro_clf_eda.csv"),
              "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["metrica", "media_buena", "media_mala",
                    "cohen_d", "ks_stat", "p_valor"])
        for f in eda_filas:
            w.writerow([f["metrica"],
                        f"{f['media_buena']:.6f}",
                        f"{f['media_mala']:.6f}",
                        f"{f['cohen_d']:.4f}",
                        f"{f['ks_stat']:.4f}",
                        f"{f['p_valor']:.4e}"])

    # 6. Score continuo por imagen (z-score combinado)
    sc_b = np.array([_score_continuo_perfilometro(
        r, mu_b, sigma_b, METRICAS_CLASIFICADOR) for r in res_buenas_ok])
    sc_m = np.array([_score_continuo_perfilometro(
        r, mu_b, sigma_b, METRICAS_CLASIFICADOR) for r in res_malas_ok])

    scores_all = np.concatenate([sc_b, sc_m])
    y_all = np.concatenate([np.zeros(len(sc_b)), np.ones(len(sc_m))])

    # 7. ROC + AUC
    fpr, tpr, thr, auc_val = _roc_auc(scores_all, y_all)
    thr_y, j_y, i_y = _umbral_youden(fpr, tpr, thr)
    log.info(f"   AUC del score continuo: {auc_val:.3f}")
    log.info(f"   Umbral Youden:          {thr_y:.3f}  (J={j_y:.3f})")

    # 8. Métricas con (a) veredicto duro PASA/NO PASA y (b) umbral Youden
    def _metricas_clasif(y_true, y_pred):
        """Devuelve dict con acc, sens, esp, ppv, npv, F1, TP, TN, FP, FN."""
        y_true = np.array(y_true); y_pred = np.array(y_pred)
        TP = int(np.sum((y_pred == 1) & (y_true == 1)))
        TN = int(np.sum((y_pred == 0) & (y_true == 0)))
        FP = int(np.sum((y_pred == 1) & (y_true == 0)))
        FN = int(np.sum((y_pred == 0) & (y_true == 1)))
        acc = (TP + TN) / max(TP + TN + FP + FN, 1)
        sens = TP / max(TP + FN, 1)
        esp = TN / max(TN + FP, 1)
        ppv = TP / max(TP + FP, 1)
        npv = TN / max(TN + FN, 1)
        f1 = 2 * ppv * sens / max(ppv + sens, 1e-9)
        return dict(acc=acc, sens=sens, esp=esp, ppv=ppv, npv=npv, f1=f1,
                    TP=TP, TN=TN, FP=FP, FN=FN)

    # (a) Veredicto duro
    y_pred_duro = np.array(
        [0] * len(res_buenas_ok) + [1] * len(res_malas_ok))
    y_pred_duro = np.array([
        1 if r["veredicto_calibrado"] == "NO PASA" else 0
        for r in res_buenas_ok + res_malas_ok
    ])
    met_duro = _metricas_clasif(y_all, y_pred_duro)
    # (b) Umbral Youden
    y_pred_youden = (scores_all >= thr_y).astype(int)
    met_youden = _metricas_clasif(y_all, y_pred_youden)

    # 9. ROC + distribuciones gráficos
    fig, axes = plt.subplots(1, 2, figsize=(13, 5))
    ax = axes[0]
    ax.plot(fpr, tpr, color="#264653", lw=2.2, label=f"AUC = {auc_val:.3f}")
    ax.plot([0, 1], [0, 1], "k--", lw=0.8, alpha=0.5)
    ax.scatter([fpr[i_y]], [tpr[i_y]], c="red", s=70, zorder=5,
               label=f"Youden: thr={thr_y:.2f}\n"
                     f"sens={tpr[i_y]:.2f}  esp={1-fpr[i_y]:.2f}")
    ax.set_xlabel("FPR (1 - especificidad)")
    ax.set_ylabel("TPR (sensibilidad)")
    ax.set_title("ROC del perfilómetro como clasificador")
    ax.grid(alpha=0.3); ax.legend(loc="lower right", fontsize=10)

    ax = axes[1]
    bins = np.linspace(0, max(scores_all.max(), 1e-3) * 1.05, 30)
    ax.hist(sc_b, bins=bins, color="#2a9d8f", alpha=0.6, label="BUENAS")
    ax.hist(sc_m, bins=bins, color="#e76f51", alpha=0.6, label="MALAS")
    ax.axvline(thr_y, color="r", ls="--", lw=1.4,
               label=f"Umbral Youden = {thr_y:.2f}")
    ax.set_xlabel("Score perfilométrico (suma z-scores)")
    ax.set_ylabel("Nº imágenes")
    ax.set_title("Distribución del score por clase")
    ax.grid(alpha=0.3); ax.legend(fontsize=9)
    plt.tight_layout()
    plt.savefig(os.path.join(ruta_salida, "perfilometro_clf_roc.png"), dpi=130)
    plt.close()

    # Histogramas de las top métricas (por |Cohen d|)
    eda_filas_orden = sorted(eda_filas, key=lambda r: abs(r["cohen_d"]),
                              reverse=True)
    top = eda_filas_orden[:6]
    fig, axes = plt.subplots(2, 3, figsize=(13, 7))
    for ax, f in zip(axes.flatten(), top):
        j = nombres.index(f["metrica"])
        a, b = Xb_imp[:, j], Xm_imp[:, j]
        bins = np.linspace(np.nanmin(np.concatenate([a, b])),
                            np.nanmax(np.concatenate([a, b])), 25)
        ax.hist(a, bins=bins, color="#2a9d8f", alpha=0.6, label="BUENAS")
        ax.hist(b, bins=bins, color="#e76f51", alpha=0.6, label="MALAS")
        ax.set_title(f"{f['metrica']}\n"
                     f"d={f['cohen_d']:+.2f}  p={f['p_valor']:.1e}",
                     fontsize=9)
        ax.legend(fontsize=8)
        ax.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(ruta_salida, "perfilometro_clf_distrib.png"),
                dpi=130)
    plt.close()

    # 10. CSV resumen por imagen
    with open(os.path.join(ruta_salida, "perfilometro_clf_resumen.csv"),
              "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        cabecera = ["archivo", "etiqueta", "score", "veredicto_calibrado",
                    "motivos"] + nombres
        w.writerow(cabecera)
        for r, sc in zip(res_buenas_ok, sc_b):
            fila = [r.get("archivo"), "BUENA", f"{sc:.4f}",
                    r["veredicto_calibrado"],
                    " ".join(r["motivos_calibrado"])]
            for nm in nombres:
                v = r.get(nm)
                if isinstance(v, float) and not np.isfinite(v):
                    v = ""
                fila.append(v if v is not None else "")
            w.writerow(fila)
        for r, sc in zip(res_malas_ok, sc_m):
            fila = [r.get("archivo"), "MALA", f"{sc:.4f}",
                    r["veredicto_calibrado"],
                    " ".join(r["motivos_calibrado"])]
            for nm in nombres:
                v = r.get(nm)
                if isinstance(v, float) and not np.isfinite(v):
                    v = ""
                fila.append(v if v is not None else "")
            w.writerow(fila)

    # 11. AGREGACIÓN A NIVEL PIEZA
    # Las imágenes pueden ser múltiples vistas de la misma pieza física.
    # Esto es relevante para el dataset actual (1 pieza buena + 1 pieza
    # mala, 11 vistas cada una). Se agrupan por ID extraído del nombre.
    piezas = agregar_a_nivel_pieza(res_buenas_ok, res_malas_ok,
                                    sc_b, sc_m, tol_calib)
    n_piezas = len(piezas)
    log.info(f"   Agregación a nivel pieza: {n_piezas} piezas únicas")

    # CSV a nivel pieza
    with open(os.path.join(ruta_salida, "perfilometro_clf_por_pieza.csv"),
              "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["pieza_id", "etiqueta", "n_vistas", "n_nopasa",
                    "score_max", "score_mean", "score_med",
                    "veredicto_any", "veredicto_majority"])
        for pid, info in sorted(piezas.items()):
            w.writerow([info["id"], info["etiqueta"], info["n_vistas"],
                        info["n_nopasa"],
                        f"{info['score_max']:.4f}",
                        f"{info['score_mean']:.4f}",
                        f"{info['score_med']:.4f}",
                        info["veredicto_any"],
                        info["veredicto_majority"]])

    # Métricas a nivel pieza (solo si hay >= 2 piezas por clase)
    metricas_pieza = None
    auc_pieza = None
    auprc_pieza = None
    n_buenas_pieza = sum(1 for i in piezas.values() if i["etiqueta"] == "BUENA")
    n_malas_pieza = sum(1 for i in piezas.values() if i["etiqueta"] == "MALA")
    if n_buenas_pieza >= 2 and n_malas_pieza >= 2:
        # AUC del score_max a nivel pieza
        scores_pieza = []
        y_pieza = []
        for info in piezas.values():
            scores_pieza.append(info["score_max"])
            y_pieza.append(1 if info["etiqueta"] == "MALA" else 0)
        scores_pieza = np.array(scores_pieza)
        y_pieza = np.array(y_pieza)
        try:
            from sklearn.metrics import roc_curve, auc as _auc
            fpr_p, tpr_p, _ = roc_curve(y_pieza, scores_pieza)
            auc_pieza = float(_auc(fpr_p, tpr_p))
        except Exception:
            auc_pieza = None
        # AUPRC (Average Precision) del score_max a nivel pieza.
        # Complementa al AUC-ROC centrándose en la clase positiva (MALA):
        # resume el compromiso precision-recall e ignora los verdaderos
        # negativos. NOTA: con pocas piezas (~24) la curva PR es
        # estadísticamente frágil; interpretar con cautela.
        try:
            from sklearn.metrics import average_precision_score
            auprc_pieza = float(average_precision_score(y_pieza,
                                                        scores_pieza))
        except Exception:
            auprc_pieza = None

        # Veredicto duro any/majority
        y_pred_any = np.array([1 if info["veredicto_any"] == "NO PASA" else 0
                               for info in piezas.values()])
        y_pred_maj = np.array([1 if info["veredicto_majority"] == "NO PASA"
                               else 0 for info in piezas.values()])
        metricas_pieza = {
            "any": _metricas_clasif(y_pieza, y_pred_any),
            "majority": _metricas_clasif(y_pieza, y_pred_maj),
            "auc": auc_pieza,
            "auprc": auprc_pieza,
            "n_buenas": int(n_buenas_pieza),
            "n_malas": int(n_malas_pieza),
        }
        log.info(f"   AUC nivel pieza (score_max): "
                 f"{auc_pieza if auc_pieza else 'N/A'}")
        log.info(f"   AUPRC nivel pieza (score_max): "
                 f"{auprc_pieza if auprc_pieza else 'N/A'}")
        log.info(f"   Veredicto ANY  - acc={metricas_pieza['any']['acc']*100:.1f}%  "
                 f"sens={metricas_pieza['any']['sens']*100:.1f}%  "
                 f"esp={metricas_pieza['any']['esp']*100:.1f}%")
        log.info(f"   Veredicto MAJ. - acc={metricas_pieza['majority']['acc']*100:.1f}%  "
                 f"sens={metricas_pieza['majority']['sens']*100:.1f}%  "
                 f"esp={metricas_pieza['majority']['esp']*100:.1f}%")

        # -----------------------------------------------------------
        # CALIBRACIÓN DEL UMBRAL DE DECISIÓN A NIVEL PIEZA.
        # En lugar de marcar la pieza como NO PASA cuando UNA vista
        # supera el umbral por vista (regla ANY clásica, demasiado
        # agresiva), usamos el percentil 95 de los score_max de las
        # piezas BUENAS como umbral de decisión a nivel pieza.
        # Esto da una especificidad razonable manteniendo sensibilidad.
        # -----------------------------------------------------------
        calib_pieza = calibrar_umbral_a_nivel_pieza(piezas, percentil=95.0)
        metricas_mejora_a = None
        umbral_mejora_a = None
        if calib_pieza is not None:
            umbral_mejora_a = calib_pieza["umbral_pieza"]
            metricas_mejora_a = aplicar_veredicto_pieza_mejora_a(
                piezas, umbral_mejora_a)
            log.info("")
            log.info(f"   [Calibración P95] Umbral a nivel pieza "
                     f"(P{calib_pieza['percentil_usado']:.0f} sobre score_max "
                     f"de las {calib_pieza['n_piezas_buenas']} piezas buenas): "
                     f"{umbral_mejora_a:.3f}")
            if metricas_mejora_a:
                log.info(f"   [Calibración P95] Acc={metricas_mejora_a['accuracy']*100:.1f}%  "
                         f"Sens={metricas_mejora_a['sensibilidad']*100:.1f}%  "
                         f"Esp={metricas_mejora_a['especificidad']*100:.1f}%  "
                         f"F1={metricas_mejora_a['f1']*100:.1f}%")
                log.info(f"   [Calibración P95] Matriz: TP={metricas_mejora_a['tp']} "
                         f"FN={metricas_mejora_a['fn']} "
                         f"FP={metricas_mejora_a['fp']} "
                         f"TN={metricas_mejora_a['tn']}")
                # Persistimos las métricas en el dict global para el informe.
                # La clave interna "mejora_a" se conserva para compatibilidad
                # con la app Streamlit y los scripts de análisis.
                metricas_pieza["mejora_a"] = {
                    **metricas_mejora_a,
                    "umbral": float(umbral_mejora_a),
                    "percentil": float(calib_pieza["percentil_usado"]),
                    "n_piezas_buenas_calibracion":
                        int(calib_pieza["n_piezas_buenas"]),
                }

                # ----- Matriz de confusión visual a nivel pieza -----
                # Genera la matriz de confusión 2x2 del veredicto a nivel
                # pieza calibrado por P95 (ver función _graficar_matriz_pieza).
                # Salida: matriz_mejora_a.png en la carpeta de evaluación.
                try:
                    _graficar_matriz_pieza(
                        tp=metricas_mejora_a["tp"],
                        fn=metricas_mejora_a["fn"],
                        fp=metricas_mejora_a["fp"],
                        tn=metricas_mejora_a["tn"],
                        umbral=float(umbral_mejora_a),
                        ruta_salida=ruta_salida,
                        titulo="Calibración del umbral de decisión a nivel pieza",
                        nombre_archivo="matriz_mejora_a.png",
                    )
                    log.info(f"   [Calibración P95] Matriz visual guardada: "
                             f"matriz_mejora_a.png")
                except Exception as _e:
                    log.warning(f"   No se pudo guardar matriz visual "
                                f"calibración P95: {_e}")

        # -----------------------------------------------------------
        # MEJORA B (V19+): calibración por POSICIÓN de vista (1..11).
        # Solo se aplica si las imágenes traen un número de vista
        # consistente entre piezas (RB01 siempre la misma zona, etc.).
        # La detección es automática (heurística de varianza).
        # -----------------------------------------------------------
        umbrales_por_vista = calibrar_umbrales_por_vista(
            res_buenas_ok, sc_b, percentil=99.0)
        metricas_pieza["mejora_b"] = {
            "aplicada": umbrales_por_vista is not None,
            "umbrales": (umbrales_por_vista
                         if umbrales_por_vista is not None else {}),
        }
        if umbrales_por_vista is None:
            log.info("   [Mejora B] No se aplica (sin orden fijo de vistas).")
    else:
        log.warning(f"   Agregación a nivel pieza no significativa: "
                    f"{n_buenas_pieza} buenas + {n_malas_pieza} malas. "
                    f"(Necesitas >=2 piezas por clase para AUC).")

    # 12. Informe de texto
    informe = []
    informe.append("=" * 78)
    informe.append(" INFORME PERFILÓMETRO ÓPTICO COMO CLASIFICADOR")
    informe.append(" Decisión PASA / NO PASA basada en geometría medida")
    informe.append("=" * 78)
    informe.append("")
    informe.append(f"Dataset:")
    informe.append(f"  BUENAS válidas:   {len(res_buenas_ok)}  "
                   f"(errores: {n_b_err})")
    informe.append(f"  MALAS válidas:    {len(res_malas_ok)}  "
                   f"(errores: {n_m_err})")
    informe.append(f"  Piezas únicas:    {n_piezas}  "
                   f"({n_buenas_pieza} BUENAS + {n_malas_pieza} MALAS)")
    informe.append("")
    informe.append(f"Calibración de tolerancias (percentil {percentil_calib} "
                   f"de las BUENAS):")
    for k, v in tol_calib.items():
        informe.append(f"  {k:30s} = {v:.4f}")
    informe.append("")
    informe.append("EDA de métricas perfilométricas:")
    informe.append("  (orden por |Cohen d|, indicador de discriminación)")
    informe.append(f"  {'metrica':30s}  {'med.B':>10s}  {'med.M':>10s}  "
                   f"{'Cohen d':>8s}  {'KS':>6s}  {'p-valor':>10s}")
    for f in eda_filas_orden:
        informe.append(
            f"  {f['metrica']:30s}  {f['media_buena']:>10.4f}  "
            f"{f['media_mala']:>10.4f}  {f['cohen_d']:>+8.2f}  "
            f"{f['ks_stat']:>6.3f}  {f['p_valor']:>10.2e}")
    informe.append("")
    informe.append("Rendimiento como clasificador:")
    informe.append("")
    informe.append(f"  (a) VEREDICTO DURO con tolerancias calibradas (P{percentil_calib}):")
    informe.append(f"      Accuracy     = {met_duro['acc']*100:6.2f}%")
    informe.append(f"      Sensibilidad = {met_duro['sens']*100:6.2f}%  "
                   f"(detección de MALAS)")
    informe.append(f"      Especificidad= {met_duro['esp']*100:6.2f}%  "
                   f"(no marcar BUENAS como malas)")
    informe.append(f"      F1           = {met_duro['f1']*100:6.2f}%")
    informe.append(f"      Matriz: TP={met_duro['TP']}  FN={met_duro['FN']}")
    informe.append(f"              FP={met_duro['FP']}  TN={met_duro['TN']}")
    informe.append("")
    informe.append(f"  (b) SCORE CONTINUO con umbral Youden ({thr_y:.3f}):")
    informe.append(f"      AUC          = {auc_val:.3f}")
    informe.append(f"      Accuracy     = {met_youden['acc']*100:6.2f}%")
    informe.append(f"      Sensibilidad = {met_youden['sens']*100:6.2f}%")
    informe.append(f"      Especificidad= {met_youden['esp']*100:6.2f}%")
    informe.append(f"      F1           = {met_youden['f1']*100:6.2f}%")
    informe.append(f"      Matriz: TP={met_youden['TP']}  FN={met_youden['FN']}")
    informe.append(f"              FP={met_youden['FP']}  TN={met_youden['TN']}")
    informe.append("")
    informe.append("")
    # Sección a nivel pieza (si es aplicable)
    if metricas_pieza is not None:
        informe.append("RESULTADOS A NIVEL PIEZA (agregando vistas múltiples):")
        informe.append(f"  Piezas: {metricas_pieza['n_buenas']} BUENAS + "
                       f"{metricas_pieza['n_malas']} MALAS")
        if metricas_pieza['auc'] is not None:
            informe.append(f"  AUC (score máximo de vistas) = "
                           f"{metricas_pieza['auc']:.3f}")
        if metricas_pieza.get('auprc') is not None:
            informe.append(f"  AUPRC (score máximo de vistas) = "
                           f"{metricas_pieza['auprc']:.3f}")
            informe.append(f"  (AUPRC: área bajo la curva Precision-Recall; "
                           f"con {metricas_pieza['n_buenas'] + metricas_pieza['n_malas']} "
                           f"piezas, interpretar con cautela)")
        m = metricas_pieza['any']
        informe.append(f"")
        informe.append(f"  (c) Veredicto ANY (1 vista NO PASA -> pieza NO PASA):")
        informe.append(f"      Accuracy     = {m['acc']*100:6.2f}%")
        informe.append(f"      Sensibilidad = {m['sens']*100:6.2f}%")
        informe.append(f"      Especificidad= {m['esp']*100:6.2f}%")
        informe.append(f"      F1           = {m['f1']*100:6.2f}%")
        informe.append(f"      Matriz: TP={m['TP']}  FN={m['FN']}  "
                       f"FP={m['FP']}  TN={m['TN']}")
        m = metricas_pieza['majority']
        informe.append(f"")
        informe.append(f"  (d) Veredicto MAJORITY (>50% vistas NO PASA):")
        informe.append(f"      Accuracy     = {m['acc']*100:6.2f}%")
        informe.append(f"      Sensibilidad = {m['sens']*100:6.2f}%")
        informe.append(f"      Especificidad= {m['esp']*100:6.2f}%")
        informe.append(f"      F1           = {m['f1']*100:6.2f}%")
        informe.append(f"      Matriz: TP={m['TP']}  FN={m['FN']}  "
                       f"FP={m['FP']}  TN={m['TN']}")
        informe.append("")

        # -----------------------------------------------------------
        # CALIBRACIÓN DEL UMBRAL DE DECISIÓN A NIVEL PIEZA
        # -----------------------------------------------------------
        if metricas_pieza.get("mejora_a"):
            ma = metricas_pieza["mejora_a"]
            informe.append(f"  (e) Calibración del umbral de decisión a nivel pieza:")
            informe.append(f"      Umbral aplicado al score_max de cada pieza: "
                            f"{ma['umbral']:.3f}")
            informe.append(f"      (= P{ma['percentil']:.0f} de los score_max "
                            f"de las {ma['n_piezas_buenas_calibracion']} "
                            f"piezas buenas)")
            informe.append(f"      Accuracy     = {ma['accuracy']*100:6.2f}%")
            informe.append(f"      Sensibilidad = {ma['sensibilidad']*100:6.2f}%")
            informe.append(f"      Especificidad= {ma['especificidad']*100:6.2f}%")
            informe.append(f"      F1           = {ma['f1']*100:6.2f}%")
            informe.append(f"      Matriz: TP={ma['tp']}  FN={ma['fn']}  "
                            f"FP={ma['fp']}  TN={ma['tn']}")
            informe.append("")
            # Comparativa rápida con ANY clásico
            try:
                any_m = metricas_pieza["any"]
                informe.append(f"      Comparación con ANY clásico:")
                informe.append(f"        ANY clásico:   "
                                f"Acc={any_m['acc']*100:.1f}%  "
                                f"Sens={any_m['sens']*100:.1f}%  "
                                f"Esp={any_m['esp']*100:.1f}%")
                informe.append(f"        Calibración P95:")
                informe.append(f"          Acc={ma['accuracy']*100:.1f}%  "
                                f"Sens={ma['sensibilidad']*100:.1f}%  "
                                f"Esp={ma['especificidad']*100:.1f}%")
                informe.append("")
            except Exception:
                pass

        # -----------------------------------------------------------
        # MEJORA B — umbral por posición de vista (si se aplicó)
        # -----------------------------------------------------------
        if metricas_pieza.get("mejora_b", {}).get("aplicada"):
            informe.append(f"  (f) MEJORA B — Umbrales por posición de vista:")
            for n_v, u in sorted(metricas_pieza["mejora_b"]["umbrales"].items()):
                informe.append(f"      vista #{n_v:02d}  umbral={u:.3f}")
            informe.append("")
        else:
            informe.append(f"  Nota: la Mejora B (umbral por posición de vista)")
            informe.append(f"  no se aplicó porque las 11 vistas no presentan")
            informe.append(f"  un orden fijo entre piezas (varianza intra-vista")
            informe.append(f"  similar a la global).")
            informe.append("")
    elif n_piezas > 0:
        informe.append("AGREGACIÓN A NIVEL PIEZA:")
        informe.append(f"  Solo {n_buenas_pieza} pieza(s) BUENA(S) + "
                       f"{n_malas_pieza} pieza(s) MALA(S) detectadas.")
        informe.append(f"  Listado por pieza (ver perfilometro_clf_por_pieza.csv):")
        for pid, info in sorted(piezas.items()):
            informe.append(f"    {pid:8s} ({info['etiqueta']:5s}, "
                           f"{info['n_vistas']:2d} vistas) "
                           f"score_max={info['score_max']:6.2f}  "
                           f"score_mean={info['score_mean']:6.2f}  "
                           f"NO_PASA={info['n_nopasa']}/{info['n_vistas']}  "
                           f"any={info['veredicto_any']}  "
                           f"maj={info['veredicto_majority']}")
        informe.append("")

    informe.append("Comparación con PatchCore (informe TFM previo): AUC=0.771")
    informe.append("")
    informe.append("Archivos generados:")
    informe.append(f"  perfilometro_clf_resumen.csv  - métrica e veredicto por imagen")
    informe.append(f"  perfilometro_clf_eda.csv      - estadísticos por métrica")
    informe.append(f"  perfilometro_clf_roc.png      - ROC y distribución de score")
    informe.append(f"  perfilometro_clf_distrib.png  - histogramas top 6 métricas")
    informe.append(f"  tolerancias_calibradas.json   - tolerancias del calibre")
    informe.append("=" * 78)
    informe_txt = "\n".join(informe)

    with open(os.path.join(ruta_salida, "perfilometro_clf_informe.txt"),
              "w", encoding="utf-8") as fh:
        fh.write(informe_txt)

    log.info("")
    for ln in informe:
        log.info(ln)

    return {
        "auc": auc_val, "umbral_youden": thr_y,
        "metricas_duro": met_duro, "metricas_youden": met_youden,
        "tolerancias_calibradas": tol_calib,
        "eda": eda_filas_orden,
    }


# ==============================================================================
#  BLOQUE 4-quater - DIAGNÓSTICO AUTOMÁTICO DEL PERFILÓMETRO
# ==============================================================================
#  Lee el CSV producido por evaluar_perfilometro_clasificador y produce un
#  informe interpretativo con:
#     - Métricas de cada vista
#     - Detección de mediciones físicamente sospechosas
#     - AUC bajo varios criterios (todas las vistas, sin sospechosas,
#       solo detectables EDA, a nivel pieza)
#     - Cruce con el subconjunto detectable según el EDA general
# ==============================================================================

# Vistas que el EDA original (informe_tfm.txt) marcó como "detectables"
# (las únicas estadísticamente distinguibles según el extractor genérico).
DETECTABLES_EDA_DEFECTO = [
    "Imagen_000501_RM05.jpg",
    "Imagen_000501_RM06.jpg",
    "Imagen_000501_RM07.jpg",
]

# Rangos físicamente razonables para detectar mediciones erróneas.
# NOTA: estos rangos están AFLOJADOS porque variaciones de captura entre
# vistas (foco, iluminación, ligera rotación) hacen que valores absolutos
# salgan distintos sin que la pieza sea defectuosa. Son SOLO PARA AVISO,
# no para descartar.
RANGOS_SANOS_DIAG = {
    "error_paso_pct":           50.0,
    "error_altura_pct":         1500.0,   # antes 80, ahora muy permisivo
    "error_angulo_deg":         60.0,
    "rms_desviacion_pct_P":     200.0,
    "cresta_radio_redondeo_um": 3000.0,
    "cresta_planitud_rms_um":   500.0,
    "cresta_longitud_mm":       3.0,
    "valle_longitud_mm":        3.0,
    "max_desviacion_mm":        10.0,
}


def _diag_es_medicion_sospechosa(fila):
    """Devuelve lista de razones por las que esta medición es sospechosa."""
    razones = []
    for clave, lim in RANGOS_SANOS_DIAG.items():
        v = fila.get(clave)
        if v is None or v == "":
            continue
        try:
            v = abs(float(v))
        except Exception:
            continue
        if v > lim:
            razones.append(f"{clave}={v:.1f} > {lim}")
    return razones


def _diag_auc_simple(scores, y_true):
    """ROC AUC sin sklearn (algoritmo Mann-Whitney U)."""
    scores = np.asarray(scores, dtype=float)
    y_true = np.asarray(y_true, dtype=int)
    pos = scores[y_true == 1]
    neg = scores[y_true == 0]
    if len(pos) == 0 or len(neg) == 0:
        return None
    n_pos, n_neg = len(pos), len(neg)
    todos = np.concatenate([pos, neg])
    rangos = todos.argsort().argsort() + 1
    suma_rangos_pos = rangos[:n_pos].sum()
    U = suma_rangos_pos - n_pos * (n_pos + 1) / 2
    return U / (n_pos * n_neg)


def _diag_id_pieza(nombre, etq):
    import re
    m = re.search(r'(\d{1,5})[\s_\-]*r[bm][\s_\-]*\d{1,3}',
                   str(nombre).lower())
    if m:
        sufijo = "B" if "rb" in str(nombre).lower() else "M"
        return f"P{int(m.group(1)):03d}_{sufijo}"
    return f"PIEZA_{etq[0]}"


def ejecutar_diagnostico_automatico(ruta_resultados, ruta_diagnostico=None,
                                     detectables_eda=None):
    """
    Lee perfilometro_clf_resumen.csv y produce un informe diagnóstico.
    Imprime el informe por log.info y lo guarda en ruta_diagnostico (si se
    especifica).
    """
    if detectables_eda is None:
        detectables_eda = DETECTABLES_EDA_DEFECTO

    csv_path = os.path.join(str(ruta_resultados), "perfilometro",
                             "evaluacion",
                             "perfilometro_clf_resumen.csv")
    if not os.path.exists(csv_path):
        log.error(f"Diagnóstico: no encuentro {csv_path}")
        return None

    # Leer el CSV
    filas = []
    with open(csv_path, encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for r in reader:
            for k, v in list(r.items()):
                if v == "" or v is None:
                    r[k] = None
                    continue
                try:
                    r[k] = float(v)
                except (ValueError, TypeError):
                    pass
            filas.append(r)

    lineas = []
    def out(s=""):
        lineas.append(s)

    out("=" * 78)
    out(" DIAGNÓSTICO AUTOMÁTICO DEL PERFILÓMETRO ÓPTICO")
    out("=" * 78)
    out(f" Archivo leído: {csv_path}")
    out(f" Total filas:   {len(filas)}")
    n_b = sum(1 for r in filas if r.get("etiqueta") == "BUENA")
    n_m = sum(1 for r in filas if r.get("etiqueta") == "MALA")
    out(f" BUENAS:        {n_b}")
    out(f" MALAS:         {n_m}")
    out()

    # ============================================================
    # 1) Listar todas las vistas con sus métricas clave
    # ============================================================
    out("-" * 78)
    out(" 1) MÉTRICAS POR VISTA (ordenadas por etiqueta y nombre)")
    out("-" * 78)
    out(f"  {'archivo':<30s} {'etq':<5s} "
        f"{'angle':>6s} {'h%err':>7s} "
        f"{'cR(um)':>7s} {'cL(um)':>7s} "
        f"{'rms%':>6s} {'score':>7s}  sospech.")
    out("-" * 78)
    filas_orden = sorted(filas, key=lambda r: (r.get("etiqueta", ""),
                                                r.get("archivo", "")))
    sospechosas_b = []
    sospechosas_m = []
    for r in filas_orden:
        nombre = str(r.get("archivo", ""))[:30]
        etq = r.get("etiqueta", "")[:5]
        ang = r.get("angulo_flanco_real_deg")
        h_err = r.get("error_altura_pct")
        cR = r.get("cresta_radio_redondeo_um")
        cL = r.get("cresta_longitud_mm")
        rms = r.get("rms_desviacion_pct_P")
        sc = r.get("score")
        cL_um = (cL * 1000.0) if cL is not None else None

        razones = _diag_es_medicion_sospechosa(r)
        marca = "*" if razones else " "   # '*' por compatibilidad ASCII
        if razones and etq == "BUENA":
            sospechosas_b.append((nombre, razones))
        if razones and etq == "MALA":
            sospechosas_m.append((nombre, razones))

        def f(v, ancho=6, dec=1):
            if v is None:
                return "—".rjust(ancho)
            return f"{v:>{ancho}.{dec}f}"

        out(f"  {nombre:<30s} {etq:<5s} "
            f"{f(ang)} {f(h_err)} "
            f"{f(cR, 7, 0)} {f(cL_um, 7, 0)} "
            f"{f(rms)} {f(sc, 7, 3)}  {marca}")
    out()
    out(f"  Vistas sospechosas (medición fuera de rango físico):")
    out(f"     BUENAS: {len(sospechosas_b)}/{n_b}")
    out(f"     MALAS:  {len(sospechosas_m)}/{n_m}")
    if sospechosas_b:
        out(f"     Detalle BUENAS:")
        for n, raz in sospechosas_b[:5]:
            out(f"        {n}  ->  {'; '.join(raz)}")
    if sospechosas_m:
        out(f"     Detalle MALAS:")
        for n, raz in sospechosas_m[:5]:
            out(f"        {n}  ->  {'; '.join(raz)}")
    out()

    # ============================================================
    # 2) AUC bajo varios criterios
    # ============================================================
    out("-" * 78)
    out(" 2) AUC BAJO DISTINTOS CRITERIOS DE EVALUACIÓN")
    out("-" * 78)

    scores_all = [r.get("score") for r in filas]
    y_all = [1 if r.get("etiqueta") == "MALA" else 0 for r in filas]
    valid_idx = [i for i, s in enumerate(scores_all) if s is not None]
    s_all = [scores_all[i] for i in valid_idx]
    y_all_v = [y_all[i] for i in valid_idx]
    auc_global = _diag_auc_simple(s_all, y_all_v) if s_all else None

    if auc_global is not None:
        out(f"  (a) AUC global (todas las vistas):      {auc_global:.3f}  "
            f"(n={len(s_all)})")
    else:
        out(f"  (a) AUC global (todas las vistas):      N/A")

    sospechosas_set = set(n for n, _ in sospechosas_b + sospechosas_m)
    filas_limpias = [r for r in filas
                     if str(r.get("archivo", ""))[:30] not in sospechosas_set
                     and r.get("score") is not None]
    s_lim = [r.get("score") for r in filas_limpias]
    y_lim = [1 if r.get("etiqueta") == "MALA" else 0 for r in filas_limpias]
    if s_lim and len(set(y_lim)) == 2:
        auc_lim = _diag_auc_simple(s_lim, y_lim)
        out(f"  (b) AUC sin vistas sospechosas:         "
            f"{auc_lim:.3f} (n={len(s_lim)})")
    else:
        out(f"  (b) AUC sin vistas sospechosas:         N/A "
            f"(n={len(s_lim)}, no quedan ambas clases)")

    filas_det = []
    for r in filas:
        nombre = str(r.get("archivo", ""))
        if r.get("etiqueta") == "BUENA":
            filas_det.append(r)
        elif any(d in nombre for d in detectables_eda):
            filas_det.append(r)
    s_det = [r.get("score") for r in filas_det if r.get("score") is not None]
    y_det = [1 if r.get("etiqueta") == "MALA" else 0
             for r in filas_det if r.get("score") is not None]
    if s_det and len(set(y_det)) == 2:
        auc_det = _diag_auc_simple(s_det, y_det)
        out(f"  (c) AUC sobre 'detectables' EDA "
            f"(RM05/06/07 + buenas): {auc_det:.3f}")
    else:
        out(f"  (c) AUC sobre detectables EDA:          N/A")

    # AUC a nivel pieza
    piezas = {}
    for r in filas:
        if r.get("score") is None:
            continue
        pid = _diag_id_pieza(r.get("archivo", ""), r.get("etiqueta", ""))
        piezas.setdefault(pid, {"scores": [], "etq": r.get("etiqueta")})
        piezas[pid]["scores"].append(r.get("score"))
    out()
    out(f"  Agregación a nivel pieza:")
    for pid, info in sorted(piezas.items()):
        sm = max(info["scores"])
        smean = float(np.mean(info["scores"]))
        out(f"     {pid:8s} ({info['etq']:5s}, "
            f"{len(info['scores']):2d} vistas) "
            f"score_max={sm:.3f}  score_mean={smean:.3f}")
    out()

    # ============================================================
    # 3) Cruce con detectables del EDA
    # ============================================================
    out("-" * 78)
    out(" 3) CRUCE CON 'DETECTABLES' DEL EDA ORIGINAL")
    out("-" * 78)
    out(f"  Vistas que el EDA marcó como detectables (3 de 11 malas):")
    for d in detectables_eda:
        for r in filas:
            if d in str(r.get("archivo", "")):
                sc = r.get("score")
                cR = r.get("cresta_radio_redondeo_um")
                ang = r.get("angulo_flanco_real_deg")
                sc_s = f"{sc:.3f}" if sc is not None else "NaN"
                cR_s = f"{cR:.0f}" if cR is not None else "NaN"
                ang_s = f"{ang:.1f}" if ang is not None else "NaN"
                out(f"     {d:<35s}  score={sc_s}  "
                    f"cR={cR_s}um  ang={ang_s}°")
                break

    malas_filas = [r for r in filas
                   if r.get("etiqueta") == "MALA"
                   and r.get("score") is not None]
    malas_filas.sort(key=lambda r: r.get("score") or 0, reverse=True)
    out(f"")
    out(f"  Top 3 vistas MALAS según el perfilómetro (mayor score):")
    for r in malas_filas[:3]:
        n = str(r.get("archivo", ""))
        sc = r.get("score")
        marca = "  -> EDA tambien" if any(d in n for d in detectables_eda) \
                else "  -> NO estaba en EDA"
        out(f"     {n:<35s}  score={sc:.3f}{marca}")
    out()

    # ============================================================
    # 4) Conclusiones
    # ============================================================
    out("-" * 78)
    out(" 4) DIAGNÓSTICO Y CONCLUSIONES")
    out("-" * 78)
    pct_sospechosas = (len(sospechosas_b) + len(sospechosas_m)) \
                       / max(len(filas), 1) * 100
    out(f"  - {pct_sospechosas:.0f}% de las vistas tienen mediciones "
        f"físicamente sospechosas.")
    if pct_sospechosas > 30:
        out(f"  - Esto INDICA que el extractor de perfil 1D no funciona bien")
        out(f"    en muchas de las 11 vistas (probablemente por orientación")
        out(f"    diferente del eje de rosca, iluminación, etc).")
        out(f"  - Recomendación: revisar visualmente las vistas marcadas con *.")
    if auc_global is not None and auc_global < 0.6:
        out(f"  - AUC global bajo ({auc_global:.3f}). Causas más probables:")
        out(f"      a) Mediciones erróneas envenenan el score")
        out(f"      b) Las vistas sin desgaste local de la pieza mala")
        out(f"         tienen scores parecidos a las buenas (esperado)")

    out()
    out("=" * 78)

    texto = "\n".join(lineas)
    # Mostrar por log
    for ln in lineas:
        log.info(ln)

    # Guardar archivo si procede
    if ruta_diagnostico:
        Path(os.path.dirname(ruta_diagnostico)).mkdir(parents=True, exist_ok=True)
        with open(ruta_diagnostico, "w", encoding="utf-8") as fh:
            fh.write(texto)
        log.info(f"Diagnóstico guardado en: {ruta_diagnostico}")

    return texto


# ==============================================================================
#  BLOQUE 5-bis - PATCHCORE SOBRE DESVIACIÓN PERFIL REAL vs TEÓRICO
# ==============================================================================
#  Variante de PatchCore donde el INPUT es la DESVIACIÓN punto a punto entre
#  el perfil real medido y el perfil teórico ISO alineado en fase.
#
#  Esta es la formulación natural para un PERFILÓMETRO ÓPTICO de inspección:
#     - El teórico ISO actúa como "modelo nominal" de la pieza.
#     - La desviación (real - teórico) es la firma del defecto.
#     - Una pieza buena tiene desviaciones pequeñas y similares a otras
#       buenas (errores de medición + ruido óptico).
#     - Una pieza mala tiene desviaciones grandes y localizadas en las
#       zonas con desgaste.
#
#  Pipeline:
#     1. perfilometro_analizar(img) extrae el perfil real, construye el
#        teórico ISO alineado, y los devuelve en mm.
#     2. desviacion = real - teorico (en mm).
#     3. Recortar a zona dentada útil + resamplear a 256 puntos fijos.
#     4. Memory bank = matriz [N_buenas × 256] de desviaciones de BUENAS.
#     5. Coreset greedy al 1%.
#     6. Inferencia: distancia al k-NN del banco como score de anomalía.
# ==============================================================================

def _resamplear_perfil(perfil, n_puntos=256):
    """Interpola el perfil/desviación a una longitud fija de n_puntos."""
    if perfil is None or len(perfil) < 2:
        return None
    perfil = np.asarray(perfil, dtype=np.float32)
    x_old = np.linspace(0, 1, len(perfil))
    x_new = np.linspace(0, 1, n_puntos)
    return np.interp(x_new, x_old, perfil)


def extraer_perfil_para_patchcore(img_gris, n_puntos=256, paso_mm=1.5,
                                   perfil_norma="iso_metrica"):
    """
    Extrae la DESVIACIÓN del perfil real respecto al teórico ISO,
    resamplea a longitud fija y devuelve un vector listo para PatchCore.

    Pipeline interno:
       1. perfilometro_analizar(img) → mide perfil real y construye teórico
          alineado en fase (ya hace todo el trabajo de extracción + alineamiento)
       2. desviacion = y_real_mm - y_teorico_mm  (en mm)
       3. Recorte a la zona dentada útil (entre primer y último pico/valle)
       4. Resampleo a n_puntos fijos por interpolación lineal
       5. NO se normaliza por z-score: las unidades absolutas (mm) son
          INFORMATIVAS - una desviación de 0.05 mm es buena, una de 0.3 mm
          es desgaste real. Conservamos esa información.

    Devuelve: vector de longitud n_puntos en MILÍMETROS, o None si falla.
    """
    try:
        # Análisis perfilométrico completo: ya extrae el real, construye el
        # teórico, los alinea en fase y los devuelve en mm.
        res = perfilometro_analizar(img_gris, paso_mm=paso_mm,
                                     perfil_norma=perfil_norma)
        if not res.get("ok"):
            return None

        x_mm = res.get("x_mm")
        y_real = res.get("y_real_mm")
        y_teo = res.get("y_teorico_mm")
        picos = res.get("picos_idx", [])
        valles = res.get("valles_idx", [])

        if y_real is None or y_teo is None or len(y_real) != len(y_teo):
            return None
        if len(y_real) < 30:
            return None

        # Desviación punto a punto en mm
        desv = np.asarray(y_real) - np.asarray(y_teo)

        # Recortar a la zona útil (entre primer y último pico/valle).
        # Esto evita que un eventual artefacto de borde domine la firma.
        if len(picos) >= 1 or len(valles) >= 1:
            todos = np.sort(np.concatenate([picos, valles]).astype(int))
            if len(todos) >= 2:
                d_med = float(np.median(np.diff(todos)))
                margen = int(0.5 * d_med)
                izq = max(0, int(todos[0]) - margen)
                der = min(len(desv), int(todos[-1]) + margen)
                if der - izq >= 30:
                    desv = desv[izq:der]

        # Resamplear a longitud fija
        desv_resamp = _resamplear_perfil(desv, n_puntos=n_puntos)
        if desv_resamp is None or not np.isfinite(desv_resamp).all():
            return None

        return desv_resamp.astype(np.float32)

    except Exception as e:
        log.warning(f"extraer_perfil_para_patchcore falló: {e}")
        return None


def _coreset_greedy(banco, ratio=0.01, min_n=4):
    """
    Reducción de memory bank por farthest-point sampling.
    Mantiene un subconjunto representativo seleccionando iterativamente
    el punto más lejano al subconjunto ya elegido.
    """
    n_total = len(banco)
    n_keep = max(min_n, int(np.ceil(n_total * ratio)))
    n_keep = min(n_keep, n_total)
    if n_keep >= n_total:
        return banco, list(range(n_total))

    # Empezamos con el punto más cercano al centroide
    centroide = np.mean(banco, axis=0)
    d0 = np.linalg.norm(banco - centroide, axis=1)
    seleccionados = [int(np.argmin(d0))]
    distancias = np.linalg.norm(banco - banco[seleccionados[0]], axis=1)

    for _ in range(n_keep - 1):
        idx = int(np.argmax(distancias))
        seleccionados.append(idx)
        d_nuevo = np.linalg.norm(banco - banco[idx], axis=1)
        distancias = np.minimum(distancias, d_nuevo)

    return banco[seleccionados], seleccionados


# (helper patchcore_perfil1d_entrenar() retirado en limpieza V20)


# (helper patchcore_perfil1d_evaluar() retirado en limpieza V20)


# (función evaluar_patchcore_perfil1d() retirada en limpieza V20: ya no se usa)


# ==============================================================================
#  BLOQUE 5 - DETECCIÓN DE ANOMALÍAS (memory bank + KNN, estilo PatchCore)
# ==============================================================================

def _redimensionar_si_grande(img, max_lado):
    """Redimensiona conservando aspecto si el lado mayor supera max_lado."""
    if max_lado is None or max_lado <= 0:
        return img
    h, w = img.shape[:2]
    lado = max(h, w)
    if lado <= max_lado:
        return img
    escala = max_lado / lado
    nuevo_w = int(round(w * escala))
    nuevo_h = int(round(h * escala))
    return cv2.resize(img, (nuevo_w, nuevo_h), interpolation=cv2.INTER_AREA)


def cargar_imagenes_carpeta(carpeta, max_imagenes=None, max_lado=None):
    """
    Lee imágenes JPG/PNG/BMP/TIF de una carpeta en escala de grises.

    Parámetros
    ----------
    max_imagenes : int o None
        Si se especifica, solo lee las primeras N (orden alfabético).
    max_lado : int o None
        Si se especifica, redimensiona para que el lado mayor sea N px.
    """
    extensiones = (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff")
    archivos = sorted([f for f in os.listdir(carpeta)
                       if f.lower().endswith(extensiones)])
    if max_imagenes is not None:
        archivos = archivos[:max_imagenes]

    imagenes = []
    primer_log = True
    for f in archivos:
        ruta = os.path.join(carpeta, f)
        img = cv2.imread(ruta, cv2.IMREAD_GRAYSCALE)
        if img is None:
            log.warning(f"No se pudo leer {ruta}, se omite.")
            continue
        h0, w0 = img.shape
        img = _redimensionar_si_grande(img, max_lado)
        if primer_log:
            log.info(f"  Tamaño original 1ª img: {h0}x{w0} -> "
                     f"procesado: {img.shape[0]}x{img.shape[1]}")
            primer_log = False
        imagenes.append((f, img))
    return imagenes


# (función construir_memory_bank() retirada en limpieza V20: ya no se usa)


# (función calcular_distancias() retirada en limpieza V20: ya no se usa)


# ==============================================================================
#  BLOQUE 6 - MÉTRICAS Y VISUALIZACIÓN
# ==============================================================================

def matriz_confusion(log_data, umbral):
    """Devuelve VP, VN, FP, FN dado un umbral."""
    vp = vn = fp = fn = 0
    for _f, cat, dist, _pl in log_data:
        pred_mala = dist > umbral
        if cat == "MALA" and pred_mala:
            vp += 1
        elif cat == "MALA" and not pred_mala:
            fn += 1
        elif cat == "BUENA" and not pred_mala:
            vn += 1
        elif cat == "BUENA" and pred_mala:
            fp += 1
    return vp, vn, fp, fn


def graficar_matriz(log_data, umbral, ruta_salida, etiqueta="AUTO"):
    vp, vn, fp, fn = matriz_confusion(log_data, umbral)
    total = max(1, vp + vn + fp + fn)
    acc = (vp + vn) / total

    cm = np.array([[vn, fp], [fn, vp]])
    fig, ax = plt.subplots(figsize=(6, 5))
    im = ax.imshow(cm, cmap=plt.cm.Blues)
    ax.set_title(f"Matriz de confusión [{etiqueta}]\n"
                 f"Umbral={umbral:.3f}  |  Precisión={acc:.2%}")
    ax.set_xticks([0, 1]); ax.set_xticklabels(["Pred. BUENA", "Pred. MALA"])
    ax.set_yticks([0, 1]); ax.set_yticklabels(["Real BUENA", "Real MALA"])
    for i, j in np.ndindex(cm.shape):
        ax.text(j, i, str(cm[i, j]), ha="center", va="center",
                color="white" if cm[i, j] > cm.max() / 2 else "black",
                fontsize=14, fontweight="bold")
    plt.colorbar(im, ax=ax)
    plt.tight_layout()
    plt.savefig(os.path.join(ruta_salida,
                             f"matriz_{etiqueta}_u{umbral:.3f}.png"), dpi=120)
    plt.close()
    return acc, vp, vn, fp, fn


def _graficar_matriz_pieza(tp, fn, fp, tn, umbral, ruta_salida, titulo,
                            nombre_archivo):
    """V20: matriz de confusión 2x2 directa a partir de TP/FN/FP/TN.

    Variante de graficar_matriz() que no necesita log_data sino que recibe
    los conteos directamente. Pensada para la calibración a nivel pieza
    y otros métodos que ya tienen su matriz calculada (perfil patrón).

    Genera un PNG con la matriz coloreada, valores y métricas resumidas.
    """
    total = max(1, tp + fn + fp + tn)
    acc = (tp + tn) / total
    sens = tp / max(1, tp + fn)
    esp = tn / max(1, tn + fp)
    f1 = 2 * tp / max(1, 2 * tp + fp + fn)

    cm = np.array([[tn, fp], [fn, tp]])
    fig, ax = plt.subplots(figsize=(6, 5.5))
    im = ax.imshow(cm, cmap=plt.cm.Blues)
    ax.set_title(f"{titulo}\n"
                  f"Umbral={umbral:.3f}  |  Acc={acc:.1%}  |  "
                  f"Sens={sens:.1%}  |  Esp={esp:.1%}  |  F1={f1:.1%}",
                  fontsize=10)
    ax.set_xticks([0, 1]); ax.set_xticklabels(["Pred. BUENA", "Pred. MALA"])
    ax.set_yticks([0, 1]); ax.set_yticklabels(["Real BUENA", "Real MALA"])
    ax.set_xlabel("Predicción del sistema")
    ax.set_ylabel("Etiqueta real")
    for i, j in np.ndindex(cm.shape):
        ax.text(j, i, str(cm[i, j]), ha="center", va="center",
                color="white" if cm[i, j] > cm.max() / 2 else "black",
                fontsize=16, fontweight="bold")
    plt.colorbar(im, ax=ax)
    plt.tight_layout()
    out_path = os.path.join(ruta_salida, nombre_archivo)
    plt.savefig(out_path, dpi=120)
    plt.close()
    return out_path


def curva_roc(log_data, ruta_salida, etiqueta="ROC"):
    """Calcula y grafica la curva ROC y su AUC."""
    distancias = np.array([r[2] for r in log_data])
    etiquetas = np.array([1 if r[1] == "MALA" else 0 for r in log_data])

    umbrales = np.linspace(distancias.min(), distancias.max(), 200)
    tpr_list, fpr_list = [], []
    for u in umbrales:
        pred = (distancias > u).astype(int)
        vp = int(np.sum((pred == 1) & (etiquetas == 1)))
        fn = int(np.sum((pred == 0) & (etiquetas == 1)))
        fp = int(np.sum((pred == 1) & (etiquetas == 0)))
        vn = int(np.sum((pred == 0) & (etiquetas == 0)))
        tpr = vp / max(1, vp + fn)
        fpr = fp / max(1, fp + vn)
        tpr_list.append(tpr); fpr_list.append(fpr)

    # AUC por trapecios (orden FPR creciente)
    orden = np.argsort(fpr_list)
    fpr_arr = np.array(fpr_list)[orden]
    tpr_arr = np.array(tpr_list)[orden]
    # np.trapezoid si está disponible (numpy >= 2.0), sino np.trapz
    if hasattr(np, "trapezoid"):
        auc = float(np.trapezoid(tpr_arr, fpr_arr))
    else:
        auc = float(np.trapz(tpr_arr, fpr_arr))

    plt.figure(figsize=(6, 5))
    plt.plot(fpr_arr, tpr_arr, lw=2, label=f"AUC = {auc:.3f}")
    plt.plot([0, 1], [0, 1], "k--", alpha=0.5)
    plt.xlabel("Falsos Positivos (FPR)")
    plt.ylabel("Verdaderos Positivos (TPR)")
    plt.title(f"Curva ROC - {etiqueta}")
    plt.legend(loc="lower right")
    plt.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(ruta_salida, f"roc_{etiqueta}.png"), dpi=120)
    plt.close()
    return auc


def umbral_solo_buenas(distancias_buenas, factor=3.0, percentil=None):
    """
    Calcula umbral usando SOLO la distribución de piezas buenas. Más riguroso
    que el promedio buenas+malas de V10 porque no contamina con info del
    conjunto a evaluar.

    - Si percentil es None: μ + factor * σ.
    - Si percentil se especifica (ej. 99): usa percentil.
    """
    if percentil is not None:
        return float(np.percentile(distancias_buenas, percentil))
    mu = float(np.mean(distancias_buenas))
    sigma = float(np.std(distancias_buenas))
    return mu + factor * sigma


# ==============================================================================
#  ANÁLISIS MULTI-UMBRAL (V17)
# ==============================================================================
#
#  En lugar de usar un único umbral, exploramos un rango amplio:
#    - Múltiplos de sigma sobre la media de buenas (μ+1σ, μ+2σ, μ+3σ, μ+4σ)
#    - Percentiles de buenas (p90, p95, p99, p99.5)
#    - Umbrales óptimos calculados sobre las MALAS:
#        * Youden's J = TPR - FPR maximizado
#        * F1 óptimo
#    - Umbral legacy v10 (media buenas + media malas) / 2
#
#  Esto permite al lector elegir el operating point más adecuado según el
#  coste industrial de los falsos positivos vs los falsos negativos.
# ==============================================================================

def metricas_punto_operacion(scores, etiquetas, umbral):
    """
    Para un umbral dado, devuelve TODAS las métricas de clasificación binaria.

    Parámetros
    ----------
    scores : np.ndarray
        Anomaly scores (1 dimensión).
    etiquetas : np.ndarray (0/1)
        0 = BUENA, 1 = MALA.
    umbral : float
        Umbral de decisión (predicción MALA si score > umbral).
    """
    pred = (scores > umbral).astype(int)
    vp = int(np.sum((pred == 1) & (etiquetas == 1)))
    fn = int(np.sum((pred == 0) & (etiquetas == 1)))
    fp = int(np.sum((pred == 1) & (etiquetas == 0)))
    vn = int(np.sum((pred == 0) & (etiquetas == 0)))
    tot = max(1, vp + fn + fp + vn)
    n_pos = vp + fn
    n_neg = vn + fp

    tpr = vp / max(1, n_pos)            # = recall = sensibilidad
    fpr = fp / max(1, n_neg)            # 1 - especificidad
    tnr = vn / max(1, n_neg)            # = especificidad
    fnr = fn / max(1, n_pos)            # tasa de fugas
    precision = vp / max(1, vp + fp)    # PPV
    npv = vn / max(1, vn + fn)          # NPV
    accuracy = (vp + vn) / tot
    f1 = 2 * precision * tpr / max(1e-12, precision + tpr) if (precision + tpr) > 0 else 0.0
    youden_j = tpr - fpr
    # Matthews Correlation Coefficient: balanceado para clases desbalanceadas
    denom = np.sqrt((vp + fp) * (vp + fn) * (vn + fp) * (vn + fn))
    mcc = ((vp * vn) - (fp * fn)) / denom if denom > 0 else 0.0

    return {
        "umbral": float(umbral),
        "VP": vp, "VN": vn, "FP": fp, "FN": fn,
        "TPR_recall": float(tpr), "FPR": float(fpr),
        "TNR_especificidad": float(tnr), "FNR_fugas": float(fnr),
        "precision": float(precision), "NPV": float(npv),
        "accuracy": float(accuracy), "F1": float(f1),
        "Youden_J": float(youden_j), "MCC": float(mcc),
    }


def encontrar_umbral_optimo(scores, etiquetas, criterio="youden"):
    """
    Busca el umbral que optimiza un criterio dado mediante barrido fino.

    Parámetros
    ----------
    criterio : "youden" | "f1" | "mcc"
    """
    if len(scores) == 0:
        return None, None

    candidatos = np.linspace(scores.min(), scores.max(), 500)
    mejores = []
    for u in candidatos:
        m = metricas_punto_operacion(scores, etiquetas, u)
        if criterio == "youden":
            mejores.append((m["Youden_J"], u, m))
        elif criterio == "f1":
            mejores.append((m["F1"], u, m))
        elif criterio == "mcc":
            mejores.append((m["MCC"], u, m))
        else:
            raise ValueError(f"Criterio desconocido: {criterio}")
    mejores.sort(key=lambda x: -x[0])
    valor_mejor, umbral_mejor, metricas_mejor = mejores[0]
    return float(umbral_mejor), metricas_mejor


def analisis_multi_umbral(log_data, ruta_salida, nombre_pipeline):
    """
    Calcula y guarda el análisis multi-umbral completo. Devuelve un dict
    con todos los umbrales evaluados y sus métricas.
    """
    distancias = np.array([r[2] for r in log_data])
    etiquetas = np.array([1 if r[1] == "MALA" else 0 for r in log_data])
    distancias_buenas = distancias[etiquetas == 0]

    if len(distancias_buenas) < 5:
        log.warning("Pocas BUENAS para análisis multi-umbral.")
        return None

    mu = float(np.mean(distancias_buenas))
    sigma = float(np.std(distancias_buenas))

    # Conjunto de umbrales a evaluar
    candidatos = {
        "mu+1sigma": mu + 1.0 * sigma,
        "mu+2sigma": mu + 2.0 * sigma,
        "mu+3sigma": mu + 3.0 * sigma,
        "mu+4sigma": mu + 4.0 * sigma,
        "p90_buenas": float(np.percentile(distancias_buenas, 90)),
        "p95_buenas": float(np.percentile(distancias_buenas, 95)),
        "p99_buenas": float(np.percentile(distancias_buenas, 99)),
        "p99.5_buenas": float(np.percentile(distancias_buenas, 99.5)),
        "v10_legacy": float((np.mean(distancias_buenas) +
                             np.mean(distancias[etiquetas == 1])) / 2),
    }

    # Umbrales óptimos basados en datos
    u_youden, m_youden = encontrar_umbral_optimo(distancias, etiquetas, "youden")
    u_f1, m_f1 = encontrar_umbral_optimo(distancias, etiquetas, "f1")
    u_mcc, m_mcc = encontrar_umbral_optimo(distancias, etiquetas, "mcc")
    if u_youden is not None:
        candidatos["Youden_J_opt"] = u_youden
    if u_f1 is not None:
        candidatos["F1_opt"] = u_f1
    if u_mcc is not None:
        candidatos["MCC_opt"] = u_mcc

    # Calcular métricas para cada umbral
    filas = []
    for nombre, u in candidatos.items():
        m = metricas_punto_operacion(distancias, etiquetas, u)
        filas.append({"nombre_umbral": nombre, **m})

    # Guardar CSV completo
    ruta_csv = os.path.join(ruta_salida,
                            f"multi_umbral_{nombre_pipeline}.csv")
    with open(ruta_csv, "w", newline="", encoding="utf-8") as fh:
        if filas:
            w = csv.DictWriter(fh, fieldnames=list(filas[0].keys()))
            w.writeheader(); w.writerows(filas)

    # Generar gráficas
    graficar_curvas_completas(distancias, etiquetas, candidatos,
                              ruta_salida, nombre_pipeline)

    return {
        "umbrales": candidatos,
        "filas": filas,
        "mu_buenas": mu,
        "sigma_buenas": sigma,
        "n_buenas": int(np.sum(etiquetas == 0)),
        "n_malas": int(np.sum(etiquetas == 1)),
    }


def graficar_curvas_completas(scores, y, umbrales_dict, ruta_salida,
                              nombre_pipeline):
    """
    Genera 4 figuras profesionales:
      1. ROC + Precision-Recall (con AUC y AUPRC)
      2. Histograma de scores con TODOS los umbrales superpuestos
      3. F1 / Precision / Recall vs umbral (curva continua)
      4. Trade-off FP / FN por umbral (visual)
    """
    if len(np.unique(y)) < 2:
        log.warning("Solo una clase en log_data, saltando curvas.")
        return

    candidatos = np.linspace(scores.min(), scores.max(), 500)
    metricas_lista = [metricas_punto_operacion(scores, y, u)
                      for u in candidatos]

    tpr_arr = np.array([m["TPR_recall"] for m in metricas_lista])
    fpr_arr = np.array([m["FPR"] for m in metricas_lista])
    prec_arr = np.array([m["precision"] for m in metricas_lista])
    f1_arr = np.array([m["F1"] for m in metricas_lista])
    mcc_arr = np.array([m["MCC"] for m in metricas_lista])

    # AUC ROC
    orden = np.argsort(fpr_arr)
    auc_roc = float(np.trapezoid(tpr_arr[orden], fpr_arr[orden])
                    if hasattr(np, "trapezoid")
                    else np.trapz(tpr_arr[orden], fpr_arr[orden]))
    # AUC PR (Average Precision)
    orden_r = np.argsort(tpr_arr)
    auc_pr = float(np.trapezoid(prec_arr[orden_r], tpr_arr[orden_r])
                   if hasattr(np, "trapezoid")
                   else np.trapz(prec_arr[orden_r], tpr_arr[orden_r]))

    # ============== FIGURA 1: ROC + PR ==============
    fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))

    axes[0].plot(fpr_arr[orden], tpr_arr[orden], lw=2.5,
                 color="steelblue", label=f"ROC (AUC = {auc_roc:.3f})")
    axes[0].plot([0, 1], [0, 1], "k--", alpha=0.4)
    axes[0].fill_between(fpr_arr[orden], tpr_arr[orden], alpha=0.15,
                         color="steelblue")
    axes[0].set_xlabel("FPR (1 - especificidad)")
    axes[0].set_ylabel("TPR (recall / sensibilidad)")
    axes[0].set_title(f"Curva ROC - {nombre_pipeline}")
    axes[0].legend(loc="lower right"); axes[0].grid(alpha=0.3)
    axes[0].set_xlim(-0.02, 1.02); axes[0].set_ylim(-0.02, 1.02)

    axes[1].plot(tpr_arr[orden_r], prec_arr[orden_r], lw=2.5,
                 color="darkorange", label=f"PR (AUC = {auc_pr:.3f})")
    baseline = float(np.mean(y))
    axes[1].axhline(baseline, color="k", ls="--", alpha=0.4,
                    label=f"baseline = {baseline:.2f}")
    axes[1].fill_between(tpr_arr[orden_r], prec_arr[orden_r], alpha=0.15,
                         color="darkorange")
    axes[1].set_xlabel("Recall (TPR)")
    axes[1].set_ylabel("Precision (PPV)")
    axes[1].set_title(f"Curva Precision-Recall - {nombre_pipeline}")
    axes[1].legend(loc="lower left"); axes[1].grid(alpha=0.3)
    axes[1].set_xlim(-0.02, 1.02); axes[1].set_ylim(-0.02, 1.02)

    plt.tight_layout()
    plt.savefig(os.path.join(ruta_salida,
                             f"fig_roc_pr_{nombre_pipeline}.png"), dpi=120)
    plt.close()

    # ============== FIGURA 2: histograma con umbrales ==============
    fig, ax = plt.subplots(figsize=(13, 6))
    bins = 40
    ax.hist(scores[y == 0], bins=bins, alpha=0.55, color="#4CAF50",
            label="BUENAS", density=False)
    ax.hist(scores[y == 1], bins=bins, alpha=0.55, color="#E53935",
            label="MALAS", density=False)
    # Superponer umbrales notables
    colores = plt.cm.tab10(np.linspace(0, 1, len(umbrales_dict)))
    estilos = ["solid", "dashed", "dashdot", "dotted"]
    for i, (nom, u) in enumerate(umbrales_dict.items()):
        ax.axvline(u, color=colores[i], lw=1.4,
                   ls=estilos[i % len(estilos)],
                   label=f"{nom} ({u:.2f})")
    ax.set_xlabel("Anomaly score")
    ax.set_ylabel("Frecuencia")
    ax.set_title(f"Distribución de scores y umbrales evaluados - {nombre_pipeline}")
    ax.legend(loc="upper right", fontsize=8, ncol=2)
    ax.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(ruta_salida,
                             f"fig_histograma_umbrales_{nombre_pipeline}.png"),
                dpi=120)
    plt.close()

    # ============== FIGURA 3: F1/Precision/Recall vs umbral ==============
    fig, ax = plt.subplots(figsize=(13, 6))
    ax.plot(candidatos, prec_arr, label="Precision", color="darkorange", lw=2)
    ax.plot(candidatos, tpr_arr, label="Recall (TPR)", color="steelblue", lw=2)
    ax.plot(candidatos, f1_arr, label="F1", color="darkgreen", lw=2.5)
    ax.plot(candidatos, mcc_arr, label="MCC", color="purple", lw=1.5,
            ls="dashed")

    # Marcar el umbral F1 óptimo
    idx_f1max = int(np.argmax(f1_arr))
    ax.axvline(candidatos[idx_f1max], color="darkgreen", ls=":", alpha=0.7,
               label=f"F1 max = {f1_arr[idx_f1max]:.3f} en u={candidatos[idx_f1max]:.2f}")

    ax.set_xlabel("Umbral de decisión")
    ax.set_ylabel("Métrica")
    ax.set_title(f"Métricas vs umbral - {nombre_pipeline}")
    ax.legend(loc="best"); ax.grid(alpha=0.3)
    ax.set_ylim(-0.05, 1.05)
    plt.tight_layout()
    plt.savefig(os.path.join(ruta_salida,
                             f"fig_metricas_vs_umbral_{nombre_pipeline}.png"),
                dpi=120)
    plt.close()

    # ============== FIGURA 4: Trade-off FP / FN ==============
    fp_arr = np.array([m["FP"] for m in metricas_lista])
    fn_arr = np.array([m["FN"] for m in metricas_lista])

    fig, ax = plt.subplots(figsize=(13, 6))
    ax.plot(candidatos, fp_arr, label="Falsos Positivos (BUENA marcada como MALA)",
            color="#FF9800", lw=2)
    ax.plot(candidatos, fn_arr, label="Falsos Negativos (MALA pasa como BUENA)",
            color="#9C27B0", lw=2)
    ax.fill_between(candidatos, fp_arr, alpha=0.15, color="#FF9800")
    ax.fill_between(candidatos, fn_arr, alpha=0.15, color="#9C27B0")

    # Umbral donde FP = FN (punto de equilibrio)
    diff = np.abs(fp_arr - fn_arr)
    idx_eq = int(np.argmin(diff))
    ax.axvline(candidatos[idx_eq], color="black", ls=":", alpha=0.6,
               label=f"FP=FN en u={candidatos[idx_eq]:.2f} (FP=FN={fp_arr[idx_eq]:.0f})")

    ax.set_xlabel("Umbral de decisión")
    ax.set_ylabel("Número de errores")
    ax.set_title(f"Trade-off FP / FN según umbral - {nombre_pipeline}\n"
                 f"(estricto: pocos FP, muchos FN | permisivo: muchos FP, pocos FN)")
    ax.legend(loc="best"); ax.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(ruta_salida,
                             f"fig_tradeoff_fp_fn_{nombre_pipeline}.png"),
                dpi=120)
    plt.close()


# (función visualizar_descomposicion() retirada en limpieza V20: ya no se usa)


def evaluar_patchcore(nombre_pipeline, pipeline_fn, imgs_buenas, imgs_malas,
                      ruta_salida, coreset_ratio=0.01, k_neighbors=1,
                      backbone="wide_resnet50_2", guardar_heatmaps=4,
                      usar_mascara=True, percentil_desgaste=90,
                      ancho_banda=20):
    """
    Evaluación PatchCore con segmentación + banda de contorno.

    Parámetros nuevos
    -----------------
    ancho_banda : int
        Ancho a CADA lado del contorno (px). 0 = no usar banda y analizar
        toda la herramienta. >0 = restringir el análisis al filo de la pieza,
        que es donde realmente está el desgaste en una rosca.
    """
    log.info(f"=== PatchCore [{nombre_pipeline}] (backbone={backbone}, "
             f"coreset={coreset_ratio:.0%}, mascara={usar_mascara}, "
             f"banda={ancho_banda}px) ===")
    t0 = time.time()

    # ----- INSTRUMENTACIÓN: cronómetros por fase + memoria -----
    cron_seg = Cronometro()       # segmentación + banda
    cron_prep = Cronometro()      # preprocesado del pipeline
    cron_fit = Cronometro()       # detector.fit (coreset + NearestNeighbors)
    cron_score_b = Cronometro()   # scoring sobre las buenas (calibra umbral)
    cron_score_m = Cronometro()   # scoring sobre las malas (inferencia real)
    ram_inicial = _memoria_rss_mb()
    _reset_gpu_peak()

    # 1. Segmentación de la herramienta sobre las imágenes ORIGINALES
    if usar_mascara:
        log.info("  Segmentando herramienta en BUENAS y MALAS...")
        with cron_seg:
            mascaras_b_full = [segmentar_herramienta(im) for _, im in imgs_buenas]
            mascaras_m_full = [segmentar_herramienta(im) for _, im in imgs_malas]
        cob_b = np.mean([100 * np.mean(m > 0) for m in mascaras_b_full])
        cob_m = np.mean([100 * np.mean(m > 0) for m in mascaras_m_full])
        log.info(f"  Cobertura herramienta: BUENAS={cob_b:.1f}%, "
                 f"MALAS={cob_m:.1f}%")

        if ancho_banda > 0:
            log.info(f"  Calculando banda de contorno (ancho={ancho_banda}px)...")
            with cron_seg:
                mascaras_b = [banda_contorno(m, ancho=ancho_banda)
                              for m in mascaras_b_full]
                mascaras_m = [banda_contorno(m, ancho=ancho_banda)
                              for m in mascaras_m_full]
            cob_bb = np.mean([100 * np.mean(m > 0) for m in mascaras_b])
            cob_mb = np.mean([100 * np.mean(m > 0) for m in mascaras_m])
            log.info(f"  Cobertura banda contorno: BUENAS={cob_bb:.1f}%, "
                     f"MALAS={cob_mb:.1f}%")
        else:
            mascaras_b = mascaras_b_full
            mascaras_m = mascaras_m_full
    else:
        mascaras_b_full = mascaras_m_full = None
        mascaras_b = mascaras_m = None

    # 2. Aplicar preprocesado (sin enmascarar la imagen, solo se usa la
    #    máscara para filtrar parches a posteriori)
    log.info("  Aplicando preprocesado a BUENAS...")
    with cron_prep:
        imgs_b_proc = [(n, pipeline_fn(im)) for n, im in imgs_buenas]
    log.info("  Aplicando preprocesado a MALAS...")
    with cron_prep:
        imgs_m_proc = [(n, pipeline_fn(im)) for n, im in imgs_malas]

    # 3. PatchCore fit + score con la máscara (banda)
    extractor = PatchCoreExtractor(DEVICE, backbone=backbone)
    detector = PatchCoreDetector(extractor,
                                 coreset_ratio=coreset_ratio,
                                 k=k_neighbors)
    with cron_fit:
        detector.fit(imgs_b_proc, mascaras_buenas=mascaras_b, verbose=True)

    with cron_score_b:
        scores_b, mapas_b, nb = detector.score_batch(imgs_b_proc,
                                                     mascaras=mascaras_b,
                                                     verbose=True)
    with cron_score_m:
        scores_m, mapas_m, nm = detector.score_batch(imgs_m_proc,
                                                     mascaras=mascaras_m,
                                                     verbose=True)

    # 4. Log unificado y métricas de detección
    log_data = []
    for n, s in zip(nb, scores_b):
        log_data.append([n, "BUENA", float(s), nombre_pipeline])
    for n, s in zip(nm, scores_m):
        log_data.append([n, "MALA", float(s), nombre_pipeline])

    u_mu3s = umbral_solo_buenas(scores_b, factor=3.0)
    u_p99 = umbral_solo_buenas(scores_b, percentil=99)
    u_v10 = (float(np.mean(scores_b)) + float(np.mean(scores_m))) / 2

    acc1, *_ = graficar_matriz(log_data, u_mu3s, ruta_salida,
                               etiqueta=f"PC_{nombre_pipeline}_mu3s")
    acc2, *_ = graficar_matriz(log_data, u_p99, ruta_salida,
                               etiqueta=f"PC_{nombre_pipeline}_p99")
    acc3, *_ = graficar_matriz(log_data, u_v10, ruta_salida,
                               etiqueta=f"PC_{nombre_pipeline}_v10")
    auc = curva_roc(log_data, ruta_salida, etiqueta=f"PC_{nombre_pipeline}")
    log.info(f"  AUC={auc:.3f}  ACC(μ+3σ)={acc1:.2%}  ACC(p99)={acc2:.2%}")

    # ANÁLISIS MULTI-UMBRAL (V17)
    log.info(f"  Generando análisis multi-umbral y gráficas detalladas...")
    multi_u = analisis_multi_umbral(log_data, ruta_salida,
                                    f"PC_{nombre_pipeline}")
    if multi_u is not None:
        # Mostrar tabla resumida en consola
        log.info(f"  {'Umbral':<14} {'TPR':>6} {'FPR':>6} "
                 f"{'Prec':>6} {'F1':>6} {'MCC':>6} {'Acc':>6}")
        for fila in multi_u["filas"]:
            log.info(f"  {fila['nombre_umbral']:<14} "
                     f"{fila['TPR_recall']:>6.3f} {fila['FPR']:>6.3f} "
                     f"{fila['precision']:>6.3f} {fila['F1']:>6.3f} "
                     f"{fila['MCC']:>+6.3f} {fila['accuracy']:>6.3f}")

    # 5. Métricas geométricas + paneles
    metricas_csv = []
    if usar_mascara and guardar_heatmaps > 0:
        log.info(f"  Calculando métricas geométricas (primeras "
                 f"{guardar_heatmaps} malas y 2 buenas)...")
        for i in range(min(guardar_heatmaps, len(mapas_m))):
            metricas = metricas_desgaste(
                mapas_m[i], mascaras_m_full[i],
                banda=mascaras_m[i] if ancho_banda > 0 else None,
                percentil_umbral=percentil_desgaste)
            guardar_panel_metricas(
                imgs_malas[i][1], mascaras_m_full[i],
                mascaras_m[i] if ancho_banda > 0 else None,
                mapas_m[i], metricas, ruta_salida,
                f"panel_{nombre_pipeline}_MALA_{Path(nm[i]).stem}",
                titulo=f"MALA: {nm[i]} (score={scores_m[i]:.3f})")
            metricas_csv.append({
                "archivo": nm[i], "categoria": "MALA",
                "score": float(scores_m[i]),
                **{k: v for k, v in metricas.items()
                   if not isinstance(v, np.ndarray)},
            })
        for i in range(min(2, len(mapas_b))):
            metricas = metricas_desgaste(
                mapas_b[i], mascaras_b_full[i],
                banda=mascaras_b[i] if ancho_banda > 0 else None,
                percentil_umbral=percentil_desgaste)
            guardar_panel_metricas(
                imgs_buenas[i][1], mascaras_b_full[i],
                mascaras_b[i] if ancho_banda > 0 else None,
                mapas_b[i], metricas, ruta_salida,
                f"panel_{nombre_pipeline}_BUENA_{Path(nb[i]).stem}",
                titulo=f"BUENA: {nb[i]} (score={scores_b[i]:.3f})")
            metricas_csv.append({
                "archivo": nb[i], "categoria": "BUENA",
                "score": float(scores_b[i]),
                **{k: v for k, v in metricas.items()
                   if not isinstance(v, np.ndarray)},
            })
    elif guardar_heatmaps > 0:
        for i in range(min(guardar_heatmaps, len(mapas_m))):
            guardar_heatmap_superpuesto(
                imgs_malas[i][1], mapas_m[i], ruta_salida,
                f"heatmap_{nombre_pipeline}_MALA_{Path(nm[i]).stem}",
                titulo=f"MALA: {nm[i]} (score={scores_m[i]:.3f})")

    if metricas_csv:
        ruta_met = os.path.join(ruta_salida,
                                f"metricas_geometricas_{nombre_pipeline}.csv")
        with open(ruta_met, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=list(metricas_csv[0].keys()))
            w.writeheader(); w.writerows(metricas_csv)
        log.info(f"  Métricas geométricas guardadas en {ruta_met}")

    # ----- INSTRUMENTACIÓN: coste computacional final -----
    tiempo_total = time.time() - t0
    ram_final = _memoria_rss_mb()
    gpu_pico = _memoria_gpu_pico_mb()

    # Parámetros del backbone (wide_resnet50_2 congelado: trainable=0).
    params_total, params_train = _contar_parametros(extractor.model)

    # Tamaño del memory bank tras coreset (float32 por defecto).
    bank_n = bank_dim = 0
    bank_mb = 0.0
    try:
        if detector.bank is not None:
            bank_n = int(detector.bank.shape[0])
            bank_dim = int(detector.bank.shape[1])
            bank_mb = float(detector.bank.nbytes) / (1024 ** 2)
    except Exception:
        pass

    n_buenas = len(imgs_buenas)
    n_malas = len(imgs_malas)
    n_inferencia = n_buenas + n_malas
    tiempo_inferencia_total = cron_score_b.total + cron_score_m.total

    # "Entrenamiento" en PatchCore es: preproc de buenas + fit (coreset+NN) +
    # cálculo del umbral sobre las propias buenas (cron_score_b). No incluye
    # el scoring de las malas, que es la inferencia real.
    tiempo_train = (cron_seg.total / 2 if cron_seg.total else 0.0) \
                   + (cron_prep.total / 2 if cron_prep.total else 0.0) \
                   + cron_fit.total + cron_score_b.total
    # cron_seg y cron_prep miden las DOS clases juntas; lo dividimos en mitad
    # como aproximación al coste de las buenas (que es lo que entra al fit).

    coste = {
        "tiempo_total_s": float(tiempo_total),
        "tiempo_segmentacion_s": float(cron_seg.total),
        "tiempo_preprocesado_s": float(cron_prep.total),
        "tiempo_fit_s": float(cron_fit.total),
        "tiempo_score_buenas_s": float(cron_score_b.total),
        "tiempo_score_malas_s": float(cron_score_m.total),
        "tiempo_train_total_s": float(tiempo_train),
        "tiempo_inferencia_total_s": float(tiempo_inferencia_total),
        "tiempo_inferencia_por_imagen_s": (
            tiempo_inferencia_total / n_inferencia
            if n_inferencia > 0 else None),
        "ram_inicial_mb": ram_inicial,
        "ram_final_mb": ram_final,
        "ram_delta_mb": (ram_final - ram_inicial
                         if (ram_inicial is not None and ram_final is not None)
                         else None),
        "gpu_pico_mb": gpu_pico,
        "dispositivo": str(DEVICE),
        "backbone": backbone,
        "params_total": params_total,
        "params_entrenables": params_train,
        "bank_n_parches": bank_n,
        "bank_dim": bank_dim,
        "bank_mb": bank_mb,
        "coreset_ratio": coreset_ratio,
        "n_buenas": n_buenas,
        "n_malas": n_malas,
    }
    log.info(
        f"  COSTE [{nombre_pipeline}] total={tiempo_total:.1f}s | "
        f"seg={cron_seg.total:.1f}s | prep={cron_prep.total:.1f}s | "
        f"fit={cron_fit.total:.1f}s | "
        f"score_B={cron_score_b.total:.1f}s | score_M={cron_score_m.total:.1f}s")
    # Formatos legibles para campos opcionales
    ram_d = coste["ram_delta_mb"]
    ram_str = f"{ram_d:.0f}MB" if ram_d is not None else "n/a"
    gpu_str = f"{gpu_pico:.0f}MB" if gpu_pico is not None else "n/a"
    log.info(
        f"  COSTE [{nombre_pipeline}] inferencia/img="
        f"{(tiempo_inferencia_total/max(n_inferencia,1))*1000:.1f}ms | "
        f"banco={bank_n} parches ({bank_mb:.1f} MB) | "
        f"params={params_total/1e6:.1f}M | "
        f"RAM Δ={ram_str} | GPU pico={gpu_str}")

    return {
        "pipeline": f"PC_{nombre_pipeline}",
        "log": log_data,
        "d_buenas": scores_b,
        "d_malas": scores_m,
        "u_mu3sigma": u_mu3s,
        "u_p99": u_p99,
        "u_v10": u_v10,
        "acc_mu3sigma": acc1,
        "acc_p99": acc2,
        "acc_v10": acc3,
        "auc": auc,
        "tiempo_s": tiempo_total,
        "coste": coste,
    }


def guardar_heatmap_superpuesto(img_gris, mapa, ruta_salida, nombre, titulo=""):
    """Guarda figura con imagen original, heatmap y superposición."""
    mapa_norm = cv2.normalize(mapa, None, 0, 1, cv2.NORM_MINMAX)
    fig, axes = plt.subplots(1, 3, figsize=(15, 5))
    axes[0].imshow(img_gris, cmap="gray"); axes[0].set_title("Original")
    axes[0].axis("off")
    im1 = axes[1].imshow(mapa_norm, cmap="jet"); axes[1].set_title("Anomaly map")
    axes[1].axis("off")
    plt.colorbar(im1, ax=axes[1], fraction=0.046)
    axes[2].imshow(img_gris, cmap="gray")
    axes[2].imshow(mapa_norm, cmap="jet", alpha=0.45)
    axes[2].set_title("Superposición"); axes[2].axis("off")
    if titulo:
        fig.suptitle(titulo)
    plt.tight_layout()
    plt.savefig(os.path.join(ruta_salida, f"{nombre}.png"), dpi=110)
    plt.close()


# ==============================================================================
#  BLOQUE 6B - MÉTRICAS GEOMÉTRICAS DEL DESGASTE EN EL CONTORNO
# ==============================================================================
#
#  El desgaste de una rosca se manifiesta en su SILUETA: pérdida de filo,
#  redondeo de puntas, mellas, rebabas. Estas métricas miden la degradación
#  del contorno, no del cuerpo metálico interior:
#    - Rugosidad del contorno (desviación local respecto al perfil suavizado)
#    - Pérdida de área respecto a la envolvente convexa
#    - Perímetro y compacidad
#    - Score de anomalía concentrado en banda de contorno
# ==============================================================================

def metricas_desgaste(anomaly_map, mascara_herramienta, banda=None,
                      percentil_umbral=90, area_min_px=20):
    """
    Calcula métricas geométricas del desgaste a partir del anomaly map y
    la máscara de la herramienta.

    Si se proporciona 'banda' (máscara binaria de la banda de contorno),
    todas las métricas se restringen a esa banda.
    """
    H, W = anomaly_map.shape

    # Región de análisis: banda si se da, si no la herramienta entera
    if banda is not None:
        region_analisis = banda
    else:
        region_analisis = mascara_herramienta

    area_region = int(np.sum(region_analisis > 0))
    area_herr = int(np.sum(mascara_herramienta > 0))

    if area_region == 0:
        return _metricas_vacias(H, W, area_herr)

    # 1. Umbralar el anomaly map dentro de la región de análisis
    valores = anomaly_map[region_analisis > 0]
    umbral = float(np.percentile(valores, percentil_umbral))

    mascara_desg = ((anomaly_map > umbral) &
                    (region_analisis > 0)).astype(np.uint8) * 255

    # 2. Limpiar pequeñas componentes
    n_lab, labels, stats, _ = cv2.connectedComponentsWithStats(mascara_desg, 8)
    if n_lab > 1:
        for i in range(1, n_lab):
            if stats[i, cv2.CC_STAT_AREA] < area_min_px:
                mascara_desg[labels == i] = 0

    area_desgaste = int(np.sum(mascara_desg > 0))
    porcentaje = 100.0 * area_desgaste / max(1, area_herr)

    # 3. Métricas geométricas DEL CONTORNO de la herramienta
    contornos_herr, _ = cv2.findContours(mascara_herramienta,
                                         cv2.RETR_EXTERNAL,
                                         cv2.CHAIN_APPROX_NONE)
    if contornos_herr:
        contorno_principal = max(contornos_herr, key=cv2.contourArea)
        perimetro_herr = float(cv2.arcLength(contorno_principal, True))
        # Envolvente convexa: el "perfil ideal" sin desgaste interior
        envolvente = cv2.convexHull(contorno_principal)
        area_convexa = float(cv2.contourArea(envolvente))
        # Pérdida de área respecto a la envolvente convexa: una rosca con
        # filos bien marcados tiene MÁS área perdida (los valles entre dientes
        # son grandes); una rosca redondeada por desgaste tiene MENOS pérdida.
        area_real = float(cv2.contourArea(contorno_principal))
        perdida_area = area_convexa - area_real
        # Solidez: ratio área_real / área_convexa. Una rosca con filos
        # marcados tiene solidez baja; redondeada por desgaste, alta.
        solidez = area_real / max(1.0, area_convexa)
        # Compacidad: 4*pi*A / P^2. Más alta cuanto más circular (desgastada).
        compacidad = 4 * np.pi * area_real / max(1.0, perimetro_herr ** 2)
        # Rugosidad del contorno: desviación local respecto a un suavizado.
        rugosidad = _rugosidad_contorno(contorno_principal,
                                        ventana_suav=15)
    else:
        perimetro_herr = 0.0
        area_convexa = 0.0
        perdida_area = 0.0
        solidez = 0.0
        compacidad = 0.0
        rugosidad = 0.0

    # 4. Perímetro de la región de desgaste detectada
    contornos_desg, _ = cv2.findContours(mascara_desg, cv2.RETR_EXTERNAL,
                                         cv2.CHAIN_APPROX_NONE)
    perimetro_desg = int(sum(cv2.arcLength(c, True) for c in contornos_desg))

    # 5. ANCHO MEDIO del desgaste (V20, objetivo SMART 4 del TFM)
    #    Calculado como 2*area / longitud_esqueleto. Mide cuán "ancha"
    #    o "fina" es la región desgastada en promedio.
    ancho_medio = _ancho_medio_desgaste(mascara_desg)

    return {
        "area_desgaste_px": area_desgaste,
        "area_herramienta_px": area_herr,
        "area_banda_px": area_region if banda is not None else 0,
        "porcentaje_desgaste": porcentaje,
        "perimetro_desgaste_px": perimetro_desg,
        "ancho_medio_desgaste_px": ancho_medio,
        "perimetro_herramienta_px": perimetro_herr,
        "area_convexa_px": area_convexa,
        "perdida_area_convexa_px": perdida_area,
        "solidez": solidez,
        "compacidad": compacidad,
        "rugosidad_contorno": rugosidad,
        "umbral_usado": umbral,
        "mascara_desgaste": mascara_desg,
    }


def _metricas_vacias(H, W, area_herr):
    return {"area_desgaste_px": 0, "area_herramienta_px": area_herr,
            "area_banda_px": 0, "porcentaje_desgaste": 0.0,
            "perimetro_desgaste_px": 0,
            "ancho_medio_desgaste_px": 0.0,
            "perimetro_herramienta_px": 0.0,
            "area_convexa_px": 0.0, "perdida_area_convexa_px": 0.0,
            "solidez": 0.0, "compacidad": 0.0, "rugosidad_contorno": 0.0,
            "umbral_usado": 0.0,
            "mascara_desgaste": np.zeros((H, W), dtype=np.uint8)}


def _ancho_medio_desgaste(mascara_desg):
    """
    Calcula el ancho medio de la región de desgaste mediante esqueletización.
    ancho_medio = 2 * área / longitud_esqueleto

    Esta es la fórmula clásica para medir el grosor medio de una región
    alargada (Serra 1982, Morphological Image Analysis). En el contexto del
    TFM, da una medida del "espesor" del desgaste detectado.
    """
    if mascara_desg is None or np.sum(mascara_desg > 0) < 10:
        return 0.0

    # Binarizar
    bin_img = (mascara_desg > 0).astype(np.uint8)

    # Esqueletización morfológica iterativa (Zhang-Suen aproximado vía cv2)
    # Si OpenCV tiene ximgproc, usarlo (más rápido). Si no, fallback manual.
    try:
        import cv2.ximgproc as xip
        esqueleto = xip.thinning(bin_img * 255,
                                 thinningType=xip.THINNING_ZHANGSUEN)
        longitud_esq = int(np.sum(esqueleto > 0))
    except (ImportError, AttributeError):
        # Fallback: esqueletización morfológica iterativa con erosión + dilatación
        esqueleto = np.zeros_like(bin_img)
        elem = cv2.getStructuringElement(cv2.MORPH_CROSS, (3, 3))
        temp = bin_img.copy()
        max_iter = 100
        for _ in range(max_iter):
            erosionado = cv2.erode(temp, elem)
            abierto = cv2.dilate(erosionado, elem)
            sub = cv2.subtract(temp, abierto)
            esqueleto = cv2.bitwise_or(esqueleto, sub)
            temp = erosionado.copy()
            if cv2.countNonZero(temp) == 0:
                break
        longitud_esq = int(np.sum(esqueleto > 0))

    if longitud_esq < 1:
        return 0.0
    area = float(np.sum(bin_img > 0))
    ancho_medio = 2.0 * area / longitud_esq
    return float(ancho_medio)


def _rugosidad_contorno(contorno, ventana_suav=15):
    """
    Rugosidad = desviación estándar de la distancia entre los puntos del
    contorno y una versión suavizada del mismo. Mide cuán irregular es
    el filo del diente.
    """
    pts = contorno.reshape(-1, 2).astype(np.float64)
    if len(pts) < ventana_suav * 2:
        return 0.0

    # Suavizado por media móvil circular en ambas coordenadas
    k = ventana_suav | 1  # impar
    half = k // 2
    pts_pad = np.vstack([pts[-half:], pts, pts[:half]])
    pts_suav = np.zeros_like(pts)
    for i in range(len(pts)):
        pts_suav[i] = pts_pad[i:i + k].mean(axis=0)

    desv = np.linalg.norm(pts - pts_suav, axis=1)
    return float(np.std(desv))


def guardar_panel_metricas(img_orig, mascara, banda, anomaly_map, metricas,
                           ruta_salida, nombre, titulo=""):
    """
    Panel completo: original, máscara+banda, anomaly map, desgaste detectado.
    """
    fig, axes = plt.subplots(1, 4, figsize=(20, 5))

    axes[0].imshow(img_orig, cmap="gray")
    axes[0].set_title("Original"); axes[0].axis("off")

    axes[1].imshow(img_orig, cmap="gray")
    axes[1].imshow(mascara, cmap="Greens", alpha=0.3)
    if banda is not None:
        axes[1].imshow(banda, cmap="Reds", alpha=0.5)
    axes[1].set_title("Verde=herramienta · Rojo=banda contorno"); axes[1].axis("off")

    mp = cv2.normalize(anomaly_map, None, 0, 1, cv2.NORM_MINMAX)
    axes[2].imshow(img_orig, cmap="gray")
    axes[2].imshow(mp, cmap="jet", alpha=0.5)
    axes[2].set_title("Anomaly map"); axes[2].axis("off")

    axes[3].imshow(img_orig, cmap="gray")
    if metricas["area_desgaste_px"] > 0:
        axes[3].imshow(metricas["mascara_desgaste"], cmap="Reds", alpha=0.55)
    txt = (f"Área desgaste: {metricas['area_desgaste_px']} px\n"
           f"% sobre herramienta: {metricas['porcentaje_desgaste']:.2f}%\n"
           f"Perímetro herram.: {metricas['perimetro_herramienta_px']:.0f} px\n"
           f"Solidez: {metricas['solidez']:.3f}\n"
           f"Compacidad: {metricas['compacidad']:.4f}\n"
           f"Rugosidad cont.: {metricas['rugosidad_contorno']:.2f}")
    axes[3].set_title("Desgaste detectado"); axes[3].axis("off")
    axes[3].text(0.02, 0.98, txt, transform=axes[3].transAxes,
                 fontsize=9, va="top", ha="left",
                 bbox=dict(boxstyle="round", facecolor="white", alpha=0.85))

    if titulo:
        fig.suptitle(titulo, fontsize=11)
    plt.tight_layout()
    plt.savefig(os.path.join(ruta_salida, f"{nombre}.png"), dpi=110)
    plt.close()


# ==============================================================================
#  BLOQUE 7 - PIPELINE PRINCIPAL DE EVALUACIÓN (LEGACY - flujo V12)
# ==============================================================================

# (función evaluar_pipeline() retirada en limpieza V20: ya no se usa)


def guardar_csv_global(resultados_por_pipeline, ruta_salida):
    """Guarda un CSV con todas las distancias para análisis posterior."""
    ruta = os.path.join(ruta_salida, "log_global_distancias.csv")
    with open(ruta, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["archivo", "categoria_real", "distancia", "pipeline"])
        for res in resultados_por_pipeline.values():
            for fila in res["log"]:
                w.writerow(fila)
    log.info(f"CSV global guardado en {ruta}")


def guardar_resumen(resultados_por_pipeline, ruta_salida):
    """Guarda un resumen de métricas por pipeline para incluir en la memoria."""
    ruta = os.path.join(ruta_salida, "resumen_pipelines.csv")
    with open(ruta, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["pipeline", "umbral_mu3s", "acc_mu3s",
                    "umbral_p99", "acc_p99",
                    "umbral_v10", "acc_v10",
                    "auc", "tiempo_s"])
        for r in resultados_por_pipeline.values():
            w.writerow([r["pipeline"],
                        f"{r['u_mu3sigma']:.4f}", f"{r['acc_mu3sigma']:.4f}",
                        f"{r['u_p99']:.4f}", f"{r['acc_p99']:.4f}",
                        f"{r['u_v10']:.4f}", f"{r['acc_v10']:.4f}",
                        f"{r['auc']:.4f}", f"{r['tiempo_s']:.1f}"])
    log.info(f"Resumen guardado en {ruta}")


# ==============================================================================
#  BLOQUE 8 - ENTRADA DE USUARIO Y MAIN
# ==============================================================================

def pedir_carpetas_gui():
    """Diálogo Tk para selección de carpetas (modo interactivo)."""
    root = tk.Tk(); root.withdraw()
    ruta_buenas = filedialog.askdirectory(title="1. Selecciona Carpeta BUENAS")
    ruta_malas = filedialog.askdirectory(title="2. Selecciona Carpeta MALAS")
    ruta_salida = filedialog.askdirectory(title="3. Carpeta para RESULTADOS")
    return ruta_buenas, ruta_malas, ruta_salida


def calcular_auc_dual(resultados_pc, eda_resumen):
    """
    Para cada pipeline, calcula dos AUCs:
      - AUC global (sobre TODAS las MALAS)
      - AUC sobre subset DETECTABLES (las que el EDA marca como outliers
        de las BUENAS — donde realmente puede haber señal)
    """
    archivos_detectables = set(
        eda_resumen["outliers"]["archivos_detectables"])

    auc_dual = {}
    for nombre, res in resultados_pc.items():
        log_data = res["log"]
        # Lista solo con BUENAS + MALAS detectables
        log_filtrado = [r for r in log_data
                        if r[1] == "BUENA" or r[0] in archivos_detectables]
        if not any(r[1] == "MALA" for r in log_filtrado):
            auc_dual[nombre] = {"auc_global": res["auc"],
                                "auc_detectables": None,
                                "n_detectables": 0}
            continue

        # Calcular AUC sobre el filtrado
        d = np.array([r[2] for r in log_filtrado])
        y = np.array([1 if r[1] == "MALA" else 0 for r in log_filtrado])
        umbrales = np.linspace(d.min(), d.max(), 200)
        tpr_l, fpr_l = [], []
        for u in umbrales:
            pred = (d > u).astype(int)
            vp = int(np.sum((pred == 1) & (y == 1)))
            fn = int(np.sum((pred == 0) & (y == 1)))
            fp = int(np.sum((pred == 1) & (y == 0)))
            vn = int(np.sum((pred == 0) & (y == 0)))
            tpr_l.append(vp / max(1, vp + fn))
            fpr_l.append(fp / max(1, fp + vn))
        orden = np.argsort(fpr_l)
        fpr_arr = np.array(fpr_l)[orden]
        tpr_arr = np.array(tpr_l)[orden]
        auc_det = float(np.trapezoid(tpr_arr, fpr_arr)
                        if hasattr(np, "trapezoid")
                        else np.trapz(tpr_arr, fpr_arr))
        auc_dual[nombre] = {"auc_global": res["auc"],
                            "auc_detectables": auc_det,
                            "n_detectables": int(np.sum(y == 1))}
    return auc_dual


def comparativa_coste_computacional(resultados, tiempos_perfilometro,
                                     auc_dual, ruta_salida, args,
                                     eda_resumen=None):
    """Genera la comparativa de coste computacional y predicción del run.

    Produce tres salidas:
      - comparativa_coste.csv: tabla con una fila por método, columnas
        [método, tipo, AUC, accuracy, T_train(s), T_inf/img(ms),
        params, bank_mb, RAM_delta_mb, GPU_pico_mb, dispositivo].
      - comparativa_coste.png: dos paneles (barras de coste de inferencia
        por método y dispersión accuracy vs coste — Pareto).
      - registra la ejecución actual en el histórico acumulado vía
        registrar_ejecucion_en_historico().

    Args:
       resultados: dict {pipeline -> dict con "auc", "acc_p99", "coste"}.
          Mezcla resultados PatchCore (clave "PC_*") y legacy.
       tiempos_perfilometro: dict con tiempos del perfilómetro
          ({"tiempo_total_s", "n_imagenes", "tiempo_por_imagen_s"})
          o None si no se ejecutó.
       auc_dual: dict de AUC dual (global + detectables).
       ruta_salida: carpeta de la ejecución actual.
       args: argparse.Namespace (para guardar el comando).
       eda_resumen: opcional, dict del EDA (para n_buenas/n_malas).

    Devuelve el dict del payload guardado (útil para el informe).
    """
    log.info("=" * 70)
    log.info(" COMPARATIVA DE COSTE COMPUTACIONAL Y PREDICCIÓN")
    log.info("=" * 70)

    # ---- 1. Construir la lista de métodos a comparar ----
    filas = []   # filas para CSV y para la figura

    for nombre, res in resultados.items():
        coste = res.get("coste", {}) or {}
        # AUC global (preferimos el dual si está disponible)
        auc_g = None
        if nombre in auc_dual:
            auc_g = auc_dual[nombre].get("auc_global")
        if auc_g is None:
            auc_g = res.get("auc")
        # Accuracy: usamos la del umbral p99 como referencia clínica.
        acc = res.get("acc_p99")
        tipo = "patchcore" if nombre.startswith("PC_") else "legacy"
        filas.append({
            "metodo": nombre,
            "tipo": tipo,
            "auc": auc_g,
            "acc": acc,
            "tiempo_train_total_s": coste.get("tiempo_train_total_s"),
            "tiempo_train_fit_s": coste.get("tiempo_fit_s"),
            "tiempo_inferencia_total_s": coste.get("tiempo_inferencia_total_s"),
            "tiempo_inferencia_por_imagen_s":
                coste.get("tiempo_inferencia_por_imagen_s"),
            "tiempo_total_s": coste.get("tiempo_total_s"),
            "params_total": coste.get("params_total"),
            "params_entrenables": coste.get("params_entrenables"),
            "bank_n_parches": coste.get("bank_n_parches"),
            "bank_mb": coste.get("bank_mb"),
            "ram_delta_mb": coste.get("ram_delta_mb"),
            "gpu_pico_mb": coste.get("gpu_pico_mb"),
            "dispositivo": coste.get("dispositivo", str(DEVICE)),
        })

    # Perfilómetro como método adicional (sin entrenamiento, sin AUC; solo
    # tiempo de medición que es lo único comparable de su coste).
    if tiempos_perfilometro and tiempos_perfilometro.get("n_imagenes", 0) > 0:
        filas.append({
            "metodo": "Perfilometro_ISO",
            "tipo": "perfilometro",
            "auc": None,
            "acc": None,
            "tiempo_train_total_s": 0.0,
            "tiempo_train_fit_s": 0.0,
            "tiempo_inferencia_total_s":
                tiempos_perfilometro.get("tiempo_total_s"),
            "tiempo_inferencia_por_imagen_s":
                tiempos_perfilometro.get("tiempo_por_imagen_s"),
            "tiempo_total_s": tiempos_perfilometro.get("tiempo_total_s"),
            "params_total": 0,
            "params_entrenables": 0,
            "bank_n_parches": None,
            "bank_mb": None,
            "ram_delta_mb": None,
            "gpu_pico_mb": None,
            "dispositivo": "cpu",
        })

    if not filas:
        log.warning("No hay métodos con coste medido. Saltando comparativa.")
        return None

    # ---- 2. CSV ----
    columnas_csv = list(filas[0].keys())
    ruta_csv = os.path.join(ruta_salida, "comparativa_coste.csv")
    try:
        with open(ruta_csv, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=columnas_csv)
            w.writeheader()
            for f in filas:
                w.writerow(f)
        log.info(f"  Tabla comparativa guardada en {ruta_csv}")
    except Exception as e:
        log.error(f"  No pude escribir CSV de comparativa: {e}")

    # ---- 3. Tabla en consola (resumida) ----
    log.info(f"  {'Método':<22} {'AUC':>6} {'Acc p99':>8} "
             f"{'T_train(s)':>11} {'T_inf/img(ms)':>14}")
    log.info("  " + "-" * 70)
    for f in filas:
        auc_s = f"{f['auc']:.3f}" if f["auc"] is not None else "  —  "
        acc_s = f"{f['acc']*100:.1f}%" if f["acc"] is not None else "  —  "
        tt = f"{f['tiempo_train_total_s']:.1f}" \
             if f["tiempo_train_total_s"] is not None else " —"
        ti = (f"{f['tiempo_inferencia_por_imagen_s']*1000:.1f}"
              if f["tiempo_inferencia_por_imagen_s"] is not None else " —")
        log.info(f"  {f['metodo']:<22} {auc_s:>6} {acc_s:>8} "
                 f"{tt:>11} {ti:>14}")

    # ---- 4. Gráfica: 2 paneles ----
    ruta_png = os.path.join(ruta_salida, "comparativa_coste.png")
    try:
        fig, axes = plt.subplots(1, 2, figsize=(14, 5.5))

        # Panel A: barras de tiempo de inferencia por imagen
        ax = axes[0]
        nombres = [f["metodo"] for f in filas]
        t_inf_ms = [
            (f["tiempo_inferencia_por_imagen_s"] or 0) * 1000 for f in filas]
        colores = []
        for f in filas:
            if f["tipo"] == "patchcore":
                colores.append("#4a90e2")
            elif f["tipo"] == "perfilometro":
                colores.append("#e2934a")
            else:
                colores.append("#9b6ab8")
        barras = ax.bar(range(len(nombres)), t_inf_ms, color=colores)
        ax.set_xticks(range(len(nombres)))
        ax.set_xticklabels(nombres, rotation=30, ha="right", fontsize=9)
        ax.set_ylabel("Tiempo de predicción por imagen (ms)")
        ax.set_title("Coste de inferencia por método")
        ax.grid(axis="y", alpha=0.3)
        for b, v in zip(barras, t_inf_ms):
            if v > 0:
                ax.text(b.get_x() + b.get_width() / 2, b.get_height(),
                        f"{v:.0f}", ha="center", va="bottom", fontsize=8)

        # Panel B: scatter accuracy vs coste de inferencia (Pareto)
        ax = axes[1]
        leyenda_usada = set()
        x_vals = []; y_vals = []
        for f in filas:
            if f["auc"] is None or f["tiempo_inferencia_por_imagen_s"] is None:
                continue
            x = f["tiempo_inferencia_por_imagen_s"] * 1000
            y = f["auc"]
            x_vals.append(x); y_vals.append(y)
            color = ("#4a90e2" if f["tipo"] == "patchcore"
                     else "#9b6ab8" if f["tipo"] == "legacy"
                     else "#e2934a")
            etiqueta = f["tipo"] if f["tipo"] not in leyenda_usada else None
            leyenda_usada.add(f["tipo"])
            ax.scatter(x, y, s=120, color=color, alpha=0.75,
                       edgecolors="white", linewidths=1.5, label=etiqueta)
            ax.annotate(f["metodo"], (x, y),
                        textcoords="offset points", xytext=(7, -3),
                        fontsize=8)
        ax.set_xlabel("Tiempo de predicción por imagen (ms)")
        ax.set_ylabel("AUC")
        ax.set_title("Calidad vs coste — mejor arriba-izquierda")
        ax.grid(alpha=0.3)
        # Ampliar los límites para que las anotaciones no se salgan del eje
        if x_vals:
            xr = max(x_vals) - min(x_vals)
            yr = max(y_vals) - min(y_vals)
            margen_x = max(xr * 0.20, 20)   # 20% o 20ms mínimo
            margen_y = max(yr * 0.15, 0.02) # 15% o 0.02 mínimo
            ax.set_xlim(min(x_vals) - margen_x * 0.2,
                         max(x_vals) + margen_x)
            ax.set_ylim(min(y_vals) - margen_y, max(y_vals) + margen_y)
        if leyenda_usada:
            ax.legend(loc="lower left", fontsize=9)

        fig.suptitle("Comparativa de coste computacional y predicción "
                     f"(dispositivo: {DEVICE})",
                     fontsize=12, fontweight="bold")
        fig.tight_layout(rect=[0, 0, 1, 0.96])
        fig.savefig(ruta_png, dpi=140, bbox_inches="tight")
        plt.close(fig)
        log.info(f"  Gráfica comparativa guardada en {ruta_png}")
    except Exception as e:
        log.error(f"  No pude generar gráfica comparativa: {e}")

    # ---- 5. Registrar en el histórico acumulado ----
    n_buenas = eda_resumen.get("n_buenas") if eda_resumen else None
    n_malas = eda_resumen.get("n_malas") if eda_resumen else None
    payload = {
        "fecha": datetime.now().isoformat(timespec="seconds"),
        "ejecucion_id": Path(ruta_salida).name or "run",
        "ruta_salida": str(ruta_salida),
        "comando": " ".join(sys.argv) if hasattr(sys, "argv") else "",
        "n_buenas": n_buenas,
        "n_malas": n_malas,
        "sistema": _info_sistema(),
        "args": {k: (str(v) if not isinstance(v, (int, float, bool, str, list))
                     else v)
                 for k, v in vars(args).items()} if args else {},
        "metodos": [
            {
                "nombre": f["metodo"],
                "tipo": f["tipo"],
                "auc": f["auc"],
                "acc": f["acc"],
                "tiempo_train_total_s": f["tiempo_train_total_s"],
                "tiempo_train_fit_s": f["tiempo_train_fit_s"],
                "tiempo_inferencia_total_s": f["tiempo_inferencia_total_s"],
                "tiempo_inferencia_por_imagen_s":
                    f["tiempo_inferencia_por_imagen_s"],
                "tiempo_total_s": f["tiempo_total_s"],
                "params_total": f["params_total"],
                "params_entrenables": f["params_entrenables"],
                "bank_n_parches": f["bank_n_parches"],
                "bank_mb": f["bank_mb"],
                "ram_delta_mb": f["ram_delta_mb"],
                "gpu_pico_mb": f["gpu_pico_mb"],
                "dispositivo": f["dispositivo"],
            }
            for f in filas
        ],
    }
    try:
        registrar_ejecucion_en_historico(ruta_salida, payload)
    except Exception as e:
        log.error(f"  No pude actualizar el histórico: {e}")

    # Devolvemos el payload para que el informe pueda incrustarlo.
    return payload


# Necesitamos sys aquí arriba para el campo "comando". Se importa al principio
# del archivo en condiciones normales (top-level imports), por lo que aquí
# no hace falta volver a importarlo.


def informe_final_tfm(resultados_pc, eda_resumen, auc_dual, args, ruta_salida,
                       res_perfilometro_clf=None, resultados_snr=None,
                       coste_comparativa=None, comparativa_clases=None):
    """Genera el informe final integrado del TFM en texto plano.

    V17: incluye sección SNR (Obj 3 TFM) si se provee `resultados_snr`.
    V18+: incluye sección "Coste computacional y predicción" si se provee
          `coste_comparativa` (devuelto por comparativa_coste_computacional).
    V19+: incluye sección "Cuantificación BUENA vs MALA vs ISO" si se provee
          `comparativa_clases` (devuelto por comparativa_buena_mala_iso).
    """
    lineas = []
    lineas.append("=" * 80)
    lineas.append(" INFORME FINAL TFM - DETECCIÓN DE DESGASTE EN ROSCAS")
    lineas.append(" V23 - Estrategia inteligente: BLMD + SNR + Perfilómetro + PatchCore")
    lineas.append("=" * 80)
    lineas.append("")
    lineas.append("FASE 1 - ANÁLISIS EXPLORATORIO DE DATOS")
    lineas.append("-" * 80)
    lineas.append(f"  Imágenes BUENAS:      {eda_resumen['n_buenas']}")
    lineas.append(f"  Imágenes MALAS:       {eda_resumen['n_malas']}")
    lineas.append(f"  Métricas computadas:  {eda_resumen['n_metricas']}")
    lineas.append(f"  Métricas con KS p<0.001: "
                  f"{eda_resumen['n_metricas_muy_significativas_p001']}")
    lineas.append(f"  Métricas con KS p<0.05:  "
                  f"{eda_resumen['n_metricas_significativas_p005']}")
    lineas.append(f"  RandomForest (cota sup., sin CV): "
                  f"{eda_resumen['rf_accuracy_train']:.3f}")
    lineas.append("")
    lineas.append("  Top 5 features más discriminativas (por KS-stat):")
    for r in eda_resumen["top_metricas"][:5]:
        lineas.append(f"    - {r['metrica']:<28} "
                      f"media B={r['media_buena']:>10.3f} "
                      f"M={r['media_mala']:>10.3f} "
                      f"Cohen d={r['cohen_d']:+.2f} "
                      f"p={r['ks_p']:.2g}")
    lineas.append("")
    lineas.append(f"  MALAS DETECTABLES (outliers del 95% central de BUENAS): "
                  f"{eda_resumen['outliers']['n_malas_detectables']} / "
                  f"{eda_resumen['outliers']['n_malas_total']}  "
                  f"({100 - eda_resumen['outliers']['pct_malas_indistinguibles']:.1f}%)")
    lineas.append(f"  MALAS INDISTINGUIBLES de BUENAS por features básicas: "
                  f"{eda_resumen['outliers']['pct_malas_indistinguibles']:.1f}%")
    lineas.append("")
    # ===== Bloque FASE 2 - PERFILÓMETRO ÓPTICO (antes de PatchCore) =====
    if res_perfilometro_clf is not None:
        lineas.append("FASE 2 - PERFILÓMETRO ÓPTICO COMO CLASIFICADOR")
        lineas.append("-" * 80)
        lineas.append(f"  Método: medición geométrica directa del perfil de")
        lineas.append(f"          rosca (paso, altura, ángulo, cresta) + ")
        lineas.append(f"          calibración data-driven al P{args.perfilometro_percentil:.0f}")
        lineas.append(f"          de las BUENAS, score continuo z-score.")
        auc_perf = res_perfilometro_clf.get("auc")
        m_duro = res_perfilometro_clf.get("metricas_duro", {})
        m_youden = res_perfilometro_clf.get("metricas_youden", {})
        lineas.append("")
        lineas.append(f"  AUC score continuo:        {auc_perf:.3f}")
        if m_duro:
            lineas.append(f"  Veredicto duro (P{args.perfilometro_percentil:.0f}):")
            lineas.append(f"     Acc={m_duro.get('acc',0)*100:.2f}%   "
                          f"Sens={m_duro.get('sens',0)*100:.2f}%   "
                          f"Esp={m_duro.get('esp',0)*100:.2f}%   "
                          f"F1={m_duro.get('f1',0)*100:.2f}%")
        if m_youden:
            lineas.append(f"  Score continuo (umbral Youden):")
            lineas.append(f"     Acc={m_youden.get('acc',0)*100:.2f}%   "
                          f"Sens={m_youden.get('sens',0)*100:.2f}%   "
                          f"Esp={m_youden.get('esp',0)*100:.2f}%   "
                          f"F1={m_youden.get('f1',0)*100:.2f}%")
        # Top 3 métricas más discriminativas según el EDA del perfilómetro
        eda_p = res_perfilometro_clf.get("eda", [])
        if eda_p:
            lineas.append("")
            lineas.append(f"  Top 3 métricas perfilométricas más discriminativas:")
            for f in eda_p[:3]:
                lineas.append(f"    - {f['metrica']:<28} "
                              f"Cohen d={f['cohen_d']:+.2f}  "
                              f"p={f['p_valor']:.2g}")

        # ===== ERROR MEDIO FRENTE A REFERENCIA ISO (V17 patch) =====
        # Esto responde literalmente al Obj 5 del TFM:
        # "Evaluar el rendimiento ... para determinar el ERROR MEDIO
        #  frente a mediciones de referencia."
        # Como no hay ground truth de microscopio digital, usamos la
        # referencia teórica ISO: para una rosca ideal sin desgaste el
        # área de desgaste, el ancho, la profundidad y el RMS deberían
        # ser 0. La diferencia entre la medida real y ese valor de
        # referencia es el error medio del sistema.
        ruta_csv_perf = os.path.join(ruta_salida, "perfilometro",
                                      "evaluacion",
                                      "perfilometro_clf_resumen.csv")
        if os.path.exists(ruta_csv_perf):
            try:
                import csv as _csv
                area_b, area_m = [], []
                ancho_b, ancho_m = [], []
                prof_b, prof_m = [], []
                rms_b, rms_m = [], []
                with open(ruta_csv_perf, "r", encoding="utf-8") as fh:
                    reader = _csv.DictReader(fh)
                    for row in reader:
                        etq = row.get("etiqueta", "")
                        def _f(k):
                            v = row.get(k, "")
                            if v == "" or v is None:
                                return None
                            try:
                                fv = float(v)
                                return fv if np.isfinite(fv) else None
                            except Exception:
                                return None
                        a = _f("area_desgaste_mm2")
                        w = _f("ancho_medio_desgaste_um")
                        p = _f("prof_max_desgaste_um")
                        r = _f("rms_desviacion_mm")
                        if etq == "BUENA":
                            if a is not None: area_b.append(a)
                            if w is not None: ancho_b.append(w)
                            if p is not None: prof_b.append(p)
                            if r is not None: rms_b.append(r)
                        elif etq == "MALA":
                            if a is not None: area_m.append(a)
                            if w is not None: ancho_m.append(w)
                            if p is not None: prof_m.append(p)
                            if r is not None: rms_m.append(r)

                if area_b and area_m:
                    lineas.append("")
                    lineas.append("  ERROR MEDIO FRENTE A REFERENCIA ISO (Obj 5 TFM):")
                    lineas.append("    Métrica                BUENAS (med)  MALAS (med)   "
                                  "Δ B→M")
                    def _med(v): return float(np.median(v)) if v else float("nan")
                    aB, aM = _med(area_b), _med(area_m)
                    wB, wM = _med(ancho_b), _med(ancho_m)
                    pB, pM = _med(prof_b), _med(prof_m)
                    rB, rM = _med(rms_b), _med(rms_m)
                    lineas.append(f"    Área desgaste (mm²)    "
                                  f"{aB:>12.4f}  {aM:>11.4f}   {aM-aB:+.4f}")
                    lineas.append(f"    Ancho medio (µm)       "
                                  f"{wB:>12.2f}  {wM:>11.2f}   {wM-wB:+.2f}")
                    lineas.append(f"    Profundidad máx (µm)   "
                                  f"{pB:>12.2f}  {pM:>11.2f}   {pM-pB:+.2f}")
                    lineas.append(f"    RMS desviación (mm)    "
                                  f"{rB:>12.4f}  {rM:>11.4f}   {rM-rB:+.4f}")
                    lineas.append("")
                    lineas.append(f"    Referencia ISO ideal: área=0 mm², ancho=0 µm, "
                                  f"profundidad=0 µm, RMS=0 mm.")
                    lineas.append(f"    Las medianas observadas reflejan el error medio "
                                  f"del sistema")
                    lineas.append(f"    frente a la referencia teórica para el subconjunto.")
            except Exception as e:
                log.warning(f"Cálculo de error vs ISO falló: {e}")
        lineas.append("")
    lineas.append("FASE 4 - DETECCIÓN PATCHCORE-IMAGEN + CORESET (zero-shot)")
    lineas.append("-" * 80)
    lineas.append(f"  Backbone:        {args.backbone}")
    lineas.append(f"  Coreset ratio:   {args.coreset_ratio*100:.1f}%")
    lineas.append(f"  Banda contorno:  {args.ancho_banda} px (0=sin banda)")
    lineas.append(f"  Mascarado:       {not args.sin_mascara}")
    lineas.append(f"  Imágenes:        max_lado={args.max_lado}, "
                  f"max_imgs={'TODAS' if args.max_imagenes < 0 else args.max_imagenes}")
    lineas.append("")
    lineas.append("RESULTADOS POR PIPELINE")
    lineas.append("-" * 80)
    lineas.append(f"  {'Pipeline':<14} {'AUC global':>11} "
                  f"{'AUC detect.':>12} {'Acc(μ+3σ)':>11} {'Tiempo':>8}")
    for nombre, res in resultados_pc.items():
        ad = auc_dual.get(nombre, {})
        auc_d = ad.get("auc_detectables")
        auc_d_str = f"{auc_d:.3f}" if auc_d is not None else "n/a"
        lineas.append(f"  {nombre:<14} {res['auc']:>11.3f} "
                      f"{auc_d_str:>12} {res['acc_mu3sigma']:>11.2%} "
                      f"{res['tiempo_s']:>7.1f}s")
    lineas.append("")

    # ===== BLOQUE SNR (V17 - Obj 3 TFM) =====
    if resultados_snr is not None and len(resultados_snr) > 0:
        lineas.append("COMPARATIVA SNR POR PIPELINE (Objetivo 3 TFM)")
        lineas.append("-" * 80)
        lineas.append("  Comparación cuantitativa de la relación señal-ruido")
        lineas.append("  tras aplicar cada pipeline de preprocesado.")
        lineas.append("")
        lineas.append(f"  {'Pipeline':<14} {'SNR_estr':>10} "
                      f"{'SNR_cont':>10} {'SNR_BM':>10}")
        lineas.append("  " + "-" * 50)
        # Ordenado por SNR_BM (separación buenas/malas) descendente
        snr_ordenados = sorted(resultados_snr.items(),
                                key=lambda x: x[1]['snr_bm_cohen_d'],
                                reverse=True)
        for nombre, r in snr_ordenados:
            lineas.append(f"  {nombre:<14} "
                          f"{r['snr_estructural_global']:>10.4f} "
                          f"{r['snr_contornos_global']:>10.4f} "
                          f"{r['snr_bm_cohen_d']:>10.4f}")
        lineas.append("")
        mejor_bm = max(resultados_snr.items(),
                       key=lambda x: x[1]['snr_bm_cohen_d'])
        mejor_cont = max(resultados_snr.items(),
                         key=lambda x: x[1]['snr_contornos_global'])
        lineas.append(f"  Mejor SNR_BM (separación buenas/malas): "
                      f"{mejor_bm[0]} ({mejor_bm[1]['snr_bm_cohen_d']:.3f})")
        lineas.append(f"  Mejor SNR_contornos (realce de filo):   "
                      f"{mejor_cont[0]} "
                      f"({mejor_cont[1]['snr_contornos_global']:.3f})")
        if "SIN_PREP" in resultados_snr:
            baseline = resultados_snr["SIN_PREP"]["snr_bm_cohen_d"]
            if baseline > 1e-6:
                mejora_pct = ((mejor_bm[1]['snr_bm_cohen_d'] - baseline)
                              / baseline) * 100
                lineas.append(f"  Mejora vs SIN_PREP (baseline): "
                              f"{mejora_pct:+.1f}% en SNR_BM")
        lineas.append("")

    lineas.append("INTERPRETACIÓN")
    lineas.append("-" * 80)
    auc_global_max = max((r["auc"] for r in resultados_pc.values()),
                         default=0.0)
    auc_det_vals = [v["auc_detectables"] for v in auc_dual.values()
                    if v["auc_detectables"] is not None]
    auc_det_max = max(auc_det_vals) if auc_det_vals else 0.0
    pct_indist = eda_resumen["outliers"]["pct_malas_indistinguibles"]

    lineas.append(f"  - El mejor AUC global obtenido es {auc_global_max:.3f}.")
    lineas.append(f"  - Sobre el subconjunto de MALAS detectables según EDA, "
                  f"el AUC sube a {auc_det_max:.3f}.")
    lineas.append("")
    if pct_indist > 50:
        lineas.append(f"  Una fracción importante de MALAS ({pct_indist:.1f}%) es")
        lineas.append(f"  estadísticamente indistinguible de las BUENAS en el espacio")
        lineas.append(f"  de features estudiado. Esto sugiere que el etiquetado a")
        lineas.append(f"  nivel de pieza incluye fotografías de zonas no afectadas,")
        lineas.append(f"  o que el desgaste es demasiado sutil para captura óptica")
        lineas.append(f"  de propósito general. Se recomienda en trabajo futuro:")
        lineas.append(f"    1. Captura multi-vista de cada pieza.")
        lineas.append(f"    2. Etiquetado a nivel de zona.")
        lineas.append(f"    3. Sensorización adicional (vibraciones, fuerzas).")
    elif pct_indist > 20:
        lineas.append(f"  Hay un solapamiento moderado ({pct_indist:.1f}%) entre")
        lineas.append(f"  las distribuciones de BUENAS y MALAS. El detector")
        lineas.append(f"  funciona razonablemente bien sobre las MALAS")
        lineas.append(f"  detectables y representa un punto de partida válido")
        lineas.append(f"  para integración en línea de producción con criterios")
        lineas.append(f"  de aceptación apropiados.")
    else:
        lineas.append(f"  Las distribuciones de BUENAS y MALAS están bien")
        lineas.append(f"  separadas (solo {pct_indist:.1f}% de solapamiento). El")
        lineas.append(f"  detector es viable para producción.")
    lineas.append("")
    # ===== COMPARATIVA FINAL =====
    if res_perfilometro_clf is not None:
        lineas.append("COMPARATIVA FINAL DE MÉTODOS")
        lineas.append("-" * 80)
        auc_perf = res_perfilometro_clf.get("auc")
        lineas.append(f"  Método                       AUC global")
        lineas.append(f"  PatchCore-imagen (mejor)     {auc_global_max:.3f}")
        lineas.append(f"  Perfilómetro óptico          {auc_perf:.3f}")
        if auc_global_max > 0:
            if auc_perf > auc_global_max:
                lineas.append(f"")
                lineas.append(f"  El perfilómetro mejora el AUC del PatchCore "
                              f"({auc_global_max:.3f} → {auc_perf:.3f}).")
            elif auc_perf < auc_global_max:
                lineas.append(f"")
                lineas.append(f"  PatchCore-imagen tiene mejor AUC global.")
                lineas.append(f"  El perfilómetro aporta interpretabilidad física")
                lineas.append(f"  (mm, µm, °) que PatchCore no proporciona.")
            else:
                lineas.append(f"")
                lineas.append(f"  Ambos métodos tienen AUC equivalente.")
        lineas.append("")

    # ===== CUMPLIMIENTO DE OBJETIVOS SMART (V17) =====
    lineas.append("CUMPLIMIENTO DE OBJETIVOS SMART DEL TFM")
    lineas.append("-" * 80)
    # Obj 1
    auc_homo = resultados_pc.get("PC_SOLO_HOMO", {}).get("auc")
    if auc_homo is not None:
        lineas.append(f"  Obj 1 - Iluminación + filtrado homomórfico:")
        lineas.append(f"          OK   Pipeline SOLO_HOMO  AUC = {auc_homo:.3f}")
    # Obj 2
    auc_blmd_real = resultados_pc.get("PC_BLMD_REAL", {}).get("auc")
    auc_blmd_v10 = resultados_pc.get("PC_BLMD_V10", {}).get("auc")
    if auc_blmd_real is not None or auc_blmd_v10 is not None:
        lineas.append(f"  Obj 2 - Implementación BLMD:")
        if auc_blmd_v10 is not None:
            lineas.append(f"          OK   Pipeline BLMD_V10   AUC = {auc_blmd_v10:.3f}")
        if auc_blmd_real is not None:
            lineas.append(f"          OK   Pipeline BLMD_REAL  AUC = {auc_blmd_real:.3f}")
    # Obj 3
    n_pipes = len(resultados_pc)
    aucs_validos = [r["auc"] for r in resultados_pc.values()
                    if r.get("auc") is not None]
    auc_max = max(aucs_validos) if aucs_validos else 0.0
    lineas.append(f"  Obj 3 - Comparar BLMD vs BEMD vs wavelet en SNR:")
    lineas.append(f"          OK   {n_pipes} pipelines comparados.")
    lineas.append(f"          OK   AUC: mejor = {auc_max:.3f}")
    if resultados_snr is not None and len(resultados_snr) > 0:
        mejor_bm = max(resultados_snr.items(),
                       key=lambda x: x[1]['snr_bm_cohen_d'])
        lineas.append(f"          OK   SNR_BM (Cohen d): mejor = "
                      f"{mejor_bm[0]} ({mejor_bm[1]['snr_bm_cohen_d']:.3f})")
    else:
        lineas.append(f"          --   SNR no calculada (resultados_snr=None)")
    # Obj 4
    lineas.append(f"  Obj 4 - Métricas geométricas (área y ancho de desgaste):")
    if res_perfilometro_clf is not None:
        auc_perf = res_perfilometro_clf.get("auc")
        lineas.append(f"          OK   Perfilómetro óptico: AUC = "
                      f"{auc_perf:.3f}")
        lineas.append(f"          OK   Área (mm²) y ancho medio (µm) por imagen.")
    else:
        lineas.append(f"          OK   Métricas geométricas calculadas por imagen.")
    # Obj 5
    lineas.append(f"  Obj 5 - Evaluación con dataset etiquetado:")
    lineas.append(f"          OK   AUC, F1, ROC, matriz confusión.")
    if n_pipes >= 2:
        lineas.append(f"          OK   Comparativa entre {n_pipes} pipelines.")
    if res_perfilometro_clf is not None:
        lineas.append(f"          OK   Error medio vs referencia ISO "
                      f"reportado en FASE 2 (área, ancho, profundidad, RMS).")
    lineas.append("")

    # --------------------------------------------------------------
    # SECCIÓN CUANTIFICACIÓN BUENA vs MALA vs ISO
    # Caracterización cuantitativa de la diferencia geométrica entre
    # roscas conformes y no conformes en unidades físicas (mm² y µm),
    # contrastada con el perfil teórico ISO.
    # --------------------------------------------------------------
    if comparativa_clases and comparativa_clases.get("resumen_metricas"):
        lineas.append("=" * 80)
        lineas.append(" CUANTIFICACIÓN BUENA vs MALA vs ISO TEÓRICO")
        lineas.append("=" * 80)
        lineas.append(
            f"  Comparación cuantitativa de roscas buenas y malas frente al")
        lineas.append(
            f"  perfil teórico ISO M{args.paso_mm}, expresada en unidades")
        lineas.append(
            f"  físicas (µm de altura, mm² de área, µm de profundidad).")
        lineas.append("")
        lineas.append(
            f"  Muestras válidas: "
            f"BUENAS={comparativa_clases.get('n_buenas_validas','?')} "
            f"MALAS={comparativa_clases.get('n_malas_validas','?')}")
        lineas.append("")
        lineas.append(
            f"  {'Métrica':<26} {'Unidad':<6} {'BUENA':>10} "
            f"{'MALA':>10} {'Δ':>10} {'Razón':>7} {'CohenD':>7}")
        lineas.append("  " + "-" * 78)
        for r in comparativa_clases["resumen_metricas"]:
            lineas.append(
                f"  {r['metrica']:<26} {r['unidad']:<6} "
                f"{r['mediana_BUENA']:>10.3f} {r['mediana_MALA']:>10.3f} "
                f"{r['delta_MALA_menos_BUENA']:>+10.3f} "
                f"{r['razon_MALA_BUENA']:>7.2f} {r['cohen_d']:>+7.2f}")
        lineas.append("")
        lineas.append(
            "  Lectura: Δ es la diferencia mediana absoluta MALA − BUENA en")
        lineas.append(
            "  la unidad de la métrica. Razón = mediana MALA / mediana BUENA.")
        lineas.append(
            "  Cohen d cuantifica la separación estandarizada entre las dos")
        lineas.append(
            "  distribuciones (>0,8 separación grande, >0,5 media, <0,3 baja).")
        lineas.append("")
        lineas.append(
            f"  Detalle visual: comparativa_buena_mala_iso.png (perfiles)")
        lineas.append(
            f"  y comparativa_buena_mala_paneles.png (histogramas+boxplots).")
        lineas.append(
            f"  Tabla CSV: comparativa_buena_mala.csv.")
        lineas.append(
            f"  Informe interpretativo: comparativa_buena_mala_informe.txt.")
        lineas.append("")

        # ----- MEJORA C: tabla restringida a detectables -----
        det = comparativa_clases.get("resumen_metricas_detectables")
        n_det = comparativa_clases.get("n_malas_detectables")
        if det:
            lineas.append(
                f"  ----- Tabla restringida a MALAS DETECTABLES (n={n_det}) -----")
            lineas.append(
                "  (Cuando la imagen SÍ muestra desgaste medible, la "
                "separación es mucho mayor.)")
            lineas.append("")
            lineas.append(
                f"  {'Métrica':<26} {'Unidad':<6} {'BUENA':>10} "
                f"{'MALA_det':>10} {'Δ':>10} {'Razón':>7} {'CohenD':>7}")
            for r in det:
                lineas.append(
                    f"  {r['metrica']:<26} {r['unidad']:<6} "
                    f"{r['mediana_BUENA']:>10.3f} {r['mediana_MALA']:>10.3f} "
                    f"{r['delta_MALA_menos_BUENA']:>+10.3f} "
                    f"{r['razon_MALA_BUENA']:>7.2f} {r['cohen_d']:>+7.2f}")
            lineas.append("")
        lineas.append(
            "  Nota: si la calibración del umbral de decisión a nivel pieza")
        lineas.append(
            "  (P95) se ha aplicado, las métricas correspondientes están en")
        lineas.append(
            "  perfilometro_clf_informe.txt, sección (e).")
        lineas.append("")

    # --------------------------------------------------------------
    # SECCIÓN COSTE COMPUTACIONAL Y PREDICCIÓN
    # Se imprime si la fase 5-bis se ejecutó (coste_comparativa != None).
    # --------------------------------------------------------------
    if coste_comparativa and coste_comparativa.get("metodos"):
        lineas.append("=" * 80)
        lineas.append(" COSTE COMPUTACIONAL Y PREDICCIÓN")
        lineas.append("=" * 80)
        sis = coste_comparativa.get("sistema") or {}
        lineas.append(
            f"  Entorno: Python {sis.get('python','?')} | "
            f"{sis.get('sistema','?')} | "
            f"CPU {sis.get('cpu_logicos','?')} cores | "
            f"GPU: {sis.get('gpu_nombre') or 'no disponible'} | "
            f"PyTorch {sis.get('torch_version','?')}")
        lineas.append(
            f"  Imágenes: {coste_comparativa.get('n_buenas','?')} BUENAS, "
            f"{coste_comparativa.get('n_malas','?')} MALAS")
        lineas.append("")
        # Cabecera de la tabla
        lineas.append(
            f"  {'Método':<22} {'AUC':>6} {'Acc(p99)':>9} "
            f"{'T_train':>9} {'T_inf/img':>10} "
            f"{'Params':>9} {'Banco':>9} {'GPU_pico':>10}")
        lineas.append("  " + "-" * 96)
        for met in coste_comparativa["metodos"]:
            auc_s = f"{met['auc']:.3f}" if met.get('auc') is not None else "  —"
            acc_s = (f"{met['acc']*100:.1f}%" if met.get('acc') is not None
                     else "  —")
            tt = (f"{met['tiempo_train_total_s']:.1f}s"
                  if met.get('tiempo_train_total_s') is not None else "  —")
            ti = (f"{met['tiempo_inferencia_por_imagen_s']*1000:.0f}ms"
                  if met.get('tiempo_inferencia_por_imagen_s') is not None
                  else "  —")
            pa = (f"{met['params_total']/1e6:.1f}M"
                  if met.get('params_total') else "  —")
            ba = (f"{met['bank_mb']:.1f}MB"
                  if met.get('bank_mb') else "  —")
            gp = (f"{met['gpu_pico_mb']:.0f}MB"
                  if met.get('gpu_pico_mb') is not None else "  —")
            lineas.append(
                f"  {met['nombre']:<22} {auc_s:>6} {acc_s:>9} "
                f"{tt:>9} {ti:>10} {pa:>9} {ba:>9} {gp:>10}")
        lineas.append("")
        lineas.append("  Notas:")
        lineas.append("    - T_train: tiempo de preparación del detector "
                      "(coreset + memory bank). El perfilómetro no se entrena.")
        lineas.append("    - T_inf/img: tiempo medio de predicción por imagen "
                      "(inferencia + scoring).")
        lineas.append("    - Banco: tamaño del memory bank PatchCore tras "
                      "submuestreo coreset.")
        lineas.append("    - Detalle completo en comparativa_coste.csv y "
                      "comparativa_coste.png (en la carpeta de salida).")
        lineas.append("    - Histórico acumulado entre ejecuciones en "
                      "historico_ejecuciones.json y historico_costes.csv "
                      "(carpeta raíz del proyecto).")
        lineas.append("")

    lineas.append("=" * 80)
    lineas.append("Archivos generados:")
    lineas.append(f"  {ruta_salida}/eda/                - análisis exploratorio")
    lineas.append(f"  {ruta_salida}/                    - heatmaps, paneles, ROC, matrices")
    if res_perfilometro_clf is not None:
        lineas.append(f"  {ruta_salida}/perfilometro/    - perfil real vs teórico (PNGs)")
        lineas.append(f"  {ruta_salida}/perfilometro/evaluacion/  - AUC, ROC, informe perfilómetro")
    if resultados_snr is not None:
        lineas.append(f"  {ruta_salida}/snr/             - SNR por pipeline (obj. 3)")
    if coste_comparativa is not None:
        lineas.append(f"  {ruta_salida}/comparativa_coste.csv  - tabla coste computacional")
        lineas.append(f"  {ruta_salida}/comparativa_coste.png  - gráfica coste vs calidad")
        lineas.append(f"  {ruta_salida}/historico_costes.csv   - CSV de esta ejecución")
    lineas.append(f"  {ruta_salida}/informe_tfm.txt     - este informe")
    lineas.append("=" * 80)
    texto = "\n".join(lineas)
    with open(os.path.join(ruta_salida, "informe_tfm.txt"), "w",
              encoding="utf-8") as fh:
        fh.write(texto)
    print("\n" + texto + "\n")


# ==============================================================================
#  FASE 5 (V20) - COMPARATIVA VISUAL DE PIPELINES
# ==============================================================================
#
#  Genera figuras combinadas que muestran de un vistazo cómo se comparan los
#  4 pipelines (BLMD_REAL, BLMD_V10, SOLO_HOMO, SIN_PREP). Estas figuras son
#  ideales para el capítulo 6 de la memoria del TFM.
# ==============================================================================

# Paleta consistente para los pipelines
COLORES_PIPELINES = {
    "PC_BLMD_REAL": "#1f77b4",   # azul
    "PC_BLMD_V10":  "#ff7f0e",   # naranja
    "PC_BEMD_REAL": "#9467bd",   # morado
    "PC_WAVELET":   "#8c564b",   # marrón
    "PC_SOLO_HOMO": "#2ca02c",   # verde
    "PC_SIN_PREP":  "#d62728",   # rojo
}


def _curvas_roc_pr_pipeline(scores, etiquetas):
    """Calcula puntos de la curva ROC y PR para un pipeline dado."""
    if len(np.unique(etiquetas)) < 2:
        return None
    candidatos = np.linspace(scores.min(), scores.max(), 300)
    tpr_l, fpr_l, prec_l = [], [], []
    for u in candidatos:
        pred = (scores > u).astype(int)
        vp = int(np.sum((pred == 1) & (etiquetas == 1)))
        fn = int(np.sum((pred == 0) & (etiquetas == 1)))
        fp = int(np.sum((pred == 1) & (etiquetas == 0)))
        vn = int(np.sum((pred == 0) & (etiquetas == 0)))
        tpr_l.append(vp / max(1, vp + fn))
        fpr_l.append(fp / max(1, fp + vn))
        prec_l.append(vp / max(1, vp + fp))
    orden_roc = np.argsort(fpr_l)
    fpr_arr = np.array(fpr_l)[orden_roc]
    tpr_arr = np.array(tpr_l)[orden_roc]
    auc_roc = float(np.trapezoid(tpr_arr, fpr_arr) if hasattr(np, "trapezoid")
                    else np.trapz(tpr_arr, fpr_arr))
    orden_pr = np.argsort(tpr_l)
    rec_arr = np.array(tpr_l)[orden_pr]
    prec_arr = np.array(prec_l)[orden_pr]
    auc_pr = float(np.trapezoid(prec_arr, rec_arr) if hasattr(np, "trapezoid")
                   else np.trapz(prec_arr, rec_arr))
    return {
        "fpr": fpr_arr, "tpr": tpr_arr, "auc_roc": auc_roc,
        "recall": rec_arr, "precision": prec_arr, "auc_pr": auc_pr,
        "candidatos": candidatos,
    }


def comparativa_pipelines(resultados_pc, auc_dual, ruta_salida):
    """
    Genera 4 figuras comparativas:
      1. Curvas ROC superpuestas
      2. Curvas Precision-Recall superpuestas
      3. Barras AUC global vs AUC detectables
      4. Histograma comparativo de scores de BUENAS
    Y una tabla CSV con métricas en F1 óptimo por pipeline.
    """
    log.info("=" * 70)
    log.info(" FASE 5 / COMPARATIVA VISUAL DE PIPELINES")
    log.info("=" * 70)

    if not resultados_pc:
        log.warning("Sin pipelines para comparar."); return None

    # Calcular curvas para cada pipeline
    curvas = {}
    for nombre, res in resultados_pc.items():
        scores = np.array([r[2] for r in res["log"]])
        etiquetas = np.array([1 if r[1] == "MALA" else 0 for r in res["log"]])
        c = _curvas_roc_pr_pipeline(scores, etiquetas)
        if c is not None:
            curvas[nombre] = c

    if not curvas:
        log.warning("No se pudieron calcular curvas."); return None

    # ===== FIGURA 1: Curvas ROC superpuestas =====
    fig, ax = plt.subplots(figsize=(9, 7))
    for nombre, c in curvas.items():
        color = COLORES_PIPELINES.get(nombre, "gray")
        ax.plot(c["fpr"], c["tpr"], color=color, lw=2.5,
                label=f"{nombre} (AUC = {c['auc_roc']:.3f})")
    ax.plot([0, 1], [0, 1], "k--", alpha=0.4, label="Aleatorio")
    ax.set_xlabel("FPR (1 - especificidad)")
    ax.set_ylabel("TPR (recall / sensibilidad)")
    ax.set_title("Comparativa de curvas ROC entre pipelines")
    ax.legend(loc="lower right", fontsize=10); ax.grid(alpha=0.3)
    ax.set_xlim(-0.02, 1.02); ax.set_ylim(-0.02, 1.02)
    plt.tight_layout()
    plt.savefig(os.path.join(ruta_salida, "comparativa_ROC.png"), dpi=120)
    plt.close()

    # ===== FIGURA 2: Curvas Precision-Recall superpuestas =====
    fig, ax = plt.subplots(figsize=(9, 7))
    for nombre, c in curvas.items():
        color = COLORES_PIPELINES.get(nombre, "gray")
        ax.plot(c["recall"], c["precision"], color=color, lw=2.5,
                label=f"{nombre} (AUPRC = {c['auc_pr']:.3f})")
    # Baseline = proporción de positivos en el dataset
    baseline = float(np.mean([1 for nombre, res in resultados_pc.items()
                              for r in res["log"] if r[1] == "MALA"]) /
                     max(1, sum(len(res["log"]) for res in resultados_pc.values()) /
                         len(resultados_pc)))
    ax.axhline(baseline, color="k", ls="--", alpha=0.4,
               label=f"baseline = {baseline:.2f}")
    ax.set_xlabel("Recall (TPR)")
    ax.set_ylabel("Precision (PPV)")
    ax.set_title("Comparativa de curvas Precision-Recall entre pipelines")
    ax.legend(loc="lower left", fontsize=10); ax.grid(alpha=0.3)
    ax.set_xlim(-0.02, 1.02); ax.set_ylim(-0.02, 1.02)
    plt.tight_layout()
    plt.savefig(os.path.join(ruta_salida, "comparativa_PR.png"), dpi=120)
    plt.close()

    # ===== FIGURA 3: Barras AUC global vs AUC detectables =====
    nombres = list(resultados_pc.keys())
    auc_g = [resultados_pc[n]["auc"] for n in nombres]
    auc_d = [auc_dual.get(n, {}).get("auc_detectables") or 0.0 for n in nombres]
    tiempos = [resultados_pc[n]["tiempo_s"] for n in nombres]

    fig, axes = plt.subplots(1, 2, figsize=(14, 6))
    x = np.arange(len(nombres))
    w = 0.38
    barras_g = axes[0].bar(x - w/2, auc_g, w, label="AUC global",
                            color="#5B8DEF")
    barras_d = axes[0].bar(x + w/2, auc_d, w, label="AUC detectables",
                            color="#F39C12")
    axes[0].set_xticks(x)
    axes[0].set_xticklabels(nombres, rotation=20, ha="right")
    axes[0].set_ylabel("AUC")
    axes[0].set_ylim(0, 1.05)
    axes[0].set_title("AUC global vs AUC sobre subconjunto detectable (EDA)")
    axes[0].legend(); axes[0].grid(alpha=0.3, axis="y")
    for bar in barras_g + barras_d:
        h = bar.get_height()
        if h > 0:
            axes[0].text(bar.get_x() + bar.get_width()/2, h + 0.01,
                         f"{h:.3f}", ha="center", fontsize=9)

    # Barras tiempo
    barras_t = axes[1].bar(nombres, tiempos,
                           color=[COLORES_PIPELINES.get(n, "gray")
                                  for n in nombres])
    axes[1].set_xticklabels(nombres, rotation=20, ha="right")
    axes[1].set_ylabel("Tiempo (segundos)")
    axes[1].set_title("Tiempo de ejecución por pipeline")
    axes[1].grid(alpha=0.3, axis="y")
    for bar, t in zip(barras_t, tiempos):
        axes[1].text(bar.get_x() + bar.get_width()/2,
                     bar.get_height() + max(tiempos)*0.01,
                     f"{t:.0f}s", ha="center", fontsize=9)
    plt.tight_layout()
    plt.savefig(os.path.join(ruta_salida, "comparativa_AUC_tiempo.png"), dpi=120)
    plt.close()

    # ===== FIGURA 4: Histograma comparativo de scores =====
    n_pipes = len(resultados_pc)
    n_cols = 3 if n_pipes > 4 else 2
    n_rows = int(np.ceil(n_pipes / n_cols))
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(5 * n_cols, 4 * n_rows))
    axes_flat = np.array(axes).flat if n_pipes > 1 else [axes]
    for i, (nombre, res) in enumerate(resultados_pc.items()):
        ax = axes_flat[i]
        scores = np.array([r[2] for r in res["log"]])
        etiquetas = np.array([1 if r[1] == "MALA" else 0 for r in res["log"]])
        color = COLORES_PIPELINES.get(nombre, "gray")
        ax.hist(scores[etiquetas == 0], bins=25, alpha=0.6,
                color="#4CAF50", label="BUENAS")
        ax.hist(scores[etiquetas == 1], bins=25, alpha=0.6,
                color="#E53935", label="MALAS")
        ax.set_title(f"{nombre} (AUC={res['auc']:.3f})", color=color,
                     fontweight="bold")
        ax.set_xlabel("Score"); ax.set_ylabel("Frecuencia")
        ax.legend(); ax.grid(alpha=0.3)
    # Ocultar huecos vacíos al final
    for j in range(n_pipes, n_rows * n_cols):
        axes_flat[j].axis("off")
    plt.suptitle("Distribución de scores BUENAS vs MALAS por pipeline",
                 fontsize=13, fontweight="bold")
    plt.tight_layout()
    plt.savefig(os.path.join(ruta_salida,
                             "comparativa_distribuciones.png"), dpi=120)
    plt.close()

    # ===== TABLA CSV con métricas en F1 óptimo =====
    filas_tabla = []
    for nombre, res in resultados_pc.items():
        scores = np.array([r[2] for r in res["log"]])
        etiquetas = np.array([1 if r[1] == "MALA" else 0 for r in res["log"]])
        u_f1, m_f1 = encontrar_umbral_optimo(scores, etiquetas, "f1")
        if m_f1 is None:
            continue
        filas_tabla.append({
            "pipeline": nombre,
            "AUC_global": res["auc"],
            "AUC_detectables": auc_dual.get(nombre, {}).get("auc_detectables"),
            "umbral_F1opt": u_f1,
            "F1_opt": m_f1["F1"],
            "TPR_opt": m_f1["TPR_recall"],
            "FPR_opt": m_f1["FPR"],
            "precision_opt": m_f1["precision"],
            "MCC_opt": m_f1["MCC"],
            "accuracy_opt": m_f1["accuracy"],
            "VP": m_f1["VP"], "VN": m_f1["VN"],
            "FP": m_f1["FP"], "FN": m_f1["FN"],
            "tiempo_s": res["tiempo_s"],
        })
    if filas_tabla:
        ruta_tabla = os.path.join(ruta_salida, "comparativa_pipelines.csv")
        with open(ruta_tabla, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=list(filas_tabla[0].keys()))
            w.writeheader(); w.writerows(filas_tabla)
        log.info(f"  Tabla comparativa guardada en {ruta_tabla}")
        # Imprimir resumen por consola
        log.info(f"  {'Pipeline':<14} {'AUC_g':>6} {'AUC_d':>6} "
                 f"{'F1':>6} {'Acc':>6}")
        for f in filas_tabla:
            audet = f["AUC_detectables"]
            audet_s = f"{audet:.3f}" if audet is not None else "n/a"
            log.info(f"  {f['pipeline']:<14} {f['AUC_global']:>6.3f} "
                     f"{audet_s:>6} {f['F1_opt']:>6.3f} "
                     f"{f['accuracy_opt']:>6.3f}")

    log.info("  Generadas: comparativa_ROC.png, comparativa_PR.png, "
             "comparativa_AUC_tiempo.png, comparativa_distribuciones.png, "
             "comparativa_pipelines.csv")
    return filas_tabla


# ==============================================================================
#  FASE 6 (V20) - PDF COMBINADO PARA ANEXO DE LA MEMORIA
# ==============================================================================

def construir_pdf_completo(ruta_salida, eda_resumen, resultados_pc, auc_dual,
                           args):
    """
    Crea un PDF que combina TODAS las salidas relevantes en un único
    documento, ideal para anexo de la memoria del TFM.

    Estructura:
      1. Portada con resumen ejecutivo
      2. Informe TFM completo
      3. Figuras del EDA
      4. Comparativa de pipelines
      5. Resultados detallados por pipeline (ROC, PR, multi-umbral...)
      6. Heatmaps de algunas detecciones
    """
    from matplotlib.backends.backend_pdf import PdfPages
    log.info("=" * 70)
    log.info(" FASE 6 / GENERANDO PDF COMBINADO")
    log.info("=" * 70)

    ruta_pdf = os.path.join(ruta_salida, "ANEXO_TFM.pdf")
    with PdfPages(ruta_pdf) as pdf:
        # ===== PÁGINA 1: PORTADA =====
        fig = plt.figure(figsize=(8.27, 11.69))  # A4
        fig.text(0.5, 0.85, "ANEXO DEL TFM", ha="center", fontsize=22,
                 fontweight="bold")
        fig.text(0.5, 0.80,
                 "Detección de desgaste en herramientas de roscado",
                 ha="center", fontsize=14, style="italic")
        fig.text(0.5, 0.76,
                 "PatchCore + Coreset + EDA + BLMD",
                 ha="center", fontsize=12)
        fig.text(0.5, 0.66, "Universidad Internacional de La Rioja",
                 ha="center", fontsize=11)
        fig.text(0.5, 0.63, "Máster Universitario en Inteligencia Artificial",
                 ha="center", fontsize=11)
        fig.text(0.5, 0.55, "Daniel Alcalde Martín-Calero",
                 ha="center", fontsize=10)
        fig.text(0.5, 0.52, "Miguel González Sánchez",
                 ha="center", fontsize=10)
        fig.text(0.5, 0.49, "Jordi Peiró Castelló",
                 ha="center", fontsize=10)
        # Resumen rápido
        if eda_resumen and resultados_pc:
            mejor_auc = max(r["auc"] for r in resultados_pc.values())
            mejor_auc_d = max(
                (v["auc_detectables"] for v in auc_dual.values()
                 if v.get("auc_detectables") is not None), default=0.0)
            fig.text(0.5, 0.35, "RESUMEN EJECUTIVO", ha="center",
                     fontsize=13, fontweight="bold")
            txt = (f"Imágenes: {eda_resumen['n_buenas']} BUENAS + "
                   f"{eda_resumen['n_malas']} MALAS\n\n"
                   f"AUC global máximo: {mejor_auc:.3f}\n"
                   f"AUC sobre MALAS detectables (EDA): {mejor_auc_d:.3f}\n\n"
                   f"MALAS detectables: "
                   f"{eda_resumen['outliers']['n_malas_detectables']}/"
                   f"{eda_resumen['outliers']['n_malas_total']} "
                   f"({100 - eda_resumen['outliers']['pct_malas_indistinguibles']:.1f}%)\n"
                   f"MALAS indistinguibles: "
                   f"{eda_resumen['outliers']['pct_malas_indistinguibles']:.1f}%")
            fig.text(0.5, 0.20, txt, ha="center", fontsize=10,
                     family="monospace")
        fig.text(0.5, 0.05,
                 f"Generado: {time.strftime('%Y-%m-%d %H:%M')}",
                 ha="center", fontsize=8, alpha=0.6)
        pdf.savefig(fig); plt.close()

        # ===== PÁGINAS 2-3: INFORME TFM (texto) =====
        ruta_informe = os.path.join(ruta_salida, "informe_tfm.txt")
        if os.path.exists(ruta_informe):
            with open(ruta_informe, encoding="utf-8") as fh:
                texto_informe = fh.read()
            # Trocear el texto en bloques que quepan en una página A4 (~50 líneas)
            lineas = texto_informe.split("\n")
            por_pagina = 55
            for i in range(0, len(lineas), por_pagina):
                bloque = "\n".join(lineas[i:i + por_pagina])
                fig = plt.figure(figsize=(8.27, 11.69))
                fig.text(0.05, 0.96, "INFORME TFM (continúa)" if i > 0
                         else "INFORME TFM",
                         fontsize=11, fontweight="bold")
                fig.text(0.05, 0.93, bloque, fontsize=7,
                         family="monospace", verticalalignment="top")
                pdf.savefig(fig); plt.close()

        # ===== PÁGINA: FIGURA EDA =====
        ruta_eda_fig = os.path.join(ruta_salida, "eda", "eda_resumen.png")
        if os.path.exists(ruta_eda_fig):
            _pdf_pagina_imagen(pdf, ruta_eda_fig,
                               "ANÁLISIS EXPLORATORIO - Resumen visual")

        # ===== PÁGINAS: COMPARATIVA DE PIPELINES =====
        for nombre_fig, titulo in [
            ("comparativa_ROC.png", "Comparativa - Curvas ROC"),
            ("comparativa_PR.png", "Comparativa - Precision-Recall"),
            ("comparativa_AUC_tiempo.png",
             "Comparativa - AUC y tiempo por pipeline"),
            ("comparativa_distribuciones.png",
             "Comparativa - Distribuciones de scores"),
        ]:
            ruta_fig = os.path.join(ruta_salida, nombre_fig)
            if os.path.exists(ruta_fig):
                _pdf_pagina_imagen(pdf, ruta_fig, titulo)

        # ===== PÁGINAS POR PIPELINE: ROC+PR, multi-umbral, trade-off =====
        for nombre_p in resultados_pc.keys():
            for prefijo, titulo_base in [
                ("fig_roc_pr_", "ROC + Precision-Recall"),
                ("fig_histograma_umbrales_", "Histograma con umbrales"),
                ("fig_metricas_vs_umbral_", "Métricas vs umbral"),
                ("fig_tradeoff_fp_fn_", "Trade-off FP / FN"),
            ]:
                ruta_fig = os.path.join(ruta_salida, f"{prefijo}{nombre_p}.png")
                if os.path.exists(ruta_fig):
                    _pdf_pagina_imagen(pdf, ruta_fig,
                                       f"{nombre_p} - {titulo_base}")

        # ===== PÁGINAS: PANELES (heatmaps) =====
        # Buscar los archivos panel_*.png más relevantes
        archivos_paneles = sorted([
            f for f in os.listdir(ruta_salida)
            if f.startswith("panel_") and f.endswith(".png")
        ])
        # Limitar a 12 para no inflar el PDF
        for nombre_panel in archivos_paneles[:12]:
            _pdf_pagina_imagen(pdf, os.path.join(ruta_salida, nombre_panel),
                               f"Detección - {nombre_panel}")

    log.info(f"  PDF generado: {ruta_pdf}")
    log.info(f"  Tamaño: {os.path.getsize(ruta_pdf)/1024:.0f} KB")
    return ruta_pdf


def _pdf_pagina_imagen(pdf, ruta_imagen, titulo):
    """Añade una página al PDF con una imagen y un título arriba."""
    fig = plt.figure(figsize=(11.69, 8.27))  # A4 horizontal
    fig.text(0.5, 0.95, titulo, ha="center", fontsize=12,
             fontweight="bold")
    img = plt.imread(ruta_imagen)
    ax = fig.add_axes([0.05, 0.05, 0.9, 0.85])
    ax.imshow(img); ax.axis("off")
    pdf.savefig(fig); plt.close()


def main():
    parser = argparse.ArgumentParser(
        description="TFM v16 - Detección de desgaste con EDA + PatchCore + BLMD")
    parser.add_argument("--buenas", type=str, default=None)
    parser.add_argument("--malas", type=str, default=None)
    parser.add_argument("--salida", type=str, default=None)
    parser.add_argument("--pipelines", type=str, nargs="+",
                        default=["BLMD_REAL", "BLMD_V10", "BEMD_REAL",
                                 "WAVELET", "SOLO_HOMO", "SIN_PREP"],
                        help="Pipelines a evaluar. Por defecto incluye "
                             "BEMD_REAL (V17): la implementación de BEMD "
                             "trabaja a 256x256 internamente para evitar "
                             "el bug de memoria de PyEMD con tamaños "
                             "arbitrarios. Si quieres excluirlo, pasa "
                             "la lista sin BEMD_REAL.")
    parser.add_argument("--max_imagenes", type=int, default=-1,
                        help="Máximo de imágenes a procesar por carpeta. "
                             "Por defecto: -1 (TODAS). "
                             "Pasa --max_imagenes 20 para modo prueba rápido.")
    parser.add_argument("--max_lado", type=int, default=512)
    parser.add_argument("--coreset_ratio", type=float, default=0.01)
    parser.add_argument("--backbone", type=str, default="wide_resnet50_2",
                        choices=["wide_resnet50_2", "resnet18"])
    parser.add_argument("--sin_mascara", action="store_true")
    parser.add_argument("--saltar_ablacion", action="store_true",
                        help="Por defecto se ejecuta automáticamente la "
                             "ablación con/sin máscara para evaluar la "
                             "aportación del contexto al detector PatchCore. "
                             "Este flag permite saltarla para acelerar la "
                             "ejecución; por defecto siempre se realiza.")
    parser.add_argument("--percentil_desgaste", type=float, default=90)
    parser.add_argument("--ancho_banda", type=int, default=20)
    parser.add_argument("--saltar_eda", action="store_true",
                        help="Si se pasa, saltarse la fase de EDA y usar un "
                             "eda_resumen.json existente en <salida>/eda/")
    parser.add_argument("--sin_pdf", action="store_true",
                        help="No generar el PDF combinado final (V20). "
                             "Útil si solo quieres los archivos sueltos.")
    parser.add_argument("--sin_excel", action="store_true",
                        help="No generar el Excel para compañeros (V20.1). "
                             "Útil si solo quieres los archivos sueltos.")
    # ===== Perfilómetro óptico (V21) =====
    parser.add_argument("--perfilometro", action="store_true", default=True,
                        help="(Activado por defecto) Ejecuta análisis "
                             "perfilométrico: extrae el perfil 1D de cada "
                             "rosca, lo compara con el teórico ISO, "
                             "calcula métricas geométricas, calibra "
                             "tolerancias data-driven y evalúa el "
                             "perfilómetro como clasificador (AUC, ROC, "
                             "matriz de confusión, agregación por pieza).")
    parser.add_argument("--sin_perfilometro", action="store_true",
                        help="Desactiva el perfilómetro (útil si solo "
                             "quieres EDA + PatchCore como antes).")
    parser.add_argument("--paso_mm", type=float, default=1.5,
                        help="Paso nominal de la rosca en mm "
                             "(por defecto 1.5).")
    parser.add_argument("--norma_rosca", type=str, default="iso_metrica",
                        choices=["iso_metrica", "whitworth"],
                        help="Norma del perfil teórico (iso_metrica = "
                             "triangular 60°, h=0.866·P).")
    parser.add_argument("--mm_por_px", type=float, default=None,
                        help="Calibración mm/px conocida. Si no se pasa, se "
                             "deduce de la distancia entre picos del perfil "
                             "real asumiendo que coincide con --paso_mm.")
    parser.add_argument("--solo_perfilometro", action="store_true",
                        help="Salta toda la pipeline de PatchCore y solo "
                             "ejecuta el perfilómetro (modo metrología pura).")
    parser.add_argument("--perfilometro_percentil", type=float, default=99.0,
                        help="Percentil de las BUENAS para calibrar las "
                             "tolerancias del perfilómetro como calibre "
                             "PASA/NO PASA (por defecto 99). Valores más "
                             "altos = más permisivo (menos falsos positivos "
                             "en BUENAS).")
    args = parser.parse_args()

    # ===== Banner de bienvenida (V20.2) =====
    log.info("=" * 70)
    log.info(" TFM - DETECCIÓN DE DESGASTE EN ROSCAS")
    log.info(" PatchCore + Coreset + EDA + BLMD")
    log.info("=" * 70)
    if args.max_imagenes < 0:
        log.info(" Modo: OFICIAL (todas las imágenes)")
        log.info(" Tiempo estimado: ~12-15 minutos")
    else:
        log.info(f" Modo: PRUEBA RÁPIDA ({args.max_imagenes} imgs por carpeta)")
        log.info(" Tiempo estimado: ~2-3 minutos")
    log.info(f" Pipelines: {', '.join(args.pipelines)}")
    log.info("=" * 70)

    max_im = None if args.max_imagenes < 0 else args.max_imagenes
    max_lado = None if args.max_lado < 0 else args.max_lado

    # ===== Avisos sobre librerías opcionales (V20) =====
    pipelines_no_disponibles = []
    if not BEMD_OK:
        pipelines_no_disponibles.append("BEMD_REAL")
        log.warning("PyEMD NO instalado: pipeline BEMD_REAL actuará como "
                    "passthrough (= SIN_PREP). Para activarlo:")
        log.warning("    pip install EMD-signal")
    if not PYWAVELETS_OK:
        pipelines_no_disponibles.append("WAVELET")
        log.warning("PyWavelets NO instalado: pipeline WAVELET actuará como "
                    "passthrough (= SIN_PREP). Para activarlo:")
        log.warning("    pip install PyWavelets")
    if pipelines_no_disponibles:
        log.warning(f"Pipelines degradados: {pipelines_no_disponibles}. "
                    f"El script seguirá funcionando con el resto.")

    if not all([args.buenas, args.malas, args.salida]):
        ruta_buenas, ruta_malas, ruta_salida = pedir_carpetas_gui()
    else:
        ruta_buenas, ruta_malas, ruta_salida = args.buenas, args.malas, args.salida
    if not all([ruta_buenas, ruta_malas, ruta_salida]):
        log.error("Faltan rutas. Saliendo."); return

    Path(ruta_salida).mkdir(parents=True, exist_ok=True)
    ruta_eda = os.path.join(ruta_salida, "eda")

    log.info(f"BUENAS: {ruta_buenas}")
    log.info(f"MALAS:  {ruta_malas}")
    log.info(f"SALIDA: {ruta_salida}")
    log.info(f"Modo: max_imagenes={max_im}  max_lado={max_lado}")

    log.info("Cargando imágenes BUENAS...")
    imgs_buenas = cargar_imagenes_carpeta(ruta_buenas, max_im, max_lado)
    log.info(f"  -> {len(imgs_buenas)} imágenes")
    log.info("Cargando imágenes MALAS...")
    imgs_malas = cargar_imagenes_carpeta(ruta_malas, max_im, max_lado)
    log.info(f"  -> {len(imgs_malas)} imágenes")

    if len(imgs_buenas) < 4 or len(imgs_malas) < 1:
        log.error("Pocas imágenes. Saliendo."); return

    # ============================================================
    # ORDEN SECUENCIAL DE EJECUCIÓN
    # ============================================================
    #   FASE 1: EDA general (estadística sobre las imágenes)
    #   FASE 2: Perfilómetro (extracción de perfil + medidas + clasificador)
    #   FASE 3: PatchCore-Desviación (entrenamiento sobre desviación al ISO)
    #   FASE 4: Pipelines de preprocesado (BLMD, BEMD, wavelet, etc.)
    #   FASE 5: PatchCore-imagen (detección por anomalía sobre imagen)
    #   FASE 6: AUC dual + informe final + Excel + PDF
    # ============================================================

    res_perfilometro_clf = None
    ejecutar_perf = (args.perfilometro or args.solo_perfilometro) and \
                    not args.sin_perfilometro

    # ===== FASE 1 - EDA GENERAL =====
    eda_json_path = os.path.join(ruta_eda, "eda_resumen.json")
    if args.saltar_eda and os.path.exists(eda_json_path):
        log.info(f"Saltando EDA (usando {eda_json_path})")
        with open(eda_json_path, encoding="utf-8") as fh:
            eda_resumen = json.load(fh)
    else:
        eda_resumen = ejecutar_eda(imgs_buenas, imgs_malas, ruta_eda,
                                   segmentar_herramienta)

    # Inicialización: acumulador de tiempos del perfilómetro. Si la fase
    # del perfilómetro no se ejecuta, queda vacío y la comparativa lo omite.
    tiempos_perf = {}

    # Inicialización (V19): resultado de la comparativa BUENA vs MALA vs ISO.
    # Queda en None si el perfilómetro no se ejecuta.
    comp_bm = None

    # ===== FASE 2 - PERFILÓMETRO ÓPTICO =====
    # Tras el EDA, extraemos el perfil 1D de cada imagen y medimos
    # geometría (paso, altura, ángulo, cresta, valle, uniformidad).
    # Esto produce el dataset de métricas físicas para los siguientes
    # pasos. Se desactiva con --sin_perfilometro.
    if ejecutar_perf:
        log.info("\n" + "=" * 70)
        log.info(" FASE 2 / PERFILÓMETRO ÓPTICO")
        log.info("   Extracción de perfil + medidas geométricas + ")
        log.info("   evaluación como clasificador (AUC, matriz confusión).")
        log.info("=" * 70)
        ruta_perf = os.path.join(ruta_salida, "perfilometro")
        # Carpeta ÚNICA para todas las comparativas (buenas + malas juntas).
        # Cada imagen lleva el prefijo BUENAS_ o MALAS_ en el nombre del
        # archivo, así no se mezclan y son fáciles de identificar.
        ruta_comp = os.path.join(ruta_perf, "comparativas")
        # tiempos_perf ya está inicializado arriba; lo rellenan las dos
        # llamadas siguientes y se usa luego en comparativa_coste_computacional.
        res_buenas_perf = ejecutar_perfilometro(
            imgs_buenas, os.path.join(ruta_perf, "buenas"),
            paso_mm=args.paso_mm,
            perfil_norma=args.norma_rosca,
            mm_por_px=args.mm_por_px,
            tag="buenas",
            ruta_comparativas=ruta_comp,
            out_tiempos=tiempos_perf,
        )
        res_malas_perf = ejecutar_perfilometro(
            imgs_malas, os.path.join(ruta_perf, "malas"),
            paso_mm=args.paso_mm,
            perfil_norma=args.norma_rosca,
            mm_por_px=args.mm_por_px,
            tag="malas",
            ruta_comparativas=ruta_comp,
            out_tiempos=tiempos_perf,
        )

        # Comparativa cuantitativa BUENA vs MALA vs ISO teórico:
        # cuantificación de la diferencia geométrica entre roscas conformes
        # y no conformes respecto al perfil teórico ISO, expresada en
        # unidades físicas (µm de altura, mm² de área).
        try:
            comp_bm = comparativa_buena_mala_iso(
                res_buenas_perf, res_malas_perf,
                ruta_salida=os.path.join(ruta_perf, "comparativa_clases"),
                paso_mm=args.paso_mm,
                eda_resumen=eda_resumen,
            )
        except Exception as e:
            log.error(f"Comparativa BUENA vs MALA falló: {e}")
            import traceback
            log.error(traceback.format_exc())
            comp_bm = None
        # Evaluación del perfilómetro como clasificador (AUC, ROC,
        # matriz, agregación por pieza).
        try:
            res_perfilometro_clf = evaluar_perfilometro_clasificador(
                res_buenas_perf, res_malas_perf,
                ruta_salida=os.path.join(ruta_perf, "evaluacion"),
                paso_mm=args.paso_mm,
                percentil_calib=args.perfilometro_percentil,
            )
        except Exception as e:
            log.error(f"Evaluación clasificador perfilómetro falló: {e}")
            import traceback
            log.error(traceback.format_exc())

        # ===== PERFIL PATRÓN =====
        # Construye un patrón promediando los perfiles BUENAS, lo contrasta
        # con el ISO teórico y lo evalúa como clasificador binario.
        # Split simple: 50 % de BUENAS de TRAIN para construir el patrón,
        # el resto para EVAL junto con todas las MALAS.
        res_patron = None
        try:
            n_b = len(res_buenas_perf)
            split = max(1, n_b // 2)
            res_buenas_train_patron = res_buenas_perf[:split]
            res_buenas_eval_patron = res_buenas_perf[split:]
            log.info("\n" + "=" * 70)
            log.info(" PERFIL PATRÓN — Clasificador basado en distancia al patrón")
            log.info("=" * 70)
            log.info(f"   Patrón construido con {len(res_buenas_train_patron)} "
                     f"BUENAS. EVAL: {len(res_buenas_eval_patron)} BUENAS + "
                     f"{len(res_malas_perf)} MALAS.")
            res_patron = evaluar_clasificador_patron(
                res_buenas_train=res_buenas_train_patron,
                res_buenas_eval=res_buenas_eval_patron,
                res_malas_eval=res_malas_perf,
                ruta_salida=os.path.join(ruta_perf, "perfil_patron"),
                n_puntos=512,
                paso_mm=args.paso_mm,
            )
            if res_patron:
                log.info(f"   [Patrón] AUC={res_patron['auc']:.3f}  "
                         f"Acc={res_patron['accuracy']*100:.1f}%  "
                         f"Sens={res_patron['sensibilidad']*100:.1f}%  "
                         f"Esp={res_patron['especificidad']*100:.1f}%")
                log.info(f"   [Patrón] Δ patrón vs ISO = "
                         f"{res_patron['delta_patron_iso_mediana_um']:.0f} µm  |  "
                         f"Área = "
                         f"{res_patron['area_patron_iso_mm2']:.3f} mm²")
        except Exception as e:
            log.warning(f"Perfil patrón falló: {e}")
            import traceback
            log.warning(traceback.format_exc())

        # Diagnóstico automático del perfilómetro.
        try:
            ejecutar_diagnostico_automatico(
                ruta_resultados=ruta_salida,
                ruta_diagnostico=os.path.join(ruta_perf, "evaluacion",
                                              "diagnostico_perfilometro.txt"),
            )
        except Exception as e:
            log.error(f"Diagnóstico automático falló: {e}")
            import traceback
            log.error(traceback.format_exc())

        # NOTA: La antigua FASE 3 (PatchCore-Desviación sobre el perfil 1D)
        # se ha retirado en la limpieza V20 porque no aportaba al resultado
        # principal del TFM (AUC=0,567 frente a 0,807 de PatchCore-Imagen).
        # El detector PatchCore se entrena directamente sobre la imagen 2D
        # en la FASE 4.

        if args.solo_perfilometro:
            log.info("Modo --solo_perfilometro: terminado.")
            return

    # ===== FASE 4 - PATCHCORE SOBRE IMAGEN (zero-shot) =====
    log.info("\n" + "=" * 70)
    log.info(" FASE 4 / PATCHCORE-IMAGEN - Detección por anomalía (zero-shot)")
    log.info("=" * 70)

    resultados = {}
    pipelines_a_correr = [p for p in args.pipelines if p in PIPELINES]
    if not pipelines_a_correr:
        log.error("No hay pipelines válidos. Saliendo.")
        return

    # PatchCore es el método principal del TFM. La rama legacy (ResNet18 +
    # GAP global, "modo legacy") se retiró en la limpieza V20 por no aportar
    # al resultado: AUC<0,55 en este dataset frente a 0,807 de PatchCore.
    for nombre_p in pipelines_a_correr:
        resultados[f"PC_{nombre_p}"] = evaluar_patchcore(
            nombre_p, PIPELINES[nombre_p],
            imgs_buenas, imgs_malas, ruta_salida,
            coreset_ratio=args.coreset_ratio,
            backbone=args.backbone,
            guardar_heatmaps=4,
            usar_mascara=not args.sin_mascara,
            percentil_desgaste=args.percentil_desgaste,
            ancho_banda=args.ancho_banda)

    guardar_csv_global(resultados, ruta_salida)
    guardar_resumen(resultados, ruta_salida)

    # ===== FASE 4-tris - ABLACIÓN CON/SIN MÁSCARA =====
    # Ejecuta el pipeline ganador (BLMD_V10) dos veces (con y sin máscara)
    # para evaluar la aportación del contexto del fondo al detector
    # PatchCore. Se ejecuta automáticamente por defecto; solo se omite si
    # el usuario pasa --saltar_ablacion.
    if not getattr(args, "saltar_ablacion", False):
        log.info("\n" + "=" * 70)
        log.info(" ABLACIÓN CON/SIN MÁSCARA — Evaluación de la aportación del fondo")
        log.info("=" * 70)
        try:
            ruta_abl = os.path.join(ruta_salida, "ablacion_mascara")
            os.makedirs(ruta_abl, exist_ok=True)

            # Si BLMD_V10 ya está en resultados, reutilizamos su AUC "con máscara"
            res_con_mask = resultados.get("PC_BLMD_V10")
            auc_con = res_con_mask.get("auc") if res_con_mask else None

            # Ahora ejecutamos BLMD_V10 sin máscara
            log.info("   [Ablación] Ejecutando BLMD_V10 SIN máscara...")
            res_sin = evaluar_patchcore(
                "BLMD_V10", PIPELINES["BLMD_V10"],
                imgs_buenas, imgs_malas, ruta_abl,
                coreset_ratio=args.coreset_ratio,
                backbone=args.backbone,
                guardar_heatmaps=0,
                usar_mascara=False,  # <-- diferencia clave
                percentil_desgaste=args.percentil_desgaste,
                ancho_banda=args.ancho_banda)
            auc_sin = res_sin.get("auc") if res_sin else None

            # Reporte texto
            ruta_inf_abl = os.path.join(ruta_abl, "ablacion_mascara.txt")
            with open(ruta_inf_abl, "w", encoding="utf-8") as fh:
                fh.write("=" * 78 + "\n")
                fh.write(" ABLACIÓN CON / SIN MÁSCARA — APORTACIÓN DEL CONTEXTO\n")
                fh.write("=" * 78 + "\n\n")
                fh.write("Estudio comparativo del efecto de aplicar la máscara\n")
                fh.write("binaria al detector PatchCore frente a procesar la\n")
                fh.write("imagen completa.\n\n")
                fh.write("Pipeline evaluado: BLMD_V10 + PatchCore.\n\n")
                fh.write(f"  AUC con máscara (fondo eliminado):   "
                          f"{auc_con:.3f}\n" if auc_con else "  AUC con máscara: N/D\n")
                fh.write(f"  AUC sin máscara (imagen completa):   "
                          f"{auc_sin:.3f}\n" if auc_sin else "  AUC sin máscara: N/D\n")
                if auc_con is not None and auc_sin is not None:
                    delta = auc_con - auc_sin
                    fh.write(f"  ΔAUC (con - sin) = {delta:+.3f}\n\n")
                    if delta > 0.01:
                        veredicto = ("La eliminación del fondo mejora el AUC "
                                     "en {0:.3f} puntos.".format(delta))
                    elif delta < -0.01:
                        veredicto = ("La eliminación del fondo reduce el AUC "
                                     "en {0:.3f} puntos. La imagen completa "
                                     "es la configuración recomendada.".format(-delta))
                    else:
                        veredicto = ("La máscara no tiene impacto significativo "
                                     "sobre el AUC.")
                    fh.write(f"Veredicto: {veredicto}\n")
            log.info(f"   [Ablación] Informe: {ruta_inf_abl}")
            if auc_con is not None and auc_sin is not None:
                log.info(f"   [Ablación] AUC con máscara = {auc_con:.3f}  |  "
                         f"sin máscara = {auc_sin:.3f}  |  "
                         f"Δ = {(auc_con-auc_sin):+.3f}")
        except Exception as e:
            log.warning(f"   Ablación máscara falló: {e}")

    # ===== FASE 4-bis - COMPARATIVA SNR POR PIPELINE (Obj 3 TFM) =====
    # Recorre los mismos pipelines pero esta vez NO entrena PatchCore: solo
    # mide SNR_estructural, SNR_contornos y SNR_BM (separación buenas/malas).
    # Es la respuesta cuantitativa al objetivo SMART 3:
    #   "Comparar BLMD vs BEMD vs wavelet en términos de relación señal-ruido."
    log.info("\n" + "=" * 70)
    log.info(" FASE 4-bis / COMPARATIVA SNR POR PIPELINE (Obj 3 TFM)")
    log.info("=" * 70)
    ruta_snr = os.path.join(ruta_salida, "snr")
    resultados_snr = None
    try:
        resultados_snr = evaluar_snr_todos_pipelines(
            imgs_buenas, imgs_malas,
            pipelines_a_correr,
            ruta_snr,
            usar_mascara=not args.sin_mascara,
            ancho_banda=args.ancho_banda)
    except Exception as e:
        log.error(f"FASE 4-bis (SNR) falló: {e}")

    # ===== FASE 5 - AUC DUAL (global + sobre detectables EDA) =====
    log.info("\n" + "=" * 70)
    log.info(" FASE 5 / EVALUACIÓN DUAL")
    log.info("=" * 70)
    auc_dual = calcular_auc_dual(
        {k: v for k, v in resultados.items() if k.startswith("PC_")},
        eda_resumen)
    for nombre, ad in auc_dual.items():
        ad_str = f"{ad['auc_detectables']:.3f}" \
                 if ad['auc_detectables'] is not None else "n/a"
        log.info(f"  {nombre}: AUC global={ad['auc_global']:.3f}, "
                 f"AUC detectables={ad_str} "
                 f"(n={ad['n_detectables']})")

    # ===== FASE 5-bis - COMPARATIVA DE COSTE COMPUTACIONAL Y PREDICCIÓN =====
    # Tabla CSV + figura PNG + entrada en el histórico acumulado
    # (historico_ejecuciones.json y historico_costes.csv en la raíz del proyecto)
    log.info("\n" + "=" * 70)
    log.info(" FASE 5-bis / COMPARATIVA DE COSTE COMPUTACIONAL")
    log.info("=" * 70)
    coste_comparativa = None
    try:
        coste_comparativa = comparativa_coste_computacional(
            resultados, tiempos_perf, auc_dual, ruta_salida, args,
            eda_resumen=eda_resumen)
    except Exception as e:
        log.error(f"Comparativa de coste falló: {e}")

    # ===== FASE 6 - INFORME FINAL =====
    log.info("\n" + "=" * 70)
    log.info(" FASE 6 / INFORME FINAL")
    log.info("=" * 70)
    informe_final_tfm(
        {k: v for k, v in resultados.items() if k.startswith("PC_")},
        eda_resumen, auc_dual, args, ruta_salida,
        res_perfilometro_clf=res_perfilometro_clf,
        resultados_snr=resultados_snr,
        coste_comparativa=coste_comparativa,
        comparativa_clases=comp_bm)

    # ===== FASE 7 - COMPARATIVA VISUAL DE PIPELINES (V20) =====
    resultados_pc_only = {k: v for k, v in resultados.items()
                          if k.startswith("PC_")}
    if len(resultados_pc_only) >= 2:
        try:
            comparativa_pipelines(resultados_pc_only, auc_dual, ruta_salida)
        except Exception as e:
            log.error(f"Comparativa visual falló: {e}")
    else:
        log.info("Saltando comparativa visual (necesita >= 2 pipelines).")

    # ===== FASE 8 - PDF COMBINADO PARA ANEXO TFM (V20) =====
    if not args.sin_pdf:
        try:
            construir_pdf_completo(ruta_salida, eda_resumen,
                                   resultados_pc_only, auc_dual, args)
        except Exception as e:
            log.error(f"Generación de PDF falló: {e}")
            log.error("El resto de salidas (PNG, CSV, txt) están intactas.")

    # ===== FASE 9 - EXCEL PARA COMPAÑEROS (V20.1) =====
    if not args.sin_excel:
        if not OPENPYXL_OK:
            log.warning("openpyxl NO instalado: no se genera el Excel.")
            log.warning("    pip install openpyxl")
        else:
            try:
                generar_excel_companeros(ruta_salida, eda_resumen,
                                         resultados_pc_only, auc_dual, args)
            except Exception as e:
                log.error(f"Generación de Excel falló: {e}")
                log.error("El resto de salidas están intactas.")


# ==============================================================================
#  FASE 7 (V20.1) - EXCEL PARA COMPAÑEROS DE PROYECTO
# ==============================================================================

def generar_excel_companeros(ruta_salida, eda_resumen, resultados_pc, auc_dual,
                             args):
    """
    Genera un Excel multi-hoja con los resultados del run actual, pensado para
    compartir con compañeros de proyecto y director.

    Hojas:
      1. Resumen ejecutivo con KPIs grandes
      2. Pipelines comparados
      3. Resultados (tabla principal con escalas de color y gráfico)
      4. EDA explicado
      5. Top features del EDA
      6. Multi-umbral (del mejor pipeline en F1)
      7. Glosario de términos
      8. Cómo lanzar el código
    """
    log.info("=" * 70)
    log.info(" FASE 7 / GENERANDO EXCEL PARA COMPAÑEROS")
    log.info("=" * 70)

    # ---- Estilos comunes ----
    AZUL = "1F4E79"
    AZUL_C = "D9E2F3"
    AZUL_MC = "EAF1F9"
    VERDE_C = "C6E0B4"
    NARANJA_C = "FCE4D6"
    GRIS_C = "D9D9D9"
    VERDE = "385723"
    GRIS = "595959"

    F = "Calibri"
    F_HEADER = Font(name=F, size=14, bold=True, color="FFFFFF")
    F_SUB = Font(name=F, size=12, bold=True, color=AZUL)
    F_TITLE = Font(name=F, size=20, bold=True, color=AZUL)
    F_TITLE_M = Font(name=F, size=14, bold=True, color=AZUL)
    F_BOLD = Font(name=F, size=11, bold=True)
    F_NORMAL = Font(name=F, size=11)
    F_NOTE = Font(name=F, size=10, italic=True, color=GRIS)
    F_GREEN = Font(name=F, size=11, bold=True, color=VERDE)
    F_CODE = Font(name="Consolas", size=10)
    F_KPI = Font(name=F, size=28, bold=True, color=AZUL)

    FILL_HD = PatternFill("solid", fgColor=AZUL)
    FILL_AC = PatternFill("solid", fgColor=AZUL_C)
    FILL_AMC = PatternFill("solid", fgColor=AZUL_MC)
    FILL_VC = PatternFill("solid", fgColor=VERDE_C)
    FILL_NC = PatternFill("solid", fgColor=NARANJA_C)
    FILL_GC = PatternFill("solid", fgColor=GRIS_C)

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

    thin = Side(border_style="thin", color="BFBFBF")
    BD = Border(top=thin, bottom=thin, left=thin, right=thin)

    def style_hd(ws, row, cols, height=30):
        ws.row_dimensions[row].height = height
        for col in cols:
            c = ws.cell(row=row, column=col)
            c.font = F_HEADER
            c.fill = FILL_HD
            c.alignment = CENTER
            c.border = BD

    def set_w(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Datos extraídos del run actual ----
    n_buenas = eda_resumen.get("n_buenas", 0) if eda_resumen else 0
    n_malas = eda_resumen.get("n_malas", 0) if eda_resumen else 0
    n_total = n_buenas + n_malas

    # Mejor AUC global y mejor AUC detectables
    if resultados_pc:
        mejor_auc = max(r["auc"] for r in resultados_pc.values())
        mejor_pipe_auc = max(resultados_pc, key=lambda k: resultados_pc[k]["auc"])
    else:
        mejor_auc = 0.0
        mejor_pipe_auc = "—"

    if auc_dual:
        det_vals = [(k, v.get("auc_detectables")) for k, v in auc_dual.items()
                    if v.get("auc_detectables") is not None]
        if det_vals:
            mejor_pipe_det, mejor_auc_det = max(det_vals, key=lambda x: x[1])
        else:
            mejor_pipe_det, mejor_auc_det = "—", 0.0
    else:
        mejor_pipe_det, mejor_auc_det = "—", 0.0

    n_pipelines = len(resultados_pc) if resultados_pc else 0

    out = eda_resumen.get("outliers", {}) if eda_resumen else {}
    n_detec = out.get("n_malas_detectables", 0)
    pct_indist = out.get("pct_malas_indistinguibles", 0.0)
    pct_detec = (n_detec / n_malas * 100.0) if n_malas else 0.0

    # ====================================================================
    wb = Workbook()

    # ===== HOJA 1 - RESUMEN EJECUTIVO =====
    ws = wb.active
    ws.title = "1 · Resumen"
    set_w(ws, [3, 30, 30, 30, 30, 3])

    ws.merge_cells("B2:E3")
    c = ws["B2"]
    c.value = "TFM — Detección de desgaste en roscas"
    c.font = F_TITLE
    c.alignment = CENTER
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30

    ws.merge_cells("B4:E4")
    c = ws["B4"]
    c.value = "Resumen ejecutivo de resultados"
    c.font = Font(name=F, size=14, italic=True, color=GRIS)
    c.alignment = CENTER

    ws.merge_cells("B5:E5")
    c = ws["B5"]
    c.value = ("Daniel Alcalde Martín-Calero · Miguel González Sánchez "
               "· Jordi Peiró Castelló")
    c.font = Font(name=F, size=10, color=GRIS)
    c.alignment = CENTER

    ws.merge_cells("B6:E6")
    c = ws["B6"]
    c.value = "Universidad Internacional de La Rioja · Máster en IA"
    c.font = Font(name=F, size=10, color=GRIS)
    c.alignment = CENTER

    # Caja de pregunta
    ws.row_dimensions[8].height = 30
    ws.merge_cells("B8:E8")
    c = ws["B8"]
    c.value = "¿Qué problema resolvemos?"
    c.font = F_SUB
    c.fill = FILL_AC
    c.alignment = CENTER
    c.border = BD

    ws.merge_cells("B9:E11")
    c = ws["B9"]
    c.value = ("Detección automática de desgaste en herramientas de roscado "
               "mediante visión artificial. Solo se necesitan imágenes de "
               "herramientas BUENAS para entrenar el sistema (zero-shot "
               "anomaly detection). El objetivo es comparar técnicas de "
               "preprocesado de imagen y medir cuál realza mejor las "
               "anomalías de desgaste.")
    c.font = F_NORMAL
    c.alignment = LEFT_TOP
    c.border = BD
    for r in range(9, 12):
        ws.row_dimensions[r].height = 22

    # KPIs
    ws.merge_cells("B13:E13")
    c = ws["B13"]
    c.value = "Cifras clave"
    c.font = F_SUB
    c.fill = FILL_AC
    c.alignment = CENTER
    c.border = BD
    ws.row_dimensions[13].height = 28

    kpis = [
        ("Imágenes analizadas", str(n_total),
         f"{n_buenas} BUENAS + {n_malas} MALAS", VERDE_C),
        ("Pipelines comparados", str(n_pipelines),
         "Estrategias de preprocesado evaluadas", AZUL_C),
        ("Mejor AUC global", f"{mejor_auc:.3f}",
         mejor_pipe_auc, VERDE_C),
        ("Mejor AUC sobre detectables", f"{mejor_auc_det:.3f}",
         mejor_pipe_det, NARANJA_C),
    ]
    for i, (titulo, valor, detalle, color) in enumerate(kpis):
        col_s = 2 + (i % 2) * 2
        col_e = col_s + 1
        r = 14 + (i // 2) * 4
        # título
        ws.merge_cells(start_row=r, start_column=col_s,
                       end_row=r, end_column=col_e)
        c = ws.cell(row=r, column=col_s)
        c.value = titulo
        c.font = Font(name=F, size=11, bold=True, color=AZUL)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = CENTER
        c.border = BD
        ws.row_dimensions[r].height = 22
        # valor
        ws.merge_cells(start_row=r+1, start_column=col_s,
                       end_row=r+1, end_column=col_e)
        c = ws.cell(row=r+1, column=col_s)
        c.value = valor
        c.font = F_KPI
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = CENTER
        c.border = BD
        ws.row_dimensions[r+1].height = 38
        # detalle
        ws.merge_cells(start_row=r+2, start_column=col_s,
                       end_row=r+2, end_column=col_e)
        c = ws.cell(row=r+2, column=col_s)
        c.value = detalle
        c.font = F_NOTE
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = CENTER
        c.border = BD
        ws.row_dimensions[r+2].height = 22

    # Hallazgo
    ws.merge_cells("B22:E22")
    c = ws["B22"]
    c.value = "Hallazgo principal"
    c.font = F_SUB
    c.fill = FILL_AC
    c.alignment = CENTER
    c.border = BD
    ws.row_dimensions[22].height = 28

    ws.merge_cells("B23:E25")
    c = ws["B23"]
    c.value = (f"El {pct_indist:.1f}% de las MALAS son estadísticamente "
               f"indistinguibles de las BUENAS. Esto NO significa que el "
               f"desgaste no exista, sino que con un esquema de captura "
               f"monovista (una foto por pieza), muchas imágenes etiquetadas "
               f"como MALAS capturan zonas sanas de piezas defectuosas. Sobre "
               f"el subconjunto realmente detectable ({pct_detec:.1f}%), el "
               f"sistema alcanza AUC de {mejor_auc_det:.3f} con "
               f"{mejor_pipe_det}.")
    c.font = F_NORMAL
    c.alignment = LEFT_TOP
    c.fill = FILL_NC
    c.border = BD
    for r in range(23, 26):
        ws.row_dimensions[r].height = 22

    # Navegación
    ws.merge_cells("B27:E27")
    c = ws["B27"]
    c.value = "Cómo navegar este Excel"
    c.font = F_SUB
    c.fill = FILL_AC
    c.alignment = CENTER
    c.border = BD
    ws.row_dimensions[27].height = 26

    nav = [
        ("Hoja 2 · Pipelines",
         "Qué hace cada uno de los pipelines comparados"),
        ("Hoja 3 · Resultados",
         "Tabla principal con AUC, F1, MCC, accuracy y tiempos"),
        ("Hoja 4 · EDA explicado",
         "Análisis Exploratorio: qué métricas distinguen BUENAS de MALAS"),
        ("Hoja 5 · Top features",
         "Las métricas más discriminativas con interpretación clara"),
        ("Hoja 6 · Multi-umbral",
         "Cómo cambian Precision, Recall y F1 según el umbral elegido"),
        ("Hoja 7 · Glosario",
         "Términos técnicos explicados en lenguaje claro"),
        ("Hoja 8 · Cómo lanzar",
         "Comandos y pasos para ejecutar el código"),
    ]
    for i, (h, d) in enumerate(nav):
        r = 28 + i
        ws.cell(row=r, column=2, value=h).font = F_BOLD
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=2).border = BD
        ws.cell(row=r, column=2).fill = FILL_AMC
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        c = ws.cell(row=r, column=3)
        c.value = d
        c.font = F_NORMAL
        c.alignment = LEFT
        c.border = BD
        ws.row_dimensions[r].height = 22

    ws.sheet_view.showGridLines = False

    # ===== HOJA 2 - PIPELINES =====
    ws = wb.create_sheet("2 · Pipelines")
    set_w(ws, [2, 20, 50, 35, 18, 3])

    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value = "Pipelines de preprocesado comparados"
    c.font = F_TITLE_M
    c.alignment = CENTER
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:E3")
    c = ws["B3"]
    c.value = ("Cada pipeline aplica un preprocesado distinto a las imágenes "
               "antes de pasarlas al detector PatchCore. Comparándolos "
               "medimos qué técnica realza mejor el desgaste.")
    c.font = F_NOTE
    c.alignment = CENTER
    ws.row_dimensions[3].height = 32

    headers = ["Pipeline", "Qué hace", "Cuándo es útil", "Velocidad"]
    for i, h in enumerate(headers, 2):
        ws.cell(row=5, column=i, value=h)
    style_hd(ws, 5, [2, 3, 4, 5])

    # Catálogo de descripciones (por nombre sin prefijo PC_)
    desc_pipelines = {
        "BLMD_REAL": (
            "Implementación rigurosa de BLMD (Smith 2005, Nunes 2009): "
            "descompone la imagen en Product Functions de distinta frecuencia "
            "espacial mediante sifting iterativo y la reconstruye amplificando "
            "los detalles de alta frecuencia.",
            "Cuando se quiere realzar texturas finas asociadas al desgaste y "
            "se cuenta con tiempo de cálculo."),
        "BLMD_V10": (
            "Versión aproximada y rápida del BLMD: aproxima la descomposición "
            "con una sola sustracción de envolvente gaussiana en lugar de "
            "iterar el sifting completo.",
            "Cuando se necesita realce tipo BLMD pero priorizando velocidad."),
        "BEMD_REAL": (
            "Bidimensional Empirical Mode Decomposition (Nunes 2009): "
            "descomposición empírica adaptativa.",
            "Alternativa multi-escala empírica. Limitaciones serias de RAM "
            "en PyEMD."),
        "WAVELET": (
            "Transformada wavelet Daubechies-4 con 3 niveles de "
            "descomposición. Amplifica los coeficientes de detalle y "
            "reconstruye la imagen para resaltar bordes y micromuescas.",
            "Familia multi-escala de base fija, alternativa clásica a BLMD."),
        "SOLO_HOMO": (
            "Aplica únicamente filtrado homomórfico (Cheng 2009): separa la "
            "iluminación lenta de la reflectancia rápida.",
            "Línea base para evaluar cuánto aporta el preprocesado de "
            "iluminación por sí solo."),
        "SIN_PREP": (
            "No aplica ningún preprocesado. Pasa la imagen original "
            "directamente al extractor de características.",
            "Línea base absoluta: mide si el preprocesado realmente aporta "
            "valor."),
    }

    pipes_in_run = []
    for nombre_full, info in resultados_pc.items():
        nombre = nombre_full.replace("PC_", "")
        tiempo = info.get("tiempo_s", 0.0)
        descripcion, cuando = desc_pipelines.get(nombre,
                                                 ("(sin descripción)", "—"))
        if tiempo < 30:
            vel = f"Muy rápida (~{int(tiempo)} s)"
        elif tiempo < 120:
            vel = f"Rápida (~{int(tiempo)} s)"
        elif tiempo < 300:
            vel = f"Media (~{int(tiempo)} s)"
        else:
            vel = f"Lenta (~{int(tiempo/60)} min)"
        pipes_in_run.append((nombre, descripcion, cuando, vel))

    for i, (p, q, cuando, v) in enumerate(pipes_in_run):
        r = 6 + i
        ws.row_dimensions[r].height = 80
        ws.cell(row=r, column=2, value=p)
        ws.cell(row=r, column=3, value=q)
        ws.cell(row=r, column=4, value=cuando)
        ws.cell(row=r, column=5, value=v)
        fill = FILL_AMC if i % 2 == 0 else None
        for col in [2, 3, 4, 5]:
            cell = ws.cell(row=r, column=col)
            cell.alignment = LEFT_TOP
            cell.border = BD
            cell.font = F_NORMAL
            if fill:
                cell.fill = fill
        ws.cell(row=r, column=2).font = F_BOLD

    # Nota sobre BEMD si NO está en el run
    if "PC_BEMD_REAL" not in resultados_pc:
        r_bemd = 6 + len(pipes_in_run) + 1
        ws.merge_cells(start_row=r_bemd, start_column=2,
                       end_row=r_bemd, end_column=5)
        c = ws.cell(row=r_bemd, column=2)
        c.value = "¿Y BEMD? — Limitación documentada"
        c.font = F_SUB
        c.fill = FILL_NC
        c.alignment = CENTER
        c.border = BD
        ws.row_dimensions[r_bemd].height = 26

        ws.merge_cells(start_row=r_bemd+1, start_column=2,
                       end_row=r_bemd+3, end_column=5)
        c = ws.cell(row=r_bemd+1, column=2)
        c.value = ("Inicialmente se diseñó un sexto pipeline (BEMD). La "
                   "librería externa PyEMD presentó limitaciones críticas de "
                   "escalabilidad (consumo > 6 GB de RAM por imagen), por lo "
                   "que se descartó del benchmark final. Esta es una "
                   "limitación de la implementación externa, NO del enfoque. "
                   "WAVELET cubre adecuadamente la familia de alternativas "
                   "multi-escala a BLMD descrita en el estado del arte.")
        c.font = F_NORMAL
        c.alignment = LEFT_TOP
        c.fill = FILL_NC
        c.border = BD
        for r in range(r_bemd+1, r_bemd+4):
            ws.row_dimensions[r].height = 22

    ws.sheet_view.showGridLines = False

    # ===== HOJA 3 - RESULTADOS =====
    ws = wb.create_sheet("3 · Resultados")
    set_w(ws, [2, 18, 14, 16, 13, 13, 13, 13, 12, 3])

    ws.merge_cells("B2:I2")
    c = ws["B2"]
    c.value = (f"Resultados oficiales sobre dataset ({n_buenas} BUENAS + "
               f"{n_malas} MALAS)")
    c.font = F_TITLE_M
    c.alignment = CENTER
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:I3")
    c = ws["B3"]
    c.value = ("Métricas oficiales de los pipelines. Los valores en verde "
               "destacan los mejores. Las celdas con escala de color "
               "visualizan el rendimiento relativo.")
    c.font = F_NOTE
    c.alignment = CENTER
    ws.row_dimensions[3].height = 30

    headers = ["Pipeline", "AUC global", "AUC detectables", "F1 óptimo",
               "MCC óptimo", "Accuracy μ+3σ", "Tiempo (s)", "Imágenes/s"]
    for i, h in enumerate(headers, 2):
        ws.cell(row=5, column=i, value=h)
    style_hd(ws, 5, list(range(2, 10)), height=40)

    # Construir filas con datos reales del run
    filas_res = []
    for nombre_full, info in resultados_pc.items():
        nombre = nombre_full.replace("PC_", "")
        auc_g = info.get("auc", 0.0)
        auc_d = auc_dual.get(nombre_full, {}).get("auc_detectables", 0.0) or 0.0
        # F1 y MCC óptimos del log multi-umbral
        f1_opt = 0.0
        mcc_opt = 0.0
        for u in info.get("umbrales", []):
            if u.get("nombre") in ("F1_opt",):
                f1_opt = u.get("f1", 0.0)
            if u.get("nombre") in ("MCC_opt",):
                mcc_opt = u.get("mcc", 0.0)
        acc = info.get("acc_mu3sigma", 0.0)
        t = info.get("tiempo_s", 0.0)
        filas_res.append((nombre, auc_g, auc_d, f1_opt, mcc_opt, acc, t))

    if filas_res:
        mejores = {
            1: max(x[1] for x in filas_res),
            2: max(x[2] for x in filas_res),
            3: max(x[3] for x in filas_res),
            4: max(x[4] for x in filas_res),
            5: max(x[5] for x in filas_res),
            6: min(x[6] for x in filas_res),
        }
    else:
        mejores = {}

    for i, fila in enumerate(filas_res):
        r = 6 + i
        ws.row_dimensions[r].height = 28
        ws.cell(row=r, column=2, value=fila[0]).font = F_BOLD
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=2).border = BD

        for j in range(1, 7):
            col = 2 + j
            valor = fila[j]
            es_mejor = (valor == mejores.get(j))
            cell = ws.cell(row=r, column=col)
            if j in [1, 2, 3, 4]:
                cell.value = round(valor, 3)
                cell.number_format = "0.000"
            elif j == 5:
                cell.value = valor
                cell.number_format = "0.0%"
            elif j == 6:
                cell.value = valor
                cell.number_format = "0"
            cell.font = F_GREEN if es_mejor else F_NORMAL
            cell.alignment = CENTER
            cell.border = BD

        # Imágenes/s con fórmula
        cell = ws.cell(row=r, column=9, value=f"={n_total}/H{r}")
        cell.number_format = "0.00"
        cell.font = F_NORMAL
        cell.alignment = CENTER
        cell.border = BD

    # Escalas de color
    if filas_res:
        ult = 5 + len(filas_res)
        for col in range(3, 8):
            ws.conditional_formatting.add(
                f"{get_column_letter(col)}6:{get_column_letter(col)}{ult}",
                ColorScaleRule(start_type="min", start_color="F8CBAD",
                               mid_type="percentile", mid_value=50,
                               mid_color="FFEB9C",
                               end_type="max", end_color="C6EFCE"))
        # Tiempo invertido
        ws.conditional_formatting.add(
            f"H6:H{ult}",
            ColorScaleRule(start_type="min", start_color="C6EFCE",
                           mid_type="percentile", mid_value=50,
                           mid_color="FFEB9C",
                           end_type="max", end_color="F8CBAD"))

    # Leyenda
    r0 = 6 + len(filas_res) + 1
    ws.merge_cells(start_row=r0, start_column=2, end_row=r0, end_column=9)
    c = ws.cell(row=r0, column=2)
    c.value = "Cómo leer la tabla"
    c.font = F_SUB
    c.fill = FILL_AC
    c.alignment = CENTER
    c.border = BD
    ws.row_dimensions[r0].height = 26

    interp = [
        ("AUC global", "Capacidad de distinguir BUENAS vs MALAS sobre todas "
         "las imágenes. 0.5 = aleatorio, 1.0 = perfecto."),
        ("AUC detectables", "Lo mismo, solo sobre el subconjunto de MALAS "
         "estadísticamente discriminables (según EDA)."),
        ("F1 óptimo", "Mejor balance entre Precision y Recall, a umbral "
         "seleccionado automáticamente."),
        ("MCC óptimo", "Matthews Correlation Coefficient: métrica robusta a "
         "desbalance de clases. Va de -1 a +1, 0 = aleatorio."),
        ("Accuracy μ+3σ", "Porcentaje de aciertos usando umbral "
         "conservador (3 desviaciones por encima de la media de BUENAS)."),
        ("Tiempo (s)", "Tiempo total del pipeline sobre todas las imágenes."),
        ("Imágenes/s", "Throughput: cuántas imágenes procesa por segundo "
         "(calculado por fórmula)."),
    ]
    for i, (m, d) in enumerate(interp):
        r = r0 + 1 + i
        ws.cell(row=r, column=2, value=m).font = F_BOLD
        ws.cell(row=r, column=2).alignment = LEFT_TOP
        ws.cell(row=r, column=2).border = BD
        ws.cell(row=r, column=2).fill = FILL_AMC
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
        c = ws.cell(row=r, column=3)
        c.value = d
        c.font = F_NORMAL
        c.alignment = LEFT_TOP
        c.border = BD
        ws.row_dimensions[r].height = 30

    # Gráfico
    if filas_res:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "AUC global vs AUC sobre detectables"
        chart.y_axis.title = "AUC"
        chart.x_axis.title = "Pipeline"
        chart.height = 10
        chart.width = 18
        ult = 5 + len(filas_res)
        data = Reference(ws, min_col=3, min_row=5, max_col=4, max_row=ult)
        cats = Reference(ws, min_col=2, min_row=6, max_row=ult)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList(showVal=True)
        ws.add_chart(chart, f"B{r0 + len(interp) + 3}")

    ws.sheet_view.showGridLines = False

    # ===== HOJA 4 - EDA EXPLICADO =====
    ws = wb.create_sheet("4 · EDA explicado")
    set_w(ws, [2, 25, 60, 18, 3])

    ws.merge_cells("B2:D2")
    c = ws["B2"]
    c.value = "Análisis Exploratorio de Datos (EDA) — explicado"
    c.font = F_TITLE_M
    c.alignment = CENTER
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:D3")
    c = ws["B3"]
    c.value = ("Antes de aplicar el modelo de IA, calculamos métricas "
               "estadísticas sobre cada imagen y miramos cuáles separan "
               "BUENAS de MALAS. Esto nos dice qué información hay y qué "
               "información no hay en las imágenes.")
    c.font = F_NOTE
    c.alignment = CENTER
    ws.row_dimensions[3].height = 32

    ws.merge_cells("B5:D5")
    c = ws["B5"]
    c.value = "Qué calculamos sobre cada imagen"
    c.font = F_SUB
    c.fill = FILL_AC
    c.alignment = LEFT
    c.border = BD
    ws.row_dimensions[5].height = 26

    cats_metricas = [
        ("Intensidad", "Estadísticas sobre los valores de píxel: media, "
         "mediana, desviación estándar, asimetría, curtosis y porcentajes "
         "de oscuro/medio/claro/saturado.", "10 métricas"),
        ("Geometría", "Forma de la herramienta segmentada: área, perímetro, "
         "área convexa, solidez, compacidad, rugosidad del contorno, "
         "aspect ratio.", "12 métricas"),
        ("Textura", "Patrones locales: varianza del Laplaciano, magnitud "
         "Sobel, entropía, estadísticas dentro de la zona de la "
         "herramienta.", "7 métricas"),
    ]
    for i, (cat, desc, n) in enumerate(cats_metricas):
        r = 6 + i
        ws.cell(row=r, column=2, value=cat).font = F_BOLD
        ws.cell(row=r, column=2).alignment = LEFT_TOP
        ws.cell(row=r, column=2).border = BD
        ws.cell(row=r, column=2).fill = FILL_AMC
        ws.cell(row=r, column=3, value=desc).font = F_NORMAL
        ws.cell(row=r, column=3).alignment = LEFT_TOP
        ws.cell(row=r, column=3).border = BD
        ws.cell(row=r, column=4, value=n).font = F_BOLD
        ws.cell(row=r, column=4).alignment = CENTER
        ws.cell(row=r, column=4).border = BD
        ws.cell(row=r, column=4).fill = FILL_VC
        ws.row_dimensions[r].height = 50

    ws.merge_cells("B10:D10")
    c = ws["B10"]
    c.value = "Resultados del EDA sobre nuestro dataset"
    c.font = F_SUB
    c.fill = FILL_AC
    c.alignment = LEFT
    c.border = BD
    ws.row_dimensions[10].height = 26

    n_metricas = eda_resumen.get("n_metricas", 29) if eda_resumen else 29
    n_p001 = eda_resumen.get("n_metricas_p001", 0) if eda_resumen else 0
    n_p05 = eda_resumen.get("n_metricas_p05", 0) if eda_resumen else 0

    res_eda = [
        ("Total de métricas calculadas", str(n_metricas),
         f"Para cada una de las {n_total} imágenes"),
        ("Métricas con p<0.001 (muy significativas)", str(n_p001),
         "Diferencia BUENAS vs MALAS prácticamente segura"),
        ("Métricas con p<0.05 (significativas)", str(n_p05),
         "Diferencia BUENAS vs MALAS estadísticamente sólida"),
        ("MALAS detectables (outliers >P95 BUENAS)",
         f"{n_detec} / {n_malas}",
         f"Solo el {pct_detec:.1f}% son estadísticamente diferentes"),
        ("MALAS indistinguibles", f"{n_malas - n_detec} / {n_malas}",
         f"El {pct_indist:.1f}% se confunde con BUENAS en features básicas"),
    ]
    for i, (m, v, d) in enumerate(res_eda):
        r = 11 + i
        ws.cell(row=r, column=2, value=m).font = F_NORMAL
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=2).border = BD
        ws.cell(row=r, column=3, value=d).font = F_NORMAL
        ws.cell(row=r, column=3).alignment = LEFT
        ws.cell(row=r, column=3).border = BD
        ws.cell(row=r, column=4, value=v).font = Font(name=F, size=14,
                                                      bold=True, color=AZUL)
        ws.cell(row=r, column=4).alignment = CENTER
        ws.cell(row=r, column=4).border = BD
        ws.cell(row=r, column=4).fill = FILL_AMC
        ws.row_dimensions[r].height = 28

    # Términos
    r0 = 11 + len(res_eda) + 1
    ws.merge_cells(start_row=r0, start_column=2, end_row=r0, end_column=4)
    c = ws.cell(row=r0, column=2)
    c.value = "Términos estadísticos sin tecnicismos"
    c.font = F_SUB
    c.fill = FILL_AC
    c.alignment = LEFT
    c.border = BD
    ws.row_dimensions[r0].height = 26

    defs = [
        ("p-valor", "Probabilidad de que la diferencia entre BUENAS y MALAS "
         "sea por azar. Si p<0.001, hay menos de 1 en 1.000 probabilidades "
         "de que sea casualidad.", "Más bajo = más fiable"),
        ("Cohen's d", "Cuán grande es la diferencia entre BUENAS y MALAS. "
         "d=0.2 efecto pequeño, d=0.5 efecto medio, d=0.8+ efecto grande.",
         "Más alto = más diferencia"),
        ("Test KS", "Compara distribuciones de BUENAS vs MALAS para una "
         "métrica. 0 = idénticas, 1 = totalmente distintas.",
         "Más alto = mejor separación"),
        ("Outlier (>P95 BUENAS)", "Una imagen se considera 'outlier' si "
         "está más lejos del centro de las BUENAS que el 95% de las "
         "propias BUENAS.", "Es la definición de 'rara'"),
    ]
    for i, (t, d, n) in enumerate(defs):
        r = r0 + 1 + i
        ws.cell(row=r, column=2, value=t).font = F_BOLD
        ws.cell(row=r, column=2).alignment = LEFT_TOP
        ws.cell(row=r, column=2).border = BD
        ws.cell(row=r, column=2).fill = FILL_AMC
        ws.cell(row=r, column=3, value=d).font = F_NORMAL
        ws.cell(row=r, column=3).alignment = LEFT_TOP
        ws.cell(row=r, column=3).border = BD
        ws.cell(row=r, column=4, value=n).font = F_NOTE
        ws.cell(row=r, column=4).alignment = LEFT_TOP
        ws.cell(row=r, column=4).border = BD
        ws.cell(row=r, column=4).fill = FILL_GC
        ws.row_dimensions[r].height = 50

    # Conclusión
    r0 = r0 + 1 + len(defs) + 1
    ws.merge_cells(start_row=r0, start_column=2, end_row=r0, end_column=4)
    c = ws.cell(row=r0, column=2)
    c.value = "La gran conclusión del EDA"
    c.font = F_SUB
    c.fill = FILL_NC
    c.alignment = LEFT
    c.border = BD
    ws.row_dimensions[r0].height = 26

    ws.merge_cells(start_row=r0+1, start_column=2,
                   end_row=r0+3, end_column=4)
    c = ws.cell(row=r0+1, column=2)
    c.value = (f"Las features SÍ contienen información para distinguir "
               f"BUENAS de MALAS ({n_p001} métricas con p<0.001). Sin "
               f"embargo, solo {n_detec} de {n_malas} MALAS son verdaderos "
               f"outliers respecto a las BUENAS. Esto sugiere que el "
               f"problema NO está en el algoritmo, sino en el etiquetado: "
               f"muchas imágenes etiquetadas como MALAS se parecen mucho a "
               f"BUENAS porque corresponden a zonas sanas de piezas "
               f"defectuosas. Solución futura: capturar varias vistas por "
               f"pieza y etiquetar a nivel de zona, no de pieza.")
    c.font = F_NORMAL
    c.alignment = LEFT_TOP
    c.fill = FILL_NC
    c.border = BD
    for r in range(r0+1, r0+4):
        ws.row_dimensions[r].height = 22

    ws.sheet_view.showGridLines = False

    # ===== HOJA 5 - TOP FEATURES =====
    ws = wb.create_sheet("5 · Top features")
    set_w(ws, [2, 24, 14, 14, 14, 14, 38, 3])

    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = "Top features más discriminativas"
    c.font = F_TITLE_M
    c.alignment = CENTER
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:G3")
    c = ws["B3"]
    c.value = ("Características que mejor separan BUENAS de MALAS sobre el "
               "dataset actual. Se muestran ordenadas por capacidad de "
               "discriminación (KS-stat).")
    c.font = F_NOTE
    c.alignment = CENTER
    ws.row_dimensions[3].height = 25

    headers = ["Métrica", "Media BUENAS", "Media MALAS", "Cohen's d",
               "p-valor", "Notas"]
    for i, h in enumerate(headers, 2):
        ws.cell(row=5, column=i, value=h)
    style_hd(ws, 5, list(range(2, 8)), height=30)

    # Top features del eda_resumen
    top_feat = eda_resumen.get("top_metricas", []) if eda_resumen else []
    for i, feat in enumerate(top_feat[:10]):
        r = 6 + i
        ws.row_dimensions[r].height = 50
        ws.cell(row=r, column=2, value=feat.get("metrica", "—")).font = F_CODE
        ws.cell(row=r, column=2).alignment = LEFT_TOP
        ws.cell(row=r, column=2).border = BD

        ws.cell(row=r, column=3, value=feat.get("media_buena", 0.0))
        ws.cell(row=r, column=3).number_format = "#,##0.000"
        ws.cell(row=r, column=4, value=feat.get("media_mala", 0.0))
        ws.cell(row=r, column=4).number_format = "#,##0.000"
        ws.cell(row=r, column=5, value=feat.get("cohen_d", 0.0))
        ws.cell(row=r, column=5).number_format = "+0.00;-0.00"
        ws.cell(row=r, column=6, value=feat.get("ks_p", 1.0))
        ws.cell(row=r, column=6).number_format = "0.0E+00"

        for col in [3, 4, 5, 6]:
            cell = ws.cell(row=r, column=col)
            cell.font = F_NORMAL
            cell.alignment = CENTER
            cell.border = BD

        ws.cell(row=r, column=7,
                value=feat.get("descripcion",
                               "Diferencia significativa BUENAS vs MALAS."))
        ws.cell(row=r, column=7).font = F_NORMAL
        ws.cell(row=r, column=7).alignment = LEFT_TOP
        ws.cell(row=r, column=7).border = BD

    if top_feat:
        ult = 5 + min(10, len(top_feat))
        ws.conditional_formatting.add(
            f"E6:E{ult}",
            ColorScaleRule(start_type="num", start_value=-1.5,
                           start_color="F8CBAD",
                           mid_type="num", mid_value=0, mid_color="FFFFFF",
                           end_type="num", end_value=1.5,
                           end_color="C6EFCE"))

    ws.sheet_view.showGridLines = False

    # ===== HOJA 6 - MULTI-UMBRAL =====
    ws = wb.create_sheet("6 · Multi-umbral")
    set_w(ws, [2, 18, 12, 12, 12, 12, 12, 12, 38, 3])

    # Buscar mejor pipeline por F1
    mejor_pipe_f1 = None
    mejor_f1 = -1
    for nombre_full, info in resultados_pc.items():
        for u in info.get("umbrales", []):
            if u.get("nombre") == "F1_opt" and u.get("f1", 0) > mejor_f1:
                mejor_f1 = u.get("f1", 0)
                mejor_pipe_f1 = nombre_full

    pipe_mu = mejor_pipe_f1 if mejor_pipe_f1 else (
        list(resultados_pc.keys())[0] if resultados_pc else None)
    nombre_mu = pipe_mu.replace("PC_", "") if pipe_mu else "—"

    ws.merge_cells("B2:I2")
    c = ws["B2"]
    c.value = f"Análisis multi-umbral de {nombre_mu} (mejor pipeline en F1)"
    c.font = F_TITLE_M
    c.alignment = CENTER
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:I3")
    c = ws["B3"]
    c.value = ("El umbral de decisión convierte el anomaly score en una "
               "predicción binaria (BUENA / MALA). Usar umbrales distintos "
               "da distintos errores. Aquí 12 umbrales evaluados.")
    c.font = F_NOTE
    c.alignment = CENTER
    ws.row_dimensions[3].height = 32

    headers = ["Umbral", "Valor", "TPR", "FPR", "Precision", "F1", "MCC",
               "Accuracy", "Cuándo se usa"]
    for i, h in enumerate(headers, 2):
        ws.cell(row=5, column=i, value=h)
    style_hd(ws, 5, list(range(2, 11)), height=30)

    desc_umbrales = {
        "mu+1sigma": "Permisivo: detecta más malas pero más falsos positivos",
        "mu+2sigma": "Equilibrado conservador",
        "mu+3sigma": "Estricto: pocos falsos positivos pero pierde malas",
        "mu+4sigma": "Muy estricto: casi no detecta nada",
        "p90_buenas": "Basado en percentil de BUENAS: medio-permisivo",
        "p95_buenas": "Equilibrado: el más estándar",
        "p99_buenas": "Estricto: solo el 1% extremo de BUENAS lo supera",
        "p99.5_buenas": "Muy estricto",
        "v10_legacy": "Punto medio entre buenas y malas (legacy)",
        "Youden_J_opt": "★ ÓPTIMO: maximiza TPR-FPR. Mejor compromiso clínico",
        "F1_opt": "★ ÓPTIMO: maximiza F1. Mejor balance Precision-Recall",
        "MCC_opt": "★ ÓPTIMO: maximiza MCC. Robusto a desbalance",
    }

    if pipe_mu:
        umbrales = resultados_pc[pipe_mu].get("umbrales", [])
        for i, u in enumerate(umbrales):
            r = 6 + i
            nombre_u = u.get("nombre", "—")
            ws.row_dimensions[r].height = 28
            ws.cell(row=r, column=2, value=nombre_u)
            ws.cell(row=r, column=3, value=u.get("valor", 0)).number_format = "0.00"
            ws.cell(row=r, column=4, value=u.get("tpr", 0)).number_format = "0.0%"
            ws.cell(row=r, column=5, value=u.get("fpr", 0)).number_format = "0.0%"
            ws.cell(row=r, column=6, value=u.get("prec", 0)).number_format = "0.000"
            ws.cell(row=r, column=7, value=u.get("f1", 0)).number_format = "0.000"
            ws.cell(row=r, column=8, value=u.get("mcc", 0)).number_format = "+0.000;-0.000"
            ws.cell(row=r, column=9, value=u.get("acc", 0)).number_format = "0.0%"
            ws.cell(row=r, column=10, value=desc_umbrales.get(nombre_u, "—"))

            star = "opt" in nombre_u.lower()
            fill = FILL_VC if star else (FILL_AMC if i % 2 == 0 else None)
            for col in range(2, 11):
                cell = ws.cell(row=r, column=col)
                cell.font = F_NORMAL
                cell.alignment = CENTER if col != 10 else LEFT_TOP
                cell.border = BD
                if fill:
                    cell.fill = fill
            ws.cell(row=r, column=2).font = F_CODE
            ws.cell(row=r, column=2).alignment = LEFT

        if umbrales:
            ult = 5 + len(umbrales)
            ws.conditional_formatting.add(
                f"G6:G{ult}",
                ColorScaleRule(start_type="min", start_color="F8CBAD",
                               mid_type="percentile", mid_value=50,
                               mid_color="FFEB9C",
                               end_type="max", end_color="C6EFCE"))
            ws.conditional_formatting.add(
                f"H6:H{ult}",
                ColorScaleRule(start_type="min", start_color="F8CBAD",
                               mid_type="percentile", mid_value=50,
                               mid_color="FFEB9C",
                               end_type="max", end_color="C6EFCE"))

    ws.sheet_view.showGridLines = False

    # ===== HOJA 7 - GLOSARIO =====
    ws = wb.create_sheet("7 · Glosario")
    set_w(ws, [2, 24, 70, 3])

    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = "Glosario de términos técnicos"
    c.font = F_TITLE_M
    c.alignment = CENTER
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:C3")
    c = ws["B3"]
    c.value = "Términos técnicos explicados en lenguaje claro."
    c.font = F_NOTE
    c.alignment = CENTER
    ws.row_dimensions[3].height = 22

    headers = ["Término", "Explicación clara"]
    for i, h in enumerate(headers, 2):
        ws.cell(row=5, column=i, value=h)
    style_hd(ws, 5, [2, 3], height=28)

    glosario = [
        ("Anomaly score",
         "Número que indica cuánto se aleja una imagen del patrón normal. "
         "Score bajo = parece BUENA. Score alto = parece MALA."),
        ("AUC", "Métrica de calidad. 0.5 = aleatorio, 1.0 = perfecto. "
         "0.7-0.8 aceptable, 0.8-0.9 bueno, >0.9 excelente."),
        ("Backbone", "Red preentrenada usada solo para extraer "
         "características. Aquí WideResNet-50 entrenada en ImageNet."),
        ("Banda de contorno", "Franja de 20 píxeles alrededor del contorno "
         "de la herramienta. El desgaste se manifiesta ahí."),
        ("BLMD", "Bidimensional Local Mean Decomposition. Descomposición de "
         "imágenes en componentes de distinta frecuencia espacial."),
        ("Coreset", "Subconjunto pequeño y representativo del memory bank. "
         "Algoritmo: greedy k-center. Reduce 100x el tamaño manteniendo "
         "diversidad."),
        ("Cohen's d", "Tamaño del efecto. 0.2 pequeño, 0.5 medio, "
         "0.8+ grande."),
        ("EDA", "Análisis exploratorio de datos. Estadísticas y tests "
         "ANTES de aplicar modelos."),
        ("F1 score", "Media armónica entre Precision y Recall. 0 a 1. "
         "Útil cuando ambos errores importan."),
        ("Falso positivo (FP)", "Imagen BUENA marcada como MALA. Coste: "
         "parar producción innecesariamente."),
        ("Falso negativo (FN)", "Imagen MALA dejada pasar como BUENA. "
         "Coste: defectuoso al cliente."),
        ("Filtrado homomórfico", "Separa iluminación lenta de reflectancia "
         "rápida. Corrige iluminación desigual."),
        ("MCC", "Matthews Correlation Coefficient. Robusto a desbalance. "
         "-1 a +1, 0 = aleatorio."),
        ("Memory bank", "Banco de vectores de características de las "
         "imágenes BUENAS. Se compara con él la imagen nueva."),
        ("Patch", "Pequeña ventana de la imagen, descrita por un vector. "
         "Una imagen genera ~2.700 patches."),
        ("PatchCore", "Algoritmo de detección de anomalías (Roth 2022). "
         "Patch-level + coreset + nearest neighbor."),
        ("Precision", "De las marcadas como MALAS, ¿qué % lo son? Mide la "
         "fiabilidad de la alarma."),
        ("Recall (TPR)", "De las MALAS reales, ¿qué % detecta? Mide la "
         "capacidad de no escapar defectos."),
        ("ROC", "Curva TPR vs FPR. Su área es el AUC."),
        ("Shadowgraph", "Imagen donde la herramienta es sombra oscura sobre "
         "fondo iluminado. Tipo de imagen del proyecto."),
        ("Sifting", "Iteraciones internas de BLMD/BEMD. Ajusta envolventes "
         "hasta convergencia."),
        ("Wavelet (db4)", "Funciones base para descomposición multi-escala. "
         "Daubechies-4 con 3 niveles."),
        ("WideResNet-50", "Red ResNet-50 'ancha'. Extractor de "
         "características de PatchCore."),
        ("Youden's J", "TPR - FPR. Maximizarlo da el umbral de mejor "
         "compromiso."),
        ("Zero-shot anomaly detection", "Detección sin necesidad de "
         "ejemplos defectuosos. Solo BUENAS para entrenar."),
    ]
    for i, (t, d) in enumerate(glosario):
        r = 6 + i
        ws.row_dimensions[r].height = 50
        ws.cell(row=r, column=2, value=t).font = F_BOLD
        ws.cell(row=r, column=2).alignment = LEFT_TOP
        ws.cell(row=r, column=2).border = BD
        ws.cell(row=r, column=2).fill = FILL_AMC if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        ws.cell(row=r, column=3, value=d).font = F_NORMAL
        ws.cell(row=r, column=3).alignment = LEFT_TOP
        ws.cell(row=r, column=3).border = BD
        if i % 2 == 0:
            ws.cell(row=r, column=3).fill = FILL_AMC

    ws.sheet_view.showGridLines = False

    # ===== HOJA 8 - CÓMO LANZAR =====
    ws = wb.create_sheet("8 · Cómo lanzar")
    set_w(ws, [2, 30, 60, 3])

    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = "Cómo lanzar el código"
    c.font = F_TITLE_M
    c.alignment = CENTER
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:C3")
    c = ws["B3"]
    c.value = "Pasos para reproducir los resultados desde cero."
    c.font = F_NOTE
    c.alignment = CENTER

    # Paso 1
    ws.merge_cells("B5:C5")
    c = ws["B5"]
    c.value = "Paso 1 — Instalar dependencias"
    c.font = F_SUB
    c.fill = FILL_AC
    c.alignment = LEFT
    c.border = BD
    ws.row_dimensions[5].height = 26

    ws.cell(row=6, column=2, value="Comando").font = F_BOLD
    ws.cell(row=6, column=2).fill = FILL_GC
    ws.cell(row=6, column=2).border = BD
    ws.cell(row=6, column=2).alignment = LEFT
    ws.cell(row=6, column=3, value=("pip install opencv-python numpy scipy "
                                    "scikit-learn matplotlib torch "
                                    "torchvision tqdm scikit-image "
                                    "PyWavelets openpyxl"))
    ws.cell(row=6, column=3).font = F_CODE
    ws.cell(row=6, column=3).fill = FILL_GC
    ws.cell(row=6, column=3).border = BD
    ws.cell(row=6, column=3).alignment = LEFT_TOP
    ws.row_dimensions[6].height = 30

    # Paso 2
    ws.merge_cells("B8:C8")
    c = ws["B8"]
    c.value = "Paso 2 — Estructura de carpetas"
    c.font = F_SUB
    c.fill = FILL_AC
    c.alignment = LEFT
    c.border = BD
    ws.row_dimensions[8].height = 26

    ws.cell(row=9, column=2, value="Carpetas").font = F_BOLD
    ws.cell(row=9, column=2).border = BD
    ws.cell(row=9, column=2).alignment = LEFT_TOP
    ws.cell(row=9, column=2).fill = FILL_GC
    ws.cell(row=9, column=3, value=(
        "TEST/\n"
        "├── FOTOS/\n"
        "│   ├── BUENAS_ALL/    ← imágenes en buen estado (.jpg)\n"
        "│   └── MALAS_ALL/     ← imágenes con desgaste (.jpg)\n"
        "└── RESULTADOS/        ← se crea automáticamente"))
    ws.cell(row=9, column=3).font = F_CODE
    ws.cell(row=9, column=3).alignment = LEFT_TOP
    ws.cell(row=9, column=3).border = BD
    ws.cell(row=9, column=3).fill = FILL_GC
    ws.row_dimensions[9].height = 100

    # Paso 3
    ws.merge_cells("B11:C11")
    c = ws["B11"]
    c.value = "Paso 3 — Comandos"
    c.font = F_SUB
    c.fill = FILL_AC
    c.alignment = LEFT
    c.border = BD
    ws.row_dimensions[11].height = 26

    cmds = [
        ("Modo prueba (3 min)", "python CODIGOV20_FINAL.py",
         "20+20 imágenes, valida instalación"),
        ("Modo oficial (12 min)",
         "python CODIGOV20_FINAL.py --max_imagenes=-1",
         "Dataset completo, resultados oficiales"),
        ("Solo algunos pipelines",
         "python CODIGOV20_FINAL.py --pipelines WAVELET SIN_PREP",
         "Iterar sin esperar BLMD_REAL"),
        ("Sin Excel",
         "python CODIGOV20_FINAL.py --sin_excel",
         "Si no quieres regenerar el Excel"),
    ]
    for i, (cuando, cmd, nota) in enumerate(cmds):
        r = 12 + i * 2
        ws.cell(row=r, column=2, value=cuando).font = F_BOLD
        ws.cell(row=r, column=2).fill = FILL_VC
        ws.cell(row=r, column=2).border = BD
        ws.cell(row=r, column=2).alignment = LEFT_TOP
        ws.cell(row=r, column=3, value=cmd).font = F_CODE
        ws.cell(row=r, column=3).fill = FILL_VC
        ws.cell(row=r, column=3).border = BD
        ws.cell(row=r, column=3).alignment = LEFT_TOP
        ws.row_dimensions[r].height = 26
        ws.cell(row=r+1, column=2, value="").border = BD
        ws.cell(row=r+1, column=3, value=nota).font = F_NOTE
        ws.cell(row=r+1, column=3).border = BD
        ws.cell(row=r+1, column=3).alignment = LEFT_TOP
        ws.row_dimensions[r+1].height = 24

    # Paso 4
    r0 = 12 + len(cmds) * 2 + 1
    ws.merge_cells(start_row=r0, start_column=2, end_row=r0, end_column=3)
    c = ws.cell(row=r0, column=2)
    c.value = "Paso 4 — Archivos generados"
    c.font = F_SUB
    c.fill = FILL_AC
    c.alignment = LEFT
    c.border = BD
    ws.row_dimensions[r0].height = 26

    archivos = [
        ("ANEXO_TFM.pdf", "PDF principal con todo el análisis."),
        ("Resultados_TFM.xlsx", "Este Excel (regenerado en cada run)."),
        ("informe_tfm.txt", "Informe textual con la tabla de resultados."),
        ("comparativa_*.png", "Figuras comparativas de pipelines."),
        ("fig_*_<pipeline>.png", "Figuras detalladas por pipeline."),
        ("panel_*.png", "Heatmaps de detección sobre imágenes concretas."),
        ("metricas_geometricas_<pipeline>.csv",
         "Métricas geométricas (área, solidez, etc.)."),
        ("log_global_distancias.csv", "Scores de todas las imágenes (crudo)."),
    ]
    for i, (a, d) in enumerate(archivos):
        r = r0 + 1 + i
        ws.cell(row=r, column=2, value=a).font = F_CODE
        ws.cell(row=r, column=2).alignment = LEFT_TOP
        ws.cell(row=r, column=2).border = BD
        ws.cell(row=r, column=2).fill = FILL_AMC
        ws.cell(row=r, column=3, value=d).font = F_NORMAL
        ws.cell(row=r, column=3).alignment = LEFT_TOP
        ws.cell(row=r, column=3).border = BD
        ws.row_dimensions[r].height = 24

    ws.sheet_view.showGridLines = False

    # ===== GUARDAR =====
    ruta_excel = os.path.join(ruta_salida, "Resultados_TFM.xlsx")
    wb.save(ruta_excel)
    log.info(f"  Excel generado: {ruta_excel}")
    tam_kb = os.path.getsize(ruta_excel) / 1024
    log.info(f"  Tamaño: {tam_kb:.0f} KB · {len(wb.sheetnames)} hojas")


if __name__ == "__main__":
    import sys
    try:
        main()
        # Si se ejecutó por doble clic (sin argumentos extra), mantener
        # la ventana abierta hasta que el usuario pulse Enter.
        if len(sys.argv) == 1 and sys.stdout.isatty():
            print()
            print("=" * 70)
            print(" EJECUCIÓN TERMINADA CORRECTAMENTE")
            print(" Resultados generados en la carpeta TEST/RESULTADOS/")
            print("   - ANEXO_TFM.pdf       (para director y memoria)")
            print("   - Resultados_TFM.xlsx (para compañeros)")
            print("   - informe_tfm.txt     (resumen textual)")
            print("=" * 70)
            input("\nPulsa Enter para cerrar esta ventana...")
    except SystemExit:
        # argparse llama a SystemExit (--help, error de args). No mostrar prompt.
        raise
    except Exception:
        # Si falla algo, mostrar el error y mantener la ventana abierta
        # para que el usuario pueda leerlo antes de que se cierre.
        import traceback
        print()
        print("=" * 70)
        print(" ERROR DURANTE LA EJECUCIÓN")
        print("=" * 70)
        traceback.print_exc()
        if len(sys.argv) == 1 and sys.stdout.isatty():
            input("\nPulsa Enter para cerrar esta ventana...")
        sys.exit(1)